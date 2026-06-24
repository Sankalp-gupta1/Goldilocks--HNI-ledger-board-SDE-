#!/usr/bin/env python3
"""
HNI Ledger Classification System — Corrected Accounting Edition
═══════════════════════════════════════════════════════════════════════
ACCOUNTING MODEL (corrected):
─────────────────────────────
Bank statement = cash book. Every entry is ALREADY settled (cash basis).

DOUBLE-ENTRY LEDGER (3 books):

  Book 1 — BALANCE SHEET (Assets / Liabilities / Equity)
    • Assets    : What you OWN  (investments, loans given, land, FDs, receivables)
    • Liabilities: What you OWE  (long-term loans, credit card outstanding)
    • Equity    : Net Worth = Assets − Liabilities + Retained Earnings

  Book 2 — INCOME & EXPENDITURE (P&L)
    • Income    : Salary, Dividend, Capital Gains, Interest, Professional
    • Expenditure: Food, Groceries, Travel, Staff, Tax, Bank Charges …
      ► These are EXPENSES PAID, not liabilities (bank debit = already settled)

  Book 3 — CAPITAL MOVEMENTS
    • Advances/Loans given or returned (asset movements)
    • Own account transfers (contra, no P&L)
    • Investment purchases / redemptions

CLASSIFICATION SOURCES:
  Step 0  → Excel ground-truth lookup (if narration matches training data exactly)
  Step 1  → Rule engine (regex patterns, high confidence)
  Step 2  → Supervised ML (TF-IDF + LinearSVC, trained on Excel data)
  Step 3  → Unsupervised KMeans fallback / anomaly detection
═══════════════════════════════════════════════════════════════════════
"""
from __future__ import annotations
import csv, io, json, os, re, sqlite3, uuid, warnings, tempfile, subprocess, shutil
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple
import joblib
import numpy as np
import pandas as pd
import hashlib
import hmac
import secrets
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
import io

print("LIVE MARKER SYSTEM 2026-04-10 16:11", flush=True)

warnings.filterwarnings("ignore")

_HERE        = os.path.dirname(os.path.abspath(__file__))
DB_PATH      = os.path.join(_HERE, "hni_ledger.db")
MODEL_PATH   = os.path.join(_HERE, "model_ledger.joblib")
CLUSTER_PATH = os.path.join(_HERE, "cluster_ledger.joblib")


# ══════════════════════════════════════════════════════════════════════════════
# LEDGER MAP  —  3-Book Accounting Structure
#
# Format: bs_key → (Book, Group, Account)
#
# Books:
#   BALANCE_SHEET   → Assets / Liabilities / Equity (snapshot of net worth)
#   INCOME_EXPENSE  → Income / Expenditure (P&L for the period)
#   CAPITAL         → Capital movements (advances, own transfers, investments)
#
# KEY RULE: Bank debit = cash ALREADY PAID = Expenditure, NOT liability
#           Only true balance-sheet items (loans, long-term assets, equity) go to BS
# ══════════════════════════════════════════════════════════════════════════════

LEDGER_MAP = {

    # ─── BALANCE SHEET : ASSETS ──────────────────────────────────────────────
    # Current Assets — things that convert to cash within 12 months
    "asset_advance_received_back":      ("BALANCE_SHEET", "Assets",      "Current Assets",      "Advance Received Back"),
    "asset_advance_returned":           ("BALANCE_SHEET", "Assets",      "Current Assets",      "Advance Returned by Party"),
    "asset_loan_repayment_received":    ("BALANCE_SHEET", "Assets",      "Current Assets",      "Loan Repayment Received"),
    "asset_loans_advances_given":       ("BALANCE_SHEET", "Assets",      "Current Assets",      "Loans & Advances Given (Receivable)"),
    "asset_own_transfer_in":            ("BALANCE_SHEET", "Assets",      "Current Assets",      "Own Account Transfer In (Contra)"),
    "asset_refund_received":            ("BALANCE_SHEET", "Assets",      "Current Assets",      "Refunds Received (Realised)"),
    "asset_broker_payout":              ("BALANCE_SHEET", "Assets",      "Current Assets",      "Broker Payout Received"),
    "asset_broker_balance":             ("BALANCE_SHEET", "Assets",      "Current Assets",      "Broker / Trading Account Balance"),
    "asset_fcy_inward":                 ("BALANCE_SHEET", "Assets",      "Current Assets",      "Foreign Currency Inward Remittance"),
    "asset_fd_maturity":                ("BALANCE_SHEET", "Assets",      "Current Assets",      "FD Maturity – Principal Returned"),
    "asset_mf_redemption":              ("BALANCE_SHEET", "Assets",      "Current Assets",      "MF Redemption – Principal Returned"),
    "asset_equity_sale_proceeds":       ("BALANCE_SHEET", "Assets",      "Current Assets",      "Equity Sale Proceeds (Principal)"),

    # Non-Current Assets — long-term holds
    "asset_land_cwip":                  ("BALANCE_SHEET", "Assets",      "Non-Current Assets",  "Capital Work-in-Progress (Land / Plot)"),
    "asset_investment_mf":              ("BALANCE_SHEET", "Assets",      "Non-Current Assets",  "Investment – Mutual Funds (SIP/Lump Sum)"),
    "asset_investment_equity":          ("BALANCE_SHEET", "Assets",      "Non-Current Assets",  "Investment – Equity Shares (Purchased)"),
    "asset_investment_fd":              ("BALANCE_SHEET", "Assets",      "Non-Current Assets",  "Investment – Fixed Deposit (Placed)"),
    "asset_investment_ppf":             ("BALANCE_SHEET", "Assets",      "Non-Current Assets",  "Investment – PPF"),
    "asset_investment_nps":             ("BALANCE_SHEET", "Assets",      "Non-Current Assets",  "Investment – NPS"),
    "asset_long_term_other":            ("BALANCE_SHEET", "Assets",      "Non-Current Assets",  "Long-Term Investments – Other"),

    # ─── BALANCE SHEET : LIABILITIES ─────────────────────────────────────────
    # Only TRUE outstanding liabilities — amounts still owed (not yet paid)
    # These keys are used when user manually classifies a PENDING outstanding balance
    # (e.g. credit card carry-forward, outstanding loan principal)
    # NOTE: on cash basis, a bank debit = ALREADY PAID → goes to Expenditure (exp_loan_emi / exp_credit_card)
    # These liability keys are for OUTSTANDING amounts still owed, not payments.
    "liability_loan_outstanding":       ("BALANCE_SHEET", "Liabilities", "Non-Current Liabilities", "Long-Term Loan Outstanding"),
    "liability_credit_card":            ("BALANCE_SHEET", "Liabilities", "Current Liabilities",     "Credit Card Outstanding Balance"),
    "liability_other":                  ("BALANCE_SHEET", "Liabilities", "Current Liabilities",     "Other Payables Outstanding"),

    # ─── INCOME & EXPENDITURE : EXPENDITURE — Financial Costs (loan/card payments) ────
    "exp_loan_emi":                     ("INCOME_EXPENSE", "Expenditure", "Financial Costs", "Loan EMI / Repayment (Paid)"),
    "exp_credit_card":                  ("INCOME_EXPENSE", "Expenditure", "Financial Costs", "Credit Card Payment (Settled)"),

    # ─── BALANCE SHEET : EQUITY ──────────────────────────────────────────────
    # Retained earnings contributed from income flows
    "equity_capital_introduced":        ("BALANCE_SHEET", "Equity",      "Owners Equity",       "Capital Introduced"),
    "equity_retained_earnings":         ("BALANCE_SHEET", "Equity",      "Retained Earnings",   "Retained Earnings (Net Income)"),

    # ─── INCOME & EXPENDITURE : INCOME ───────────────────────────────────────
    "income_salary":                    ("INCOME_EXPENSE", "Income",     "Employment Income",   "Salary Received"),
    "income_professional":              ("INCOME_EXPENSE", "Income",     "Employment Income",   "Professional / Consulting Fees"),
    "income_rental":                    ("INCOME_EXPENSE", "Income",     "Investment Income",   "Rental Income"),
    "income_dividend":                  ("INCOME_EXPENSE", "Income",     "Investment Income",   "Dividend Income"),
    "income_interest":                  ("INCOME_EXPENSE", "Income",     "Investment Income",   "Interest Income (SB / FD)"),
    "income_capital_gains":             ("INCOME_EXPENSE", "Income",     "Investment Income",   "Capital Gains (Equity / MF Trading)"),
    "income_other":                     ("INCOME_EXPENSE", "Income",     "Other Income",        "Miscellaneous Receipts"),
    "income_gift_family":               ("BALANCE_SHEET", "Equity",     "Owners Equity",       "Family Transfer / Capital Introduced"),
    "income_inward_payment":            ("INCOME_EXPENSE", "Income",     "Other Income",        "Miscellaneous Receipts"),

    # ─── INCOME & EXPENDITURE : EXPENDITURE ──────────────────────────────────
    # Food & Dining
    "exp_food":                         ("INCOME_EXPENSE", "Expenditure", "Household Expenses",  "Food & Dining"),
    "exp_grocery":                      ("INCOME_EXPENSE", "Expenditure", "Household Expenses",  "Groceries & Quick Commerce"),
    "exp_household":                    ("INCOME_EXPENSE", "Expenditure", "Household Expenses",  "Household Supplies"),
    # Shopping
    "exp_shopping_online":              ("INCOME_EXPENSE", "Expenditure", "Personal Expenses",   "Online Shopping"),
    "exp_clothing":                     ("INCOME_EXPENSE", "Expenditure", "Personal Expenses",   "Clothing & Apparel"),
    "exp_personal_care":                ("INCOME_EXPENSE", "Expenditure", "Personal Expenses",   "Personal Care & Beauty"),
    "exp_gifts":                        ("INCOME_EXPENSE", "Expenditure", "Personal Expenses",   "Gifts & Stationery"),
    # Health & Wellness
    "exp_health":                       ("INCOME_EXPENSE", "Expenditure", "Health & Wellness",   "Health & Medical Expenses"),
    # Travel & Transport
    "exp_travel":                       ("INCOME_EXPENSE", "Expenditure", "Travel & Transport",  "Travel, Transport & Cab"),
    # Entertainment
    "exp_entertainment":                ("INCOME_EXPENSE", "Expenditure", "Entertainment",       "Entertainment & Subscriptions"),
    # Education
    "exp_education":                    ("INCOME_EXPENSE", "Expenditure", "Education",           "Education & Tuition Fees"),
    # Home
    "exp_home_decor":                   ("INCOME_EXPENSE", "Expenditure", "Home & Property",     "Home Décor & Furnishings"),
    "exp_maintenance":                  ("INCOME_EXPENSE", "Expenditure", "Home & Property",     "Maintenance & Repair"),
    # Utilities
    "exp_utilities":                    ("INCOME_EXPENSE", "Expenditure", "Utilities & Bills",   "Electricity, Gas, Water, Internet"),
    # Staff & Professional Payments
    "exp_staff_wages":                  ("INCOME_EXPENSE", "Expenditure", "Employment Costs",    "Staff / Domestic Wages Paid"),
    "exp_consultant":                   ("INCOME_EXPENSE", "Expenditure", "Employment Costs",    "Consultant / Contractor Fees Paid"),
    # Financial Costs
    "exp_bank_charges":                 ("INCOME_EXPENSE", "Expenditure", "Financial Costs",     "Bank & Demat Charges"),
    "exp_broker_charges":               ("INCOME_EXPENSE", "Expenditure", "Financial Costs",     "Broker Contract Bill (STT + Fees)"),
    "exp_broker_dp":                    ("INCOME_EXPENSE", "Expenditure", "Financial Costs",     "Demat / DP Annual Charges"),
    "exp_broker_interest":              ("INCOME_EXPENSE", "Expenditure", "Financial Costs",     "Broker Delayed Payment Interest"),
    "exp_insurance":                    ("INCOME_EXPENSE", "Expenditure", "Financial Costs",     "Insurance Premium Paid"),
    # Taxation
    "exp_tax":                          ("INCOME_EXPENSE", "Expenditure", "Taxation",            "Income Tax / Advance Tax / GST Paid"),
    # Personal Transfers
    "exp_personal_transfer":            ("INCOME_EXPENSE", "Expenditure", "Personal Transfers",  "Personal Transfer – Family / Friends"),
    # Miscellaneous
    "exp_misc":                         ("INCOME_EXPENSE", "Expenditure", "Miscellaneous",       "Miscellaneous Payments"),
    "exp_cheque_misc":                  ("INCOME_EXPENSE", "Expenditure", "Miscellaneous",       "Cheque Payments (Misc)"),
    "exp_fcy_outward":                  ("INCOME_EXPENSE", "Expenditure", "Miscellaneous",       "Foreign Currency Outward Remittance"),

    # ─── SUSPENSE ─────────────────────────────────────────────────────────────
    "suspense_credit":                  ("SUSPENSE",       "Suspense",   "Unclassified",         "Credit – Requires Review"),
    "suspense_debit":                   ("SUSPENSE",       "Suspense",   "Unclassified",         "Debit – Requires Review"),

    # ─── TRADING ACCOUNT (5paisa / broker ledger) ─────────────────────────────
    "trading_funds_added":              ("BALANCE_SHEET",  "Assets",     "Current Assets",       "Trading Account – Funds Added"),
    "trading_payout":                   ("BALANCE_SHEET",  "Assets",     "Current Assets",       "Trading Account – Payout to Bank"),
    "trading_bill_equity":              ("INCOME_EXPENSE", "Expenditure","Financial Costs",      "Equity Brokerage & Transaction Costs"),
    "trading_bill_fno":                 ("INCOME_EXPENSE", "Income",     "Investment Income",    "F&O Net Realised P&L"),
    "trading_dp_charges":               ("INCOME_EXPENSE", "Expenditure","Financial Costs",      "DP & Demat Charges"),
    "trading_nb_charges":               ("INCOME_EXPENSE", "Expenditure","Financial Costs",      "Net Banking / Gateway Charges"),
}

# ── Asset attribution (who owns / owes) ───────────────────────────────────
ATTRIBUTION = {
    "asset_advance_received_back":      "Account Holder (Self) – advance returned by counterparty",
    "asset_advance_returned":           "Account Holder (Self) – advance refunded",
    "asset_loan_repayment_received":    "Account Holder (Self) – loan recovered from borrower",
    "asset_loans_advances_given":       "Counterparty (Borrower) – amount lent, receivable",
    "asset_own_transfer_in":            "Account Holder (Self) – own bank to bank transfer",
    "asset_refund_received":            "Account Holder (Self) – refund from vendor / UPI reversal",
    "asset_broker_payout":              "Account Holder (Self) – broker releasing trade proceeds",
    "asset_broker_balance":             "Account Holder (Self) – held in broker trading pool",
    "asset_fcy_inward":                 "Account Holder (Self) – foreign inward remittance",
    "asset_fd_maturity":                "Account Holder (Self) – FD principal returned",
    "asset_mf_redemption":              "Account Holder (Self) – MF units redeemed",
    "asset_equity_sale_proceeds":       "Account Holder (Self) – equity shares sold",
    "asset_land_cwip":                  "Account Holder (Self) – land/plot CWIP",
    "asset_investment_mf":              "Account Holder (Self) – MF portfolio",
    "asset_investment_equity":          "Account Holder (Self) – equity portfolio",
    "asset_investment_fd":              "Account Holder (Self) – bank FD",
    "asset_investment_ppf":             "Account Holder (Self) – PPF account at SBI",
    "asset_investment_nps":             "Account Holder (Self) – NPS CRA account",
    "asset_long_term_other":            "Account Holder (Self) – long-term investment",
    "exp_loan_emi":                     "Lending Institution – EMI paid, no longer a liability",
    "exp_credit_card":                  "Credit Card Issuer – bill settled in full",
    "liability_loan_outstanding":       "Lending Institution – Outstanding loan principal still owed",
    "liability_credit_card":            "Credit Card Issuer – outstanding balance still owed",
    "liability_other":                  "Counterparty – other outstanding payable",
    "income_salary":                    "Account Holder (Self) – from employer",
    "income_professional":              "Account Holder (Self) – from clients",
    "income_rental":                    "Account Holder (Self) – from tenant",
    "income_dividend":                  "Account Holder (Self) – from investee company",
    "income_interest":                  "Account Holder (Self) – from bank",
    "income_capital_gains":             "Account Holder (Self) – from trading / MF",
    "income_other":                     "Account Holder (Self) – miscellaneous",
    "income_gift_family":               "Family Member / Friend (Capital Introduced)",
    "income_inward_payment":            "Account Holder (Self) – unidentified inward credit treated as other income",
    "exp_food":                         "Merchant (Restaurant / Delivery)",
    "exp_grocery":                      "Merchant (Grocery / Quick Commerce)",
    "exp_household":                    "Merchant (Household Goods)",
    "exp_shopping_online":              "Merchant (E-commerce)",
    "exp_clothing":                     "Merchant (Apparel)",
    "exp_personal_care":                "Merchant (Beauty / Wellness)",
    "exp_gifts":                        "Merchant (Gifts / Stationery)",
    "exp_health":                       "Healthcare Provider / Pharmacy",
    "exp_travel":                       "Travel / Transport Service",
    "exp_entertainment":                "Entertainment / Subscription Service",
    "exp_education":                    "Educational Institution",
    "exp_home_decor":                   "Merchant (Home Goods)",
    "exp_maintenance":                  "Service Provider (Maintenance)",
    "exp_utilities":                    "Utility Provider",
    "exp_staff_wages":                  "Employee / Domestic Staff",
    "exp_consultant":                   "Consultant / Contractor",
    "exp_bank_charges":                 "Bank (Charges)",
    "exp_broker_charges":               "Stock Exchange / Broker (STT, fees, GST)",
    "exp_broker_dp":                    "CDSL / NSDL (Depository)",
    "exp_broker_interest":              "Broker (Delayed Payment Interest)",
    "exp_insurance":                    "Insurance Company",
    "exp_tax":                          "Government (Income Tax / GST Authority)",
    "exp_personal_transfer":            "Family Member / Friend",
    "exp_misc":                         "Various Counterparties",
    "exp_cheque_misc":                  "Payee (Cheque)",
    "exp_fcy_outward":                  "Foreign Beneficiary",
    "suspense_credit":                  "Review Required",
    "suspense_debit":                   "Review Required",
}

# ── Map old Excel line items → new ledger keys ─────────────────────────────
EXCEL_LINE_TO_KEY = {
    "Advance Received Back":                    "asset_advance_received_back",
    "Advance Returned":                         "asset_advance_returned",
    "Loan Repayment Received":                  "asset_loan_repayment_received",
    "Loans & Advances Given (Receivable)":      "asset_loans_advances_given",
    "Own Account Transfer In":                  "asset_own_transfer_in",
    "Refunds Receivable (Realised)":            "asset_refund_received",
    "Capital Work-in-Progress (Land)":          "asset_land_cwip",
    "Long-Term Investments":                    "asset_long_term_other",
    "Investment – Mutual Funds (Purchase)":     "asset_investment_mf",
    "Investment – Equity Shares (Purchase)":    "asset_investment_equity",
    "Investment – Fixed Deposit":               "asset_investment_fd",
    "Investment – PPF":                         "asset_investment_ppf",
    "Investment – NPS":                         "asset_investment_nps",
    "MF Redemption – Principal Return":         "asset_mf_redemption",
    "Equity Sale – Principal Proceeds":         "asset_equity_sale_proceeds",
    "FD Maturity – Principal Return":           "asset_fd_maturity",
    "Bank Charges Payable":                     "exp_bank_charges",
    "Broker Contract Bill (STT + Fees)":        "exp_broker_charges",
    "Demat DP Charges / AMC":                  "exp_broker_dp",
    "Credit Card Payable (Discharged)":         "exp_credit_card",    # FIXED: FIX-1
    "Other Payables – Cheque Payment":          "exp_cheque_misc",
    "Other Payables – Miscellaneous":           "exp_misc",
    "Personal Transfer – Family":               "exp_personal_transfer",
    "Staff Wages Payable (Discharged)":         "exp_staff_wages",
    "Consultant Fees Payable (Discharged)":     "exp_consultant",
    "Tax Payable (Discharged)":                 "exp_tax",
    "Trade Payables – Clothing & Apparel":      "exp_clothing",
    "Trade Payables – Education":               "exp_education",
    "Trade Payables – Entertainment":           "exp_entertainment",
    "Trade Payables – Food & Dining":           "exp_food",
    "Trade Payables – Gifts & Stationery":      "exp_gifts",
    "Trade Payables – Groceries":               "exp_grocery",
    "Trade Payables – Health & Medical":        "exp_health",
    "Trade Payables – Home Décor":              "exp_home_decor",
    "Trade Payables – Household Supplies":      "exp_household",
    "Trade Payables – Maintenance & Repair":    "exp_maintenance",
    "Trade Payables – Online Shopping":         "exp_shopping_online",
    "Trade Payables – Personal Care & Beauty":  "exp_personal_care",
    "Trade Payables – Travel & Transport":      "exp_travel",
    "Utility Bills Payable":                    "exp_utilities",
    "Long-Term Loan Repayment":                 "exp_loan_emi",       # FIXED: FIX-1
    "Income – Capital Gains (Trading)":         "income_capital_gains",
    "Income – Capital Gains":                   "income_capital_gains",
    "Income – Dividend":                        "income_dividend",
    "Income – Interest on Savings":             "income_interest",
    "Income – Interest on Savings / FD":        "income_interest",
    "Income – Other Investment":                "income_other",
    "Income – Employment":                      "income_professional",
    "Income – Professional / Consulting":       "income_professional",
    "Income – Miscellaneous Receipts":          "income_other",
    "Income – Salary":                          "income_salary",
    "Suspense – Credit (Review Required)":      "suspense_credit",
    "Suspense – Debit (Review Required)":       "suspense_debit",
}

# ── Category to key (for ML output) ───────────────────────────────────────
CAT_TO_KEY = {k: k for k in LEDGER_MAP}  # direct keys
# Plus legacy aliases
CAT_TO_KEY.update({
    "income_dividend":                "income_dividend",
    "banking_interest":               "income_interest",
    "income_salary":                  "income_salary",
    "professional_income":            "income_professional",
    "rental_income":                  "income_rental",
    "income_capital_gains":           "income_capital_gains",
    "capital_gains":                  "income_capital_gains",
    "income_misc":                    "income_other",
    "investment_mutual_fund_purchase":"asset_investment_mf",
    "investment_mutual_fund":         "asset_investment_mf",
    "investment_equity_purchase":     "asset_investment_equity",
    "investment_equity":              "asset_investment_equity",
    "investment_fd_creation":         "asset_investment_fd",
    "ppf_investment":                 "asset_investment_ppf",
    "nps_investment":                 "asset_investment_nps",
    "ppf_nps":                        "asset_investment_ppf",
    "capital_work_land":              "asset_land_cwip",
    "loans_advances_given":           "asset_loans_advances_given",
    "loan_repayment_received":        "asset_loan_repayment_received",
    "advance_received_back":          "asset_advance_received_back",
    "advance_returned":               "asset_advance_returned",
    "own_account_transfer_in":        "asset_own_transfer_in",
    "refunds_receivable":             "asset_refund_received",
    "broker_payout":                  "asset_broker_payout",
    "broker_transfer":                "asset_broker_balance",
    "investment_mf_redemption":       "asset_mf_redemption",
    "investment_equity_sale":         "asset_equity_sale_proceeds",
    "investment_fd_maturity":         "asset_fd_maturity",
    "fcy_inward":                     "asset_fcy_inward",
    "loan_repayment":                 "exp_loan_emi",
    "credit_card_payable":            "exp_credit_card",
    "loan_outstanding":               "liability_loan_outstanding",
    "credit_card_outstanding":        "liability_credit_card",
    "liability_other_payable":        "liability_other",
    "trade_payable_food":             "exp_food",
    "trade_payable_grocery":          "exp_grocery",
    "trade_payable_shopping":         "exp_shopping_online",
    "trade_payable_clothing":         "exp_clothing",
    "trade_payable_health":           "exp_health",
    "trade_payable_travel":           "exp_travel",
    "trade_payable_entertainment":    "exp_entertainment",
    "trade_payable_education":        "exp_education",
    "trade_payable_personal_care":    "exp_personal_care",
    "trade_payable_home":             "exp_home_decor",
    "trade_payable_household":        "exp_household",
    "trade_payable_gifts":            "exp_gifts",
    "trade_payable_maintenance":      "exp_maintenance",
    "utility_payable":                "exp_utilities",
    "staff_wages":                    "exp_staff_wages",
    "consultant_payable":             "exp_consultant",
    "tax_payable":                    "exp_tax",
    "bank_charges_payable":           "exp_bank_charges",
    "broker_contract_bill":           "exp_broker_charges",
    "broker_dp_charges":              "exp_broker_dp",
    "broker_interest_charge":         "exp_broker_interest",
    "insurance_premium":              "exp_insurance",
    "personal_transfer_family":       "exp_personal_transfer",
    "other_payables_misc":            "exp_misc",
    "other_payables_cheque":          "exp_cheque_misc",
    "fcy_outward":                    "exp_fcy_outward",
    "food":                           "exp_food",
    "grocery":                        "exp_grocery",
    "quick_commerce":                 "exp_grocery",
    "shopping":                       "exp_shopping_online",
    "clothing":                       "exp_clothing",
    "health":                         "exp_health",
    "travel":                         "exp_travel",
    "entertainment":                  "exp_entertainment",
    "education":                      "exp_education",
    "personal_care":                  "exp_personal_care",
    "home_decor":                     "exp_home_decor",
    "household_supplies":             "exp_household",
    "gifts_stationery":               "exp_gifts",
    "maintenance_repair":             "exp_maintenance",
    "utility":                        "exp_utilities",
    "tax_payment":                    "exp_tax",
    "fees":                           "exp_bank_charges",
    "sms_alert_charges":              "exp_bank_charges",
    "cheque_misc":                    "exp_cheque_misc",
    "transfer":                       "asset_own_transfer_in",
    "unclassified":                   "suspense_debit",
    "suspense_credit":                "suspense_credit",
    "suspense_debit":                 "suspense_debit",
    "ach_credit_other":               "income_other",
})


# ══════════════════════════════════════════════════════════════════════════════
# TEXT NORMALIZATION HELPERS
# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
# CANONICAL DIRECTION DERIVATION  (Part 2 — Req 1)
# Single authoritative function; call BEFORE classification.
# ══════════════════════════════════════════════════════════════════════════════

def derive_txn_type(
    raw_debit:   Any = None,
    raw_credit:  Any = None,
    raw_type:    Any = None,
    raw_amount:  Any = None,
    narration:   str = "",
    parser_txn_type: str = "",
) -> str:
    """
    Determine the canonical transaction direction ('debit' or 'credit').

    Precedence (highest → lowest):
      a. Explicit debit-column / credit-column values (non-zero beats zero).
      b. Explicit CR / DR markers in the type column or amount suffix.
      c. Signed amount (negative → debit, positive → credit).
      d. parser_txn_type already derived by the parser (trusted if not 'unknown').
      e. Narration-based heuristic (_infer_txn_type).
      f. Conservative fallback → 'debit'.

    Never returns anything other than 'debit' or 'credit'.
    """
    # ── a. Separate debit/credit columns ─────────────────────────────────────
    def _amt(v) -> float:
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return abs(float(v))
        s = re.sub(r'[₹$£€,\s]', '', str(v)).upper()
        s = re.sub(r'(DR|CR)\.?$', '', s).strip()
        s = re.sub(r'[^\d.]', '', s)
        try:
            return abs(float(s))
        except (ValueError, TypeError):
            return 0.0

    dr_val = _amt(raw_debit)
    cr_val = _amt(raw_credit)
    if dr_val > 0 and cr_val == 0:
        return "debit"
    if cr_val > 0 and dr_val == 0:
        return "credit"

    # ── a+. HDFC WDL TFR / DEP TFR prefix (embedded direction) ──────────────
    if narration:
        _n = narration.strip().upper()
        if re.match(r'^WDL\b', _n):
            return "debit"
        if re.match(r'^DEP\b', _n):
            return "credit"

    # ── b. Explicit CR/DR markers in type column or amount suffix ─────────────
    tc = str(raw_type or "").strip().upper()
    if tc in ("CR", "CREDIT"):
        return "credit"
    if tc in ("DR", "DEBIT"):
        return "debit"

    # Check amount-string suffix (e.g. "1234.56 Cr" or "1234.56Dr")
    for val in (raw_debit, raw_credit, raw_amount):
        sv = str(val or "").strip().upper()
        if re.search(r'\bCR\.?$', sv):
            return "credit"
        if re.search(r'\bDR\.?$', sv):
            return "debit"

    # ── c. Signed amount — ONLY negative signals debit; positive is ambiguous ──
    # A positive magnitude does NOT mean credit: single-amount columns store
    # absolute values for both directions.  Only a genuinely negative raw value
    # (e.g. "-1234.56") is a reliable debit signal.
    for val in (raw_amount, raw_debit):
        sv = str(val or "").strip()
        try:
            f = float(re.sub(r'[^\d.\-]', '', sv))
            if f < 0:
                return "debit"
            # Positive: fall through — not enough information to conclude credit
        except (ValueError, TypeError):
            pass

    # ── d. Parser-derived txn_type — elevated above narration heuristic ───────
    # The Excel/PDF parser already validated direction from debit/credit columns
    # or from a type column.  Trust it before falling back to narration keywords.
    if parser_txn_type in ("credit", "debit"):
        return parser_txn_type

    # ── e. Narration heuristic ────────────────────────────────────────────────
    inferred = _infer_txn_type((narration or "").upper())
    if inferred in ("credit", "debit"):
        return inferred

    # ── f. Conservative fallback ──────────────────────────────────────────────
    return "debit"


def normalize_parsed_record(raw: Dict[str, Any], source: str = "excel") -> Dict[str, Any]:
    """
    Normalise a raw parser output dict into a canonical structure before
    classification.  Accepts dicts from Excel, CSV, or PDF parsers.

    Output keys (always present):
      txn_date       – ISO YYYY-MM-DD string
      narration      – clean string
      amount         – float ≥ 0
      txn_type       – 'debit' or 'credit'
      source         – parser tag (e.g. 'excel', 'pdf')
      parser_notes   – diagnostic string (empty if clean)
      raw_debit      – raw debit cell value (for audit)
      raw_credit     – raw credit cell value (for audit)
    """
    narration = str(raw.get("description") or raw.get("narration") or "").strip()
    txn_date  = str(raw.get("txn_date") or "").strip()
    amount_in = raw.get("amount", 0)
    raw_debit  = raw.get("raw_debit")
    raw_credit = raw.get("raw_credit")
    raw_type   = raw.get("type_col") or raw.get("raw_type") or ""
    # Only use raw_amount from an explicit signed-amount column cell.
    # Do NOT pass the already-normalised positive `amount` field here —
    # a positive magnitude is ambiguous (could be debit or credit).
    raw_amount = raw.get("raw_amount")   # None unless parser explicitly set it
    parser_type = str(raw.get("txn_type") or "").strip().lower()
    if parser_type not in ("credit", "debit"):
        parser_type = ""

    # Re-derive direction using the canonical helper
    txn_type = derive_txn_type(
        raw_debit=raw_debit,
        raw_credit=raw_credit,
        raw_type=raw_type,
        raw_amount=raw_amount,
        narration=narration,
        parser_txn_type=parser_type,
    )

    # Amount magnitude
    try:
        amount = abs(float(amount_in or 0))
    except (TypeError, ValueError):
        amount = 0.0

    notes = []
    if not narration:
        notes.append("empty_narration")
    if amount == 0:
        notes.append("zero_amount")
    if not txn_date:
        notes.append("missing_date")
        txn_date = ""

    return {
        "txn_date":    txn_date,
        "narration":   narration,
        "description": narration,        # alias kept for downstream compat
        "amount":      amount,
        "txn_type":    txn_type,
        "source":      source,
        "parser_notes": "; ".join(notes),
        "raw_debit":   raw_debit,
        "raw_credit":  raw_credit,
    }


def normalize_txn_text(narration: str) -> str:
    """
    Normalize a bank narration for classification.
    - Uppercase, strip whitespace
    - Remove UPI/IMPS reference IDs (long digit sequences after channel keywords)
    - Remove masked card numbers (XXXXXX1234, XX1234)
    - Remove date stamps (DD-MM-YYYY, DD/MM/YYYY, DD.MM.YYYY)
    - Collapse multiple spaces
    - Preserve meaningful merchant words and channel words
    """
    s = (narration or "").strip().upper()
    if not s:
        return ""
    # Strip PDF/CC statement artefacts (card holder name prefix, border chars, charge suffix)
    s = re.sub(r'^[A-Z][A-Z\s]{3,25}\s+(?://|\|)\s*', '', s).strip()
    s = re.sub(r'^[|/\s:]+', '', s).strip()
    s = re.sub(r'\s*\+?\s*C\s+L\s*$', '', s).strip()
    s = re.sub(r'\s*\(REF#\s*[A-Z0-9]+\)', '', s).strip()
    # Remove long reference/transaction IDs (8+ digits) that appear after known prefixes
    s = re.sub(r'(?<=UPI)\s*/\s*\d{8,}', '', s)
    s = re.sub(r'\b(?:REF|TXN|TXNID|REFNO|REFERENCE)\s*(?:NO|ID|NUMBER)?\s*[:\-]?\s*\d{6,}\b', '', s)
    # Remove standalone long digit sequences (12+ digits) — UPI ref numbers, IMPS IDs
    s = re.sub(r'\b\d{12,}\b', '', s)
    # Remove masked card numbers: XXXXXX1234, XX1234, ****1234
    s = re.sub(r'\b[X*]{2,}\d{3,6}\b', '', s)
    # Remove date stamps: DD-MM-YYYY, DD/MM/YYYY, DD.MM.YYYY, YYYY-MM-DD
    s = re.sub(r'\b\d{1,2}[/\-.](\d{1,2}|[A-Z]{3})[/\-.]\d{2,4}\b', '', s)
    s = re.sub(r'\b\d{4}[/\-.]\d{2}[/\-.]\d{2}\b', '', s)
    # Collapse multiple spaces
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def tokenize_txn_text(narration: str) -> set:
    """
    Tokenize a normalized narration into a set of meaningful tokens.
    - Calls normalize_txn_text first
    - Splits on whitespace and punctuation
    - Drops tokens ≤ 2 chars and pure numeric tokens
    """
    norm = normalize_txn_text(narration)
    if not norm:
        return set()
    # Split on whitespace and common punctuation
    raw_tokens = re.split(r'[\s/\-_.,;:@#|()\[\]{}]+', norm)
    # Keep tokens > 2 chars that are not pure digits
    return {t for t in raw_tokens if len(t) > 2 and not t.isdigit()}


# ══════════════════════════════════════════════════════════════════════════════
# RULE ENGINE  —  Pattern-based classification
# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
# POST-CLASSIFICATION SANITY VALIDATION  (Part 2 — Req 4)
# ══════════════════════════════════════════════════════════════════════════════

# Keys that are inherently credit-direction (inflows)
_CREDIT_ONLY_KEYS: frozenset = frozenset({
    "income_salary", "income_professional", "income_rental",
    "income_dividend", "income_interest", "income_capital_gains",
    "income_other", "income_gift_family", "income_inward_payment",
    "asset_own_transfer_in", "asset_refund_received", "asset_broker_payout",
    "asset_fcy_inward", "asset_fd_maturity", "asset_mf_redemption",
    "asset_equity_sale_proceeds", "asset_loan_repayment_received",
    "asset_advance_received_back", "asset_advance_returned",
    "suspense_credit",
    "trading_payout",
})

# Keys that are inherently debit-direction (outflows)
_DEBIT_ONLY_KEYS: frozenset = frozenset({
    "asset_investment_mf", "asset_investment_equity", "asset_investment_fd",
    "asset_investment_ppf", "asset_investment_nps", "asset_land_cwip",
    "asset_loans_advances_given",
    "exp_loan_emi", "exp_credit_card",
    "exp_food", "exp_grocery", "exp_household", "exp_shopping_online",
    "exp_clothing", "exp_personal_care", "exp_gifts", "exp_health",
    "exp_travel", "exp_entertainment", "exp_education", "exp_home_decor",
    "exp_maintenance", "exp_utilities", "exp_staff_wages", "exp_consultant",
    "exp_bank_charges", "exp_broker_charges", "exp_broker_dp",
    "exp_broker_interest", "exp_insurance", "exp_tax",
    "exp_personal_transfer", "exp_misc", "exp_cheque_misc", "exp_fcy_outward",
    "suspense_debit",
    "trading_funds_added",
})

# Keys that can go either way depending on context
_BIDIRECTIONAL_KEYS: frozenset = frozenset({
    "asset_broker_balance",
    "asset_long_term_other",
    "equity_capital_introduced",
    "equity_retained_earnings",
    "income_gift_family",      # can rarely appear as debit (gift returned)
    "trading_bill_equity",
    "trading_bill_fno",
    "trading_dp_charges",
    "trading_nb_charges",
    # Liabilities: credit = outstanding balance grows; debit = payment reduces it
    "liability_loan_outstanding",
    "liability_credit_card",
    "liability_other",
})


def validate_classification(
    result: Dict[str, Any],
    forced_type: Optional[str],
) -> Dict[str, Any]:
    """
    Post-classification sanity check.
    If the chosen ledger_key is directionally incompatible with forced_type:
      - Try to map to the correct directional suspense key.
      - Or correct to a known compatible key when unambiguous.
      - Always add a conflict diagnostic in 'note'.

    Returns a (possibly updated) result dict.
    Mutates a copy — does not mutate the original.
    """
    if not forced_type or forced_type not in ("credit", "debit"):
        return result   # nothing to check without a known direction

    key = result.get("ledger_key", "")
    conflict = False
    fix_note = ""

    if forced_type == "debit" and key in _CREDIT_ONLY_KEYS:
        conflict = True
        # Attempt a sensible correction
        if key == "asset_own_transfer_in":
            # Debit own transfer = funds leaving this account to own account elsewhere
            # This is fine to keep as asset_own_transfer_in but direction should be debit
            # Treat as a contra debit movement.
            pass  # keep key, just fix direction
        else:
            fix_note = f"CONFLICT: debit txn matched credit-only key '{key}'. Downgraded to suspense_debit."
            result = dict(result)
            result["ledger_key"] = "suspense_debit"
            result["book"] = "SUSPENSE"
            result["section"] = "Suspense"
            result["grp"] = "Unclassified"
            result["account"] = "Debit – Requires Review"
            result["confidence"] = min(result.get("confidence", 0.5), 0.40)
        result = dict(result)
        result["txn_type"] = "debit"
        if fix_note:
            existing_note = result.get("note", "")
            result["note"] = (existing_note + " | " + fix_note).strip(" | ")

    elif forced_type == "credit" and key in _DEBIT_ONLY_KEYS:
        conflict = True
        fix_note = f"CONFLICT: credit txn matched debit-only key '{key}'. Downgraded to suspense_credit."
        result = dict(result)
        result["ledger_key"] = "suspense_credit"
        result["book"] = "SUSPENSE"
        result["section"] = "Suspense"
        result["grp"] = "Unclassified"
        result["account"] = "Credit – Requires Review"
        result["confidence"] = min(result.get("confidence", 0.5), 0.40)
        result["txn_type"] = "credit"
        existing_note = result.get("note", "")
        result["note"] = (existing_note + " | " + fix_note).strip(" | ")

    elif forced_type == "credit" and key == "suspense_debit":
        # Suspense direction fix
        result = dict(result)
        result["ledger_key"] = "suspense_credit"
        ledger_entry = LEDGER_MAP.get("suspense_credit", LEDGER_MAP["suspense_debit"])
        result["book"] = ledger_entry[0]
        result["section"] = ledger_entry[1]
        result["grp"] = ledger_entry[2]
        result["account"] = ledger_entry[3]
        result["txn_type"] = "credit"

    elif forced_type == "debit" and key == "suspense_credit":
        result = dict(result)
        result["ledger_key"] = "suspense_debit"
        ledger_entry = LEDGER_MAP.get("suspense_debit", LEDGER_MAP["suspense_debit"])
        result["book"] = ledger_entry[0]
        result["section"] = ledger_entry[1]
        result["grp"] = ledger_entry[2]
        result["account"] = ledger_entry[3]
        result["txn_type"] = "debit"

    narr_for_guard = result.get("narration", "")
    current_key = result.get("ledger_key", "")
    current_type = forced_type or result.get("txn_type", "")
    if (
        current_key in ("asset_loans_advances_given", "asset_loan_repayment_received")
        and _is_family_loan_style_narration(narr_for_guard)
        and not _is_explicit_self_given_loan_narration(narr_for_guard)
        and not _has_broker_trading_marker(narr_for_guard)
    ):
        result = dict(result)
        result["ledger_key"] = "liability_loan_outstanding"
        ledger_entry = LEDGER_MAP["liability_loan_outstanding"]
        result["book"] = ledger_entry[0]
        result["section"] = ledger_entry[1]
        result["grp"] = ledger_entry[2]
        result["account"] = ledger_entry[3]
        if current_type in ("debit", "credit"):
            result["txn_type"] = current_type
        result["confidence"] = min(max(float(result.get("confidence", 0.60) or 0.60), 0.60), 0.90)
        existing_note = result.get("note", "")
        correction_note = (
            "Family-style transfer cannot remain in self-given loan asset bucket "
            "without explicit loan-given / repayment-received markers. "
            "Corrected to liability_loan_outstanding."
        )
        result["note"] = (existing_note + " | " + correction_note).strip(" | ")


    # Final anti-suspense rule for credits:
    # unidentified / low-confidence inward credits should still hit Income so
    # the bank balance and equity reconcile on cash basis.
    if forced_type == "credit":
        current_key = result.get("ledger_key", "")
        if current_key == "suspense_credit":
            result = dict(result)
            result["ledger_key"] = "income_other"
            ledger_entry = LEDGER_MAP["income_other"]
            result["book"] = ledger_entry[0]
            result["section"] = ledger_entry[1]
            result["grp"] = ledger_entry[2]
            result["account"] = ledger_entry[3]
            result["txn_type"] = "credit"
            result["confidence"] = min(max(float(result.get("confidence", 0.35) or 0.35), 0.35), 0.75)
            existing_note = result.get("note", "")
            extra = "FINAL CREDIT FALLBACK: unresolved inward credit posted to income_other to prevent suspense / balance-sheet mismatch."
            result["note"] = (existing_note + " | " + extra).strip(" | ")

    # Always ensure txn_type on result matches forced_type
    if not conflict and forced_type:
        result = dict(result)
        result["txn_type"] = forced_type

    return result


def _u(s): return (s or "").strip().upper()


FAMILY_NAME_PAT = r'GARG|KUMAR|SHARMA|SINGH|GUPTA|AGARWAL|JAIN|PATEL|MOTHER|FATHER|MOM|DAD|BROTHER|SISTER|UNCLE|AUNT|MAMA|PAPA|HUF|YADAV|VERMA|MEHTA|CHAUHAN|RAJPUT|SONI|BAHADUR|SAHU|SHETH|MODI|KAPOOR|KHAN|KHANNA|MALHOTRA|MALIK|MISHRA|TRIVEDI|TIWARI'


def _is_family_loan_style_narration(narration: str) -> bool:
    u = _u(narration)
    if not u:
        return False
    return bool(
        re.search(r'\bTPT\b|IB\s+FUNDS\s+TRANSFER|\bU/G\b|UNDER\s+GUARDIAN|\bHUF\b', u)
        or re.search(FAMILY_NAME_PAT, u)
    )


def _is_explicit_self_given_loan_narration(narration: str) -> bool:
    u = _u(narration)
    if not u:
        return False
    return bool(
        re.search(
            r'\b('
            r'LOAN\s+GIVEN|'
            r'LENT\s+TO|'
            r'ADVANCE\s+PAID|'
            r'ADVANCE\s+TO|'
            r'LOAN\s+RECOVERY\s+RECEIVED|'
            r'REPAYMENT\s+RECEIVED\s+FROM\s+BORROWER'
            r')\b',
            u,
        )
    )


PROTECTED_CLASSIFICATION_KEYS = {
    "income_salary",
    "income_dividend",
    "income_interest",
    "asset_investment_ppf",
    "asset_investment_nps",
    "exp_tax",
    "trading_funds_added",
    "trading_payout",
    "exp_broker_charges",
    "exp_broker_dp",
    "exp_broker_interest",
}

_BROKER_TRADING_MARKER_RE = re.compile(
    r'\b(?:'
    r'ZERODHA|KITE|UPSTOX|5PAISA|GROWW|IIFL|VENTURA|LDK|'
    r'SHARES?|SECURITIES|NSE|BSE|CLEARING|SETTLEMENT|BROKER|'
    r'DEMAT|DP|TRADING\s+ACCOUNT|HOLDING\s+ACCOUNT'
    r')\b|L\s*D\s*K\s+SHARES?',
    re.I,
)

_BROKER_CASH_MOVEMENT_RE = re.compile(
    r'\b(?:'
    r'LDK|ZERODHA|KITE|UPSTOX|5PAISA|GROWW|IIFL|VENTURA|'
    r'BROKER|BROKING|TRANSFER|TRADING\s+ACCOUNT|HOLDING\s+ACCOUNT|'
    r'FUNDS?\s+(?:ADDED|TRANSFER|PAYOUT|PAY\s*OUT)|PAYOUT|PAY\s*OUT|'
    r'NSE|BSE|CLEARING|SETTLEMENT'
    r')\b|L\s*D\s*K\s+(?:SHARES?|SECURITIES?)',
    re.I,
)

_BROKER_COST_RE = re.compile(
    r'\b(?:'
    r'CONTRACT\s+(?:BILL|COPY|NOTE)|CONTRACT\s+NOTE|STT|GST|BROKERAGE|'
    r'DP\s+CHARGES?|DEPOSITORY\s+AMC|DEMAT\s+AMC|AMC|'
    r'DELAYED\s+PAYMENT\s+INTEREST|BROKER\s+INTEREST'
    r')\b',
    re.I,
)

_PROTECTED_INVESTMENT_INCOME_RE = re.compile(
    r'\b(?:'
    r'PPF|PUBLIC\s+PROVIDENT\s+FUND|NPS|NATIONAL\s+PENSION|NSDL|CRA|POP|POP[-\s]*SP|'
    r'PLOT|SCHEME|TRUST|REDEMPTION|DIVIDEND|DIV|INTDIV|FINDIV|FINALDIV|SPLDIV|'
    r'INTEREST|INT\.?PD|FD|FIXED\s+DEPOSIT|SIP|MUTUAL\s+FUND|MF\b|'
    r'RENT|SALARY|PAYROLL|CBDT|ADVANCE\s+TAX|TDS\s+PAYMENT|GST\s+CHALLAN|SELF\s+ASSESSMENT\s+TAX'
    r')\b',
    re.I,
)

_DIVIDEND_ISSUER_GUARD_RE = re.compile(
    r'\b(?:'
    r'SALARY|PAYROLL|INTEREST|INT\s*PD|INT\.PD|REFUND|REVERSAL|CASHBACK|'
    r'LOAN|TPT|TRANSFER|NEFT\s+DR|RTGS\s+DR|UPI\s*/?\s*DR|'
    r'BROKER|ZERODHA|NSE|BSE|CLEARING|REDEMPTION|MATURITY'
    r')\b',
    re.I,
)

_DIVIDEND_ISSUER_NAME_RE = re.compile(
    r'\b(?:'
    r'COAL\s+INDIA|OIL\s+INDIA|OIL\s+(?:AND|&)\s+NATURAL\s+GAS|ONGC|NHPC|SJVN|HUDCO|'
    r'MAHANAGAR\s+GAS|HERO\s+MOTOCORP|WIPRO|TCS|TATA\s+CONSULTANCY|HDB\s+FINAN|'
    r'GENERAL\s+INSURANCE|ADITYA\s+BIRLA|LAURUS\s+LABS|BIKAJI\s+FOODS|NLC\s+INDIA|'
    r'LIFE\s+INSURANCE\s+CORPO|STATE\s+(?:BANK\s+)?OF\s+INDIA|'
    r'BAJAJ\s+FINSERV|BAJAJ\s+HOLDINGS|IRCTC|INDIAN\s+RAILWAY\s+CATERING|'
    r'ENGINEERS\s+INDIA|HINDUSTAN\s+AERONAUTICS|HOUSING\s+(?:AND|&)\s+URBAN\s+DEVELOPMENT|'
    r'EQUITAS\s+SMALL\s+FINANCE|WONDERLA|BIOCON|BHARTI\s+HEXACOM|SHIPPING\s+CORPORATION|'
    r'IRCON|UNITED\s+NILGIRI|PNB\s+HOU|POWER\s+GRID|NTPC|REC\s+LIMITED|'
    r'REC\s+LTD|PFC|POWER\s+FINANCE|GAIL|SAIL|NMDC|BEL|BHEL|IOC|'
    r'INDIAN\s+OIL|HINDUSTAN\s+ZINC|VEDANTA|INFOSYS|HCL\s+TECH|'
    r'TATA\s+STEEL|TATA\s+MOTORS|HINDUSTAN\s+UNILEVER|ITC|'
    r'RELIANCE\s+INDUSTRIES|MARUTI|BAJAJ\s+AUTO|ASIAN\s+PAINTS'
    r')\b',
    re.I,
)

_DIVIDEND_COMPACT_ISSUERS = {
    "MAHANAGARGAS": "Mahanagar Gas",
    "LAURUSLABSLIMITED": "Laurus Labs Limited",
    "LAURUSLABS": "Laurus Labs",
    "BIKAJIFOODSINTLTD": "Bikaji Foods Intl Ltd",
    "BIKAJIFOODS": "Bikaji Foods",
    "HEROMOTOCORPLTD": "Hero Motocorp Ltd",
    "HEROMOTOCORP": "Hero Motocorp",
    "BAJAJFINSERVLIMITED": "Bajaj Finserv Limited",
    "BAJAJFINSERVLIMITE": "Bajaj Finserv Limited",
    "BAJAJFINSERV": "Bajaj Finserv",
    "BAJAJHOLDINGSLTD": "Bajaj Holdings Ltd",
    "BAJAJHOLDINGS": "Bajaj Holdings",
    "HUDCO": "HUDCO",
    "HDBFINAN": "HDB Finan",
    "HDFC": "HDFC",
    "IRCTC": "IRCTC",
    "INDIANRAILWAYCATERING": "Indian Railway Catering",
    "ENGINEERSINDIA": "Engineers India",
    "HINDUSTANAERONAUTICS": "Hindustan Aeronautics",
    "HOUSINGANDURBANDEVELOPMENT": "Housing And Urban Development",
    "EQUITASSMALLFINANCE": "Equitas Small Finance",
    "WONDERLA": "Wonderla",
    "BIOCON": "Biocon",
    "BHARTIHEXACOM": "Bharti Hexacom",
    "SHIPPINGCORPORATION": "Shipping Corporation",
    "OILINDIA": "Oil India",
    "COALINDIA": "Coal India",
    "OILANDNATURALGAS": "Oil And Natural Gas",
    "ONGC": "ONGC",
    "NHPC": "NHPC",
    "SJVN": "SJVN",
    "WIPRO": "Wipro",
    "TCS": "TCS",
    "STATEBANKOFINDIA": "State Bank Of India",
    "STATEOFINDIA": "State Of India",
    "SBI": "SBI",
    "LIFEINSURANCECORPO": "Life Insurance Corpo",
    "LIC": "LIC",
}

_DIVIDEND_ISSUER_STYLE_RE = re.compile(
    r'(?:ACH\s*C[-\s]*|ACH[-\s]*CR|NACH\s+(?:CR|CREDIT)|ECS\s+CR|CMS\s+CR).{0,80}'
    r'\b(?:LTD|LIMITED|LIMITE|CORP(?:ORATION)?|COMPANY|FINAN|LABS|MOTOCORP|INSURANCE|BANK\s+OF\s+INDIA)\b',
    re.I,
)

_ACH_DIVIDEND_PREFIX_RE = re.compile(r'^\s*(?:ACH\s*C[-\s]*|ACH[-\s]*CR|NACH\s+(?:CR|CREDIT))', re.I)


def _has_broker_trading_marker(narration: str) -> bool:
    return bool(_BROKER_TRADING_MARKER_RE.search(narration or ""))


def _compact_dividend_issuer_text(narration: str) -> str:
    compact = re.sub(r'^\s*(?:ACH\s*C[-\s]*|ACH[-\s]*CR|NACH\s+(?:CR|CREDIT))', '', narration or "", flags=re.I)
    compact = re.sub(r'(?:[-\s/]*[A-Z0-9]*\d{5,}[A-Z0-9]*)+$', '', compact, flags=re.I)
    compact = re.sub(r'[^A-Z]+', '', compact.upper())
    compact = re.sub(r'(?:INTDIV|FINDIV|FINALDIV|SPLDIV|ANNUALDIV|DIVIDEND|DIV|FNL|INT\d+).*$',
                     '', compact)
    return compact


def _match_compact_dividend_issuer(narration: str) -> str:
    compact = _compact_dividend_issuer_text(narration)
    if not compact:
        return ""
    for token in sorted(_DIVIDEND_COMPACT_ISSUERS, key=len, reverse=True):
        if compact.startswith(token) or token in compact:
            return _DIVIDEND_COMPACT_ISSUERS[token]
    return ""


def _looks_like_dividend_issuer_credit(narration: str) -> bool:
    text = narration or ""
    if _DIVIDEND_ISSUER_GUARD_RE.search(text):
        return False
    if re.search(r'\b(?:PVT|PRIVATE|PAYROLL|SALARY)\b', text, re.I):
        return False
    return bool(
        _DIVIDEND_ISSUER_NAME_RE.search(text)
        or _DIVIDEND_ISSUER_STYLE_RE.search(text)
        or (_ACH_DIVIDEND_PREFIX_RE.search(text) and _match_compact_dividend_issuer(text))
    )


def _append_note(result: Dict[str, Any], marker: str) -> Dict[str, Any]:
    if not marker:
        return result
    note = result.get("note", "") or ""
    if marker not in note:
        result = dict(result)
        result["note"] = (note + " | " + marker).strip(" | ")
    return result


def _protected_lock(result: Dict[str, Any]) -> Dict[str, Any]:
    if result.get("ledger_key") in PROTECTED_CLASSIFICATION_KEYS:
        return _append_note(result, "protected_key_locked=1")
    return result


# Rules: (regex_pattern, ledger_key, forced_txn_type_or_None)
RULES: List[Tuple[str, str, Optional[str]]] = [

    # ── ASSET MOVEMENTS (Credits that are not income) ─────────────────────────
    (r"TPT-PLOT\b",                             "asset_land_cwip",                  "debit"),
    # TPT-LOAN / TPT-TXFR handled in rule_classify() pre-loop block  # FIXED: FIX-5

    # IB FUNDS TRANSFER DR handled in rule_classify() special-case below  # FIXED: FIX-2

    # Own account transfers (NEFT/RTGS where sender = receiver)
    (r"OWN\s+ACCOUNT|SELF\s+TRANSFER",         "asset_own_transfer_in",            "credit"),

    # Refunds & reversals
    (r"REFUND|REVERSAL|REVERSED|CHARGEBACK|FAILED\s+TXN",
                                                "asset_refund_received",            "credit"),

    # Broker payouts / credits
    (r"BEING\s+PAYOUT\s+RELEASED|FUNDS\s+PAYOUT|PAYOUT.*IIFL",
                                                "asset_broker_payout",              "credit"),
    (r"TRADING\s+ACCOUNT|HOLDING\s+ACCOUNT|NBSM/|BROKER\s+TRANSFER|VENTURA\s+SECURITIES|IMPS.*IIFL|IMPS.*ZERODHA",
                                                "asset_broker_balance",             None),

    # FX inward
    (r"INW\s+\d|FCY\s+INWARD|INWARD\s+REMITTANCE|TT\s+INWARD|SWIFT\s+INWARD",
                                                "asset_fcy_inward",                 "credit"),

    
    # LDK Shares & Securities — debit=funds sent to broker, credit=payout received back
    (r"L\s*D\s*K\s+SHARES",                    "trading_funds_added",              "debit"),
    (r"L\s*D\s*K\s+SHARES",                    "trading_payout",                   "credit"),

    # Dividend — ACH/NACH credit patterns from listed companies / depositories
    (r"ACH[-\s]*CR.*DIV|\bNACH\s+(?:CR|CREDIT).*DIV|ACH\s*C[-\s]*.*DIV",
                                                "income_dividend",                  "credit"),
    (r"\b(?:DIV|DIVIDEND|INTDIV|FINDIV|FINALDIV|1STINTDIV|2NDINTDIV|SPLDIV|ANNUALDIV)\b",
                                                "income_dividend",                  "credit"),
    (r"UNITED\s+NILGIRI.*DIV|IRCON.*DIV|OIL\s+AND\s+NATURAL\s+GAS.*DIV|PNB\s+HOU.*DIV|HDB\s+FINAN.*DIV",
                                                "income_dividend",                  "credit"),

    # ── INVESTMENT CREATION (Asset out, Investment in) ────────────────────────
    (r"NACH\s+DR.*(?:SIP|MF\b|FUND\b)|SIP\s+(?:MANDATE|DEBIT|PAYMENT|INSTALMENT)|ACH\s+DR.*(?:SIP|MF\b)|ECS\s+DR.*SIP|NACH-DR.*SIPMANDATEREF",
                                                "asset_investment_mf",              "debit"),
    (r"MF\s+PURCHASE|MUTUAL\s+FUND\s+(?:PURCHASE|INVEST|LUMP)",
                                                "asset_investment_mf",              "debit"),
    (r"MF\s+REDEMPTION|MUTUAL\s+FUND\s+REDEMPTION|REDEMPTION\s+CREDIT|NEFT\s+CR.*REDEMPTION",
                                                "asset_mf_redemption",              "credit"),
    (r"FD\s+(?:BOOKED|BOOKING|CREATION|RENEWAL)|FIXED\s+DEPOSIT\s+(?:PLACED|BOOKED|CREATION)|TERM\s+DEPOSIT",
                                                "asset_investment_fd",              "debit"),
    (r"FD\s+MATURITY|RD\s+(?:CREDIT|MATURITY)|RECURRING\s+DEPOSIT\s+CREDIT|SWEEP.?IN\s+CR",
                                                "asset_fd_maturity",                "credit"),
    (r"\bPPF\b|PUBLIC\s+PROVIDENT\s+FUND",     "asset_investment_ppf",             "debit"),
    (r"\bNPS\b|NATIONAL\s+PENSION|\bNSDL\b|\bCRA\b|\bPOP\b|POP[-\s]*SP|PENSION\s+CONTRIBUTION",
                                                "asset_investment_nps",             "debit"),

    # Equity investment / sale
    (r"\bSHARES\s+PURCHASE|SHARE\s+PURCHASE|BUY\s+TRADE\s+OBLIGATION|EQUITY\s+(?:BUY|PURCHASE)",
                                                "asset_investment_equity",          "debit"),
    (r"\d{15,}/ZERODHA",                        "asset_investment_equity",          "debit"),
    (r"SALE\s+OF\s+SHARES|EQUITY\s+SETTLEMENT\s+CREDIT",
                                                "asset_equity_sale_proceeds",       "credit"),

    # ── INCOME ────────────────────────────────────────────────────────────────
    (r"TPT-SAL\b|ECS\s+CR\s+SALARY|\bSALARY\s+CREDIT\b|\bPAYROLL\s+CREDIT\b|SAL\s+CR\b",
                                                "income_salary",                    "credit"),
    (r"RETAINER\s+CREDIT|CONSULTING\s+FEE\s+(?:RECEIVED|CREDIT)|ADVISORY\s+FEE|PROFESSIONAL\s+FEE\s+(?:RECEIVED|CREDIT)|DIRECTOR\s+FEE",
                                                "income_professional",              "credit"),
    (r"RENT\s+RECEIVED|PROPERTY\s+RENT|RENTAL\s+INCOME",
                                                "income_rental",                    "credit"),

    # Interest income — Int.Pd is credited BY the bank TO the account holder.
    # NB: Int.Coll is handled above in rule_classify() special-case block.
    (r"INT\.PD|INT\s+PD\b|INTEREST\s+PAID\s+TILL|INTEREST\s+CREDIT\s+PERIOD|SB\s+\d+\s+INT|SB:\d+:Int\.Pd",
                                                "income_interest",                  "credit"),
    (r"FD\s+INTEREST|INT\s+PD\s+ON\s+FD|INTEREST\s+INCOME\s+FD|CREDIT\s+INTEREST\s+CAPITALISED",
                                                "income_interest",                  "credit"),

    # Capital gains — broker payouts from Zerodha / NSE clearing
    (r"NEFT\s+CR.*ZERODHA\s+BROKING|NEFT\s+CR.*YESB0000001.*ZERODHA",
                                                "income_capital_gains",             "credit"),
    # NSE/BSE CLEARING handled in rule_classify() pre-loop block  # FIXED: FIX-3b
    (r"\bLTCG\b|\bSTCG\b|CAPITAL\s+GAIN",      "income_capital_gains",             "credit"),

    # ── EXPENDITURE: Loan EMIs and Credit Card payments (cash basis = already paid) ─
    (r"NACH\s+DR.*LOAN|EMI\s+DEBIT|LOAN\s+REPAYMENT|HOME\s+LOAN\s+EMI|AUTO\s+LOAN\s+EMI|PERSONAL\s+LOAN\s+EMI|LOAN\s+INSTALLMENT",
                                                "exp_loan_emi",                     "debit"),  # FIXED: FIX-1
    (r"(?:NEFT|RTGS)\s+DR.*\bLOAN\b|(?:NEFT|RTGS)\s+DR.*\bREPAY\b",
                                                "exp_loan_emi",                     "debit"),  # FIXED: FIX-1
    (r"CREDIT\s+CARD\s+PAYMENT|CC\s+PAYMENT|\bAMEX\b|CARD\s+PAYMENT\s+\d|SBI\s+CARD|BHDF3S|AUTOPAY\s+THANK\s+YOU|AUTOPAY\s+DONE",
                                                "exp_credit_card",                  "debit"),  # FIXED: FIX-1
    # GST on credit card transactions
    (r"IGST[-\s]VPS|IGST[-\s]RATE|CGST[-\s]VPS|SGST[-\s]VPS|FCY\s*MARKUP|CONSOLIDATED\s*FCY|FOREIGN\s+CURRENCY\s+MARKUP|MARKUP\s+FEE",
                                                "exp_bank_charges",                 "debit"),
    # Google Play and App stores
    (r"GOOGLE\s*PLAY|APPLE\.COM/BILL|APPLE\s*ITUNES|MICROSOFT\s+STORE",
                                                "exp_entertainment",                "debit"),
    # Gas stations / petrol pumps by city name pattern
    (r"\bGAS\s*(?:NOIDA|DELHI|MUMBAI|PUNE|BANGALORE|BENGALURU|GURGAON|GURUGRAM|CHENNAI|HYDERABAD|KOLKATA)\b|\bGASNOIDA\b",
                                                "exp_travel",                       "debit"),
    # Cloud / digital hosting
    (r"DIGITALOCEAN|DIGITAL\s+OCEAN|AMAZON\s+WEB\s+SERVICES|\bAWS\.AMAZON\b|GOOGLE\s+CLOUD|MICROSOFT\s+AZURE|CLOUDFLARE|GODADDY|NAMECHEAP|LINODE|VULTR",
                                                "exp_utilities",                    "debit"),
    # Foreign tax payments
    (r"GENERAL\s+DEPARTMENT\s+OF\s+TAX|DEPT\s+OF\s+TAX",
                                                "exp_tax",                          "debit"),

    # ── EXPENDITURE (bank debit = already paid expense) ───────────────────────
    # Staff (NEFT tagged with SAL / PAYROLL going to individuals)
    (r"NEFT\s+DR.*NETBANK.*(SAL\b|PAYROLL|STIP)|NACH\s+DR.*NETBANK.*(SAL\b|PAYROLL)",
                                                "exp_staff_wages",                  "debit"),
    (r"NEFT\s+DR.*NETBANK.*(?:CON\b|CONTRACT\b)|CONSULTANT\s+PAYMENT|CONTRACTOR\s+PAYMENT",
                                                "exp_consultant",                   "debit"),

    # Tax
    (r"\bCBDT\b|ADVANCE\s+TAX|TDS\s+PAYMENT|GST\s+CHALLAN|INCOME\s+TAX\s+CHALLAN|SELF\s+ASSESSMENT\s+TAX",
                                                "exp_tax",                          "debit"),

    # Bank & broker charges
    (r"INSTAALERTCHG|SMS\s+ALERT|BANK\s+CHARGES|ATM\s+FEE|ANNUAL\s+FEE|PROCESSING\s+FEE|DEBIT\s+CARD\s+ANNUAL\s+FEE",
                                                "exp_bank_charges",                 "debit"),
    (r"CDSL\s+DP\s+BILL|DEPOSITORY\s+AMC|DP\s+CHARGES|DEMAT\s+AMC",
                                                "exp_broker_dp",                    "debit"),
    (r"BEING\s+CONTRACT\s+NOTE\s+BILL|CONTRACT\s+(?:BILL|COPY|NOTE)\s+(?:NM|NB|BB|NZ)\d+|BILL\s+FOR\s+(?:NM|NB|BB|NZ)\d+",
                                                "exp_broker_charges",               "debit"),
    (r"DELAYED\s+PAYMENT\s+INTEREST|BROKER\s+INTEREST",
                                                "exp_broker_interest",              "debit"),
    (r"INSURANCE|PREMIUM\s+PAYMENT|LIC\s+PREMIUM|HDFC\s+LIFE|SBI\s+LIFE|MAX\s+LIFE|ICICI\s+PRU",
                                                "exp_insurance",                    "debit"),

    # Utilities
    (r"ELECTRICITY|BESCOM|MSEDCL|BSES|TATA\s+POWER|WATER\s+BILL|GAS\s+BILL|MAHANAGAR\s+GAS|AIRTEL|JIO|BROADBAND|DTH|TATA\s+PLAY",
                                                "exp_utilities",                    "debit"),

    # Food
    (r"SWIGGY|ZOMATO|ETERNAL\s+LIMITED|DOMINOS|PIZZA\b|KFC|MCDONALD|STARBUCKS|CAFE\b|RESTAURANT|BARBEQUE|BURGER\s+KING|CHAAYOS",
                                                "exp_food",                         "debit"),

    # Grocery
    (r"BLINKIT|ZEPTO|ZEPTONOW|INSTAMART|BIGBASKET|DUNZO|MILKBASKET|GROFERS|DMART|GROCERY|COUNTRYDELIGHT|COUNTRY\s+DELIGHT",
                                                "exp_grocery",                      "debit"),

    # Shopping
    (r"AMAZON|FLIPKART|MEESHO|SNAPDEAL|MYNTRA|AJIO|NYKAA|TATA\s+CLIQ",
                                                "exp_shopping_online",              "debit"),
    (r"LIFESTYLE|WESTSIDE|ZARA|H&?M\b|UNIQLO|MANYAVAR|RAYMOND|ARROW|PETER\s+ENGLAND|BIBA|FABINDIA|ZUDIO|TRENT",
                                                "exp_clothing",                     "debit"),

    # Entertainment
    (r"NETFLIX|SPOTIFY|PRIME\s+VIDEO|HOTSTAR|DISNEY|ZEE5|BOOKMYSHOW|PVR|INOX|CINEPOLIS",
                                                "exp_entertainment",                "debit"),

    # Health
    (r"APOLLO\s+PHARMACY|MEDPLUS|NETMEDS|1MG|PHARMEASY|PRACTO|HOSPITAL|HEALTHIANS|THYROCARE|DR\s+LAL|PHARMACY",
                                                "exp_health",                       "debit"),

    # Travel
    (r"UBER|OLA|RAPIDO|IRCTC|MAKEMYTRIP|GOIBIBO|CLEARTRIP|INDIGO|AIR\s+INDIA|VISTARA|SPICEJET|REDBUS|IXIGO|YATRA",
                                                "exp_travel",                       "debit"),

    # Education
    (r"COURSERA|UDEMY|BYJU|UNACADEMY|VEDANTU|SCHOOL\s+FEE|COLLEGE\s+FEE|UNIVERSITY|TUITION",
                                                "exp_education",                    "debit"),
    (r"IIT\s+(?:MADRAS|DELHI|BOMBAY|KANPUR|ROORKEE|KHARAGPUR|GUWAHATI|HYDERABAD|BHU|INDORE|MANDI|JODHPUR|TIRUPATI|PALAKKAD|GANDHINAGAR|DHARWAD|JAMMU|BHILAI|GOA|VARANASI)|IITM\b|IITD\b|IITB\b|IITK\b|IITR\b|NIT\s+[A-Z]+|BITS\s+PILANI|VIT\s+[A-Z]+|MANIPAL|AMITY|SYMBIOSIS|NIRMA|SRM\s+[A-Z]+",
                                                "exp_education",                    "debit"),

    # Personal care
    (r"SALON|SPA\b|BEAUTY|MAMAEARTH|SUGAR\s+COSMET|MINIMALIST|WOW\s+SKIN|TIRA",
                                                "exp_personal_care",                "debit"),

    # Home
    (r"IKEA|URBAN\s+LADDER|PEPPERFRY|HOME\s+DECOR|HOME\s+CENTRE",
                                                "exp_home_decor",                   "debit"),

    # FCY outward
    (r"FCY\s+OUTWARD|OUTWARD\s+REMITTANCE|TT\s+OUTWARD|WIRE\s+TRANSFER\s+OUTWARD",
                                                "exp_fcy_outward",                  "debit"),

    # SIP via HDFC mandate codes — EBA/MFP and SEBA/MFP are SIP debit formats
    (r"^S?EBA/MFP[-/]?\d|^MFP[-/]\d",          "asset_investment_mf",              "debit"),

    # ECS/auto-debit fallback
    (r"^ECS\s+DR|^SI\s+DR",                     "exp_misc",                         "debit"),

    # ── AU bank-specific patterns ─────────────────────────────────────────────
    (r"^BPAY\b",                                        "exp_utilities",            "debit"),
    (r"^OSKO\s+PAYMENT|^NPP\s+",                        "asset_own_transfer_in",    "credit"),
    (r"ATO\s+PAYMENT|AUSTRALIAN\s+TAX\s+OFFICE",        "exp_tax",                  "debit"),
    (r"CENTRELINK|SERVICES\s+AUSTRALIA",                "income_other",             "credit"),
    (r"MEDICARE|PRIVATE\s+HEALTH",                      "exp_health",               "debit"),
    (r"SUPERANNUATION|SUPER\s+FUND|HESTA|HOSTPLUS|AUSTRALIAN\s+SUPER",
                                                        "asset_long_term_other",    "debit"),
    (r"^PAYMENT\s+TO\b",                                "exp_misc",                 "debit"),
    (r"^DEPOSIT\s+FROM\b",                              "income_other",             "credit"),
    (r"^TRANSFER\s+TO\b",                               "exp_personal_transfer",    "debit"),
    (r"^TRANSFER\s+FROM\b",                             "income_other",             "credit"),
    (r"UBER\s*\*|GOOGLE\s*\*|APPLE\.COM|AMAZON\s*AU",   "exp_misc",                 "debit"),
]


def rule_classify(narration: str, forced_type: Optional[str] = None) -> Optional[Dict[str, Any]]:
    u = _u(narration)

    _HDFC_PREFIX = re.compile(r'^(?:WDL\s+TFR|DEP\s+TFR|WDL\s+CASH|DEP\s+CASH)\s+', re.IGNORECASE)
    _stripped = _HDFC_PREFIX.sub('', narration).strip()
    _u_stripped = _stripped.upper()

    if not forced_type:
        if u.startswith('WDL'):
            forced_type = 'debit'
        elif u.startswith('DEP'):
            forced_type = 'credit'

    # HARD DIVIDEND OVERRIDE:
    # Any CREDIT narration that starts with ACH / NACH, or contains DIV anywhere,
    # should be forced to dividend before any loan / transfer / fallback logic.
    if forced_type == 'credit' and (
        re.match(r'^\s*(?:ACH|NACH)\b', u, re.I)
        or re.search(r'DIV', u, re.I)
    ):
        return _protected_lock(_build(
            narration,
            'income_dividend',
            'credit',
            0.995,
            note='Hard dividend override: ACH/NACH prefix or DIV token detected.'
        ))

    if forced_type == 'debit' and re.search(
        r'\bNPS\b|NATIONAL\s+PENSION|\bNSDL\b|\bCRA\b|\bPOP\b|POP[-\s]*SP|PENSION\s+CONTRIBUTION',
        u,
        re.I,
    ):
        return _protected_lock(_build(
            narration,
            'asset_investment_nps',
            'debit',
            0.97,
            note='Explicit NPS investment debit.'
        ))

    if _stripped != narration and _u_stripped.startswith(('UPI', 'NEFT', 'NACH', 'IMPS', 'ACH', 'IB ', 'BIL/')):
        result = rule_classify(_stripped, forced_type)
        if result:
            result['narration'] = narration
            result['note'] = ((result.get('note') or '') + ' | hdfc_prefix_stripped').strip(' | ')
            return result

    # Priority 1: protected investment / broker / income / tax intents.
    # These must resolve before any family-loan, own-account, or transfer fallback.
    if forced_type == 'credit' and re.search(
        r'\b(?:SALARY|PAYROLL|TPT-SAL|SAL\s+CR|WAGES|STIPEND)\b',
        u,
        re.I,
    ):
        return _protected_lock(_build(
            narration,
            'income_salary',
            'credit',
            0.98,
            note='Explicit salary/payroll credit pattern.'
        ))

    if forced_type == 'debit' and re.search(r'\bPPF\b|PUBLIC\s+PROVIDENT\s+FUND', u, re.I):
        return _protected_lock(_build(
            narration,
            'asset_investment_ppf',
            'debit',
            0.98,
            note='Explicit PPF investment debit.'
        ))

    if forced_type == 'debit' and re.search(
        r'\bNPS\b|NATIONAL\s+PENSION|\bNSDL\b|\bCRA\b|\bPOP\b|POP[-\s]*SP|PENSION\s+CONTRIBUTION',
        u,
        re.I,
    ):
        return _protected_lock(_build(
            narration,
            'asset_investment_nps',
            'debit',
            0.98,
            note='Explicit NPS investment debit.'
        ))

    if forced_type == 'debit' and re.search(
        r'\bCBDT\b|ADVANCE\s+TAX|TDS\s+PAYMENT|GST\s+CHALLAN|INCOME\s+TAX\s+CHALLAN|SELF\s+ASSESSMENT\s+TAX',
        u,
        re.I,
    ):
        return _protected_lock(_build(
            narration,
            'exp_tax',
            'debit',
            0.98,
            note='Explicit tax payment debit.'
        ))

    broker_marker = _has_broker_trading_marker(u)
    if broker_marker and _BROKER_CASH_MOVEMENT_RE.search(u) and not _BROKER_COST_RE.search(u):
        broker_dir = forced_type or _infer_txn_type(u)
        if broker_dir == 'credit':
            return _protected_lock(_build(
                narration,
                'trading_payout',
                'credit',
                0.995,
                note='Broker/exchange cash movement credit. hard_broker_guard=1 broker_priority_override=1'
            ))
        return _protected_lock(_build(
            narration,
            'trading_funds_added',
            'debit',
            0.995,
            note='Broker/exchange cash movement debit. hard_broker_guard=1 broker_priority_override=1'
        ))

    if broker_marker:
        if re.search(r'DELAYED\s+PAYMENT\s+INTEREST|BROKER\s+INTEREST|INT\.COLL|INT\s+COLL\b|INTEREST\s+COLLECTED|INTEREST\s+CHARGED', u, re.I):
            return _protected_lock(_build(
                narration,
                'exp_broker_interest',
                forced_type if forced_type in ('debit', 'credit') else 'debit',
                0.96,
                note='Broker/trading interest cost. broker_priority_override=1'
            ))
        if re.search(r'\b(?:DP\s+CHARGES?|DEPOSITORY\s+AMC|DEMAT\s+AMC|DEMAT|AMC)\b', u, re.I):
            return _protected_lock(_build(
                narration,
                'exp_broker_dp',
                forced_type if forced_type in ('debit', 'credit') else 'debit',
                0.95,
                note='Demat / DP / AMC broker cost. broker_priority_override=1'
            ))
        if _BROKER_COST_RE.search(u):
            return _protected_lock(_build(
                narration,
                'exp_broker_charges',
                forced_type if forced_type in ('debit', 'credit') else 'debit',
                0.95,
                note='Broker contract bill / statutory trading cost. broker_priority_override=1'
            ))
        if _BROKER_CASH_MOVEMENT_RE.search(u):
            broker_dir = forced_type or _infer_txn_type(u)
            if broker_dir == 'credit':
                return _protected_lock(_build(
                    narration,
                    'trading_payout',
                    'credit',
                    0.97,
                    note='Broker/exchange cash movement credit. broker_priority_override=1'
                ))
            return _protected_lock(_build(
                narration,
                'trading_funds_added',
                'debit',
                0.97,
                note='Broker/exchange cash movement debit. broker_priority_override=1'
            ))

    if forced_type == 'credit' and _ACH_DIVIDEND_PREFIX_RE.search(u) and _looks_like_dividend_issuer_credit(u):
        return _protected_lock(_build(
            narration,
            'income_dividend',
            'credit',
            0.94,
            note='ACH/NACH issuer-company credit inferred as dividend.'
        ))
    
        # ── REIT / TRUST / DISTRIBUTION INCOME (VERY IMPORTANT FIX) ──
    if forced_type == 'credit' and re.search(r'\b(TRUST|REIT|DISTRIBUTION)\b', u, re.I):
        return _protected_lock(_build(
            narration,
            'income_investment',
            'credit',
            0.92,
            note='REIT/trust/distribution income detected.'
        ))

    if forced_type == 'credit' and _looks_like_dividend_issuer_credit(u):
        return _protected_lock(_build(
            narration,
            'income_dividend',
            'credit',
            0.90,
            note='Issuer/company credit inferred as dividend.'
        ))

    int_coll_match = re.search(r'INT\.COLL|INT\s+COLL\b|INTEREST\s+COLLECTED|INTEREST\s+CHARGED', u)
    if int_coll_match:
        if forced_type == 'credit':
            return _build(
                narration,
                'suspense_credit',
                'credit',
                0.45,
                note='Int.Coll pattern on a credit row — review.'
            )
        return _protected_lock(_build(
            narration,
            'exp_broker_interest',
            'debit',
            0.92,
            note='Int.Coll = bank charging interest on OD/loan (debit).'
        ))

    ib = re.match(r'^IB\s+FUNDS\s+TRANSFER\s+(CR|DR)', u)
    if ib and not broker_marker:
        is_credit = ib.group(1) == 'CR'
        if re.search(r'\bU/G\b|UNDER\s+GUARDIAN', u):
            return _build(
                narration,
                'liability_loan_outstanding',
                'credit' if is_credit else 'debit',
                0.95,
                note="IB Transfer U/G treated as family loan flow. Credits increase outstanding liability, debits reduce it."
            )
        # Extract counterparty name from IB FUNDS TRANSFER CR/DR narration
        # Format: IB FUNDS TRANSFER CR-<acct_no>  -<NAME>
        _party = re.sub(
            r'^IB\s+FUNDS\s+TRANSFER\s+(?:CR|DR)[-\s]+\S+\s*[-–]\s*', '',
            narration, flags=re.I
        ).strip()

        if is_credit:
            # Money received INTO this account from a named individual via IB Transfer.
            # This is a LOAN received (liability) — not an own-account contra.
            # Own-account transfers via IB would have no separate named counterparty
            # or would be between accounts held by the same account holder.
            return _build(
                narration,
                'liability_loan_outstanding',
                'credit',
                0.82,
                note=f"IB Funds Transfer CR from '{_party}' — loan received. Outstanding liability."
            )
        # DR side: money going back to the named individual who previously sent funds.
        # Treat as loan repayment — reduces the outstanding liability balance.
        return _build(
            narration,
            'liability_loan_outstanding',
            'debit',
            0.80,
            note=f"IB Funds Transfer DR to '{_party}' — loan repayment. Reduces outstanding liability."
        )

    if re.search(r'(?:RTGS|NEFT)\s+(?:CR|DR)', u):
        parts = u.split('-')
        if len(parts) >= 4:
            sender = parts[2].strip()
            receiver = parts[3].strip()
            is_cr = 'CR' in parts[0]
            family_evidence = _is_family_loan_style_narration(u)
            if sender and receiver and sender == receiver and len(sender) > 4 and not broker_marker and not family_evidence:
                direction = 'credit' if is_cr else 'debit'
                return _build(
                    narration,
                    'asset_own_transfer_in',
                    direction,
                    0.92,
                    note='Own account transfer (sender = receiver name). No P&L impact. own_account_override=1'
                )
            # Sender and receiver are DIFFERENT parties — genuine inter-party transfer.
            # Only treat as a loan if the sender is a person, not a corporate/institution.
            # Institutions (brokers, banks, clearing corps) are income/asset events, not loans.
            _INST_TOKENS = {
                'CLEARING', 'LIMITED', 'LTD', 'BANK', 'NSE', 'BSE', 'SEBI',
                'NSDL', 'CDSL', 'CORP', 'CORPORATION', 'FINANCE', 'FINANCIAL',
                'BROKING', 'SECURITIES', 'INSURANCE', 'MUTUAL', 'FUND',
                'SERVICES', 'TECHNOLOGIES', 'SYSTEMS', 'HOLDINGS', 'ZERODHA',
                'TRUST', 'SCHEME', 'AUTHORITY', 'GOVT', 'GOVERNMENT', 'PLOTS',
                'RESIDENTIAL', 'DEVELOPMENT',
            }
            _sender_tokens = set(sender.split())
            _sender_is_person = not bool(_sender_tokens & _INST_TOKENS)
            if is_cr and sender and len(sender) > 4 and _sender_is_person and not broker_marker and not _PROTECTED_INVESTMENT_INCOME_RE.search(u):
                return _build(
                    narration,
                    'liability_loan_outstanding',
                    'credit',
                    0.82,
                    note=f"RTGS/NEFT CR from '{sender.title()}' — loan received from individual. Outstanding liability. family_loan_override=1"
                )
            # DR to a person = loan repayment → reduces outstanding liability.
            # BUT: first check for known asset/investment keywords in the narration
            # (e.g. PPF, NPS, SIP) — those are investments, not loan repayments.
            _INVESTMENT_KEYWORDS = re.compile(
                r'\bPPF\b|PUBLIC\s+PROVIDENT\s+FUND|\bNPS\b|NATIONAL\s+PENSION|'
                r'\bSIP\b|MUTUAL\s+FUND|\bFD\b|FIXED\s+DEPOSIT|\bRD\b|RECURRING',
                re.I
            )
            _receiver_tokens = set(receiver.split())
            _receiver_is_person = not bool(_receiver_tokens & _INST_TOKENS)
            if not is_cr and not _INVESTMENT_KEYWORDS.search(u) and receiver and len(receiver) > 4 and _receiver_is_person and not broker_marker and not _PROTECTED_INVESTMENT_INCOME_RE.search(u):
                return _build(
                    narration,
                    'liability_loan_outstanding',
                    'debit',
                    0.80,
                    note=f"RTGS/NEFT DR to '{receiver.title()}' — loan repayment. Reduces outstanding liability. family_loan_override=1"
                )

    if re.match(r'^BIL/INFT/', u):
        return _build(
            narration,
            'income_other',
            'credit',
            0.93,
            note='HDFC BIL/INFT inward credit treated as other income pending review.'
        )

    hdfc_upi = re.match(r'^UPI/(DR|CR)/(\d+)/([^/]+)/([A-Z0-9]{2,8})/([^/\s]+)(?:/(.+))?$', u, re.IGNORECASE)
    if hdfc_upi:
        direction = hdfc_upi.group(1).upper()
        payee = (hdfc_upi.group(3) or '').strip()
        bank_code = (hdfc_upi.group(4) or '').strip()
        vpa = (hdfc_upi.group(5) or '').strip().lower()
        memo = (hdfc_upi.group(6) or '').strip().upper()
        canonical_type = forced_type or ('debit' if direction == 'DR' else 'credit')

        if canonical_type == 'credit':
            payee_u = payee.upper()
            if re.search(r'\bONE97\b|\bONE97\s*CO\b|PAYTM', payee_u):
                return _build(
                    narration,
                    'asset_refund_received',
                    'credit',
                    0.92,
                    note='Paytm (ONE97) UPI credit — cashback/refund.'
                )
            if re.search(r'NPCI|BHIM', payee_u):
                return _build(
                    narration,
                    'asset_refund_received',
                    'credit',
                    0.97,
                    note='NPCI BHIM cashback received.'
                )
            if re.search(r'ZERODHA|IIFL|5PAISA|UPSTOX|GROWW', payee_u):
                return _build(
                    narration,
                    'asset_broker_payout',
                    'credit',
                    0.92,
                    note=f'Broker payout: {payee}.'
                )
            if re.search(FAMILY_NAME_PAT, payee_u):
                return _build(
                    narration,
                    'income_gift_family',
                    'credit',
                    0.88,
                    note=f'UPI credit from family member: {payee}.'
                )
            return _build(
                narration,
                'income_other',
                'credit',
                0.78,
                note=f'UPI/CR from {payee} via {bank_code}.'
            )

        payee_u = payee.upper()
        memo_u = memo.upper()

        branded_rules = [
            (r'SWIGGY|ZOMATO|DOMINOS|PIZZA|KFC|MCDONALD|STARBUCKS|CHAAYOS|BURGER\s*KING|BARBEQUE|HALDIRAM|SUBWAY|BEHROUZ|FAASOS|BIRYANI', 'exp_food'),
            (r'BLINKIT|ZEPTO|BIGBASKET|DUNZO|GROFERS|DMART|INSTAMART|MILKBASKET|COUNTRYDELIGHT|RELIANCE\s*FRESH|MORE\s*SUPER', 'exp_grocery'),
            (r'AMAZON|FLIPKART|MEESHO|MYNTRA|AJIO|NYKAA|TATA\s*CLIQ', 'exp_shopping_online'),
            (r'UBER|OLA|RAPIDO|IRCTC|REDBUS|DMR|DMRC|METRO|DELHI\s*METRO', 'exp_travel'),
            (r'NETFLIX|SPOTIFY|PVR|INOX|BOOKMYSHOW|DISNEY|HOTSTAR', 'exp_entertainment'),
            (r'APOLLO|MEDPLUS|NETMEDS|1MG|PHARMEASY|PRACTO', 'exp_health'),
            (r'AIRTEL|JIO|BESCOM|MSEDCL|BSES|TATA\s*PLAY', 'exp_utilities'),
        ]
        for pat, key in branded_rules:
            if re.search(pat, payee_u) or re.search(pat, memo_u):
                return _build(
                    narration,
                    key,
                    'debit',
                    0.95,
                    note=f'HDFC UPI/DR branded: {payee}/{memo[:20]}'
                )

        if re.search(r'\b(?:FOOD|ROLL|ROLLS|MOMO|MOMOS|NOODLE|NOODLES|BURGER|BURG|DOSA|ICE|CREAM|JUICE|JUIC|CHAAT|CHAT|PAAN|KACHORI|CORN|GAVA|NIMBU|CHOC|CHOCOLATE|SWEET|SWEETS|SNACK|SNACKS|BIRYANI|EGG|PAPAD|CHAI|CHAAI|TEA|PASTA|SANDWICH|WRAP|EAT|EATS|GOL)\b', memo_u) or re.search(r'\b(?:ROLL|BURGER|DOSA|JUICE|ICE|MOMO|WRAPPERZ|BIG\s*SMOK|DOSAJEE|SWEET\s*MO)\b', payee_u):
            return _build(
                narration,
                'exp_food',
                'debit',
                0.87,
                note=f'Street vendor QR payment (food): {payee}/{memo[:20]}'
            )

        if re.search(r'\b(?:MILK|MILKS|DAIRY|CURD|PANEER|VEG|SABZI|RATION|GROCERY|KIRANA|PROVISION|OIL|GHEE|FLOUR|DAL|LENTIL|MASALA|VEGETABLE|FRUIT|FRUITS|BANANA|ONION|TOMATO|POTATO)\b', memo_u) or re.search(r'\b(?:GROCE|GROCERY|KIRANA|DAIRY|FARM|STORE|MART|OM\s*GROCE|FRESH|SABZI|MARKET)\b', payee_u):
            return _build(
                narration,
                'exp_grocery',
                'debit',
                0.87,
                note=f'Street vendor QR payment (grocery): {payee}/{memo[:20]}'
            )

        if re.search(r'\b(?:MEDICINE|MEDICINES|EYE\s*DROP|EYEDROP|DROP|TABLET|TABS|PHARMACY|PHARMA|SYRUP|BANDAGE|DRESSING|INJECTION|CHEMIST)\b', memo_u) or re.search(r'\b(?:PHARMACY|MEDICAL|CHEMIST|EYE)\b', payee_u):
            return _build(
                narration,
                'exp_health',
                'debit',
                0.87,
                note=f'QR medicine/pharmacy payment: {payee}/{memo[:20]}'
            )

        if re.match(r'^(?:paytmqr|paytm\.s|paytm-|pay\.t|bharatpe\.|bharat\.pe|phonepe\.|gpay\.|googlepay\.|vyapar\.)', vpa, re.I):
            if re.search(r'\b(?:GROCE|KIRANA|DAIRY|FARM|STORE|MART|OM\s*GROCE|FRESH|SABZI|MARKET)\b', payee_u):
                return _build(
                    narration,
                    'exp_grocery',
                    'debit',
                    0.82,
                    note=f'Local grocery QR: {payee}'
                )
            if re.search(r'\b(?:NAAN|NAN|NOODLE|ROLL|ROTI|SWEET|DOSA|JUICE|ICE|MOMO|BURGER|BIJU|JIBA|HIDAYAT|DOSAJEE|OMKAR|AKASH|WRAPPERZ|BIG\s*SMOK|MR\s*NITE|JASVEER)\b', payee_u):
                return _build(
                    narration,
                    'exp_food',
                    'debit',
                    0.82,
                    note=f'Local food vendor QR: {payee}'
                )
            return _build(
                narration,
                'exp_misc',
                'debit',
                0.80,
                note=f'Local merchant QR payment: {payee} via {vpa[:15]}'
            )

        if re.search(r'@(ybl|oksbi|ibl|pts|paytm|upi|fbl|sbi|boi|pnb|rbl|barodampay|ikwik|okicici|okhdfcbank|okaxis|axl|airtel|apl|federal|aubank)$', vpa, re.I):
            if re.search(r'\bSENT\b', memo_u):
                return _build(
                    narration,
                    'exp_personal_transfer',
                    'debit',
                    0.82,
                    note=f'UPI/DR personal transfer to {payee}.'
                )
            return _build(
                narration,
                'exp_misc',
                'debit',
                0.82,
                note=f'UPI/DR to individual: {payee}'
            )

        return _build(
            narration,
            'exp_misc',
            'debit',
            0.75,
            note=f'HDFC UPI/DR: {payee} via {bank_code} — classified as misc'
        )

    upi_credit_slash = re.match(r'^UPI/([^/]+)/([^/]*)/', u)
    if upi_credit_slash and forced_type == 'credit':
        payee = upi_credit_slash.group(1).strip()
        vpa = upi_credit_slash.group(2).strip()
        full = u

        if re.search(r'CASHBACK|REFUND|REVERSED|REVERSAL', full):
            return _build(
                narration,
                'asset_refund_received',
                'credit',
                0.97,
                note='UPI credit refund / cashback.'
            )

        if re.search(r'COLLECT\s+RE', full):
            if re.search(r'KUKUFM|CASHBACK|REWARD', full):
                return _build(
                    narration,
                    'asset_refund_received',
                    'credit',
                    0.88,
                    note='UPI collect-request credit — refund/cashback.'
                )
            return _build(
                narration,
                'income_other',
                'credit',
                0.85,
                note='UPI Collect request credit — verify source.'
            )

        merchant_pat = re.compile(r'SWIGGY|ZOMATO|AMAZON|FLIPKART|MYNTRA|NYKAA|BLINKIT|ZEPTO|NETFLIX|SPOTIFY|UBER|OLA|IRCTC|MAKEMYTRIP|BOOKMYSHOW|PVR|INOX|MEDPLUS|1MG|AIRTEL|JIO', re.I)
        if merchant_pat.search(payee):
            return _build(
                narration,
                'asset_refund_received',
                'credit',
                0.90,
                note='UPI credit from known merchant — likely refund.'
            )

        corporate_pat = re.compile(r'LTD|PVT|PRIVATE|LIMITED|BANK|FINANCE|TECH|SOLUTIONS|SERVICES|ENTERPRISE|CORP|PAY|PAYMENTS|DIGITAL|PLATFORM|MART|FOODS|LOGIS', re.I)
        family_pat = re.compile(FAMILY_NAME_PAT, re.I)

        if family_pat.search(payee) or family_pat.search(full):
            return _build(
                narration,
                'income_gift_family',
                'credit',
                0.88,
                note=f"UPI credit from family identified as capital introduced by '{payee.title()}'."
            )

        if not corporate_pat.search(payee):
            return _build(
                narration,
                'income_other',
                'credit',
                0.78,
                note=f"UPI credit from individual '{payee.title()}' treated as other income pending review."
            )

        return _build(
            narration,
            'income_other',
            'credit',
            0.80,
            note='UPI credit treated as other income until nature is confirmed.'
        )

    upi = re.match(r'^UPI[/-]([^/\-]+)[/-]([^/\-]+)[/-]([^/\-]*)[/-]?(.*)', u)
    if upi:
        payee = upi.group(1).strip()
        vpa = upi.group(2).strip()
        rem = upi.group(3).strip() + ' ' + upi.group(4).strip()

        if any(w in payee + vpa + rem for w in ('CASHBACK', 'BHIMCASHBACK', 'REFUND', 'REVERSED', 'REVERSAL')):
            return _build(
                narration,
                'asset_refund_received',
                'credit',
                0.97,
                note='UPI refund / cashback. Not income — contra to earlier expense.'
            )

        if forced_type == 'credit' and re.search(r'COLLECT\s*RE|KUKUFM|CASHBACK|REWARD', rem + payee + vpa, re.I):
            return _build(
                narration,
                'asset_refund_received',
                'credit',
                0.88,
                note='UPI collect-request credit — refund/cashback.'
            )

        merchant_rules = [
            (r'SWIGGY|ZOMATO|ETERNAL|DOMINOS|PIZZA|KFC|MCDONALD|STARBUCKS|CAFE|RESTAURANT|BARBEQUE|BURGER\s+KING|CHAAYOS', 'exp_food'),
            (r'BLINKIT|ZEPTO|ZEPTONOW|INSTAMART|BIGBASKET|DUNZO|MILKBASKET|GROFERS|DMART|GROCERY|COUNTRYDELIGHT', 'exp_grocery'),
            (r'AMAZON|FLIPKART|MEESHO|SNAPDEAL|MYNTRA|AJIO|NYKAA|TATA\s+CLIQ', 'exp_shopping_online'),
            (r'LIFESTYLE|WESTSIDE|ZARA|H&?M\b|UNIQLO|MANYAVAR|RAYMOND|BIBA|FABINDIA|ZUDIO|TRENT', 'exp_clothing'),
            (r'NETFLIX|SPOTIFY|PRIME\s+VIDEO|HOTSTAR|DISNEY|ZEE5|BOOKMYSHOW|PVR|INOX|CINEPOLIS', 'exp_entertainment'),
            (r'APOLLO|MEDPLUS|NETMEDS|1MG|PHARMEASY|PRACTO|HOSPITAL|THYROCARE|DR\s+LAL', 'exp_health'),
            (r'UBER|OLA|RAPIDO|IRCTC|MAKEMYTRIP|GOIBIBO|CLEARTRIP|INDIGO|AIR\s+INDIA|VISTARA|SPICEJET|REDBUS', 'exp_travel'),
            (r'SALON|SPA\b|BEAUTY|MAMAEARTH|SUGAR\s+COSMET|MINIMALIST|WOW\s+SKIN|TIRA', 'exp_personal_care'),
            (r'ELECTRICITY|BESCOM|AIRTEL|JIO|BROADBAND|DTH|TATA\s+PLAY|MAHANAGAR\s+GAS', 'exp_utilities'),
            (r'IIT|NIT|AIIMS|IGNOU|UNIVERSITY|COLLEGE|SCHOOL|INSTITUTE|ACADEMY', 'exp_education'),
            (r'\bMILK\b|\bMILKS\b|\bDAIRY\b', 'exp_grocery'),
            (r'\bROLL\b|\bROLLS\b|\bMOMO\b|\bNOODLE\b|\bBURGER\b|\bBURG\b|\bDOSA\b|\bICE\b|\bJUICE\b|\bCHAAT\b|\bPAAN\b|\bKACHORI\b|\bCORN\b|\bGAVA\b|\bNIMBU\b|\bCHOC\b|\bCHOCOLATE\b|\bSWEET\b|\bSNACK\b|\bFOOD\b|\bEAT\b|\bEATS\b|\bPAPAD\b|\bVERI\b|\bCHAAI\b|\bTEA\b', 'exp_food'),
            (r'\bGROCERY\b|\bKIRANA\b|\bPROVISION\b|\bSUPERMARKET\b|\bMARKET\b|\bSABZI\b|\bVEG\b', 'exp_grocery'),
            (r'\bFARMACY\b|\bMEDICINE\b|\bCHEMIST\b|\bEYE\b|\bEYEDROP\b|\bDROP\b', 'exp_health'),
        ]
        for pat, key in merchant_rules:
            if re.search(pat, payee) or re.search(pat, vpa) or re.search(pat, rem):
                return _build(narration, key, 'debit', 0.97)

        if forced_type == 'credit' or (not forced_type and 'CREDIT' in rem):
            if re.search(FAMILY_NAME_PAT, payee + ' ' + rem, re.I):
                return _build(
                    narration,
                    'income_gift_family',
                    'credit',
                    0.86,
                    note='UPI credit from family / known person treated as capital introduced.'
                )
            return _build(
                narration,
                'income_other',
                'credit',
                0.65,
                note='UPI credit from individual treated as other income pending review.'
            )

        personal_vpa = re.compile(r'@(ybl|oksbi|ibl|pts|paytm|upi|fbl|sbi|boi|pnb|rbl|barodampay|ikwik)$', re.I)
        if personal_vpa.search(vpa):
            if re.search(r'\bSENT\b', rem, re.I):
                return _build(
                    narration,
                    'exp_personal_transfer',
                    'debit',
                    0.85,
                    note='UPI debit to individual (personal transfer).'
                )
            return _build(
                narration,
                'exp_misc',
                'debit',
                0.85,
                note='UPI debit to individual (personal VPA). Misc expenditure.'
            )

        return _build(narration, 'exp_misc', 'debit', 0.75)

    if forced_type == 'credit' and re.search(r'\bONE97\b|\bONE97\s*CO\b', u):
        return _build(
            narration,
            'asset_refund_received',
            'credit',
            0.94,
            note='ONE97 CO (Paytm) credit = cashback/reward/refund.'
        )

    if forced_type == 'credit' and re.search(r'\bNPCI\s*BHIM\b|\bBHIM\s*CASHBACK\b|\bBHIMCASHBACK\b', u):
        return _build(
            narration,
            'asset_refund_received',
            'credit',
            0.97,
            note='NPCI BHIM cashback received.'
        )

    if re.match(r'^NEFT\*', u) or re.match(r'^DEP\s+TFR\s+NEFT', u, re.I):
        parts = [p.strip() for p in re.split(r'[*\-/]+', u) if p.strip()]
        payee_part = parts[-1].strip() if parts else ''
        parts_joined = ' '.join(parts)
        if forced_type == 'credit':
            explicit_salary = re.search(r'\bSAL\b|SALARY|PAYROLL|WAGES|STIPEND|EMPLOYEE\b|EMPLOYER\b', payee_part, re.I) or re.search(r'\bSAL\b|SALARY|PAYROLL|WAGES|STIPEND|EMPLOYEE\b|EMPLOYER\b', parts_joined, re.I)
            payroll_batch = re.search(r'\bHDFC0{4,}1\b|\bHDFCH\d{6,}\b', parts_joined, re.I)
            corp_salary_like = re.search(r'\b(PRIV(?:ATE)?|PVT|LIMITED|LTD|LLP|TECH|TECHNOLOGIES|SOLUTIONS|SERVICES|LABS|SYSTEMS|SOFTWARE|CONSULT|CONSULTING|ENTERPRISES?)\b', payee_part, re.I)
            non_salary_credit = re.search(r'DIV|DIVIDEND|INTEREST|INT\.?PD|REFUND|REVERSAL|CASHBACK|REWARD|LOAN|EMI|REPAY|MATURITY|REDEMPTION|BROKER|ZERODHA|UPSTOX|GROWW|NSE|BSE|MF|MUTUAL\s*FUND|PPF|NPS|FD\b|FIXED\s*DEPOSIT|RENT|SALE\b|INVOICE|VENDOR|PROFESSIONAL\s*FEE|CONSULTANCY\s*FEE', u, re.I)
            if explicit_salary or (payroll_batch and corp_salary_like and not non_salary_credit):
                return _build(
                    narration,
                    'income_salary',
                    'credit',
                    0.93 if explicit_salary else 0.86,
                    note='DEP TFR NEFT payroll-like corporate credit treated as salary.'
                )
            return _build(
                narration,
                'income_other',
                'credit',
                0.85,
                note=f'DEP TFR NEFT inward: {payee_part[:30]}'
            )

    # ── Broker funding / payout transfers (bank <-> broker) ─────────────────
    # Keep actual trading transfers in trading receivable netting flow.
    # Exclude contract note / DP / interest style lines which belong elsewhere.
    if re.search(r'\b(ZERODHA|KITE|BROKER|UPSTOX|5PAISA|GROWW|IIFL|VENTURA)\b', u):
        if not re.search(r'CONTRACT|BILL|DP\s+CHARGES|DEMAT|INTEREST|MARGIN|STT|GST|CHARGES?', u):
            _broker_dir = forced_type or _infer_txn_type(u)
            if _broker_dir == 'debit':
                return _build(
                    narration,
                    'trading_funds_added',
                    'debit',
                    0.95,
                    note='Broker funding transfer from bank to trading account.'
                )
            if _broker_dir == 'credit':
                return _build(
                    narration,
                    'trading_payout',
                    'credit',
                    0.95,
                    note='Broker payout from trading account back to bank.'
                )

    if re.search(r'\b(?:NSE|BSE)\b', u) and re.search(r'\b(?:CLEARING|SETTLEMENT)\b', u):
        _clearing_dir = forced_type or _infer_txn_type(u)
        if _clearing_dir == 'credit':
            return _build(
                narration,
                'trading_payout',
                'credit',
                0.95,
                note='NSE/BSE clearing settlement credit = trading payout cash flow.'
            )
        return _build(
            narration,
            'trading_funds_added',
            'debit',
            0.95,
            note='NSE/BSE clearing settlement debit = trading funds added cash flow.'
        )

    
        # ── ACH / NACH dividend credits ───────────────────────────────────────────
    if forced_type == 'credit' and re.search(
        r'^(?:ACH|NACH)\b.*(?:DIV|DIVIDEND|INTDIV|FINDIV|FINALDIV|SPLDIV|ANNUALDIV)',
        u,
        re.I,
    ):
        return _build(
            narration,
            'income_dividend',
            'credit',
            0.97,
            note='Forced rule: ACH/NACH dividend credit.'
        )
            

    # ── SPECIFIC TPT rules (must run BEFORE the generic TPT pattern) ─────────────
    #
    # LOAN ACCOUNTING MODEL for TPT transfers with family/individual names:
    #
    #   CREDIT (money IN from person):
    #     → liability_loan_outstanding (credit, +amt)
    #     → The person lent you money. Liability GROWS.
    #
    #   DEBIT (money OUT to person) = loan repayment back to the lender:
    #     → liability_loan_outstanding (debit, -amt)
    #     → You are repaying them. Liability SHRINKS.
    #     → Both credit and debit rows land in the SAME liability account.
    #     → Net = total received − total repaid = amount still outstanding.
    #     → When fully repaid: net ≤ 0 → liability disappears from Balance Sheet.
    #     → Trace on Liabilities shows full history (both directions).
    #
    #   EXCEPTION — TPT-GIFT: genuine gift, not a loan. Different keys.
    # ─────────────────────────────────────────────────────────────────────────────

    if re.search(r'TPT-REPAY\b', u) and not broker_marker:
        # Treat named-person TPT repayment flows on the same loan liability axis.
        # Credit = fresh loan / return from lender side into bank
        # Debit  = repayment back to lender
        if forced_type == 'credit':
            return _build(
                narration,
                'liability_loan_outstanding',
                'credit',
                0.92,
                note='TPT-REPAY credit — loan-related inward flow kept in outstanding liability bucket. family_loan_override=1'
            )
        return _build(
            narration,
            'liability_loan_outstanding',
            'debit',
            0.92,
            note='TPT-REPAY debit — loan repayment made. Reduces outstanding liability balance. family_loan_override=1'
        )

    if re.search(r'TPT-GIFT\b', u) and not broker_marker:
        # Gifts are the only TPT variant that is NOT a loan
        if forced_type == 'credit':
            return _build(
                narration,
                'income_gift_family',
                'credit',
                0.95,
                note='TPT-GIFT credit — gift received from family / capital introduced.'
            )
        return _build(
            narration,
            'exp_personal_transfer',
            'debit',
            0.82,
            note='TPT-GIFT debit — gift / personal transfer to family.'
        )

    if re.search(r'TPT-SARJU\s+HUF\b|TPT-.*\bHUF\b', u) and not broker_marker:
        if forced_type == 'credit':
            return _build(
                narration,
                'liability_loan_outstanding',
                'credit',
                0.90,
                note='TPT HUF credit — loan received from HUF. Outstanding liability. family_loan_override=1'
            )
        return _build(
            narration,
            'liability_loan_outstanding',
            'debit',
                0.90,
                note='TPT HUF debit — loan repayment to HUF. Reduces outstanding liability. family_loan_override=1'
            )

    if forced_type == 'debit' and re.search(r'\bPPF\b|PUBLIC\s+PROVIDENT\s+FUND', u):
        return _protected_lock(_build(
            narration,
            'asset_investment_ppf',
            'debit',
            0.96,
            note='Explicit PPF investment debit.'
        ))

    if forced_type == 'debit' and re.search(r'\bNPS\b|NATIONAL\s+PENSION|\bNSDL\b|\bCRA\b|\bPOP\b|POP[-\s]*SP|PENSION\s+CONTRIBUTION', u):
        return _protected_lock(_build(
            narration,
            'asset_investment_nps',
            'debit',
            0.96,
            note='Explicit NPS investment debit.'
        ))

    if forced_type == 'credit' and re.search(r'\b(TRUST|SCHEME|PLOT|YEIDA|AUTHORITY|RESIDENTIAL\s+PLOTS)\b', u):
        return _build(
            narration,
            'income_other',
            'credit',
            0.70,
            note='Institutional / scheme credit should not be treated as personal loan.'
        )

    if forced_type == 'debit' and re.search(r'\b(PLOT|SCHEME|YEIDA|AUTHORITY|RESIDENTIAL\s+PLOTS)\b', u):
        return _build(
            narration,
            'asset_land_cwip',
            'debit',
            0.88,
            note='Plot / scheme related payment treated as land / CWIP.'
        )

    if forced_type in ('credit', 'debit') and re.search(r'\bU/G\b|UNDER\s+GUARDIAN', u) and re.search(r'GARG|KUMAR|SHARMA|SINGH|PATEL|JAIN|HUF', u) and not broker_marker:
        return _build(
            narration,
            'liability_loan_outstanding',
            forced_type,
            0.96,
            note='Guardian / family transfer treated as loan movement on liability axis. family_loan_override=1'
        )

    if re.search(r'\b(?:U/G|UNDER\s+GUARDIAN)\b', u) and not broker_marker:
        return _build(
            narration,
            'liability_loan_outstanding',
            forced_type or ('credit' if re.search(r'\b(?:CR|CREDIT)\b', u) else 'debit'),
            0.92,
            note='U/G flow treated as family loan movement on outstanding liability axis. family_loan_override=1'
        )

    if forced_type == 'debit' and _is_explicit_self_given_loan_narration(u):
        return _build(
            narration,
            'asset_loans_advances_given',
            'debit',
            0.92,
            note='Explicit loan / advance given by account holder — receivable asset.'
        )

    if forced_type == 'credit' and _is_explicit_self_given_loan_narration(u):
        return _build(
            narration,
            'asset_loan_repayment_received',
            'credit',
            0.92,
            note='Explicit recovery / repayment received on loan given by account holder.'
        )

    if (
        _is_family_loan_style_narration(u)
        and re.search(r'(?:IB\s+FUNDS\s+TRANSFER|TPT-|NEFT|RTGS|IMPS|UPI)', u)
        and not broker_marker
        and not re.search(
            r'\b(PPF|PUBLIC\s+PROVIDENT\s+FUND|NPS|PLOT|SCHEME|TRUST|REDEMPTION|DIVIDEND|INTEREST|FD|FIXED\s+DEPOSIT|SIP|MUTUAL\s+FUND|ZERODHA|BROKER|NSE|BSE|CLEARING|RENT|SALARY)\b',
            u
        )
        and not _PROTECTED_INVESTMENT_INCOME_RE.search(u)
        and not _is_explicit_self_given_loan_narration(u)
    ):
        return _build(
            narration,
            'liability_loan_outstanding',
            forced_type or ('credit' if re.search(r'\b(?:CR|CREDIT)\b', u) else 'debit'),
            0.84,
            note='Named family transfer treated as loan movement on outstanding liability axis. family_loan_override=1'
        )

    # Explicit loan/advance GIVEN by you → receivable asset
    if forced_type == 'debit' and _is_explicit_self_given_loan_narration(u):
        return _build(
            narration,
            'asset_loans_advances_given',
            'debit',
            0.90,
            note='Explicit loan / advance given by account holder — receivable asset.'
        )

    if re.search(r'TPT-TXFR\b', u) and not broker_marker:
        if re.search(r'L\s*D\s*K\s+SHARES', u):
            key = 'trading_funds_added' if (forced_type or 'debit') == 'debit' else 'trading_payout'
            return _build(
                narration,
                key,
                forced_type or 'debit',
                0.92,
                note='TPT-TXFR to/from LDK Shares — trading account transfer.'
            )

        # Treat TPT transfer credits as loan received unless clearly own-account
        if forced_type == 'credit':
            if re.search(r'\b(SELF|OWN\s+ACCOUNT|OWN\s+A/C)\b', u):
                return _build(
                    narration,
                    'asset_own_transfer_in',
                    'credit',
                    0.72,
                    note='TPT-TXFR credit — own account transfer in (contra). own_account_override=1'
                )
            return _build(
                narration,
                'liability_loan_outstanding',
                'credit',
                0.88,
                note='TPT-TXFR credit — treated as loan received. Outstanding liability until repaid. family_loan_override=1'
            )

        # Treat TPT transfer debits as repayment / transfer back on same liability axis
        return _build(
            narration,
            'liability_loan_outstanding',
            'debit',
            0.88,
            note='TPT-TXFR debit — treated as loan repayment. Reduces outstanding liability. family_loan_override=1'
        )

    if re.search(r'TPT-LOAN\b', u) and not broker_marker:
        if forced_type == 'credit':
            return _build(
                narration,
                'liability_loan_outstanding',
                'credit',
                0.92,
                note='TPT-LOAN credit — loan received. Outstanding liability until repaid. family_loan_override=1'
            )
        return _build(
            narration,
            'liability_loan_outstanding',
            'debit',
            0.90,
            note='TPT-LOAN debit — loan repayment. Reduces outstanding liability balance. family_loan_override=1'
        )

    # ── GENERIC TPT with named person (runs AFTER all specific TPT patterns) ────
    # Default stance: named TPT person flows are loan movements, not assets,
    # unless explicitly tagged as loan/advance given by the account holder.
    if re.search(r'\bTPT-[A-Z0-9]+-[A-Z][A-Z\s]+$', u) and not broker_marker:
        if forced_type == 'credit':
            return _build(
                narration,
                'liability_loan_outstanding',
                'credit',
                0.86,
                note='Generic TPT credit — treated as loan received from named individual. Outstanding liability. family_loan_override=1'
            )
        return _build(
            narration,
            'liability_loan_outstanding',
            'debit',
            0.84,
            note='Generic TPT debit — treated as loan repayment to named individual. Reduces outstanding liability balance. family_loan_override=1'
        )

    if forced_type == 'debit':
        _salary_pat = re.compile(r'TPT-SAL\b|ECS\s+CR\s+SALARY|SALARY\s+CREDIT|PAYROLL\s+CREDIT|SAL\s+CR\b', re.I)
        _div_pat = re.compile(r'ACH[-\s]*CR.*DIV|\bNACH\s+(?:CR|CREDIT).*DIV|ACH\s*C[-\s]*.*DIV|\b(?:DIV|DIVIDEND|INTDIV|FINDIV|FINALDIV|SPLDIV|ANNUALDIV)\b|\bDIV(?:\d{2,}|[/-])', re.I)
        _int_pat = re.compile(r'INT\.PD|INT\s+PD|SB\s+\d+\s+INT|INTEREST\s+CREDIT|CREDIT\s+INTEREST', re.I)
        if _salary_pat.search(u):
            return _build(
                narration,
                'suspense_debit',
                'debit',
                0.50,
                note='Salary pattern in a debit txn — moved to suspense for review.'
            )
        if _div_pat.search(u):
            return _build(
                narration,
                'suspense_debit',
                'debit',
                0.50,
                note='Dividend pattern in a debit txn — moved to suspense for review.'
            )
        if _int_pat.search(u):
            return _build(
                narration,
                'suspense_debit',
                'debit',
                0.50,
                note='Interest pattern in a debit txn — moved to suspense for review.'
            )

    for pat, key, type_hint in RULES:
        if re.search(pat, u):
            if (type_hint == 'debit' and forced_type == 'credit') or (type_hint == 'credit' and forced_type == 'debit'):
                continue
            if key == "asset_own_transfer_in" and (broker_marker or _is_family_loan_style_narration(u)):
                continue
            txn_type = forced_type if forced_type in ('credit', 'debit') else (type_hint or 'debit')
            return _protected_lock(_build(narration, key, txn_type, 0.97))

    return None

_INDIAN_BANK_PREFIX_PAT = re.compile(
    r'^(?:'
    r'INB(?:\s*FUNDS?\s*TFR)?[/ ]*|'
    r'NETBANKING[/ ]*|'
    r'BY\s+TRANSFER[-–\s]*|'
    r'TO\s+TRANSFER[-–\s]*|'
    r'BY\s+CLG\s+|BY\s+TRF\s+|'
    r'IMPS/\d{6,}/|'
    r'NEFT/[^/]*/|'
    r'UPI-(?=\w)|'
    r'(?:NFS|VPS)/\d+/|'
    r'BIL/INFT/[^/]+/|'
    r'WDL\s+TFR\s+UPI/(?:DR|CR)?/?|'
    r'(?:AUTOPAY|AUTODEBIT|NACH|ACH)\s+(?:CR|DR)[-–\s]*|'
    r'(?:ACH\s+C-|ACH\s+D-)'
    r')',
    re.IGNORECASE
)

def _normalize_indian_bank_narration(narration: str) -> str:
    """
    Strip leading bank/channel prefixes so counterparty extraction sees the
    human-readable portion first. Preserve TPT and ACH C- formats for their
    dedicated branches below.
    """
    s = (narration or "").strip()
    if re.search(r'\bTPT-', s, re.I):
        return s
    if re.match(r'ACH\s+C-', s, re.I):
        return s
    stripped = _INDIAN_BANK_PREFIX_PAT.sub('', s).strip(' /-')
    return stripped if stripped else s

def _extract_counterparty(narration: str) -> str:
    """
    Extract a human-useful counterparty / merchant name from common bank narrations.
    Works on both raw narrations (with / separators) and normalized narrations
    (space-separated text used by some staging paths).
    """
    s = (narration or "").strip()
    if not s:
        return "—"

    fallback_text = re.sub(r'\s+', ' ', s).strip(" /-")
    s = _normalize_indian_bank_narration(s)
    u = s.upper().strip()



        # PNB UPI pattern: UPI/txn/P2V/vpa/Actual Name
    if "/" in s and ("/P2V/" in s.upper() or "/P2M/" in s.upper()):
        parts = [p.strip() for p in s.split("/") if p.strip()]
        if len(parts) >= 5:
            name = parts[-1]
            name = re.sub(r"[^A-Za-z ]+", " ", name)
            name = re.sub(r"\s+", " ", name).strip().title()
            if name and name.upper() not in ("P2V", "P2M"):
                return name
            




    bank_tokens = {
        "HDFC", "BANK", "ICICI", "AXIS", "YES", "YESB", "SBIN", "STATE",
        "PUNJAB", "FEDERAL", "BA", "NAT", "INDIA", "OF", "IN", "L", "BANKL"
    }
    skip_tokens = {
        "UPI", "DR", "CR", "NO", "REMARKS", "SENT", "USING", "PAY", "FOR",
        "IN", "COLLECT", "RE", "TFR", "WDL", "DEP", "TO", "FROM"
    }
    noise_tokens = {
        "UPI", "NEFT", "RTGS", "IMPS", "ACH", "NACH", "TXN", "TXNID", "ID",
        "CR", "DR", "CREDIT", "DEBIT", "REF", "UTR", "RRN", "SEQ", "NETBANK",
        "HDFC", "HDFCBANK", "ICICI", "ICICIBANK", "AXIS", "AXISBANK", "YESB",
        "YESBANK", "SBIN", "STATEBANK", "BANK", "BANKL"
    }
    month_tokens = {"JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"}
    banned_counterparties = {
        "tpt", "sal", "salary", "ach", "cr", "dr", "unknown", "transaction",
        "bank", "hdfc", "icici", "remarks", "sent using", "sent", "using"
    }
    bank_only_pattern = re.compile(r'^(?:Hdfc|Icici|Axis|Sbi|Yes|Kotak|Federal|Indusind|Idfc|Au)(?: Bank)?$', re.I)

    def _strip_noise(x: str) -> str:
        x = (x or "").strip()
        if not x:
            return ""
        x = re.sub(r'@\S+', ' ', x)
        x = re.sub(r'\b(?:UPI|NEFT|RTGS|IMPS|ACH|NACH|TXN|TXNID|ID|UTR|RRN|REF|SEQ|NETBANK|USCNB|HDFCBANK|HDFCB)\b', ' ', x, flags=re.I)
        x = re.sub(r'\b(?:HDFC|ICICI|AXIS|YESB?|SBIN|KOTAK|FEDERAL|IDFC|AUBANK|BANDHAN|INDUSIND|PNB|CANARA)\b(?:\s*BANKL?|\s*BANK)?', ' ', x, flags=re.I)
        x = re.sub(r'\b[A-Z]{2,6}\d{4,}\b', ' ', x, flags=re.I)
        x = re.sub(r'\b[A-Z0-9]{12,}\b', ' ', x, flags=re.I)
        x = re.sub(r'(?<![A-Za-z])\d{7,}(?![A-Za-z])', ' ', x)
        x = re.sub(r'\b(?:INB|IB|TFR|WDL|DEP|CLG|ECS|NACH|AUTOPAY|AUTODEBIT)\b', ' ', x, flags=re.I)
        x = re.sub(r'[/\-]+', ' ', x)
        x = re.sub(r'[^A-Z0-9&. ]', ' ', x, flags=re.I)
        x = re.sub(r'\s+', ' ', x).strip()
        return x

    def _normalize_private_limited(x: str) -> str:
        x = re.sub(r'\bP\s*L\b', 'Private Limited', x, flags=re.I).strip()
        x = re.sub(r'\bPRI\b', 'Private Limited', x, flags=re.I).strip()
        return x

    def _is_bad_candidate(x: str) -> bool:
        compact = re.sub(r'\s+', ' ', (x or '')).strip()
        if not compact:
            return True
        if compact.lower() in banned_counterparties:
            return True
        if bank_only_pattern.fullmatch(compact):
            return True
        return False

    def _meaningful_fragments(x: str) -> list[str]:
        cleaned = _strip_noise(x)
        if not cleaned:
            return []
        fragments = []
        for chunk in re.split(r'\s{2,}|/| - |\|', cleaned):
            chunk = chunk.strip(" /-")
            if not chunk:
                continue
            words = [
                w for w in chunk.split()
                if w and w.upper() not in noise_tokens and w.upper() not in month_tokens and not re.fullmatch(r'\d+', w)
            ]
            if not words:
                continue
            fragment = _normalize_private_limited(' '.join(words).title())
            if not _is_bad_candidate(fragment):
                fragments.append(fragment)
        return fragments

    def _clean_fallback_text(x: str) -> str:
        fragments = _meaningful_fragments(x)
        if fragments:
            return fragments[-1]
        cleaned = _normalize_private_limited(_strip_noise(x).title())
        cleaned = re.sub(r'\bHuf\b', '', cleaned, flags=re.I).strip()
        cleaned = re.sub(r'\s+', ' ', cleaned)
        if cleaned and not _is_bad_candidate(cleaned):
            return cleaned
        fallback_fragments = _meaningful_fragments(fallback_text)
        if fallback_fragments:
            return fallback_fragments[-1]
        return _normalize_private_limited((fallback_text or "Unknown").title())

    def _clean_name(x: str) -> str:
        x = _strip_noise(x)
        if not x:
            return ""
        chunks = [c.strip() for c in re.split(r'\s{2,}|/', x) if c.strip()]
        if not chunks:
            chunks = [x]
        meaningful = []
        for chunk in chunks:
            words = [w for w in chunk.split() if w and w.upper() not in noise_tokens and not re.fullmatch(r'\d+', w)]
            if words:
                meaningful.append(' '.join(words))
        candidate = meaningful[-1] if meaningful else x
        candidate = re.sub(r'\b(?:NO|REMARKS|SENT|USING|PAY|FOR|COLLECT|RE|TFR|WDL|DEP|TO|FROM)\b', ' ', candidate, flags=re.I)
        candidate = re.sub(r'\s+', ' ', candidate).strip(" /-")
        candidate = _normalize_private_limited(candidate.title()) if candidate else ""
        candidate = re.sub(r'\bHuf\b', '', candidate, flags=re.I).strip()
        candidate = re.sub(r'\s+', ' ', candidate)
        if _is_bad_candidate(candidate):
            fragments = _meaningful_fragments(x)
            return fragments[-1] if fragments else ""
        return candidate

    # ── HDFC-specific extractions ──────────────────────────────────────────
    if re.search(r'L\s*D\s*K\s+(?:SHARES?|SECURITIES?)', u):
        return 'LDK Shares'
    if re.search(r'\bZERODHA\s+BROKING\s+LTD\b', u):
        return 'Zerodha Broking Ltd'
    if re.search(r'\bZERODHA\b|\bKITE\b', u):
        return 'Zerodha'
    if re.search(r'\bUPSTOX\b', u):
        return 'Upstox'
    if re.search(r'\b5PAISA\b', u):
        return '5Paisa'
    if re.search(r'\bGROWW\b', u):
        return 'Groww'
    if re.search(r'\bIIFL\b', u):
        return 'IIFL'
    if re.search(r'\bVENTURA\b', u):
        return 'Ventura'
    if re.search(r'\bBROKER\b', u):
        m = re.search(r'\b([A-Z][A-Z0-9&.\s]{2,}BROKER(?:ING)?(?:\s+LTD)?)\b', u)
        if m:
            cleaned = _clean_name(m.group(1))
            if cleaned:
                return cleaned

    if re.search(r'\bACH\s*C[-\s]', u):
        compact_issuer = _match_compact_dividend_issuer(u)
        if compact_issuer:
            return compact_issuer
        m = re.search(r'ACH\s*C[-\s]*([A-Z][A-Z0-9 &.\-/]+?)(?:[-\s/]*[A-Z0-9]*\d{5,}[A-Z0-9]*\s*$|$)', u)
        if m:
            issuer_raw = m.group(1)
            issuer_raw = re.sub(r'(?:INTDIV|FINDIV|FINALDIV|SPLDIV|ANNUALDIV|DIVIDEND|DIV|FNL|INT\d+).*$', '', issuer_raw, flags=re.I)
            issuer_raw = re.sub(r'\b(?:LIMITE|LIMITED)\b', 'Limited', issuer_raw, flags=re.I)
            issuer_raw = re.sub(r'\bLTD\b', 'Ltd', issuer_raw, flags=re.I)
            cleaned = _clean_name(issuer_raw)
            if cleaned != "—":
                return cleaned

    tpt_source = re.sub(r'^\d{4}-\d{2}-\d{2}(?:\d{8,})?-', '', s, flags=re.I)
    tpt_source = re.sub(r'^\d{10,}-', '', tpt_source)
    if re.match(r'^TPT-', tpt_source, re.I):
        tpt_salary = re.match(r'^TPT-SAL(?:[\s-]+)?(.+)$', tpt_source, re.I)
        if tpt_salary:
            remainder = (tpt_salary.group(1) or '').strip(" -/")
            remainder = re.sub(r'^(?:' + '|'.join(sorted(month_tokens)) + r')\b[\s-]*', '', remainder, flags=re.I)
            cleaned = _clean_name(remainder)
            if cleaned:
                return cleaned
        tpt_generic = re.match(r'^TPT-[^-]+-(.+)$', tpt_source, re.I)
        if tpt_generic:
            cleaned = _clean_name((tpt_generic.group(1) or '').strip(" -/"))
            if cleaned:
                return cleaned

    if re.search(r'\bTPT-(?:SAL|LOAN|REPAY|RETURN|GIFT)-', u):
        m = re.search(r'TPT-(?:SAL|LOAN|REPAY|RETURN|GIFT)(?:\s+[A-Z]{3,9})?-(.+)$', u)
        if m:
            cleaned = _clean_name(m.group(1))
            if cleaned:
                return cleaned

    if re.search(r'\bIB FUNDS TRANSFER\b', u):
        m = re.search(r'IB\s+FUNDS\s+TRANSFER\s+(?:CR|DR)-\S+\s+-([A-Z][A-Z0-9 &.\-/]+)$', u)
        if m:
            cleaned = _clean_name(m.group(1))
            if cleaned:
                return cleaned

    if re.search(r'\b(?:NEFT|RTGS)\s+(?:CR|DR)-', u):
        parts = [p.strip() for p in re.split(r'-+', s) if p.strip()]
        if len(parts) >= 4:
            mids = parts[2:-1]
            for p in mids:
                pu = p.upper()
                if len(pu) >= 4 and not re.fullmatch(r'[A-Z0-9]{8,}', pu):
                    if not re.search(r'NETBANK|MUM|YESB|UTIB|HDFC|SBIN|KARB|ICIC|AXIS', pu):
                        cleaned = _clean_name(p)
                        if cleaned:
                            return cleaned
                        
    m = re.search(r'(?:^|\b)(?:WDL\s+TFR\s+)?UPI/(?:DR|CR/)?(?:\d+/)?([^/]+)/', s, re.I)
    if m:
        cleaned = _clean_name(m.group(1))
        if cleaned:
            return cleaned

    if 'UPI' in u:
        tokens = [t for t in re.split(r'[\s/\-]+', u) if t]
        try:
            i = tokens.index('UPI')
        except ValueError:
            i = -1
        if i >= 0:
            name_parts = []
            for tok in tokens[i+1:]:
                if tok in skip_tokens or tok in bank_tokens:
                    break
                if re.fullmatch(r'\d{6,}', tok) or re.fullmatch(r'[A-Z0-9]{10,}', tok):
                    if name_parts:
                        break
                    continue
                if tok.endswith('QR') or tok.startswith('PAYTMQR') or tok.startswith('BHARATPE'):
                    break
                name_parts.append(tok)
                if len(name_parts) >= 4:
                    break
            if name_parts:
                cleaned = _clean_name(' '.join(name_parts))
                if cleaned:
                    return cleaned

    # ── TPT salary credits: extract employer/company name ───────────────────
    # Examples:
    #   50200014927657-TPT-SAL FEB-GOLDILOCKS TECH SOLUTION PRI
    #   TPT-SAL JAN-GOLDILOCKS TECH SOLUTION PRIVATE LIMITED
    #   SALARY GOLDILOKS TECH SOLUTION P L
    m = re.search(
        r'(?:^|-)TPT-SAL\s+[A-Z]{3,9}\s*-\s*([A-Z][A-Z0-9 &.\-]+)$',
        u,
        flags=re.I
    )
    if m:
        cleaned = _clean_name(m.group(1))
        cleaned = re.sub(r'\bP\s*L\b$', 'Private Limited', cleaned, flags=re.I).strip()
        cleaned = re.sub(r'\bPRI\b$', 'Private Limited', cleaned, flags=re.I).strip()
        if cleaned:
            return cleaned

    m = re.search(
        r'^(?:SALARY|SAL)\s+([A-Z][A-Z0-9 &.\-]+)$',
        u,
        flags=re.I
    )
    if m:
        cleaned = _clean_name(m.group(1))
        cleaned = re.sub(r'\bP\s*L\b$', 'Private Limited', cleaned, flags=re.I).strip()
        cleaned = re.sub(r'\bPRI\b$', 'Private Limited', cleaned, flags=re.I).strip()
        if cleaned:
            return cleaned
        
    m = re.match(r'^BIL/INFT/[^/]+/(.+)$', s, re.I)
    if m:
        cleaned = _clean_name(m.group(1))
        if cleaned:
            return cleaned

    # ── TPT transfers — full HDFC format: [date-][txnid-]TPT-[subtype|token]-NAME
    # Handles both:
    #   TPT-REPAY-RAVI KUMAR
    #   2026-02-2100271140017574-TPT-NBBFJHJP7LBJJMUX-SARJU GARG
    m = re.match(
        r'^(?:\d{4}-\d{2}-\d{2})?(?:[\d]+[-])?TPT-[A-Z0-9]+-([A-Z][A-Z\s&.]+?)(?:\s*-\s*USCNB\b.*)?$',
        s, re.I
    )
    if m:
        candidate = m.group(1).strip()
        if not re.fullmatch(r'[A-Z0-9]{8,}', candidate.replace(' ', '')):
            cleaned = _clean_name(candidate)
            if cleaned:
                return cleaned

    m = re.match(
        r'^(?:[\d-]+)?TPT-[A-Z]+\s+[A-Z]{2,4}\s*-\s*([A-Z][A-Z0-9 &.]+)$',
        s, re.I
    )
    if m:
        cleaned = _clean_name(m.group(1).strip())
        cleaned = re.sub(r'\bP\s*L\b$', 'Private Limited', cleaned, flags=re.I).strip()
        cleaned = re.sub(r'\bPRI\b$', 'Private Limited', cleaned, flags=re.I).strip()
        if cleaned:
            return cleaned

    m = re.match(r'^(?:ACH|NACH)\s+[A-Z-]*\s*(STATE\s+BANK\s+OF\s+INDIA|SBI|HDFC\s+BANK|ICICI\s+BANK|AXIS\s+BANK|YES\s+BANK|KOTAK\s+MAHINDRA\s+BANK)(?:[-/ ]\d|$)', u, re.I)
    if m:
        cleaned = _clean_name(m.group(1))
        if cleaned:
            return cleaned

    m = re.match(r'^IB\s+FUNDS\s+TRANSFER\s+(?:CR|DR)[-/ ]+\S+\s*[-–]\s*(.+)$', s, re.I)
    if m:
        cleaned = _clean_name(m.group(1))
        if cleaned:
            return cleaned

    if any(k in u for k in ("NEFT", "RTGS", "IMPS")):
        parts = [p.strip() for p in re.split(r'[*\/-]+', s) if p.strip()]
        stop_words = {"NEFT", "RTGS", "IMPS", "CR", "DR", "DEP", "TFR", "NETBANK", "UPI"}
        for p in parts:
            pu = p.upper().strip()
            if pu in stop_words:
                continue
            if re.fullmatch(r'[A-Z0-9]{8,}', pu) or re.fullmatch(r'\d{6,}', pu):
                continue
            if any(bt in pu.split() for bt in bank_tokens):
                continue
            if re.fullmatch(r'[A-Z0-9]+', pu) and len(pu) >= 6:
                continue
            if len(pu) >= 4 and re.search(r'[A-Z]', pu):
                cleaned = _clean_name(p)
                if cleaned:
                    return cleaned

    m = re.match(r'^(?:ACH|NACH)\s+[A-Z-]*\s*([A-Z][A-Z0-9 &.]+?)(?:[-/ ]\d|$)', u, re.I)
    if m:
        cleaned = _clean_name(m.group(1))
        if cleaned:
            return cleaned

    # ── ICICI Bank: VPS/CR/txnid/MerchantName or VPS/DR/txnid/MerchantName ──────
    m = re.match(r'^VPS/(?:CR|DR)/\d+/(.+)$', s, re.I)
    if m:
        cleaned = _clean_name(m.group(1))
        if cleaned: return cleaned

    # ── ICICI: ICICIB/txnid/MerchantName or similar ──────────────────────────
    m = re.match(r'^[A-Z]{4,6}/\d{6,}/(.{4,})$', s, re.I)
    if m:
        cleaned = _clean_name(m.group(1))
        if cleaned and not cleaned.upper().startswith(("UPI","NEFT","IMPS","RTGS","NACH","ACH")):
            return cleaned

    # ── Axis Bank: UPI-PayeeName-VPA@bank-ref (hyphen separator) ─────────────
    m = re.match(r'^UPI-([^@\-]{3,})-[^@\-]+@[^@\-]+-', s, re.I)
    if m:
        cleaned = _clean_name(m.group(1))
        if cleaned:
            return cleaned

    # ── Axis compact UPI / beneficiary formats ───────────────────────────────
    m = re.match(r'^UPI-([^@\-]{3,})-[^@]+@[^@]+-\d+', s, re.I)
    if m:
        cleaned = _clean_name(m.group(1))
        if cleaned:
            return cleaned

    m = re.match(r'^(?:IMPS|NEFT|RTGS)[-/].*?[-/ ]([A-Z][A-Z0-9 &.]{3,})$', s, re.I)
    if m:
        cleaned = _clean_name(m.group(1))
        if cleaned:
            return cleaned

    # ── HDFC NFS: NFS/txnid/MERCHANT ─────────────────────────────────────────
    m = re.match(r'^NFS/\d+/(.+)$', s, re.I)
    if m:
        cleaned = _clean_name(m.group(1))
        if cleaned: return cleaned

    # ── Kotak: TRF FROM Name / TRF TO Name ───────────────────────────────────
    m = re.match(r'^TRF\s+(?:FROM|TO)\s+(.+?)(?:\s+ON\s+|\s+\d{2}[/-]|\s+REF|\s*$)', s, re.I)
    if m:
        cleaned = _clean_name(m.group(1))
        if cleaned: return cleaned

    # ── SBI: TO TRANSFER-Name / BY TRANSFER-Name ─────────────────────────────
    m = re.match(r'^(?:TO|BY)\s+(?:TRANSFER|TRF)\s*[-–]\s*(.+?)(?:\s+ON\s+|\s+\d{2}[/-]|\s+REF|\s*$)', s, re.I)
    if m:
        cleaned = _clean_name(m.group(1))
        if cleaned: return cleaned

    # ── SBI: INB/FUNDS TFR/Name or NETBANKING/Name ───────────────────────────
    m = re.match(r'^(?:INB|NETBANKING)\s*/\s*(?:FUNDS?\s*TFR\s*/\s*)?(.+?)(?:\s*/|\s*$)', s, re.I)
    if m:
        cleaned = _clean_name(m.group(1))
        if cleaned and len(cleaned) >= 3: return cleaned

    # ── Yes Bank / Federal Bank: IMPS/txnid/Name/BankCode ────────────────────
    m = re.match(r'^IMPS/\d{6,}/([^/]{3,})', s, re.I)
    if m:
        cleaned = _clean_name(m.group(1))
        if cleaned: return cleaned

    if re.search(r'INT\.COLL|INT\.PD', u):
        return "Bank / Broker Interest"
    if re.search(r'RNWL', u):
        return "Renewal"

    parts = [p.strip() for p in re.split(r'[/\-]+', s) if p.strip()]
    for p in parts:
        pu = p.upper()
        if pu in skip_tokens:
            continue
        if re.fullmatch(r'\d{6,}', pu) or re.fullmatch(r'[A-Z0-9]{12,}', pu):
            continue
        if any(bt in pu.split() for bt in bank_tokens):
            continue
        cleaned = _clean_name(p)
        if cleaned:
            return cleaned

    fallback = _clean_fallback_text(s)
    return fallback if not _is_bad_candidate(fallback) else _clean_fallback_text(fallback_text)

def _build(narration: str, key: str, txn_type: str, confidence: float,
           note: str = "", source: str = "rule") -> Dict[str, Any]:
    rec = LEDGER_MAP.get(key, LEDGER_MAP["suspense_debit"])
    book, section, group, account = rec
    attr = ATTRIBUTION.get(key, "Account Holder (Self)")
    return {
        "narration":      narration,
        "ledger_key":     key,
        "book":           book,
        "section":        section,
        "group":          group,
        "account":        account,
        "txn_type":       txn_type,
        "counterparty":   _extract_counterparty(narration),
        "attribution":    attr,
        "confidence":     confidence,
        "note":           note,
        "source":         source,
        "amount":         0.0,
        "txn_date":       "",
    }


# ══════════════════════════════════════════════════════════════════════════════
# EVERYDAY EXPENSE CLASSIFIER  —  Token-set matching for common debits
# ══════════════════════════════════════════════════════════════════════════════

# Each entry: (ledger_key, [multi-word phrases checked first, then single keywords])
# Multi-word phrases are checked via substring match on normalized text.
# Single keywords are checked via token-set intersection.

_EVERYDAY_RULES: List[Tuple[str, List[str], List[str]]] = [
    # (ledger_key, multi_word_phrases, single_keywords)

    # Grocery — supermarkets, quick-commerce, provision stores
    ("exp_grocery", [
        "RELIANCE SMART", "RELIANCE FRESH", "STAR BAZAAR", "NATURES BASKET",
        "NATURE BASKET", "MORE RETAIL", "MORE SUPERMARKET", "EASY DAY",
        "COUNTRY DELIGHT", "MILK BASKET", "BIG BASKET", "LE MARCHE",
        "RATNADEEP", "HERITAGE FRESH",
    ], [
        "SUPERMARKET", "PROVISION", "KIRANA", "GROCERY",
        "GROFERS", "SPENCER", "SPAR", "EASYDAY", "RATNADEEP",
        "HERITAGE", "NILGIRIS", "NEEDS", "FRESHMART",
        "HYPERMART", "HYPERMARKET", "DEPARTMENTAL",
    ]),

    # Food — restaurants, cafés, bakeries, food delivery, QSR
    ("exp_food", [
        "BURGER KING", "TACO BELL", "PIZZA HUT", "BASKIN ROBBINS",
        "DUNKIN DONUTS", "CHAI POINT", "THIRD WAVE", "BLUE TOKAI",
        "MAD OVER DONUTS", "FOOD COURT", "FOOD HALL",
    ], [
        "BIRYANI", "HALDIRAM", "BAKERY", "CAKE", "FOOD", "EATERY",
        "CANTEEN", "MESS", "TIFFIN", "DHABA", "PAAN", "TEA",
        "JUICE", "SNACK", "SUBWAY", "BASKIN", "DINER",
        "BIRYANI", "MITHAI", "SWEET", "CONFECTION",
        "CHAAYOS", "BARISTA", "COSTA", "CCD", "BARBEQUE",
        "BEHROUZ", "FAASOS", "FRESHMEN", "ROLLS", "KEBAB",
        "WOK", "NOODLE", "SUSHI", "GRILL", "BISTRO",
    ]),

    # Pharmacy / Health — medical stores, diagnostics, hospitals
    ("exp_health", [
        "APOLLO PHARMACY", "DR LAL", "DR PATH", "MAX HOSPITAL",
        "FORTIS HOSPITAL", "MANIPAL HOSPITAL", "MEDANTA",
        "NARAYANA HEALTH", "ASTER PHARMACY", "FRANK ROSS",
    ], [
        "PHARMACY", "PHARMA", "MEDICAL", "CHEMIST", "DRUG",
        "WELLNESS", "DIAGNOSTIC", "PATHOLOGY", "CLINIC",
        "DENTAL", "OPTICAL", "HOSPITAL", "HEALTH",
        "MEDICINES", "MEDPLUS", "NETMEDS", "PHARMEASY",
        "PRACTO", "HEALTHIANS", "THYROCARE", "LYBRATE",
    ]),

    # Fuel — petrol pumps, CNG, EV charging
    ("exp_travel", [
        "BHARAT PETROLEUM", "INDIAN OIL", "HINDUSTAN PETROLEUM",
        "HP PETROL", "RELIANCE PETRO", "NAYARA ENERGY",
        "ADANI GAS", "MAHANAGAR GAS", "EV CHARGE", "CHARGE POINT",
        "ATHER GRID", "TATA POWER EZ",
    ], [
        "PETROL", "PETROLEUM", "DIESEL", "FUEL", "BPCL", "HPCL",
        "IOCL", "SHELL", "NAYARA", "CNG", "PUMP",
        "FILLING", "SERVO",
    ]),

    # Utilities — electricity, water, gas, telecom, DTH, broadband
    ("exp_utilities", [
        "TATA POWER", "ADANI ELECTRICITY", "ADANI GAS", "MAHANAGAR GAS",
        "BILL PAY", "BILL PAYMENT", "BILLDESK",
        "TATA PLAY", "DISH TV", "SUN DIRECT", "HATHWAY", "ACT FIBERNET",
        "TORRENT POWER", "CESC LTD",
    ], [
        "ELECTRICITY", "WATER", "GAS", "RECHARGE", "POSTPAID",
        "PREPAID", "MOBILE", "INTERNET", "WIFI", "CABLE",
        "LPG", "PIPELINE", "SEWAGE", "TORRENT", "ADANI",
        "BESCOM", "BSES", "MSEDCL", "AIRTEL", "JIO",
        "BROADBAND", "DTH", "VODAFONE", "IDEA",
    ]),

    # E-commerce — online retailers, electronics, specialty
    ("exp_shopping_online", [
        "RELIANCE DIGITAL", "VIJAY SALES", "CROMA RETAIL",
        "TATA CLIQ", "JIO MART", "PAYTM MALL",
    ], [
        "ECOM", "ONLINE", "PAYTMMALL", "JIOMART", "TATACLIQ",
        "CROMA", "LENSKART", "FIRSTCRY", "MAMAEARTH",
        "PURPLLE", "BEWAKOOF", "SOULED", "SHOPPERS",
        "DECATHLON", "IKEA",
    ]),
]


def classify_everyday_expense(narration: str, forced_type: Optional[str] = None) -> Optional[Dict[str, Any]]:
    """
    Classify common everyday debit transactions using normalized text and
    token-set matching. Only fires for debit or unknown direction — never
    for confirmed credits.

    Returns a _build() result dict or None if no everyday pattern matches.
    """
    # Never classify credits as everyday expenses
    if forced_type == "credit":
        return None

    norm = normalize_txn_text(narration)
    if not norm:
        return None
    tokens = tokenize_txn_text(narration)

    for ledger_key, phrases, keywords in _EVERYDAY_RULES:
        # Check multi-word phrases first (substring match on normalized text)
        for phrase in phrases:
            if phrase in norm:
                return _build(narration, ledger_key, forced_type or "debit", 0.88,
                              note=f"Everyday expense: matched phrase '{phrase}'",
                              source="everyday")
        # Check single keywords via token-set intersection
        for kw in keywords:
            if kw in tokens:
                return _build(narration, ledger_key, forced_type or "debit", 0.88,
                              note=f"Everyday expense: matched keyword '{kw}'",
                              source="everyday")

    return None


# ══════════════════════════════════════════════════════════════════════════════
# TRANSFER HEURISTIC  —  NEFT/RTGS/IMPS debits to individuals
# ══════════════════════════════════════════════════════════════════════════════

# Corporate / institutional keywords — if any matches, do NOT classify as personal transfer
_CORPORATE_KEYWORDS = re.compile(
    r'\b(?:LTD|PVT|PRIVATE|LIMITED|BANK|FINANCE|FINSERV|TECH|SOLUTIONS|SERVICES'
    r'|ENTERPRISE|CORP|CORPORATION|COMPANY|PAY|PAYMENTS|DIGITAL|PLATFORM|MART'
    r'|FOODS|LOGIS|LOGISTICS|VENTURES|CAPITAL|SECURITIES|BROKING|INSURANCE'
    r'|HOSPITAL|PHARMA|REALTY|INFRA|CONSTRUCTION|HOUSING|LOAN|EMI|CREDIT'
    r'|CBDT|GST|GOVERNMENT|AUTHORITY|COUNCIL|BOARD|BUREAU|MINISTRY'
    r'|MUTUAL\s+FUND|MF\b|SIP\b|PREMIUM|CHALLAN|TAX|EDUCATION|SCHOOL'
    r'|UNIVERSITY|COLLEGE|INSTITUTE|ACADEMY|TRUST|FOUNDATION|SOCIETY)\b',
    re.I
)

def classify_transfer_heuristic(narration: str, forced_type: Optional[str] = None) -> Optional[Dict[str, Any]]:
    """
    Catch NEFT/RTGS/IMPS debit transfers to individual persons that would
    otherwise fall to suspense. Conservative: skips anything that looks
    corporate, institutional, or is already a known merchant.

    Returns a _build() result dict or None.
    """
    if forced_type == "credit":
        return None

    u = _u(narration)
    if _has_broker_trading_marker(u) or _PROTECTED_INVESTMENT_INCOME_RE.search(u):
        return None

    # Must be an outward NEFT/RTGS/IMPS transfer
    m = re.match(
        r'^(?:NEFT|RTGS|IMPS)\s+DR[-/\s]+([A-Z0-9]+)[-/\s]+(.+?)(?:[-/\s]+NETBANK|[-/\s]+UPI|$)',
        u
    )
    if not m:
        return None

    beneficiary = m.group(2).strip()
    # Must have a meaningful name (> 3 chars)
    if len(beneficiary) <= 3:
        return None

    # Skip if looks corporate / institutional
    if _CORPORATE_KEYWORDS.search(beneficiary):
        return None

    # Skip if it looks like a known expense pattern already handled by rules
    if re.search(r'LOAN|EMI|CREDIT\s*CARD|INSURANCE|PREMIUM|CHALLAN|TAX|RENT|FEE|SALARY|PAYROLL', u):
        return None

    # Family/individual name → loan repayment (debit reduces the outstanding liability)
    # Both the original loan credit AND this repayment debit land in liability_loan_outstanding,
    # so they net correctly on the Balance Sheet.
    FAMILY_NAME_PAT = r'GARG|KUMAR|SHARMA|SINGH|GUPTA|AGARWAL|JAIN|PATEL|YADAV|VERMA|MEHTA|CHAUHAN|RAJPUT|SONI|BAHADUR|SAHU|SHETH|MODI|KAPOOR|KHAN|KHANNA|MALHOTRA|MALIK|MISHRA|TRIVEDI|TIWARI'
    if re.search(FAMILY_NAME_PAT, beneficiary):
        return _build(narration, "liability_loan_outstanding", forced_type or "debit", 0.80,
                      note=f"NEFT/RTGS debit to '{beneficiary[:30]}' — loan repayment. Reduces outstanding liability. Reclassify to exp_personal_transfer if routine family payment.",
                      source="transfer_heuristic")

    return _build(narration, "exp_personal_transfer", forced_type or "debit", 0.78,
                  note=f"Transfer heuristic: NEFT/RTGS/IMPS debit to individual '{beneficiary[:30]}'",
                  source="transfer_heuristic")


    # ── AU Bank narration prefix stripping ───────────────────────────────────────
_AU_PREFIX = re.compile(
    r'^(?:'
    r'EFTPOS\s+|'
    r'DIRECT\s+DEBIT\s+|'
    r'DIRECT\s+CREDIT\s+|'
    r'BPAY\s+|'
    r'OSKO\s+PAYMENT\s+|'
    r'NPP\s+|'
    r'CARD\s+PURCHASE\s+|'
    r'ATM\s+WITHDRAWAL\s+|'
    r'PAYMENT\s+TO\s+|'
    r'DEPOSIT\s+FROM\s+|'
    r'TRANSFER\s+TO\s+|'
    r'TRANSFER\s+FROM\s+'
    r')',
    re.IGNORECASE
)

def _strip_au_prefix(narration: str) -> str:
    """Strip Australian bank narration prefixes before classification."""
    return _AU_PREFIX.sub('', narration).strip()


# ══════════════════════════════════════════════════════════════════════════════
# CLASSIFICATION STATS  —  Suspense tracking and logging
# ══════════════════════════════════════════════════════════════════════════════

from collections import Counter as _Counter

_classification_stats = {
    "total": 0,
    "suspense": 0,
    "suspense_narrations": _Counter(),
}


def _track_classification(result: Dict[str, Any]) -> None:
    """Accumulate classification stats for suspense tracking."""
    _classification_stats["total"] += 1
    if result.get("book") == "SUSPENSE" or result.get("ledger_key", "").startswith("suspense_"):
        _classification_stats["suspense"] += 1
        narr = (result.get("narration") or "")[:80].strip()
        if narr:
            _classification_stats["suspense_narrations"][narr] += 1


def log_classification_stats() -> str:
    """
    Print and return suspense percentage + top 50 suspense narrations.
    Resets counters after logging.
    """
    total = _classification_stats["total"]
    susp = _classification_stats["suspense"]
    lines = []
    lines.append("")
    lines.append("═" * 80)
    lines.append("  CLASSIFICATION STATS")
    lines.append("═" * 80)
    if total == 0:
        lines.append("  No transactions classified in this batch.")
    else:
        pct = (susp / total) * 100
        lines.append(f"  Total classified : {total}")
        lines.append(f"  Suspense         : {susp}  ({pct:.1f}%)")
        lines.append(f"  Classified OK    : {total - susp}  ({100 - pct:.1f}%)")
        lines.append("")
        top50 = _classification_stats["suspense_narrations"].most_common(50)
        if top50:
            lines.append("  TOP 50 SUSPENSE NARRATIONS:")
            lines.append("  " + "─" * 76)
            for i, (narr, cnt) in enumerate(top50, 1):
                lines.append(f"  {i:3d}. [{cnt:3d}x]  {narr}")
        else:
            lines.append("  ✅ No transactions went to suspense!")
    lines.append("═" * 80)
    output = "\n".join(lines)
    print(output)
    # Reset for next batch
    _classification_stats["total"] = 0
    _classification_stats["suspense"] = 0
    _classification_stats["suspense_narrations"] = _Counter()
    return output


def reset_classification_stats() -> None:
    """Reset stats without logging (for testing)."""
    _classification_stats["total"] = 0
    _classification_stats["suspense"] = 0
    _classification_stats["suspense_narrations"] = _Counter()


# ══════════════════════════════════════════════════════════════════════════════
# SUPERVISED ML  —  TF-IDF + Calibrated LinearSVC
# ══════════════════════════════════════════════════════════════════════════════

class SupervisedModel:
    def __init__(self):
        self.vectorizer = None
        self.model = None
        self.labels: List[str] = []
        self.ready = False
        self._load_or_train()

    def _load_or_train(self):
        if os.path.exists(MODEL_PATH):
            try:
                b = joblib.load(MODEL_PATH)
                self.vectorizer = b["vectorizer"]
                self.model = b["model"]
                self.labels = b["labels"]
                self.ready = True
                print(f"  ✅ ML model loaded — {len(self.labels)} categories")
                return
            except Exception as e:
                print(f"  ⚠️ Model load error: {e}")

        # Build training data from Excel
        EXCEL = "/mnt/user-data/uploads/BalanceSheet_Classified_MANI_GARG__1_.xlsx"
        X, y = [], []

        if os.path.exists(EXCEL):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(EXCEL, data_only=True)
                ws = wb["BS Classified Transactions"]
                for row in ws.iter_rows(values_only=True):
                    if not row[0] or not str(row[0]).strip().isdigit():
                        continue
                    narr = str(row[2]).strip()
                    line_item = str(row[7]).strip()
                    mapped = EXCEL_LINE_TO_KEY.get(line_item)
                    if narr and mapped:
                        # Augment: repeat 3× to weight real data higher
                        X.extend([narr]*3); y.extend([mapped]*3)
                print(f"  📊 Loaded {len(X)//3} training samples from Excel")
            except Exception as e:
                print(f"  ⚠️ Excel load error: {e}")

        # Also load dataset.csv if present
        dataset = os.path.join(_HERE, "dataset.csv")
        if os.path.exists(dataset):
            with open(dataset, encoding="utf-8") as f:
                for row in csv.DictReader(f):
                    d = (row.get("description") or "").strip()
                    c = (row.get("category") or "").strip()
                    if d and c:
                        mapped = CAT_TO_KEY.get(c, c)
                        if mapped in LEDGER_MAP:
                            X.append(d); y.append(mapped)

        # Load user corrections (highest weight — repeat 5×)
        corrections_path = os.path.join(_HERE, "user_corrections.csv")
        if os.path.exists(corrections_path):
            with open(corrections_path, encoding="utf-8") as f:
                for row in csv.DictReader(f, fieldnames=["description", "category"]):
                    d = (row.get("description") or "").strip()
                    c = (row.get("category") or "").strip()
                    if d and c and c in LEDGER_MAP:
                        X.extend([d] * 5); y.extend([c] * 5)

        if len(X) < 10:
            print("  ⚠️ Insufficient training data — ML disabled")
            return

        print(f"  🔄 Training ML on {len(X)} samples …")
        from sklearn.feature_extraction.text import TfidfVectorizer
        from sklearn.svm import LinearSVC
        from sklearn.calibration import CalibratedClassifierCV

        vec = TfidfVectorizer(lowercase=True, analyzer="char_wb",
                              ngram_range=(3, 5), min_df=1, max_features=200000)
        Xtr = vec.fit_transform(X)
        base = LinearSVC(class_weight="balanced", max_iter=3000)
        model = CalibratedClassifierCV(estimator=base, method="sigmoid", cv=min(3, len(set(y))))
        model.fit(Xtr, y)

        self.vectorizer = vec
        self.model = model
        self.labels = list(model.classes_)
        self.ready = True
        joblib.dump({"vectorizer": vec, "model": model, "labels": self.labels}, MODEL_PATH)
        print(f"  ✅ ML trained — {len(self.labels)} categories")

    def predict(self, text: str) -> Tuple[str, float]:
        if not self.ready:
            return "suspense_debit", 0.0
        X = self.vectorizer.transform([text])
        proba = self.model.predict_proba(X)[0]
        idx = int(proba.argmax())
        return self.labels[idx], float(proba[idx])

    def add_correction(self, narration: str, correct_key: str) -> None:
        """Persist a user reclassification and retrain incrementally."""
        if not self.ready:
            return
        corrections_path = os.path.join(_HERE, "user_corrections.csv")
        with open(corrections_path, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([narration.strip(), correct_key.strip()])
        self._load_or_train()

class UnsupervisedClusters:
    def __init__(self):
        self.vectorizer = None
        self.kmeans = None
        self.cluster_labels: Dict[int, str] = {}
        self.ready = False
        self._load_or_train()

    def _load_or_train(self):
        if os.path.exists(CLUSTER_PATH):
            try:
                b = joblib.load(CLUSTER_PATH)
                self.vectorizer = b["vectorizer"]
                self.kmeans = b["kmeans"]
                self.cluster_labels = b["cluster_labels"]
                self.ready = True
                return
            except Exception:
                pass

        EXCEL = "/mnt/user-data/uploads/BalanceSheet_Classified_MANI_GARG__1_.xlsx"
        X_text, y_cat = [], []
        if os.path.exists(EXCEL):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(EXCEL, data_only=True)
                ws = wb["BS Classified Transactions"]
                for row in ws.iter_rows(values_only=True):
                    if not row[0] or not str(row[0]).strip().isdigit():
                        continue
                    narr = str(row[2]).strip()
                    mapped = EXCEL_LINE_TO_KEY.get(str(row[7]).strip())
                    if narr and mapped:
                        X_text.append(narr); y_cat.append(mapped)
            except Exception:
                pass

        if len(X_text) < 20:
            return

        from sklearn.feature_extraction.text import TfidfVectorizer
        from sklearn.cluster import MiniBatchKMeans
        from collections import Counter

        vec = TfidfVectorizer(lowercase=True, analyzer="word", ngram_range=(1,2),
                              max_features=30000, min_df=1)
        Xmat = vec.fit_transform(X_text)
        k = min(30, len(X_text) // 5)
        km = MiniBatchKMeans(n_clusters=k, random_state=42, n_init=5)
        cids = km.fit_predict(Xmat)
        cluster_labels = {}
        for cid in range(k):
            cats = [y_cat[i] for i, c in enumerate(cids) if c == cid]
            cluster_labels[cid] = Counter(cats).most_common(1)[0][0] if cats else "suspense_debit"

        self.vectorizer = vec
        self.kmeans = km
        self.cluster_labels = cluster_labels
        self.ready = True
        joblib.dump({"vectorizer": vec, "kmeans": km, "cluster_labels": cluster_labels}, CLUSTER_PATH)

    def predict(self, text: str) -> Tuple[int, str, float]:
        if not self.ready:
            return -1, "suspense_debit", 0.0
        X = self.vectorizer.transform([text])
        cid = int(self.kmeans.predict(X)[0])
        center = self.kmeans.cluster_centers_[cid]
        dist = float(np.linalg.norm(X.toarray()[0] - center))
        return cid, self.cluster_labels.get(cid, "suspense_debit"), dist

    def is_anomaly(self, text: str, threshold: float = 0.85) -> bool:
        _, _, dist = self.predict(text)
        return dist > threshold


_ml: Optional[SupervisedModel] = None
_cl: Optional[UnsupervisedClusters] = None

def get_ml():
    global _ml
    if _ml is None: _ml = SupervisedModel()
    return _ml

def get_cl():
    global _cl
    if _cl is None: _cl = UnsupervisedClusters()
    return _cl


# ══════════════════════════════════════════════════════════════════════════════
# MAIN CLASSIFY FUNCTION
# ══════════════════════════════════════════════════════════════════════════════

# ── Misc rescue: reclassify exp_misc/suspense_debit when counterparty is known ──
_MISC_RESCUE_RULES: List[Tuple[str, str]] = [
    # Food & dining
    (r"SWIGGY|ZOMATO|ETERNAL|DOMINOS|PIZZA|KFC|MCDONALD|STARBUCKS|CAFE|RESTAURANT|BURGER\s*KING|CHAAYOS|BEHROUZ|FAASOS|HALDIRAM|BIRYANI|BAKERY|DHABA|FOOD\s*COURT|SUBWAY|BARISTA|COSTA|CCD|WINGSTOP|SUSHI|GRILL|BISTRO|BARBEQUE", "exp_food"),
    # Grocery
    (r"BLINKIT|ZEPTO|INSTAMART|BIGBASKET|DUNZO|MILKBASKET|GROFERS|DMART|BLINKIT|RELIANCE\s*FRESH|RELIANCE\s*SMART|STAR\s*BAZAAR|BIG\s*BASKET|COUNTRY\s*DELIGHT|NATURES?\s*BASKET|SPENCER|SPAR|EASYDAY|HERITAGE\s*FRESH|KIRANA|PROVISION|GROCERY|SUPERMARKET", "exp_grocery"),
    # Entertainment
    (r"NETFLIX|SPOTIFY|PRIME\s*VIDEO|HOTSTAR|DISNEY|ZEE5|BOOKMYSHOW|PVR|INOX|CINEPOLIS|GAMESKRAFT|DREAM11|MPL|WINZO", "exp_entertainment"),
    # Shopping
    (r"AMAZON|FLIPKART|MEESHO|SNAPDEAL|MYNTRA|AJIO|NYKAA|TATA\s*CLIQ|RELIANCE\s*DIGITAL|CROMA|LENSKART|FIRSTCRY|MAMAEARTH|PURPLLE|BEWAKOOF|DECATHLON|IKEA|SHOPPERS\s*STOP|WESTSIDE|LIFESTYLE|ZARA|H&?M|UNIQLO|MANYAVAR|RAYMOND|BIBA|FABINDIA|ZUDIO", "exp_shopping_online"),
    # Travel
    (r"UBER|OLA|RAPIDO|IRCTC|MAKEMYTRIP|GOIBIBO|CLEARTRIP|INDIGO|AIR\s*INDIA|VISTARA|SPICEJET|REDBUS|IXIGO|YATRA|EASEMYTRIP|ABHIBUS|METRO\s*RAIL|BMTC|KSRTC", "exp_travel"),
    # Health
    (r"APOLLO\s*PHARMACY|MEDPLUS|NETMEDS|1MG|PHARMEASY|PRACTO|HEALTHIANS|THYROCARE|DR\s*LAL|MEDANTA|FORTIS|MANIPAL\s*HOSPITAL|MAX\s*HOSPITAL|AIIMS|PHARMACY|CHEMIST|CLINIC|DIAGNOSTIC", "exp_health"),
    # Utilities
    (r"AIRTEL|JIO|VODAFONE|IDEA|BSNL|TATA\s*PLAY|DISH\s*TV|HATHWAY|BESCOM|BSES|MSEDCL|TATA\s*POWER|ADANI\s*(?:ELEC|GAS)|MAHANAGAR\s*GAS|TORRENT\s*POWER|CESC", "exp_utilities"),
    # Education
    (r"IIT|NIT|BITS\s*PILANI|COURSERA|UDEMY|BYJU|UNACADEMY|VEDANTU|SCHOOL\s*FEE|COLLEGE\s*FEE|UNIVERSITY|TUITION|INSTITUTE|ACADEMY|NIIT|APTECH", "exp_education"),
    # Personal care
    (r"SALON|SPA|BEAUTY|MAMAEARTH|SUGAR\s*COSMET|MINIMALIST|WOW\s*SKIN|TIRA|LAKME|NYKAA\s*BEAUTY|PLUM|DOT\s*&\s*KEY", "exp_personal_care"),
    # Insurance
    (r"LIC|HDFC\s*LIFE|SBI\s*LIFE|MAX\s*LIFE|ICICI\s*PRU|BAJAJ\s*ALLIANZ|STAR\s*HEALTH|NIVA\s*BUPA|CARE\s*HEALTH|RELIANCE\s*GENERAL|NEW\s*INDIA\s*ASSURANCE|TATA\s*AIG", "exp_insurance"),
]

def _rescue_misc_by_counterparty(narration: str, result: Dict[str, Any], forced_type: Optional[str]) -> Dict[str, Any]:
    """
    If classification ended at exp_misc or suspense_debit, try to recover
    a specific ledger key by pattern-matching the full narration against
    known merchant / brand patterns.
    """
    u = (narration or "").upper()
    for pattern, key in _MISC_RESCUE_RULES:
        if re.search(pattern, u, re.I):
            rescued = _build(narration, key, forced_type or "debit", 0.82,
                             note=f"Misc rescue: matched pattern for '{key}'",
                             source="rescue")
            rescued["amount"]     = result.get("amount", 0)
            rescued["txn_date"]   = result.get("txn_date", "")
            rescued["cluster_id"] = result.get("cluster_id", -1)
            rescued["is_anomaly"] = result.get("is_anomaly", False)
            return rescued
    return result


def classify(narration: str, forced_type: Optional[str] = None,
             amount: float = 0.0, txn_date: str = "") -> Dict[str, Any]:
    """
    Classify a bank narration into the 3-book ledger.
    Returns a rich dict with book / section / group / account / attribution.

    Pipeline order:
      Step 1 → Rule engine (regex patterns, high confidence)
      Step 2 → Everyday expense classifier (token-set matching)
      Step 3 → Transfer heuristic (NEFT/RTGS/IMPS to individuals)
      Step 4 → Supervised ML (TF-IDF + LinearSVC)
      Step 5 → Unsupervised cluster fallback
      Step 6 → Suspense (last resort)

    FIXED: Symmetric direction guards (Issues A1-A9).
    forced_type from the parser is treated as source of truth (Issue A3).
    """
    narration = (narration or "").strip()
    # ── Strip PDF/credit-card statement artefacts ─────────────────────────────
    # Remove card holder name prefix: "SARJU GARG //| : MERCHANT" → "MERCHANT"
    narration = re.sub(r'^[A-Za-z][A-Za-z\s]{3,25}\s+(?://|\|)\s*', '', narration).strip()
    # Remove "//| :" or "| :" or "//" prefix from PDF table border bleeding
    narration = re.sub(r'^[|/\s:]+', '', narration).strip()
    # Remove "+ C l" or "C l" credit card charge indicator suffix
    narration = re.sub(r'\s*\+?\s*C\s+[Ll]\s*$', '', narration).strip()
    # Remove "(Ref# VT...)" or "(Ref# ST...)" reference numbers
    narration = re.sub(r'\s*\(Ref#\s*[A-Z0-9]+\)', '', narration).strip()
    # Remove trailing isolated "C l" or "+ C"
    narration = re.sub(r'\s+\+?\s*C\s*$', '', narration).strip()
    u = _u(narration)

    # Strip AU bank prefixes if present
    if _AU_PREFIX.match(narration):
        narration = _strip_au_prefix(narration)

    # ── Step 1: Rule engine ──────────────────────────────────────────────────
    hit = rule_classify(narration, forced_type)
    if hit:
        # Run post-classification sanity validation
        hit = validate_classification(hit, forced_type)
        key = hit["ledger_key"]
        if key in PROTECTED_CLASSIFICATION_KEYS:
            hit = _protected_lock(hit)
            hit["amount"] = amount
            hit["txn_date"] = txn_date
            hit["cluster_id"] = -1
            hit["is_anomaly"] = False
            _track_classification(hit)
            return hit
        # Secondary direction guard: if a credit-only key is still present on a debit txn
        # after validation it means the validator allowed it (bidirectional key).
        # Only fall through to everyday if it's a clear case that everyday should handle.
        if forced_type == "debit" and key in _CREDIT_ONLY_KEYS:
            print(f"  [CLASSIFY] Post-validate guard (debit→credit key '{key}'): '{narration[:60]}' → everyday")
            hit = None
        elif forced_type == "credit" and key.startswith("exp_") and key not in _BIDIRECTIONAL_KEYS:
            print(f"  [CLASSIFY] Post-validate guard (credit→expense key '{key}'): '{narration[:60]}' → suspense_credit")
            hit = _build(narration, "income_other", "credit", 0.45,
                         note=f"Direction guard: credit txn matched expense key '{key}'. Posted to other income to avoid suspense.")
    if hit:
        hit["amount"] = amount
        hit["txn_date"] = txn_date
        hit["cluster_id"] = -1
        hit["is_anomaly"] = False
        _track_classification(hit)
        return hit

    # ── Step 2: Everyday expense classifier ──────────────────────────────────
    everyday = classify_everyday_expense(narration, forced_type)
    if everyday:
        everyday = validate_classification(everyday, forced_type)
        everyday["amount"] = amount
        everyday["txn_date"] = txn_date
        everyday["cluster_id"] = -1
        everyday["is_anomaly"] = False
        _track_classification(everyday)
        return everyday

    # ── Step 3: Transfer heuristic ───────────────────────────────────────────
    transfer = classify_transfer_heuristic(narration, forced_type)
    if transfer:
        transfer = validate_classification(transfer, forced_type)
        transfer["amount"] = amount
        transfer["txn_date"] = txn_date
        transfer["cluster_id"] = -1
        transfer["is_anomaly"] = False
        _track_classification(transfer)
        return transfer

    # ── Step 4: Supervised ML ────────────────────────────────────────────────
    ml = get_ml()
    pred_key, conf = ml.predict(narration)
    pred_key = CAT_TO_KEY.get(pred_key, pred_key)
    if pred_key not in LEDGER_MAP:
        pred_key = "suspense_debit"

    # Symmetric ML direction guards using canonical sets
    INCOME_KEY_PREFIXES = ("income_",)
    CREDIT_ASSET_KEYS = {
        "asset_own_transfer_in", "asset_refund_received", "asset_broker_payout",
        "asset_fcy_inward", "asset_fd_maturity", "asset_mf_redemption",
        "asset_equity_sale_proceeds", "asset_loan_repayment_received",
        "asset_advance_received_back", "asset_advance_returned",
    }
    if forced_type == "credit" and pred_key in _DEBIT_ONLY_KEYS:
        pred_key = "income_other"
        conf = min(max(conf, 0.35), 0.40)
        print(f"  [ML] Credit txn → debit-only key blocked → income_other fallback: '{narration[:60]}'")
    elif forced_type == "debit" and (
        any(pred_key.startswith(p) for p in INCOME_KEY_PREFIXES)
        or pred_key in CREDIT_ASSET_KEYS
    ):
        pred_key = "suspense_debit"
        conf = min(conf, 0.40)
        print(f"  [ML] Debit txn → income/credit-asset key blocked → suspense_debit: '{narration[:60]}'")

    # ── Step 5: Cluster fallback for low confidence ──────────────────────────
    cl = get_cl()
    cid, cl_key, dist = cl.predict(narration)
    is_anomaly = cl.is_anomaly(narration)
    if conf < 0.45 and cl_key in LEDGER_MAP:
        if not (forced_type == "credit" and cl_key in _DEBIT_ONLY_KEYS):
            if not (forced_type == "debit" and (
                any(cl_key.startswith(p) for p in INCOME_KEY_PREFIXES)
                or cl_key in CREDIT_ASSET_KEYS
            )):
                pred_key = cl_key
                conf = max(conf, 0.35)

    # ── Step 6: Suspense (last resort) ───────────────────────────────────────
    if conf < 0.30:
        pred_key = "income_other" if forced_type == "credit" else "suspense_debit"
        print(f"  [ML] Low confidence ({conf:.0%}) → fallback: '{narration[:60]}'")

    # Hard overrides: certain narration patterns must never go to wrong keys
    u_check = u  # already uppercased
    if re.search(r'AUTOPAY', u_check) and pred_key in ('liability_loan_outstanding','exp_loan_emi'):
        pred_key = 'exp_credit_card'; conf = 0.92
    elif re.search(r'IGST|CGST|SGST|FCY.MARKUP|MARKUP.FEE', u_check) and pred_key not in ('exp_bank_charges',):
        pred_key = 'exp_bank_charges'; conf = 0.88
    elif re.search(r'GOOGLE.PLAY|APPLE.*BILL|NETFLIX', u_check) and pred_key not in ('exp_entertainment',):
        pred_key = 'exp_entertainment'; conf = 0.90

    resolved_type = forced_type if forced_type in ("credit", "debit") else _infer_type(u)
    result = _build(narration, pred_key, resolved_type, conf,
                    note=f"ML: {pred_key} ({conf:.0%})", source="ml")
    result = validate_classification(result, forced_type)
    result["amount"] = amount
    result["txn_date"] = txn_date
    result["cluster_id"] = cid
    result["is_anomaly"] = is_anomaly

    # ── Post-classification misc rescue ──────────────────────────────────────
    # If we landed on exp_misc or suspense, try to recover a specific category
    # by matching the counterparty name against known brand patterns.
    if result.get("ledger_key") in ("exp_misc", "suspense_debit") and resolved_type == "debit":
        result = _rescue_misc_by_counterparty(narration, result, forced_type)

    _track_classification(result)
    return result


def _infer_type(u: str) -> str:
    CR_HINTS = ("NEFT CR","RTGS CR","IMPS CR","ACH CR","ACH-CR","ACH C-","NACH CR",
                "ECS CR","CREDIT","INWARD","REFUND","SALARY","PAYROLL","DIVIDEND",
                "INT.PD","INT PD","RECEIVED","PAYOUT","PROCEEDS","MATURITY",
                "CASHBACK","REVERSAL","TPT-REPAY","TPT-RETURN","TPT-MANI")
    DR_HINTS = ("NEFT DR","NACH DR","ACH DR","EMI DEBIT","DEBIT","PAYMENT",
                "PURCHASE","WITHDRAWAL","SIP","PREMIUM","CHALLAN","ADVANCE TAX",
                "BOOKING","RECHARGE","TPT-LOAN","TPT-PLOT")
    for w in CR_HINTS:
        if w in u: return "credit"
    for w in DR_HINTS:
        if w in u: return "debit"
    return "debit"


# ══════════════════════════════════════════════════════════════════════════════
# DATABASE
# ══════════════════════════════════════════════════════════════════════════════

class DBStore:
    def __init__(self, db_path=DB_PATH):
        self.db_path = db_path
        self._init()

    def _conn(self):
        c = sqlite3.connect(self.db_path, check_same_thread=False)
        c.row_factory = sqlite3.Row
        return c

    def _init(self):
        with self._conn() as c:
            c.executescript("""
                CREATE TABLE IF NOT EXISTS users(
                  id TEXT PRIMARY KEY, name TEXT, email TEXT, phone TEXT, created_at TEXT);
                CREATE TABLE IF NOT EXISTS ledger(
                  id TEXT PRIMARY KEY, user_id TEXT, account_id TEXT,
                  txn_date TEXT, narration TEXT, amount REAL DEFAULT 0,
                  txn_type TEXT,
                  ledger_key TEXT, book TEXT, section TEXT, grp TEXT, account TEXT,
                  attribution TEXT, counterparty TEXT DEFAULT '',
                  confidence REAL DEFAULT 0, source TEXT,
                  cluster_id INTEGER DEFAULT -1, is_anomaly INTEGER DEFAULT 0,
                  note TEXT DEFAULT '', created_at TEXT,
                  FOREIGN KEY(user_id) REFERENCES users(id));
                CREATE INDEX IF NOT EXISTS idx_uid   ON ledger(user_id);
                CREATE INDEX IF NOT EXISTS idx_book  ON ledger(book);
                CREATE INDEX IF NOT EXISTS idx_date  ON ledger(txn_date);
                CREATE UNIQUE INDEX IF NOT EXISTS ux_ledger_dedupe
                ON ledger(user_id, account_id, txn_date, narration, amount, txn_type, counterparty, source);
            """)

    def upsert_user(self, uid, name, email="", phone=""):
        now = datetime.utcnow().isoformat()
        with self._conn() as c:
            c.execute("""INSERT INTO users(id,name,email,phone,created_at) VALUES(?,?,?,?,?)
                         ON CONFLICT(id) DO UPDATE SET name=excluded.name""",
                      (uid, name, email, phone, now))

    def find_user(self, name="", phone="", email="") -> Optional[Dict]:
        with self._conn() as c:
            for col, val in [("phone",phone),("email",email),("name",name)]:
                if val:
                    r = c.execute(f"SELECT * FROM users WHERE LOWER({col})=LOWER(?) LIMIT 1", (val,)).fetchone()
                    if r: return dict(r)
        return None

    def insert_txn(self, row: Dict[str, Any]):
        now = datetime.utcnow().isoformat()
        with self._conn() as c:
            c.execute("""INSERT OR IGNORE INTO ledger(
                id,user_id,account_id,txn_date,narration,amount,txn_type,
                ledger_key,book,section,grp,account,attribution,counterparty,
                confidence,source,cluster_id,is_anomaly,note,created_at)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (
                row.get("id", str(uuid.uuid4())), row["user_id"], row.get("account_id","main"),
                row.get("txn_date",""), row.get("narration",""), float(row.get("amount",0)),
                row.get("txn_type","debit"), row.get("ledger_key","suspense_debit"),
                row.get("book","SUSPENSE"), row.get("section","Suspense"),
                row.get("group", row.get("grp","Unclassified")),
                row.get("account","Debit – Requires Review"),
                row.get("attribution",""), row.get("counterparty",""),
                float(row.get("confidence",0)), row.get("source",""),
                int(row.get("cluster_id",-1)), int(row.get("is_anomaly",0)),
                row.get("note",""), now))

    def get_txns(self, user_id: str, limit=1000) -> List[Dict]:
        with self._conn() as c:
            rows = c.execute("""SELECT * FROM ledger WHERE user_id=?
                               ORDER BY txn_date DESC, created_at DESC LIMIT ?""",
                             (user_id, limit)).fetchall()
            return [dict(r) for r in rows]

    def get_users(self) -> List[Dict]:
        with self._conn() as c:
            return [dict(r) for r in c.execute("SELECT * FROM users ORDER BY created_at DESC")]

    def get_ledger_summary(self, user_id: str) -> List[Dict]:
        with self._conn() as c:
            rows = c.execute("""SELECT book, section, grp, account, txn_type,
                               SUM(amount) as total, COUNT(*) as cnt
                               FROM ledger WHERE user_id=?
                               GROUP BY book, section, grp, account, txn_type
                               ORDER BY book, section, grp, account""",
                             (user_id,)).fetchall()
            return [dict(r) for r in rows]


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL PARSER
# ══════════════════════════════════════════════════════════════════════════════

def _parse_xlsx_xml(file_bytes: bytes) -> List[List[Any]]:
    """
    Low-level XLSX reader using raw XML parsing.
    Works even when openpyxl fails (e.g. non-standard purl.oclc.org namespace used by HDFC exports).
    Returns list of rows from the sheet with the most transaction-like content.
    Each row is a list of cell values (str or float). Sparse columns are None.

    Shared strings: XLSX stores unique strings in xl/sharedStrings.xml and
    references them by integer index. Each <si> element contains either a
    single <t> (plain string) or multiple <r><t> sub-elements (rich text).

    Worksheet rows: xl/worksheets/sheet*.xml contains <row> elements each
    holding <c> (cell) elements.  The 'r' attribute on <c> gives the cell
    reference (e.g. "B3"); 't' attribute tells the type ('s'=sharedString,
    'inlineStr'=inline text, 'n' or absent=numeric/date).

    rows_raw format: List[List[Any]] — outer list = rows, inner = column
    values left-padded with None so that column indices are stable.
    """
    import zipfile
    from xml.etree import ElementTree as ET

    try:
        z = zipfile.ZipFile(io.BytesIO(file_bytes))
    except Exception as e:
        raise ValueError(f"_parse_xlsx_xml: cannot open ZIP: {e}")

    names = z.namelist()

    # ── Shared strings (may be absent in small or formula-only workbooks) ───────
    shared: List[str] = []
    if 'xl/sharedStrings.xml' in names:
        try:
            ss_raw = z.read('xl/sharedStrings.xml')
            ss_tree = ET.fromstring(ss_raw)
            # Strip all namespace prefixes so we can use plain tag names
            for el in ss_tree.iter():
                el.tag = re.sub(r'\{[^}]+\}', '', el.tag)
            for si in ss_tree.findall('.//si'):
                t_el = si.find('t')
                if t_el is not None:
                    shared.append(t_el.text or '')
                else:
                    # Rich text: concatenate all <r><t> fragments
                    shared.append(''.join(t.text or '' for t in si.findall('.//t')))
        except Exception as e:
            print(f"  [XML] sharedStrings.xml parse warning: {e} — continuing without shared strings")

    # ── Collect all worksheet paths ───────────────────────────────────────────
    sheet_paths = sorted(
        [n for n in names if re.match(r'xl/worksheets/sheet\d+\.xml', n)],
        key=lambda x: int(re.search(r'\d+', x.split('/')[-1]).group())
    )
    if not sheet_paths:
        raise ValueError("_parse_xlsx_xml: no worksheet XML files found in workbook")

    print(f"  [XML] sheets found in ZIP: {sheet_paths}")

    def col_idx(col_str: str) -> int:
        """Convert Excel column letter(s) to 0-based index. A→0, B→1, Z→25, AA→26."""
        idx = 0
        for ch in col_str.upper():
            idx = idx * 26 + (ord(ch) - ord('A') + 1)
        return idx - 1

    def _parse_sheet_rows(sheet_path: str) -> List[List[Any]]:
        sheet_raw = z.read(sheet_path)
        sheet_tree = ET.fromstring(sheet_raw)
        for el in sheet_tree.iter():
            el.tag = re.sub(r'\{[^}]+\}', '', el.tag)
        rows_out = []
        for row_el in sheet_tree.findall('.//row'):
            cells_dict: Dict[int, Any] = {}
            max_col = 0
            for c in row_el.findall('c'):
                ref = c.get('r', '')
                col_letters = ''.join(filter(str.isalpha, ref))
                if not col_letters:
                    continue
                cidx = col_idx(col_letters)
                max_col = max(max_col, cidx)
                t_attr = c.get('t', '')
                v_el = c.find('v')
                is_el = c.find('is')   # inlineStr container

                if t_attr == 'inlineStr' and is_el is not None:
                    # Inline string: text lives in <is><t>...</t></is>
                    val = ''.join(t.text or '' for t in is_el.findall('.//t'))
                elif t_attr == 's' and v_el is not None:
                    # Shared string reference
                    try:
                        val = shared[int(v_el.text)]
                    except (IndexError, ValueError, TypeError):
                        val = v_el.text  # fall back to raw index text
                elif v_el is not None:
                    # Numeric or date serial; keep as float if possible
                    try:
                        val = float(v_el.text)
                    except (ValueError, TypeError):
                        val = v_el.text
                else:
                    val = None
                cells_dict[cidx] = val
            if cells_dict:
                row = [cells_dict.get(i) for i in range(max_col + 1)]
                rows_out.append(row)
        return rows_out

    # ── Try all sheets; pick the one with the most non-empty rows ────────────
    best_rows: List[List[Any]] = []
    best_sheet = ""
    last_exc: Optional[Exception] = None

    for sp in sheet_paths:
        try:
            rows = _parse_sheet_rows(sp)
            non_empty = sum(
                1 for r in rows
                if any(c is not None and str(c).strip() for c in r)
            )
            print(f"  [XML] {sp}: {len(rows)} rows, {non_empty} non-empty")
            if non_empty > sum(
                1 for r in best_rows
                if any(c is not None and str(c).strip() for c in r)
            ):
                best_rows = rows
                best_sheet = sp
        except Exception as e:
            print(f"  [XML] {sp}: parse error — {e}")
            last_exc = e

    if not best_rows:
        exc_msg = f" Last error: {last_exc}" if last_exc else ""
        raise ValueError(f"_parse_xlsx_xml: no readable rows found in any worksheet.{exc_msg}")

    print(f"  [XML] selected sheet: {best_sheet} ({len(best_rows)} rows)")
    return best_rows



def _infer_txn_type(narration: str) -> str:
    """
    Infer credit/debit from narration text when no amount columns are present.
    HDFC bank statements embed CR/DR in the narration prefix.
    """
    u = narration.upper().strip()
    u_stripped = re.sub(r'^(?:WDL\s+TFR|DEP\s+TFR|WDL\s+CASH|DEP\s+CASH)\s+', '', u, flags=re.I)
    if u_stripped != u:
        if u.startswith('WDL'):
            return 'debit'
        if u.startswith('DEP'):
            return 'credit'
    u = u_stripped
    if re.match(r'NEFT\s+CR\b|RTGS\s+CR\b|IMPS.*-CR-|SWIFT\s+CR\b', u): return 'credit'
    if re.match(r'NEFT\s+DR\b|RTGS\s+DR\b|ACH\s+DR\b|NACH\s+DR\b|ECS\s+DR\b', u): return 'debit'
    if re.match(r'IB\s+FUNDS\s+TRANSFER\s+CR', u): return 'credit'
    if re.match(r'IB\s+FUNDS\s+TRANSFER\s+DR', u): return 'debit'
    if re.match(r'^UPI/CR/', u): return 'credit'
    if re.match(r'^UPI/DR/', u): return 'debit'
    if u.startswith('UPI-'): return 'debit'
    if re.search(r'TPT-(SAL|SALARY|REPAY|RETURN|MANI|LOAN\s+CR|TXFR\s+CR)\b', u): return 'credit'
    if re.search(r'TPT-(LOAN|TXFR|PLOT|NB[A-Z0-9]+)\b', u): return 'debit'
    if re.match(r'ACH\s*C-|ACH-CR', u): return 'credit'
    if re.search(r'INT\.PD|INT\s+PD|SB\s+\d+\s+INT|INTEREST\s+CREDIT|CREDIT\s+INTEREST|INTEREST\s+PAID\s+TILL', u): return 'credit'
    if re.search(r'\bDIV\d{2}|\bDIVIDEND\b|DIV\s*2[0-9]|DIVINTDIV|FINDIV|ANNUALDIV|ORDI\s+DIV|FINAL\s+DIV|SPL\s+INT\s+DIV|SPECIAL\s+DIV', u): return 'credit'
    if re.search(r'TATA\s+MOTORS.*DIV|HDFC\s+BANK.*DIV|UNITED\s+NIL.*DIV|UPL\s+LTD.*DIV|EIH\s+LIMITED.*DIV|BAYER.*DIV|PNB\s+HOU.*DIV|HDB\s+FINAN.*DIV', u): return 'credit'
    if re.search(r'YEIDA\s+REFUND|REFUND\s+RPS', u): return 'credit'
    if re.search(r'YAMUNAEXPRESSWAY|TOLL\s+PLAZA|GADPURI|FASTAG', u): return 'debit'
    if re.search(r'/RAZP[A-Z]+|RAZORPAY', u): return 'debit'
    if re.search(r'^TRAVELFG/|CMUPLFCM/', u): return 'debit'
    if re.search(r'HDFC\s+BANK.*INT|BANK.*INTEREST|INT\s+PAID', u): return 'credit'
    if re.search(r'NACH\s+DR|SIP\s+MANDATE|SIPMANDATEREF|^S?EBA/MFP[-/]?\d|^MFP[-/]\d', u): return 'debit'
    if re.search(r'CHQ\s+PAID|CHEQUE\s+PAID|CTS-RK', u): return 'debit'
    if re.match(r'IMPS-', u): return 'credit' if re.search(r'-CR-|-CREDIT-', u) else 'debit'
    return 'debit'


def _is_xls_bytes(file_bytes: bytes) -> bool:
    return file_bytes.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1")


def _is_xlsx_bytes(file_bytes: bytes) -> bool:
    if not file_bytes:
        return False
    if file_bytes.startswith(b"PK\x03\x04"):
        return True
    try:
        import zipfile, io
        return zipfile.is_zipfile(io.BytesIO(file_bytes))
    except Exception:
        return False


def _convert_xls_bytes_to_xlsx_bytes(file_bytes: bytes) -> bytes:
    """
    Convert legacy .xls bytes to .xlsx using LibreOffice/soffice.
    Raises a descriptive ValueError at each failure point so callers can
    surface the exact reason to the user.
    """
    # Locate executable
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        raise ValueError(
            "LibreOffice executable not found. "
            "Legacy .xls files require LibreOffice (soffice) to convert. "
            "Install LibreOffice or convert the file to .xlsx before uploading."
        )
    print(f"  [XLS→XLSX] Using LibreOffice: {soffice}")

    with tempfile.TemporaryDirectory(prefix="hni_xls_") as td:
        xls_path = os.path.join(td, "input.xls")
        with open(xls_path, "wb") as f:
            f.write(file_bytes)

        cmd = [soffice, "--headless", "--convert-to", "xlsx", "--outdir", td, xls_path]
        print(f"  [XLS→XLSX] Running: {' '.join(cmd)}")
        try:
            proc = subprocess.run(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                timeout=60,
            )
        except subprocess.TimeoutExpired:
            raise ValueError("LibreOffice conversion timed out after 60 seconds.")
        except Exception as e:
            raise ValueError(f"LibreOffice conversion subprocess error: {e}")

        if proc.returncode != 0:
            detail = (proc.stderr or proc.stdout or "").strip()
            raise ValueError(
                f"LibreOffice conversion failed with exit code {proc.returncode}. "
                f"stderr/stdout: {detail or '(no output)'}"
            )

        out_path = os.path.join(td, "input.xlsx")
        if not os.path.exists(out_path):
            detail = (proc.stderr or proc.stdout or "").strip()
            raise ValueError(
                f"Converted xlsx file was not created at {out_path}. "
                f"LibreOffice output: {detail or '(no output)'}"
            )

        size = os.path.getsize(out_path)
        if size == 0:
            raise ValueError(
                f"LibreOffice produced an empty xlsx file (0 bytes). "
                "The source .xls may be corrupt or password-protected."
            )

        print(f"  [XLS→XLSX] Conversion succeeded — output {size} bytes")
        with open(out_path, "rb") as f:
            return f.read()

# ══════════════════════════════════════════════════════════════════════════════
# PDF SUPPORT (NEW) — Bank-agnostic PDF statement parser
# ══════════════════════════════════════════════════════════════════════════════
# Install deps:
#   pip install pdfplumber pymupdf pdfminer.six pytesseract
#   For OCR: apt-get install tesseract-ocr  (Linux)
#             brew install tesseract          (macOS)
# ══════════════════════════════════════════════════════════════════════════════

# ── PDF SUPPORT (NEW) ── Check PDF magic bytes
def _is_pdf_bytes(file_bytes: bytes) -> bool:
    """Returns True if file_bytes starts with the %PDF magic header."""
    return file_bytes[:4] == b'%PDF'


def _try_decrypt_pdf_bytes(file_bytes: bytes, password: Optional[str] = None) -> bytes:
    """Decrypt password protected PDFs when a password is supplied."""
    if not _is_pdf_bytes(file_bytes):
        return file_bytes
    try:
        from pypdf import PdfReader, PdfWriter  # type: ignore
    except Exception:
        return file_bytes
    try:
        reader = PdfReader(io.BytesIO(file_bytes))
    except Exception:
        return file_bytes
    if not getattr(reader, 'is_encrypted', False):
        return file_bytes
    if not password:
        raise ValueError('PDF is password-protected. Enter the password in the upload box and try again.')
    try:
        decrypt_result = reader.decrypt(password)
    except Exception as e:
        raise ValueError(f'PDF password decryption failed: {e}')
    if decrypt_result == 0:
        raise ValueError('Incorrect PDF password. Please check the password and try again.')
    writer = PdfWriter()
    for page in reader.pages:
        writer.add_page(page)
    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


# ── PDF SUPPORT (NEW) ── pdfplumber table extractor
def _extract_pdf_tables_pdfplumber(file_bytes: bytes) -> List[List[List[str]]]:
    """
    Extract tables from all PDF pages using pdfplumber.
    Returns a list-of-pages where each page is a list-of-rows,
    and each row is a list of cell strings.
    """
    try:
        import pdfplumber  # type: ignore
    except ImportError:
        return []

    all_page_tables: List[List[List[str]]] = []
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            if pdf.metadata.get("Encrypted"):
                raise ValueError("PDF is password-protected. Please remove the password before uploading.")
            for page in pdf.pages:
                tables = page.extract_tables()
                for tbl in (tables or []):
                    for row in (tbl or []):
                        clean = [str(c or "").strip() for c in row]
                        all_page_tables.append(clean)
    except ValueError:
        raise
    except Exception:
        pass
    return all_page_tables


# ── PDF SUPPORT (NEW) ── PyMuPDF text extractor
def _extract_pdf_text_fitz(file_bytes: bytes) -> List[str]:
    """
    Extract raw text per page using PyMuPDF (fitz).
    Returns a list of strings, one per page.
    """
    try:
        import fitz  # type: ignore  (PyMuPDF)
    except ImportError:
        return []

    pages_text: List[str] = []
    try:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        if doc.is_encrypted:
            raise ValueError("PDF is password-protected. Please remove the password before uploading.")
        for page in doc:
            pages_text.append(page.get_text("text"))
    except ValueError:
        raise
    except Exception:
        pass
    return pages_text


# ── PDF SUPPORT (NEW) ── pdfminer text extractor
def _extract_pdf_text_pdfminer(file_bytes: bytes) -> List[str]:
    """
    Extract text per page using pdfminer.six.
    Returns a list of strings, one per page.
    Uses extract_pages() (element-level) first, which handles more encodings
    than the stream-based extract_text_to_fp().
    """
    try:
        from pdfminer.high_level import extract_pages  # type: ignore
        from pdfminer.layout import LTTextContainer    # type: ignore
        from pdfminer.layout import LAParams           # type: ignore
    except ImportError:
        return []

    pages_text: List[str] = []

    # ── Method A: element-level extraction (handles more custom fonts) ─────
    try:
        for page_layout in extract_pages(io.BytesIO(file_bytes), laparams=LAParams()):
            lines: List[str] = []
            for element in page_layout:
                if isinstance(element, LTTextContainer):
                    lines.append(element.get_text())
            page_str = "".join(lines)
            if page_str.strip():
                pages_text.append(page_str)
    except Exception:
        pass

    if pages_text:
        return pages_text

    # ── Method B: stream-based extraction (original fallback) ─────────────
    try:
        from pdfminer.high_level import extract_text_to_fp  # type: ignore
        buf = io.StringIO()
        extract_text_to_fp(io.BytesIO(file_bytes), buf, laparams=LAParams())
        full = buf.getvalue()
        parts = full.split("\x0c")
        pages_text = [p for p in parts if p.strip()]
    except Exception:
        pass

    return pages_text


# ── PDF SUPPORT (NEW) ── pdfplumber raw-text extractor (separate from table mode)
def _extract_pdf_text_pdfplumber(file_bytes: bytes) -> List[str]:
    """
    Extract raw text per page using pdfplumber (no table heuristics).
    This often decodes custom bank fonts better than pdfminer because
    pdfplumber uses pdfminer internally but applies its own glyph mapping.
    Returns a list of strings, one per page.
    """
    try:
        import pdfplumber  # type: ignore
    except ImportError:
        return []

    pages_text: List[str] = []
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            if pdf.metadata.get("Encrypted"):
                raise ValueError("PDF is password-protected. Please remove the password before uploading.")
            for page in pdf.pages:
                text = page.extract_text(x_tolerance=3, y_tolerance=3) or ""
                if text.strip():
                    pages_text.append(text)
    except ValueError:
        raise
    except Exception:
        pass
    return pages_text


# ── PDF SUPPORT (NEW) ── OCR fallback using pytesseract
def _extract_pdf_text_ocr(file_bytes: bytes) -> List[str]:
    """
    Convert PDF pages to images and run OCR via pytesseract.
    Last-resort for scanned/image-only PDFs.
    Returns a list of strings, one per page.
    """
    try:
        import fitz          # type: ignore
        import pytesseract   # type: ignore
        from PIL import Image  # type: ignore
    except ImportError:
        return []

    pages_text: List[str] = []
    try:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        if doc.is_encrypted:
            raise ValueError("PDF is password-protected. Please remove the password before uploading.")
        for page in doc:
            mat = fitz.Matrix(2.0, 2.0)   # 2× zoom → better OCR accuracy
            pix = page.get_pixmap(matrix=mat)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            text = pytesseract.image_to_string(img, lang="eng")
            pages_text.append(text)
    except ValueError:
        raise
    except Exception:
        pass
    return pages_text


# ── PDF SUPPORT (NEW) ── Regex-based row reconstruction from raw text
_DATE_PAT  = re.compile(
    r"""
    \b
    (?:
        \d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4}       # DD/MM/YYYY  DD.MM.YYYY  D/M/YYYY
      | \d{1,2}[\s\-]\w{3}[\s\-]\d{2,4}            # DD Mon YYYY / DD-Mon-YY
      | \d{4}[/\-]\d{2}[/\-]\d{2}                  # YYYY-MM-DD
      | \d{8}                                       # DDMMYYYY compact
    )
    \b
    """,
    re.VERBOSE | re.IGNORECASE,
)

_AMT_PAT = re.compile(
    r"""
    (?:[\u20b9Rs\.]+\s*)?                           # optional ₹ / Rs prefix
    \d{1,3}(?:,\d{2,3})*(?:\.\d{2})?               # 1,23,456.78 or 1234.56
    |\d+\.\d{2}                                     # plain 0.00
    """,
    re.VERBOSE,
)


def _parse_text_to_rows(page_text: str) -> List[List[str]]:
    """
    Reconstruct tabular rows from space-aligned plain text.

    Two-pass strategy:
      Pass 1 — Date-anchored: lines that START with a date begin a new row.
               Continuation lines (no leading date) are merged into the
               previous row's narration.
      Pass 2 — Date-anywhere: if Pass 1 finds < 2 rows, scan every line for
               an embedded date token and split around it.  This handles
               formats where date appears mid-line or after a ref number.

    For each row the output columns are:
      [date_str, narration, amt1, amt2, ...]
    """
    rows: List[List[str]] = []

    # ── Pass 1: date at start of line ────────────────────────────────────────
    current: List[str] = []
    for raw_line in page_text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        date_match = _DATE_PAT.match(line)
        # ICICI: line may start with S.No ("1 01.04.2025 ...")
        if date_match is None:
            sno = re.match(r'^\d{1,4}\s+', line)
            if sno:
                date_match = _DATE_PAT.match(line, sno.end())
        if date_match:
            if current:
                rows.append(current)
            date_str  = date_match.group()
            remainder = line[date_match.end():].strip()
            amounts   = _AMT_PAT.findall(remainder)
            narration = _AMT_PAT.sub("", remainder).strip()
            narration = re.sub(r"\s{2,}", " ", narration)
            current   = [date_str, narration] + amounts
        else:
            if current:
                narration_part = _AMT_PAT.sub("", line).strip()
                if narration_part:
                    current[1] = (current[1] + " " + narration_part).strip()
                extra_amts = _AMT_PAT.findall(line)
                current.extend(extra_amts)
    if current:
        rows.append(current)

    if len(rows) >= 2:
        return rows

    # ── Pass 2: date anywhere in line (fallback) ──────────────────────────────
    rows = []
    for raw_line in page_text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        date_match = _DATE_PAT.search(line)
        if not date_match:
            continue
        date_str  = date_match.group()
        # Everything outside the date token
        before    = line[:date_match.start()].strip()
        after     = line[date_match.end():].strip()
        combined  = (before + " " + after).strip()
        amounts   = _AMT_PAT.findall(combined)
        narration = _AMT_PAT.sub("", combined).strip()
        narration = re.sub(r"\s{2,}", " ", narration)
        rows.append([date_str, narration] + amounts)

    if len(rows) >= 2:
        return rows

    # ── Pass 3: split entire page on whitespace boundaries into token rows ────
    # Last resort for very tightly packed text with no clear row structure.
    # Emit one "row" per line that has at least one date AND one amount token.
    rows = []
    for raw_line in page_text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if not _DATE_PAT.search(line):
            continue
        amounts = _AMT_PAT.findall(line)
        if not amounts:
            continue
        date_str  = _DATE_PAT.search(line).group()
        narration = _DATE_PAT.sub("", _AMT_PAT.sub("", line)).strip()
        narration = re.sub(r"\s{2,}", " ", narration)
        rows.append([date_str, narration] + amounts)

    return rows


# ── PDF SUPPORT (NEW) ── Amount normaliser
def _normalise_amount(raw: str) -> float:
    """
    Parse Indian-format amount strings to float.
    Handles: "1,23,456.78", "₹ 500.00", "Rs.1000", blanks, dashes.
    """
    if not raw or str(raw).strip() in ("", "-", "–", "—", "nil", "NIL"):
        return 0.0
    # Strip currency symbols and whitespace
    s = re.sub(r"[^\d.,\-]", "", str(raw))
    # Remove thousand separators (Indian grouping uses commas)
    s = s.replace(",", "")
    if not s or s == "-":
        return 0.0
    try:
        return abs(float(s))
    except ValueError:
        return 0.0


# ── PDF SUPPORT (NEW) ── Transaction type detector
def _detect_txn_type(
    debit_val: str,
    credit_val: str,
    type_col_val: str = "",
) -> str:
    """
    Determine 'debit' or 'credit'.

    Priority:
      1. Explicit DR/CR column (e.g. ICICI single-amount + Dr/Cr suffix).
      2. Non-zero debit cell  → debit.
      3. Non-zero credit cell → credit.
      4. Fallback → debit.
    """
    tc = (type_col_val or "").strip().lower()
    if tc in ("cr", "credit"):
        return "credit"
    if tc in ("dr", "debit"):
        return "debit"

    dr = _normalise_amount(debit_val)
    cr = _normalise_amount(credit_val)

    # ICICI-style: amount column may have "Dr" / "Cr" suffix baked in
    dv = str(debit_val or "").strip()
    cv = str(credit_val or "").strip()
    if re.search(r'\bCr\.?$', dv, re.I):
        return "credit"
    if re.search(r'\bDr\.?$', dv, re.I):
        return "debit"

    if dr > 0 and cr == 0:
        return "debit"
    if cr > 0 and dr == 0:
        return "credit"
    return "debit"


# ── PDF SUPPORT (NEW) ── Date normaliser (ISO output)
def _normalise_date(raw: str) -> str:
    """
    Parse a date string in any Indian bank format and return YYYY-MM-DD.
    Supports:
      DD/MM/YYYY  DD-MM-YYYY  DD/MM/YY
      DD Mon YYYY  DD-Mon-YYYY  DD-Mon-YY
      YYYY-MM-DD
      D/M/YYYY  (no leading zeros)
      DDMMYYYY  (compact 8-digit)
    """
    if not raw:
        return datetime.now().strftime("%Y-%m-%d")
    s = str(raw).strip()
    formats = [
        "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y",
        "%d %b %Y", "%d-%b-%Y", "%d %b %y", "%d-%b-%y",
        "%Y-%m-%d", "%Y/%m/%d",
        "%-d/%-m/%Y",   # no leading zero (Linux/Mac)
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except (ValueError, TypeError):
            pass
    # Compact DDMMYYYY
    if re.fullmatch(r'\d{8}', s):
        try:
            return datetime.strptime(s, "%d%m%Y").strftime("%Y-%m-%d")
        except ValueError:
            pass
    # Partial match — extract digits
    m = re.search(r'(\d{1,2})[/\-\s](\d{1,2}|\w{3})[/\-\s](\d{2,4})', s)
    if m:
        candidate = f"{m.group(1)} {m.group(2)} {m.group(3)}"
        for fmt in ("%d %m %Y", "%d %m %y", "%d %b %Y", "%d %b %y"):
            try:
                return datetime.strptime(candidate, fmt).strftime("%Y-%m-%d")
            except (ValueError, TypeError):
                pass
    return datetime.now().strftime("%Y-%m-%d")


# ── PDF SUPPORT (NEW) ── Header-row detector
_HEADER_DATE_WORDS  = {"date", "txn date", "value date", "posting date", "tran date",
                        "transaction date", "trans date", "chq date"}
_HEADER_DESC_WORDS  = {"narration", "description", "particulars", "details",
                        "remarks", "transaction details", "trans details",
                        "transaction narration"}
_HEADER_DEBIT_WORDS = {"debit", "withdrawal", "withdrawals", "dr", "debit amount",
                        "withdrawal amt", "withdrawal amt.", "dr amount"}
_HEADER_CREDIT_WORDS = {"credit", "deposit", "deposits", "cr", "credit amount",
                         "deposit amt", "deposit amt.", "cr amount"}
_HEADER_BALANCE_WORDS = {"balance", "closing balance", "running balance",
                          "available balance", "bal"}


def _detect_header_row(rows: List[List[str]]) -> Optional[int]:
    """
    Scan rows for the header row by checking that it contains recognisable
    keywords for date, description, and at least one amount concept.
    Returns the 0-based row index, or None if not found.
    """
    for i, row in enumerate(rows):
        cells = {str(c or "").strip().lower() for c in row}

        has_date   = bool(cells & _HEADER_DATE_WORDS or
                          any("date" in c for c in cells))
        has_desc   = bool(cells & _HEADER_DESC_WORDS or
                          any(w in c for c in cells
                              for w in ("narrat", "descri", "particu", "detail", "remark")))
        has_amount = bool(
            (cells & _HEADER_DEBIT_WORDS or any("withdraw" in c or "debit" in c for c in cells)) and
            (cells & _HEADER_CREDIT_WORDS or any("deposit" in c or "credit" in c for c in cells))
        ) or bool(
            any("amount" in c for c in cells)
        )

        if has_date and has_desc and has_amount:
            return i
    return None


def _map_columns(header_row: List[str]) -> Dict[str, Optional[int]]:
    """
    Build a positional column-index mapping from a detected header row.
    Returns dict with keys: date, desc, debit, credit, amount, type, balance.
    """
    h = [str(c or "").strip().lower() for c in header_row]

    def find(word_set, extra_check=None) -> Optional[int]:
        # Exact match first
        for i, cell in enumerate(h):
            if cell in word_set:
                return i
        # Substring match
        for i, cell in enumerate(h):
            if any(w in cell for w in word_set):
                return i
        if extra_check:
            for i, cell in enumerate(h):
                if extra_check(cell):
                    return i
        return None

    date_col    = find(_HEADER_DATE_WORDS)
    desc_col    = find(_HEADER_DESC_WORDS)
    debit_col   = find(_HEADER_DEBIT_WORDS)
    credit_col  = find(_HEADER_CREDIT_WORDS)
    balance_col = find(_HEADER_BALANCE_WORDS)
    type_col    = find({"type", "txn type", "dr/cr", "dr / cr", "transaction type", "cr/dr"})

    # Fallback: single "amount" column
    amount_col: Optional[int] = None
    if debit_col is None and credit_col is None:
        for i, cell in enumerate(h):
            if "amount" in cell or cell in ("amt", "txn amount", "net amount"):
                amount_col = i
                break

    return {
        "date":    date_col,
        "desc":    desc_col,
        "debit":   debit_col,
        "credit":  credit_col,
        "amount":  amount_col,
        "type":    type_col,
        "balance": balance_col,
    }


def _is_junk_row(row: List[str], col_map: Dict[str, Optional[int]]) -> bool:
    """
    Return True for rows that should be skipped:
    repeated headers, page-number footers, opening/closing balance lines, etc.
    """
    joined = " ".join(str(c or "") for c in row).strip().lower()
    if not joined:
        return True
    junk_patterns = [
        r"^\s*page\s+\d",
        r"continued\s+on\s+next",
        r"opening\s+balance",
        r"closing\s+balance",
        r"brought\s+forward",
        r"b/?f\b",
        r"total\s+credit",
        r"total\s+debit",
        r"statement\s+summary",
        r"micr\s+code",
        r"ifsc\s+code",
        r"^\s*[\*\-=_]+\s*$",
    ]
    for pat in junk_patterns:
        if re.search(pat, joined):
            return True

    # Repeated header row detection
    date_cell = str(row[col_map["date"]] if col_map.get("date") is not None
                    and col_map["date"] < len(row) else "").lower()
    if date_cell in _HEADER_DATE_WORDS or "date" in date_cell:
        return True

    return False


def _parse_pdf_multiline_bank_statement(raw_text_pages: List[str]) -> Dict[str, Any]:
    """
    Parse multiline PDF statements of this shape:

      01-03-2026 B/F 4,794.26
      01-03-2026
      UPI/Brijesh /9029183746@pty/NO REMARKS/INDIAN
      BAN/200206884733/UPI4fe947...
      50.00 4,744.26

    Also handles one-line entries like:
      09-03-2026 ICICI EBA/MFP-... 100.00 4,539.26

    Returns:
      {
        "records": [...],
        "opening_balance": float | None,
      }
    """
    import re

    DATE_PAT = re.compile(r'^\s*(\d{2}[-/]\d{2}[-/]\d{4})\s*(.*)$')
    AMT_BAL_PAT = re.compile(r'^\s*([0-9,]+\.\d{2})\s+([0-9,]+\.\d{2})\s*$')
    ONE_LINE_PAT = re.compile(r'^\s*(\d{2}[-/]\d{2}[-/]\d{4})\s+(.+?)\s+([0-9,]+\.\d{2})\s+([0-9,]+\.\d{2})\s*$')
    BF_PAT = re.compile(r'^\s*(\d{2}[-/]\d{2}[-/]\d{4})\s+(?:B/?F|B\.F\.|BROUGHT\s+FORWARD|OPENING\s+BALANCE)\s+([0-9,]+\.\d{2})\s*$', re.I)

    def _amt(x: str) -> float:
        return float(str(x).replace(",", "").strip())

    def _norm_date(x: str) -> str:
        x = x.strip().replace("/", "-")
        dd, mm, yyyy = x.split("-")
        return f"{yyyy}-{mm}-{dd}"

    def _is_noise(line: str) -> bool:
        s = line.strip()
        if not s:
            return True
        su = s.upper()
        if re.match(r'^(PAGE\s+\d+|TOTAL:|MICR\s+CODE|IFSC\s+CODE|STATEMENT\s+SUMMARY|CORPORATE\s+OFFICE:)', su):
            return True
        if re.match(r'^[*\-_=]{4,}$', s):
            return True
        return False

    records = []
    opening_balance = None
    prev_balance = None    

    current_date = None
    current_desc_lines: List[str] = []

    def _flush(amount: float, balance: float):
        nonlocal current_date, current_desc_lines, prev_balance, records, opening_balance
        if not current_date:
            return

        desc = " ".join(x.strip() for x in current_desc_lines if x.strip())
        desc = re.sub(r'\s+', ' ', desc).strip(" /|-")

        if not desc:
            current_date = None
            current_desc_lines = []
            prev_balance = balance
            return

        # infer direction from balance movement
        if prev_balance is None:
           txn_type = _infer_txn_type(desc)
        else:
            if balance < prev_balance:
                txn_type = "debit"
            elif balance > prev_balance:
                txn_type = "credit"
            else:
                txn_type = _infer_txn_type(desc)

        records.append({
            "txn_date": _norm_date(current_date),
            "description": desc,
            "narration": desc,
            "amount": amount,
            "txn_type": txn_type,
            "balance": balance,
            "raw_debit": amount if txn_type == "debit" else None,
            "raw_credit": amount if txn_type == "credit" else None,
        })

        if opening_balance is None and prev_balance is None:
            # first transaction
            if txn_type == "debit":
                opening_balance = balance + amount
            else:
                opening_balance = balance - amount
        prev_balance = balance
        current_date = None
        current_desc_lines = []

    for page_text in raw_text_pages:
        for raw_line in page_text.splitlines():
            line = re.sub(r'\s+', ' ', str(raw_line or "")).strip()
            if _is_noise(line):
                continue

            # Opening balance line
            m_bf = BF_PAT.match(line)
            if m_bf:
                current_date = None
                current_desc_lines = []
                opening_balance = _amt(m_bf.group(2))
                prev_balance = opening_balance
                continue

            # Full one-line transaction
            m_one = ONE_LINE_PAT.match(line)
            if m_one and " B/F " not in f" {line.upper()} ":
                dt = m_one.group(1)
                desc = m_one.group(2).strip()
                amt = _amt(m_one.group(3))
                bal = _amt(m_one.group(4))

                if prev_balance is None:
                    txn_type = _infer_txn_type(desc)
                else:
                    if bal < prev_balance:
                        txn_type = "debit"
                    elif bal > prev_balance:
                        txn_type = "credit"
                    else:
                        txn_type = _infer_txn_type(desc)

                records.append({
                    "txn_date": _norm_date(dt),
                    "description": desc,
                    "narration": desc,
                    "amount": amt,
                    "txn_type": txn_type,
                    "balance": bal,
                    "raw_debit": amt if txn_type == "debit" else None,
                    "raw_credit": amt if txn_type == "credit" else None,
                })

                if opening_balance is None and prev_balance is None:
                    if txn_type == "debit":
                        opening_balance = bal + amt
                    else:
                        opening_balance = bal - amt
                prev_balance = bal
                current_date = None
                current_desc_lines = []
                continue

            # New dated line
            m_date = DATE_PAT.match(line)
            if m_date:
                # if another txn was in progress but not closed, drop it
                current_date = m_date.group(1)
                tail = m_date.group(2).strip()
                current_desc_lines = [tail] if tail else []
                continue

            # Amount + balance terminator line
            m_amt_bal = AMT_BAL_PAT.match(line)
            if m_amt_bal and current_date:
                amt = _amt(m_amt_bal.group(1))
                bal = _amt(m_amt_bal.group(2))
                _flush(amt, bal)
                continue

            # narration continuation
            if current_date:
                current_desc_lines.append(line)

    return {
        "records": records,
        "opening_balance": opening_balance,
        "closing_balance": _resolve_statement_closing_balance(
            records, opening_balance, None, "_parse_pdf_multiline_bank_statement"
        ),
    }

def _parse_axis_pdf_text(raw_text_pages: List[str]) -> Dict[str, Any]:
    """
    Axis PDF text parser.
    Typical pattern:
      01-03-2026 01-03-2026 UPI-ABC-abc@okhdfcbank-123456 250.00 Dr 12000.50
    """
    import re

    line_pat = re.compile(
        r'^\s*(\d{2}[-/\.]\d{2}[-/\.]\d{4})\s+'
        r'(?:\d{2}[-/\.]\d{2}[-/\.]\d{4}\s+)?'
        r'(.+?)\s+'
        r'([0-9,]+\.\d{2})\s+'
        r'(Dr|Cr|DR|CR)\s+'
        r'([0-9,]+\.\d{2})\s*$'
    )

    records = []
    opening_balance = None

    def _amt(x):
        return float(str(x).replace(",", "").strip())

    def _norm_date(x):
        x = x.replace(".", "-").replace("/", "-")
        dd, mm, yyyy = x.split("-")
        return f"{yyyy}-{mm}-{dd}"

    for page_text in raw_text_pages:
        for raw_line in page_text.splitlines():
            line = re.sub(r"\s+", " ", str(raw_line or "")).strip()
            m = line_pat.match(line)
            if not m:
                continue

            dt = _norm_date(m.group(1))
            desc = m.group(2).strip()
            amt = _amt(m.group(3))
            typ = m.group(4).lower()
            bal = _amt(m.group(5))
            txn_type = "credit" if typ == "cr" else "debit"

            records.append({
                "txn_date": dt,
                "description": desc,
                "narration": desc,
                "amount": amt,
                "txn_type": txn_type,
                "balance": bal,
                "raw_debit": amt if txn_type == "debit" else None,
                "raw_credit": amt if txn_type == "credit" else None,
            })

            if opening_balance is None:
                opening_balance = bal - amt if txn_type == "credit" else bal + amt

    statement_from_date = None
    statement_to_date = None
    closing_balance = _resolve_statement_closing_balance(
        records, opening_balance, None, "_parse_axis_pdf_text"
    )

    def _safe_parse_statement_date(text: Any) -> Optional[str]:
        s = str(text or "").strip()
        if not s:
            return None

        s = s.replace(".", "-").replace("/", "-")
        s = re.sub(r"\s+", "-", s)

        # Prefer day-first for Indian bank statements
        for fmt in (
            "%d-%m-%Y",
            "%d-%m-%y",
            "%d-%b-%Y",
            "%d-%b-%y",
            "%d-%B-%Y",
            "%d-%B-%y",
        ):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except Exception:
                pass

        # Explicit month-first fallback only if day-first failed
        for fmt in (
            "%m-%d-%Y",
            "%m-%d-%y",
        ):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except Exception:
                pass

        try:
            return _safe_parse_excel_date(s)
        except Exception:
            return None
    
    def _extract_statement_period_from_header_rows(header_rows: List[List[Any]]) -> Tuple[Optional[str], Optional[str]]:
        lines = []
        for row in header_rows or []:
            txt = " ".join(str(c or "").strip() for c in row if str(c or "").strip())
            if txt:
                lines.append(re.sub(r"\s+", " ", txt).strip())

        candidates: List[Tuple[str, str, str]] = []

        date_pat = r'(\d{1,2}[./-](?:\d{1,2}|[A-Za-z]{3,9})[./-]\d{2,4})'

        strong_patterns = [
            rf'(?:statement\s*period|period)\s*[:\-]?\s*{date_pat}\s*(?:to|\-)\s*[:\-]?\s*{date_pat}',
            rf'(?:from)\s*[:\-]?\s*{date_pat}\s*(?:to)\s*[:\-]?\s*{date_pat}',
        ]
        for line in lines:
            s = re.sub(r'\s+', ' ', line).strip()
            for pat in strong_patterns:
                m = re.search(pat, s, flags=re.I)
                if m:
                    d1 = _safe_parse_statement_date(m.group(1))
                    d2 = _safe_parse_statement_date(m.group(2))
                    if d1 and d2 and d1 <= d2:
                        candidates.append((d1, d2, s))

        if not candidates:
            for line in lines:
                s = re.sub(r'\s+', ' ', line).strip()
                m = re.search(
                    rf'{date_pat}\s*(?:to|\-)\s*[:\-]?\s*{date_pat}',
                    s,
                    flags=re.I
                )
                if m:
                    d1 = _safe_parse_statement_date(m.group(1))
                    d2 = _safe_parse_statement_date(m.group(2))
                    if d1 and d2 and d1 <= d2:
                        candidates.append((d1, d2, s))

        if not candidates:
            return None, None

        def _score(item):
            d1, d2, line = item
            start_day1 = 1 if d1.endswith("-01") else 0
            try:
                span_days = (datetime.strptime(d2, "%Y-%m-%d") - datetime.strptime(d1, "%Y-%m-%d")).days
            except Exception:
                span_days = -1
            return (start_day1, span_days, len(line))

        best = max(candidates, key=_score)
        print(f"  [PERIOD] Selected header period line: {best[2]}")
        return best[0], best[1]

    try:
        header_rows = (locals().get("rows") or locals().get("rows_raw") or [])[:25]
        statement_from_date, statement_to_date = _extract_statement_period_from_header_rows(header_rows)
        print(f"  [PERIOD] Header-derived period: {statement_from_date} to {statement_to_date}")
    except Exception as e:
        print(f"  [PERIOD] Header extraction failed: {e}")
        statement_from_date, statement_to_date = None, None


    return {
        "records": records,
        "opening_balance": opening_balance,
        "closing_balance": closing_balance,
        "statement_from_date": statement_from_date,
        "statement_to_date": statement_to_date,
    }

def _parse_au_pdf_text(raw_text_pages: List[str]) -> Dict[str, Any]:
    """
    AU PDF parser for common retail statement lines.
    """
    import re

    line_pat = re.compile(
        r'^\s*(\d{2}[-/\.]\d{2}[-/\.]\d{4})\s+(.+?)\s+([0-9,]+\.\d{2})(?:\s+([0-9,]+\.\d{2}))?\s+([0-9,]+\.\d{2})\s*$'
    )

    records = []
    opening_balance = None
    prev_balance = None

    def _amt(x):
        return float(str(x).replace(",", "").strip())

    def _norm_date(x):
        x = x.replace(".", "-").replace("/", "-")
        dd, mm, yyyy = x.split("-")
        return f"{yyyy}-{mm}-{dd}"

    for page_text in raw_text_pages:
        for raw_line in page_text.splitlines():
            line = re.sub(r"\s+", " ", str(raw_line or "")).strip()
            m = line_pat.match(line)
            if not m:
                continue

            dt = _norm_date(m.group(1))
            desc = _strip_au_prefix(m.group(2).strip())
            a1 = _amt(m.group(3))
            a2 = _amt(m.group(4)) if m.group(4) else 0.0
            bal = _amt(m.group(5))

            if a2 > 0:
                amt = a1 or a2
            else:
                amt = a1

            if prev_balance is None:
                txn_type = _infer_txn_type(desc)
            else:
                txn_type = "debit" if bal < prev_balance else "credit" if bal > prev_balance else _infer_txn_type(desc)

            records.append({
                "txn_date": dt,
                "description": desc,
                "narration": desc,
                "amount": amt,
                "txn_type": txn_type,
                "balance": bal,
                "raw_debit": amt if txn_type == "debit" else None,
                "raw_credit": amt if txn_type == "credit" else None,
            })

            if opening_balance is None:
                opening_balance = bal - amt if txn_type == "credit" else bal + amt

            prev_balance = bal

    return {
        "records": records,
        "opening_balance": opening_balance,
        "closing_balance": _resolve_statement_closing_balance(
            records, opening_balance, None, "_parse_au_pdf_text"
        ),
    }

def _parse_hdfc_pdf_text(raw_text_pages: List[str]) -> Dict[str, Any]:
    """
    HDFC PDF parser that extracts ALL transactions by looking for the
    exact HDFC transaction line pattern.
    """
    import re

    records: List[Dict[str, Any]] = []
    statement_from_date: Optional[str] = None
    statement_to_date: Optional[str] = None
    opening_balance: Optional[float] = None
    closing_balance: Optional[float] = None
    
    # Combine all pages into one text
    full_text = "\n".join(raw_text_pages)
    
    # Extract period
    period_pattern = re.compile(r'Statement\s+From\s*:\s*(\d{2}/\d{2}/\d{4})\s+To\s*:\s*(\d{2}/\d{2}/\d{4})', re.I)
    period_match = period_pattern.search(full_text)
    if period_match:
        statement_from_date = _norm_date_hdfc(period_match.group(1))
        statement_to_date = _norm_date_hdfc(period_match.group(2))
        print(f'[HDFC] Period: {statement_from_date} to {statement_to_date}', flush=True)
    
    # Extract opening and closing balance from summary
    summary_pattern = re.compile(
        r'Opening\s+Balance\s+([0-9,]+\.\d{2})\s+\d+\s+\d+\s+([0-9,]+\.\d{2})\s+([0-9,]+\.\d{2})\s+([0-9,]+\.\d{2})',
        re.I
    )
    summary_match = summary_pattern.search(full_text)
    if summary_match:
        opening_balance = float(summary_match.group(1).replace(',', ''))
        closing_balance = float(summary_match.group(4).replace(',', ''))
        print(f'[HDFC] Opening balance: {opening_balance}, Closing: {closing_balance}', flush=True)
    
    # HDFC transaction pattern - matches a single line with:
    # Date, Narration, Ref No, Value Date, Withdrawal Amt, Deposit Amt, Balance
    # Pattern matches lines like:
    # 02/04/24 NEFTDR-KARB0000225-MANIGARG-NETBANK, M UM-N093242963932427-FORLOCKER 02/04/24 40,000.00 21,494.67
    txn_pattern = re.compile(
        r'^(\d{2}/\d{2}/\d{2})\s+'           # Date (DD/MM/YY)
        r'([^0-9]+?)'                         # Narration (non-greedy, stop before numbers)
        r'\s+(\d{2}/\d{2}/\d{2})\s+'          # Value Date
        r'([0-9,]+\.\d{2})\s+'                # Withdrawal Amt (debit)
        r'([0-9,]+\.\d{2})\s+'                # Deposit Amt (credit)
        r'([0-9,]+\.\d{2})\s*$'               # Closing Balance
    , re.MULTILINE)
    
    # Also match lines with DR/CR indicator in narration
    txn_pattern2 = re.compile(
        r'^(\d{2}/\d{2}/\d{2})\s+'           # Date
        r'(.+?)\s+'                           # Narration
        r'\b(DR|CR)\b\s+'                     # DR/CR indicator
        r'([0-9,]+\.\d{2})\s+'                # Amount
        r'([0-9,]+\.\d{2})\s*$'               # Balance
    , re.MULTILINE | re.IGNORECASE)
    
    # Split text into lines and process each line
    lines = full_text.split('\n')
    processed_count = 0
    
    for line_num, line in enumerate(lines):
        line = line.strip()
        if not line:
            continue
        
        # Skip header/footer lines
        skip_patterns = [
            r'^Page No\.', r'^Statement of account', r'^MRS\.', r'^MANI GARG',
            r'^FLAT NO', r'^SECTOR', r'^NEAR GYM', r'^FARIDABAD', r'^HARYANA',
            r'^JOINT HOLDERS', r'^Nomination', r'^Registered', r'^Account Branch',
            r'^Address', r'^City', r'^State', r'^Phone no', r'^OD Limit',
            r'^Currency', r'^Email', r'^Cust ID', r'^Account No', r'^A/C Open Date',
            r'^Account Status', r'^RTGS/NEFT IFSC', r'^MICR', r'^Branch Code',
            r'^Account Type', r'^HDFC BANK LIMITED', r'^\*Closing balance',
            r'^Contents of this statement', r'^State account branch GSTN',
            r'^HDFC Bank GSTIN', r'^Registered Office Address', r'^Generated On',
            r'^Generated By', r'^Requesting Branch Code', r'^This is a computer generated',
            r'^STATEMENT SUMMARY', r'^Opening Balance', r'^Dr Count', r'^Cr Count',
            r'^Debits', r'^Credits', r'^Closing Bal'
        ]
        skip = False
        for pat in skip_patterns:
            if re.match(pat, line, re.I):
                skip = True
                break
        if skip:
            continue
        
        # Try pattern 1 (HDFC standard format)
        match = txn_pattern.match(line)
        if match:
            raw_date = match.group(1)
            narration_raw = match.group(2).strip()
            # value_date = match.group(3)
            withdrawal_amt = float(match.group(4).replace(',', ''))
            deposit_amt = float(match.group(5).replace(',', ''))
            balance = float(match.group(6).replace(',', ''))
            
            if withdrawal_amt > 0:
                txn_type = 'debit'
                amount = withdrawal_amt
            else:
                txn_type = 'credit'
                amount = deposit_amt
            
            narration = _clean_hdfc_narration_simple(narration_raw)
            processed_count += 1
            
        else:
            # Try pattern 2 (with DR/CR indicator)
            match2 = txn_pattern2.match(line)
            if match2:
                raw_date = match2.group(1)
                narration_raw = match2.group(2).strip()
                dr_cr = match2.group(3).upper()
                amount = float(match2.group(4).replace(',', ''))
                balance = float(match2.group(5).replace(',', ''))
                
                txn_type = 'debit' if dr_cr == 'DR' else 'credit'
                narration = _clean_hdfc_narration_simple(narration_raw)
                processed_count += 1
            else:
                # Try to find date at start of line and extract manually
                date_match = re.match(r'^(\d{2}/\d{2}/\d{2})\s+', line)
                if date_match:
                    raw_date = date_match.group(1)
                    remaining = line[date_match.end():].strip()
                    
                    # Look for amount patterns
                    amounts = re.findall(r'([0-9,]+\.\d{2})', remaining)
                    if len(amounts) >= 2:
                        withdrawal_amt = float(amounts[0].replace(',', ''))
                        deposit_amt = float(amounts[1].replace(',', '')) if len(amounts) > 1 else 0
                        balance = float(amounts[-1].replace(',', '')) if len(amounts) > 2 else None
                        
                        if withdrawal_amt > 0 and withdrawal_amt > deposit_amt:
                            txn_type = 'debit'
                            amount = withdrawal_amt
                        elif deposit_amt > 0:
                            txn_type = 'credit'
                            amount = deposit_amt
                        else:
                            continue
                        
                        narration_raw = remaining
                        for amt in amounts:
                            narration_raw = narration_raw.replace(amt, '')
                        narration_raw = re.sub(r'\s+', ' ', narration_raw).strip()
                        narration = _clean_hdfc_narration_simple(narration_raw)
                        processed_count += 1
                    else:
                        continue
                else:
                    continue
        
        # Convert date
        try:
            parts = raw_date.split('/')
            if len(parts) == 3:
                dd, mm, yy = parts
                yyyy = '20' + yy if len(yy) == 2 else yy
                txn_date = f"{yyyy}-{mm}-{dd}"
            else:
                txn_date = raw_date
        except:
            txn_date = raw_date
        
        records.append({
            'txn_date': txn_date,
            'description': narration,
            'narration': narration,
            'amount': amount,
            'txn_type': txn_type,
            'balance': balance if balance else None,
            'raw_debit': amount if txn_type == 'debit' else None,
            'raw_credit': amount if txn_type == 'credit' else None,
            'source': 'pdf',
        })
    
    print(f'[HDFC] Processed {processed_count} lines, extracted {len(records)} transactions', flush=True)
    
    # Print sample
    print('[HDFC] Sample transactions:')
    for i, r in enumerate(records[:20]):
        print(f'  [{i}] {r["txn_date"]} | {r["txn_type"]:6} | {r["amount"]:>12,.2f} | {r["narration"]}')
    
    return {
        'records': records,
        'opening_balance': opening_balance,
        'closing_balance': _resolve_statement_closing_balance(
            records, opening_balance, closing_balance, "_parse_hdfc_pdf_text"
        ),
        'statement_from_date': statement_from_date,
        'statement_to_date': statement_to_date,
    }


def _clean_hdfc_narration_simple(narration: str) -> str:
    """
    Preserve the real HDFC narration while removing only obvious ref junk.
    Do NOT collapse to generic labels like SALARY / ZERODHA / TPT.
    """
    if not narration:
        return "Unknown"

    s = str(narration).strip()

    # normalize whitespace only
    s = re.sub(r'\s+', ' ', s).strip()

    # remove trailing long refs and masked ids, but keep meaningful words
    s = re.sub(r'\b(?:UTR|RRN|REF|TXN|TXNID|SEQ|CHQ|CHEQUE|NETBANK)\s*[:#-]?\s*[A-Z0-9-]{6,}\b', ' ', s, flags=re.I)
    s = re.sub(r'\b[A-Z]{2,}\d{6,}\b', ' ', s)
    s = re.sub(r'(?<![A-Za-z])\d{10,}(?![A-Za-z])', ' ', s)
    s = re.sub(r'\b[X*]{2,}\d{3,6}\b', ' ', s)

    # keep separators readable
    s = re.sub(r'[_|]+', ' ', s)
    s = re.sub(r'\s*[/]+\s*', '/', s)
    s = re.sub(r'\s+', ' ', s).strip(' -/')

    # Never downgrade rich narrations to generic placeholders
    if not s:
        return "Unknown"
    return s[:180]


def _norm_date_hdfc(date_str: str) -> str:
    """Convert DD/MM/YYYY to YYYY-MM-DD"""
    try:
        parts = date_str.split('/')
        if len(parts) == 3:
            dd, mm, yyyy = parts
            return f"{yyyy}-{mm}-{dd}"
    except:
        pass
    return date_str

def _extract_counterparty_from_hdfc(remaining_text: str, narration_clean: str) -> str:
    """
    Extract the actual counterparty name from HDFC transaction text.
    Examples:
      - "UTIB0002172-MANI GARG" -> "Mani Garg"
      - "UTIB0002172-SARJU GARG" -> "Sarju Garg"  
      - "TPT-MANISH GARG" -> "Manish Garg"
      - "TPT-SARJU GARG HUF" -> "Sarju Garg HUF"
      - "ICIC0002122-CHANDRA BHAN KUMAR" -> "Chandra Bhan Kumar"
    """
    import re

    generic_tokens = {
        "TPT", "LOAN", "REPAY", "RETURN", "GIFT", "TXFR", "SAL", "SALARY",
        "ACH", "NACH", "CREDIT", "DEBIT", "CR", "DR", "UNKNOWN", "TRANSFER",
        "FUNDS", "FUND", "IB", "UPI", "NEFT", "RTGS", "IMPS", "PAYROLL",
        "NETBANK", "REF", "UTR", "RRN", "ID", "SEQ",
        "BANK", "HDFC", "UTIB", "ICIC", "YESB", "SBIN", "STATE", "INDIA",
    }
    banned_values = {"tpt", "salary", "ach credit", "cr", "dr", "unknown"}

    def _format_name(value: str) -> str:
        words = []
        for word in re.split(r"\s+", str(value or "").strip()):
            if not word:
                continue
            def _format_name(value: str) -> str:
                words = []
                for word in re.split(r"\s+", str(value or "").strip()):
                    if not word:
                        continue
                    if word.upper() == "HUF":
                        continue   # REMOVE HUF FROM DISPLAY
                    words.append(word.title())
    
                return " ".join(words).strip()

    def _clean_name_chunk(value: str) -> str:
        s = str(value or "").upper()
        s = re.sub(r'\b(?:UTIB|ICIC|YESB|SBIN|HDFC)\d{4,}\b', ' ', s)
        s = re.sub(r'\b[A-Z]{2,6}\d{4,}\b', ' ', s)
        s = re.sub(r'(?<![A-Z])\d{5,}(?![A-Z])', ' ', s)
        s = re.sub(r'[^A-Z\s&.]', ' ', s)
        s = re.sub(r'\s+', ' ', s).strip()
        return s

    def _human_name(value: str) -> str:
        cleaned = _clean_name_chunk(value)
        if not cleaned:
            return ""
        words = [w for w in cleaned.split() if w]
        core_words = [w for w in words if w != "HUF"]
        if len(core_words) < 2:
            return ""
        if any(w in generic_tokens for w in core_words):
            return ""
        candidate = _format_name(" ".join(words))
        return "" if candidate.lower() in banned_values else candidate

    def _final_hyphen_name(value: str) -> str:
        parts = [p.strip() for p in re.split(r'[-–]+', str(value or "")) if p.strip()]
        for part in reversed(parts):
            candidate = _human_name(part)
            if candidate:
                return candidate
        return ""

    sources = []
    for source in (
        remaining_text or "",
        narration_clean or "",
        f"{remaining_text or ''} {narration_clean or ''}".strip(),
    ):
        text = str(source or "").strip()
        if text and text not in sources:
            sources.append(text)

    explicit_patterns = [
        r'(?:UTIB|ICIC|YESB|SBIN|HDFC)\d{4,}-([A-Z][A-Z\s]+(?:\s+HUF)?)\b',
        r'TPT-(?:LOAN|REPAY|RETURN|GIFT|TXFR|SAL)(?:-[A-Z0-9]+)?-([A-Z][A-Z\s]+(?:\s+HUF)?)\b',
        r'TPT-([A-Z][A-Z\s]+(?:\s+HUF)?)\b',
    ]

    for source in sources:
        upper = str(source or "").upper().strip()
        if not upper:
            continue
        for pattern in explicit_patterns:
            match = re.search(pattern, upper)
            if match:
                candidate = _human_name(match.group(1))
                if candidate:
                    return candidate

    for source in sources:
        if "TPT" in str(source or "").upper():
            candidate = _final_hyphen_name(source)
            if candidate:
                return candidate

    for source in sources:
        upper = _clean_name_chunk(source)
        if not upper:
            continue
        words = upper.split()
        for size in range(min(4, len(words)), 1, -1):
            for start in range(0, len(words) - size + 1):
                candidate = _human_name(" ".join(words[start:start + size]))
                if candidate:
                    return candidate

    fallback = _final_hyphen_name(narration_clean or remaining_text or "")
    return fallback or "Unknown"






def parse_pdf_statement(file_bytes: bytes, filename: str = "", password: Optional[str] = None) -> Dict[str, Any]:
    """
    HDFC-focused PDF parser - the version that gave 236 transactions.
    """
    import re
    from datetime import datetime

    print(f"  [PDF:{filename}] Starting HDFC-focused parser", flush=True)

    # Filename based PNB hint
    is_pnb_file = bool(re.search(r'PNB|Punjab', filename or '', re.I))

    if is_pnb_file:
        full_text = ""

        try:
            import io
            from pypdf import PdfReader

            reader = PdfReader(io.BytesIO(file_bytes), strict=False)

            if reader.is_encrypted:
                try:
                    reader.decrypt(password or "")
                except Exception:
                    pass

            try:
                reader._override_encryption = True
            except Exception:
                pass

            pages_text = []
            for page in reader.pages:
                try:
                    pages_text.append(page.extract_text() or "")
                except Exception as pe:
                    print(f"  [PDF:PNB] page extract skipped: {pe}", flush=True)

            full_text = "\n".join(pages_text)
            print("  [PDF:PNB] pypdf force text extracted", flush=True)

        except Exception as e:
            print(f"  [PDF:PNB] pypdf force failed: {e}", flush=True)

        print("  [PDF] PNB statement detected by filename", flush=True)
        return _parse_pnb_pdf_text(full_text)

    file_bytes = _try_decrypt_pdf_bytes(file_bytes, password=password)

  
   
    
    file_bytes = _try_decrypt_pdf_bytes(file_bytes, password=password)
    
    records = []
    opening_balance = None
    closing_balance = None
    statement_from_date = None
    statement_to_date = None
    
    try:
        import fitz  # PyMuPDF
    except ImportError:
        print("  [PDF] PyMuPDF not installed. Run: pip install pymupdf", flush=True)
        return {"records": [], "opening_balance": None, "closing_balance": None}
    
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    
    # Extract all text
    all_text = []
    for page_num, page in enumerate(doc):
        text = page.get_text()
        all_text.append(text)
    
    doc.close()
    
    full_text = "\n".join(all_text)





        # PNB PDF support
    if is_pnb_file or re.search(r'Punjab\s+National\s+Bank|PUNB\d+|Generated\s+through\s+mPassBook', full_text, re.I):
        print("  [PDF] PNB statement detected", flush=True)
        return _parse_pnb_pdf_text(full_text)
    
    # Union Bank PDF support
    if re.search(r'Union|UBIN|unionbankofindia', full_text, re.I):
        print("  [PDF] Union Bank statement detected", flush=True)
        return _parse_union_bank_pdf_text(full_text)
    
        # PNB PDF support
    if re.search(r'Punjab\s+National\s+Bank|PUNB\d+|Generated\s+through\s+mPassBook', full_text, re.I):
        print("  [PDF] PNB statement detected", flush=True)
        return _parse_pnb_pdf_text(full_text)
 

    
    # Extract period from statement header
    period_pattern = re.compile(r'Statement\s+From\s*:\s*(\d{2}/\d{2}/\d{4})\s+To\s*:\s*(\d{2}/\d{2}/\d{4})', re.I)
    period_match = period_pattern.search(full_text)
    if period_match:
        statement_from_date = _norm_date_hdfc(period_match.group(1))
        statement_to_date = _norm_date_hdfc(period_match.group(2))
        print(f'  [PDF] Period: {statement_from_date} to {statement_to_date}', flush=True)
    
    # Extract opening and closing balance from summary
    summary_pattern = re.compile(
        r'Opening\s+Balance\s+([0-9,]+\.\d{2})\s+\d+\s+\d+\s+([0-9,]+\.\d{2})\s+([0-9,]+\.\d{2})\s+([0-9,]+\.\d{2})',
        re.I
    )
    summary_match = summary_pattern.search(full_text)
    if summary_match:
        opening_balance = float(summary_match.group(1).replace(',', ''))
        closing_balance = float(summary_match.group(4).replace(',', ''))
        print(f'  [PDF] Opening balance: {opening_balance}, Closing: {closing_balance}', flush=True)
    
    # Split into lines and process each line
    lines = full_text.split('\n')
    
    # HDFC transaction pattern - matches lines with date at start
    date_pattern = re.compile(r'^(\d{2}/\d{2}/\d{2})\s+')
    
    transactions = []
    current_txn = None
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Skip header/footer lines
        if re.match(r'^(Page No\.|Statement of account|MRS\.|MANI GARG|FLAT NO|SECTOR|NEAR GYM|FARIDABAD|HARYANA|JOINT HOLDERS|Nomination|Registered|Account Branch|Address|City|State|Phone no|OD Limit|Currency|Email|Cust ID|Account No|A/C Open Date|Account Status|RTGS/NEFT IFSC|MICR|Branch Code|Account Type|HDFC BANK LIMITED|\*Closing balance|Contents of this statement|State account branch GSTN|HDFC Bank GSTIN|Registered Office Address|Generated On|Generated By|Requesting Branch Code|This is a computer generated|STATEMENT SUMMARY|Opening Balance|Dr Count|Cr Count|Debits|Credits|Closing Bal)', line, re.I):
            continue
        
        if date_pattern.match(line):
            if current_txn:
                transactions.append(current_txn)
            current_txn = line
        elif current_txn:
            current_txn += ' ' + line
    
    if current_txn:
        transactions.append(current_txn)
    
    print(f'  [PDF] Found {len(transactions)} potential transactions', flush=True)
    
    # Parse each transaction
    for txn in transactions:
        # Extract date
        date_match = re.match(r'^(\d{2}/\d{2}/\d{2})\s+', txn)
        if not date_match:
            continue
        
        raw_date = date_match.group(1)
        remaining = txn[date_match.end():].strip()
        
        # Check for CR/DR indicator
        cr_match = re.search(r'\bCR\b', remaining, re.IGNORECASE)
        dr_match = re.search(r'\bDR\b', remaining, re.IGNORECASE)
        
        # Find all amounts in the line
        amount_pattern = re.compile(r'([0-9]{1,3}(?:,[0-9]{3})*\.[0-9]{2})')
        amounts = amount_pattern.findall(remaining)
        
        if len(amounts) < 2:
            continue
        
        # Remove amounts from remaining to get narration
        narration = remaining
        for amt in amounts:
            narration = narration.replace(amt, '')
        
        # Clean narration
        narration = re.sub(r'\s+', ' ', narration).strip()
        
        def parse_amt(s):
            return float(s.replace(',', ''))
        
        withdrawal_amt = parse_amt(amounts[0]) if len(amounts) >= 1 else 0
        deposit_amt = parse_amt(amounts[1]) if len(amounts) >= 2 else 0
        balance = parse_amt(amounts[-1]) if len(amounts) >= 3 else None
        
        # Determine transaction type
        if cr_match and not dr_match:
            txn_type = 'credit'
            amount = deposit_amt if deposit_amt > 0 else withdrawal_amt
        elif dr_match:
            txn_type = 'debit'
            amount = withdrawal_amt if withdrawal_amt > 0 else deposit_amt
        else:
            if withdrawal_amt > 0 and withdrawal_amt > deposit_amt:
                txn_type = 'debit'
                amount = withdrawal_amt
            elif deposit_amt > 0:
                txn_type = 'credit'
                amount = deposit_amt
            else:
                continue
        
        # Simple narration cleaning - keep it short
        narration_clean = narration
        narration_clean = re.sub(r'^(NEFT|RTGS|IMPS|IB\s+FUNDS\s+TRANSFER)\s*(?:CR|DR)?[-/]*\s*', '', narration_clean, flags=re.I)
        narration_clean = re.sub(r'\s+', ' ', narration_clean).strip()
        
        if not narration_clean or len(narration_clean) < 2:
            narration_clean = "Transaction"
        
        # Convert date
        try:
            parts = raw_date.split('/')
            if len(parts) == 3:
                dd, mm, yy = parts
                yyyy = '20' + yy if len(yy) == 2 else yy
                txn_date = f"{yyyy}-{mm}-{dd}"
            else:
                txn_date = raw_date
        except:
            txn_date = raw_date
        
        records.append({
            "txn_date": txn_date,
            "description": narration_clean,
            "narration": narration_clean,
            "amount": amount,
            "txn_type": txn_type,
            "balance": balance,
            "counterparty": "",
            "raw_debit": amount if txn_type == 'debit' else None,
            "raw_credit": amount if txn_type == 'credit' else None,
        })
    
    print(f'  [PDF] Parsed {len(records)} transactions', flush=True)
    
    # Print sample
    print('  [PDF] Sample transactions:')
    for i, r in enumerate(records[:10]):
        print(f'    [{i}] {r["txn_date"]} | {r["txn_type"]:6} | {r["amount"]:>12,.2f} | {r["narration"][:50]}')
    
    return {
        "records": records,
        "opening_balance": opening_balance,
        "closing_balance": _resolve_statement_closing_balance(
            records, opening_balance, closing_balance, "parse_pdf_statement"
        ),
        "statement_from_date": statement_from_date,
        "statement_to_date": statement_to_date,
    }

def _parse_union_bank_pdf_text(full_text: str) -> Dict[str, Any]:
    import re
    from datetime import datetime

    records = []
    statement_from_date = None
    statement_to_date = None

    period_match = re.search(
        r'Statement\s+Period\s+From\s*-?\s*(\d{2}/\d{2}/\d{4})\s+To\s+(\d{2}/\d{2}/\d{4})',
        full_text,
        re.I
    )
    if period_match:
        statement_from_date = _norm_date_hdfc(period_match.group(1))
        statement_to_date = _norm_date_hdfc(period_match.group(2))

    text = re.sub(r'\s+', ' ', full_text)

    pattern = re.compile(
        r'([SC]\d{6,})\s+'
        r'(\d{2}/\d{2}/\d{4})\s+'
        r'(.*?)\s+'
        r'([0-9,]+\.\d{2})\s*\((Dr|Cr)\)\s+'
        r'([0-9,]+\.\d{2})',
        re.I
    )

    for m in pattern.finditer(text):
        tran_id, raw_date, narration, amount_raw, drcr, balance_raw = m.groups()

        try:
            txn_date = datetime.strptime(raw_date, "%d/%m/%Y").strftime("%Y-%m-%d")
        except Exception:
            txn_date = raw_date

        amount = float(amount_raw.replace(",", ""))
        balance = float(balance_raw.replace(",", ""))
        txn_type = "debit" if drcr.lower() == "dr" else "credit"

        narration = re.sub(r'\s+', ' ', narration).strip()

        records.append({
            "txn_date": txn_date,
            "description": narration,
            "narration": narration,
            "amount": amount,
            "txn_type": txn_type,
            "balance": balance,
            "counterparty": "",
            "raw_debit": amount if txn_type == "debit" else None,
            "raw_credit": amount if txn_type == "credit" else None,
        })

    print(f"  [PDF:Union] Parsed {len(records)} transactions", flush=True)

    return {
        "records": records,
        "opening_balance": None,
        "closing_balance": _resolve_statement_closing_balance(
            records, None, None, "union_bank_pdf"
        ),
        "statement_from_date": statement_from_date,
        "statement_to_date": statement_to_date,
    }







def _parse_pnb_pdf_text(full_text: str) -> Dict[str, Any]:
    import re
    from datetime import datetime

    records = []

    period_match = re.search(
        r'Statement\s+For:\s*(\d{4}/\d{2}/\d{2})\s*to\s*(\d{4}/\d{2}/\d{2})',
        full_text,
        re.I
    )

    statement_from_date = None
    statement_to_date = None

    if period_match:
        statement_from_date = period_match.group(1).replace("/", "-")
        statement_to_date = period_match.group(2).replace("/", "-")

    pattern = re.compile(
        r'(\d{2}/\d{2}/\d{4})\s+'
        r'([0-9]+\.[0-9]{2})?\s*'
        r'([0-9]+\.[0-9]{2})?\s*'
        r'([0-9]+\.[0-9]{2})\s+Cr\.?\s+'
        r'(.*?)(?=\d{2}/\d{2}/\d{4}|Page\s+\d+|$)',
        re.S
    )

    for m in pattern.finditer(full_text):
        raw_date, withdrawal, deposit, balance, narration = m.groups()

        try:
            txn_date = datetime.strptime(raw_date, "%d/%m/%Y").strftime("%Y-%m-%d")
        except:
            txn_date = raw_date

        narration = re.sub(r"\s+", " ", narration).strip()





        counterparty = ""

        parts = narration.split("/")
        if len(parts) >= 2:
           counterparty = parts[-1].strip()

        counterparty = re.sub(r"[^A-Za-z ]+", " ", counterparty)
        counterparty = re.sub(r"\s+", " ", counterparty).strip().title()

        if not counterparty or counterparty.upper() in ("P2M", "P2V"):
           counterparty = ""










        # if deposit:
        #     amount = float(deposit)
        #     txn_type = "credit"
        # elif withdrawal:
        #     amount = float(withdrawal)
        #     txn_type = "debit"
        # else:
        #     continue
        if deposit:
            amount = float(deposit)
            txn_type = "credit"
        elif withdrawal:
             amount = float(withdrawal)

    # PNB PDF extraction shifts deposit values into withdrawal column.
     # IMPS-IN rows are inward credits in this statement format.
             if re.search(r'\bIMPS\s*-\s*IN\b|\bIMPS[-\s]*IN\b', narration, re.I):
               txn_type = "credit"
             else:
                  txn_type = "debit"
        else:
          continue


 





        records.append({
            "txn_date": txn_date,
            "description": narration,
            "narration": narration,
            "amount": amount,
            "txn_type": txn_type,
            "balance": float(balance),
            "counterparty": counterparty,
            "raw_debit": amount if txn_type=="debit" else None,
            "raw_credit": amount if txn_type=="credit" else None,
        })

    print(f"  [PDF:PNB] Parsed {len(records)} transactions", flush=True)

    return {
        "records": records,
        "opening_balance": None,
        "closing_balance": _resolve_statement_closing_balance(
            records, None, None, "pnb_pdf"
        ),
        "statement_from_date": statement_from_date,
        "statement_to_date": statement_to_date,
    }





def _norm_date_hdfc(date_str: str) -> str:
    """Convert DD/MM/YYYY to YYYY-MM-DD"""
    try:
        parts = date_str.split('/')
        if len(parts) == 3:
            dd, mm, yyyy = parts
            return f"{yyyy}-{mm}-{dd}"
    except:
        pass
    return date_str

# ── PDF SUPPORT (NEW) ── Library availability checker
def _pdf_library_status() -> Dict[str, bool]:
    """Return availability of each PDF parsing library."""
    status: Dict[str, bool] = {}

    try:
        import pdfplumber  # type: ignore  # noqa: F401
        status["pdfplumber"] = True
    except ImportError:
        status["pdfplumber"] = False

    try:
        import fitz  # type: ignore  # noqa: F401
        status["fitz"] = True
    except ImportError:
        status["fitz"] = False

    try:
        from pdfminer.high_level import extract_text_to_fp  # type: ignore  # noqa: F401
        status["pdfminer"] = True
    except ImportError:
        status["pdfminer"] = False

    try:
        import pytesseract  # type: ignore  # noqa: F401
        import shutil as _shutil
        status["tesseract"] = bool(_shutil.which("tesseract"))
    except ImportError:
        status["tesseract"] = False

    return status


# ══════════════════════════════════════════════════════════════════════════════
# END PDF SUPPORT (NEW)
# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL PARSE HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _normalize_header_label(text: Any) -> str:
    """Lowercase, strip punctuation, collapse spaces for loose header matching."""
    s = str(text or "").lower()
    s = re.sub(r'[^\w\s]', ' ', s)   # punctuation → space
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def _safe_parse_amount(value: Any) -> float:
    """
    Parse a cell value to a float amount.
    Handles: commas, currency symbols (₹$£€), CR/DR suffixes,
    parentheses for negatives, blank strings, merged-cell artifacts.
    Returns 0.0 on failure instead of raising.
    """
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return abs(float(value))
    s = str(value).strip()
    if not s:
        return 0.0
    su = s.upper()
    # Strip currency symbols and commas
    cleaned = re.sub(r'[₹$£€,\s]', '', su)
    # Parentheses = negative magnitude
    if cleaned.startswith('(') and cleaned.endswith(')'):
        cleaned = cleaned[1:-1]
    # Remove CR/DR suffixes
    cleaned = re.sub(r'(CR|DR)$', '', cleaned).strip()
    # Remove any remaining non-numeric except dot and minus
    cleaned = re.sub(r'[^\d.\-]', '', cleaned)
    if not cleaned:
        return 0.0
    # Handle multiple dots (e.g. "1,234.56" after partial cleaning)
    parts = cleaned.split('.')
    if len(parts) > 2:
        cleaned = parts[0] + '.' + ''.join(parts[1:])
    try:
        return abs(float(cleaned))
    except ValueError:
        return 0.0


def _safe_parse_excel_date(value: Any) -> str:
    """
    Convert any Excel date representation to 'YYYY-MM-DD' string.
    Handles: datetime objects, date objects, numeric serials, formatted strings.
    Returns today's date string on failure.
    """
    from datetime import date as _date_type
    today = datetime.now().strftime('%Y-%m-%d')
    if value is None:
        return today
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, _date_type):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).strftime('%Y-%m-%d')
        except Exception:
            return today
    s = str(value).strip()
    if not s:
        return today
    # Normalise single-digit day/month
    s = re.sub(
        r'(?<![:\d])(\d)(?=[/\-])|(?<=[/\-])(\d)(?![:\d])',
        lambda m: '0' + (m.group(1) or m.group(2)), s
    )
    for fmt in (
        '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d',
        '%d %b %Y', '%d-%b-%Y',
        '%d/%m/%y', '%d-%b-%y',
        '%d%m%Y',
        '%Y/%m/%d',
    ):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except Exception:
            pass
    return today


def _is_transaction_like_row(row: List[Any]) -> bool:
    """
    Heuristic: True if the row looks like a transaction data row.
    Criteria: at least one date-like cell, at least one non-empty text cell,
    at least one numeric cell, and the row is not mostly empty.
    """
    non_empty = [c for c in row if c is not None and str(c).strip()]
    if len(non_empty) < 3:
        return False
    has_date = any(
        re.search(r'\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}|\d{5,}', str(c))
        for c in non_empty
    )
    has_numeric = any(
        isinstance(c, (int, float)) or
        bool(re.search(r'^\d[\d,. ]*$', str(c).strip()))
        for c in non_empty
    )
    has_text = any(
        isinstance(c, str) and len(c.strip()) > 3 and not c.strip().replace('.', '').replace(',', '').isdigit()
        for c in non_empty
    )
    return has_date and has_numeric and has_text


def _select_best_sheet_from_workbook(wb: Any) -> Any:
    """
    Given an openpyxl workbook, return the worksheet with the most
    transaction-like rows. Falls back to the active sheet.
    """
    best_ws = wb.active
    best_score = 0
    for name in wb.sheetnames:
        try:
            ws = wb[name]
            score = 0
            for row in ws.iter_rows(max_row=200, values_only=True):
                if _is_transaction_like_row(list(row)):
                    score += 1
            print(f"  [OPENPYXL] Sheet '{name}': {score} transaction-like rows")
            if score > best_score:
                best_score = score
                best_ws = ws
        except Exception as e:
            print(f"  [OPENPYXL] Sheet '{name}' inspect error: {e}")
    print(f"  [OPENPYXL] Selected sheet: '{best_ws.title}' (score={best_score})")
    return best_ws


def _looks_like_encrypted_office_file(file_bytes: bytes, filename: str = "") -> bool:
    """Heuristic for password protected Office workbooks."""
    lower = (filename or '').lower()
    return lower.endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')) and _is_xls_bytes(file_bytes) and not _is_xlsx_bytes(file_bytes)


def _safe_parse_excel_date(v: Any) -> str:
    """
    Parse Excel/openpyxl/xlrd date-ish values into ISO YYYY-MM-DD.
    """
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")

    s = str(v).strip()
    if not s:
        return ""

    for fmt in (
        "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y",
        "%d/%m/%y", "%d-%m-%y", "%d.%m.%y",
        "%Y-%m-%d", "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass

    try:
        return _normalise_date(s)
    except Exception:
        return ""


def _valid_balance_value(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, str) and not value.strip():
        return None
    try:
        val = float(value)
    except Exception:
        return None
    if not np.isfinite(val):
        return None
    return val


def _sum_statement_rows(records: List[Dict[str, Any]]) -> Tuple[float, float]:
    total_credits = 0.0
    total_debits = 0.0
    for r in records or []:
        amt = float(r.get("amount") or 0)
        txn_type = str(r.get("txn_type") or "").strip().lower()
        if txn_type == "credit":
            total_credits += amt
        elif txn_type == "debit":
            total_debits += amt
    return total_credits, total_debits


def _last_valid_running_balance(records: List[Dict[str, Any]]) -> Optional[float]:
    for r in reversed(records or []):
        bal = _valid_balance_value(r.get("balance"))
        if bal is not None:
            return bal
    return None


def _resolve_statement_closing_balance(
    records: List[Dict[str, Any]],
    opening_balance: Any,
    explicit_closing_balance: Any = None,
    source: str = "statement",
) -> Optional[float]:
    """
    Authoritative closing balance precedence:
      1. explicit parsed statement closing balance,
      2. opening + credits - debits for parsed statement rows,
      3. last valid running balance captured from rows.
    """
    explicit = _valid_balance_value(explicit_closing_balance)
    opening = _valid_balance_value(opening_balance)
    computed = None

    if opening is not None:
        total_credits, total_debits = _sum_statement_rows(records)
        computed = opening + total_credits - total_debits

    if explicit is not None:
        if computed is not None and abs(explicit - computed) > 0.01:
            print(
                f"  [CLOSING BALANCE NOTE] {source}: explicit closing {explicit:.2f} "
                f"differs from computed {computed:.2f}; keeping explicit parsed value.",
                flush=True,
            )
        return explicit

    if computed is not None:
        return computed

    return _last_valid_running_balance(records)


def _extract_explicit_closing_balance_from_rows(rows: List[List[Any]]) -> Optional[float]:
    """
    Capture statement-summary closing balances without using date ranges or row
    positions as inference. Returns None if no clear summary line is found.
    """
    for row in rows or []:
        cells = [str(c or "").strip() for c in row]
        joined = " ".join(cells)
        if not re.search(r"\bclosing\s+(?:bal(?:ance)?|balance)\b", joined, re.I):
            continue
        nums = [_safe_parse_amount(c) for c in cells if _safe_parse_amount(c) > 0]
        if nums:
            return nums[-1]
    return None


def _looks_like_axis_statement(rows: List[List[Any]], filename: str = "") -> bool:
    name = (filename or "").upper()
    if "AXIS" in name:
        return True

    text = " ".join(str(c or "") for row in rows[:40] for c in row[:8]).upper()
    axis_signals = [
        "AXIS BANK",
        "TRAN DATE",
        "PARTICULARS",
        "CHQ NO",
        "AMOUNT",
        "DR/CR",
        "BALANCE",
        "BRANCH NAME",
    ]
    return sum(1 for s in axis_signals if s in text) >= 2


def _parse_axis_rows(rows_raw: List[List[Any]], filename: str = "") -> Dict[str, Any]:
    """
    Axis specific parser for Excel-like row grids.
    Typical columns:
      Tran Date | Value Date | Particulars | Chq No | Amount | Dr/Cr | Balance | Branch Name
    """
    import re

    header_idx = None

    def _norm(x):
        return re.sub(r"\s+", " ", str(x or "")).strip().lower()

    for i, row in enumerate(rows_raw[:60]):
        vals = [_norm(c) for c in row]
        if "tran date" in vals or ("particulars" in vals and "balance" in vals):
            header_idx = i
            break

    if header_idx is None:
        return {"records": [], "opening_balance": None}

    hdr = [_norm(c) for c in rows_raw[header_idx]]

    def _find(*needles):
        for n in needles:
            for j, h in enumerate(hdr):
                if n == h or n in h:
                    return j
        return None

    date_col    = _find("tran date", "transaction date", "date")
    desc_col    = _find("particulars", "narration", "description", "details")
    amount_col  = _find("amount", "txn amount")
    type_col    = _find("dr/cr", "cr/dr", "type")
    balance_col = _find("balance", "running balance", "closing balance")

    if date_col is None or desc_col is None or amount_col is None:
        return {"records": [], "opening_balance": None}

    records = []
    opening_balance = None

    def _amt(v):
        s = str(v or "").replace(",", "").strip()
        s = re.sub(r"[^\d.\-]", "", s)
        try:
            return abs(float(s))
        except Exception:
            return 0.0

    for row in rows_raw[header_idx + 1:]:
        if not any(str(c or "").strip() for c in row):
            continue

        date_val = str(row[date_col] if date_col < len(row) else "").strip()
        desc = str(row[desc_col] if desc_col < len(row) else "").strip()
        amt_raw = row[amount_col] if amount_col < len(row) else ""
        typ_raw = str(row[type_col] if type_col is not None and type_col < len(row) else "").strip().lower()
        bal_raw = row[balance_col] if balance_col is not None and balance_col < len(row) else ""

        if not date_val or not desc:
            continue

        amount = _amt(amt_raw)
        balance = _amt(bal_raw) if str(bal_raw or "").strip() else None

        if typ_raw in ("cr", "credit"):
            txn_type = "credit"
        elif typ_raw in ("dr", "debit"):
            txn_type = "debit"
        else:
            txn_type = _infer_txn_type(desc)

        records.append({
            "txn_date": _safe_parse_excel_date(date_val),
            "description": desc,
            "narration": desc,
            "amount": amount,
            "txn_type": txn_type,
            "raw_debit": amount if txn_type == "debit" else None,
            "raw_credit": amount if txn_type == "credit" else None,
            "balance": balance,
        })

    if opening_balance is None and records:
        first = next(
            (r for r in records
             if r.get("balance") is not None and r.get("amount") and r.get("txn_type") in ("debit", "credit")),
            None
        )
        if first:
            bal = float(first["balance"])
            amt = float(first["amount"])
            opening_balance = bal - amt if first["txn_type"] == "credit" else bal + amt

    statement_from_date = None
    statement_to_date = None
    explicit_closing_balance = _extract_explicit_closing_balance_from_rows(rows_raw[:60])
    closing_balance = _resolve_statement_closing_balance(
        records, opening_balance, explicit_closing_balance, "_parse_axis_rows"
    )

    return {
        "records": records,
        "opening_balance": opening_balance,
        "closing_balance": closing_balance,
        "statement_from_date": statement_from_date,
        "statement_to_date": statement_to_date,
    }


def _looks_like_au_statement(rows: List[List[Any]], filename: str = "") -> bool:
    name = (filename or "").upper()
    if "AUBANK" in name or "AU BANK" in name or "AU SMALL FINANCE" in name:
        return True

    text = " ".join(str(c or "") for row in rows[:50] for c in row[:8]).upper()
    signals = [
        "DATE",
        "DESCRIPTION",
        "DEBIT",
        "CREDIT",
        "BALANCE",
        "BPAY",
        "OSKO",
        "NPP",
        "DIRECT DEBIT",
        "DIRECT CREDIT",
    ]
    return sum(1 for s in signals if s in text) >= 3

def _parse_au_rows(rows_raw: List[List[Any]], filename: str = "") -> Dict[str, Any]:
    import re

    header_idx = None

    def _norm(x):
        return re.sub(r"\s+", " ", str(x or "")).strip().lower()

    def _safe_parse_statement_date(text: Any) -> Optional[str]:
        s = str(text or "").strip()
        if not s:
            return None

        s = s.replace(".", "-").replace("/", "-")
        s = re.sub(r"\s+", "-", s)

        for fmt in (
            "%d-%m-%Y",
            "%d-%m-%y",
            "%d-%b-%Y",
            "%d-%b-%y",
            "%d-%B-%Y",
            "%d-%B-%y",
        ):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except Exception:
                pass

        for fmt in (
            "%m-%d-%Y",
            "%m-%d-%y",
        ):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except Exception:
                pass

        try:
            return _safe_parse_excel_date(s)
        except Exception:
            return None

    def _extract_statement_period_from_header_rows(header_rows: List[List[Any]]) -> Tuple[Optional[str], Optional[str]]:
        lines = []
        for row in header_rows or []:
            txt = " ".join(str(c or "").strip() for c in row if str(c or "").strip())
            if txt:
                lines.append(re.sub(r"\s+", " ", txt).strip())

        candidates: List[Tuple[str, str, str]] = []

        strong_patterns = [
            r'(?:statement\s*period|period)\s*[:\-]?\s*(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})\s*(?:to|\-)\s*[:\-]?\s*(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})',
            r'(?:from)\s*[:\-]?\s*(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})\s*(?:to)\s*[:\-]?\s*(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})',
        ]

        for line in lines:
            s = re.sub(r'\s+', ' ', line).strip()
            for pat in strong_patterns:
                m = re.search(pat, s, flags=re.I)
                if m:
                    d1 = _safe_parse_statement_date(m.group(1))
                    d2 = _safe_parse_statement_date(m.group(2))
                    if d1 and d2 and d1 <= d2:
                        candidates.append((d1, d2, s))

        if not candidates:
            for line in lines:
                s = re.sub(r'\s+', ' ', line).strip()
                m = re.search(
                    r'(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})\s*(?:to|\-)\s*[:\-]?\s*(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})',
                    s,
                    flags=re.I
                )
                if m:
                    d1 = _safe_parse_statement_date(m.group(1))
                    d2 = _safe_parse_statement_date(m.group(2))
                    if d1 and d2 and d1 <= d2:
                        candidates.append((d1, d2, s))

        if not candidates:
            return None, None

        def _score(item):
            d1, d2, line = item
            start_day1 = 1 if d1.endswith("-01") else 0
            try:
                span_days = (datetime.strptime(d2, "%Y-%m-%d") - datetime.strptime(d1, "%Y-%m-%d")).days
            except Exception:
                span_days = -1
            return (start_day1, span_days, len(line))

        best = max(candidates, key=_score)
        print(f"  [AU PERIOD] Selected header period line: {best[2]}")
        return best[0], best[1]

    for i, row in enumerate(rows_raw[:60]):
        vals = [_norm(c) for c in row]
        if "description" in vals and "balance" in vals and ("debit" in vals or "credit" in vals or "amount" in vals):
            header_idx = i
            break

    if header_idx is None:
        return {
            "records": [],
            "opening_balance": None,
            "closing_balance": None,
            "statement_from_date": None,
            "statement_to_date": None,
        }

    hdr = [_norm(c) for c in rows_raw[header_idx]]

    def _find(*needles):
        for n in needles:
            for j, h in enumerate(hdr):
                if n == h or n in h:
                    return j
        return None

    date_col    = _find("date")
    desc_col    = _find("description", "details", "narration")
    debit_col   = _find("debit")
    credit_col  = _find("credit")
    amount_col  = _find("amount")
    balance_col = _find("balance")

    records = []
    opening_balance = None

    def _amt(v):
        s = str(v or "").replace(",", "").strip()
        s = re.sub(r"[^\d.\-]", "", s)
        try:
            return abs(float(s))
        except Exception:
            return 0.0

    for row in rows_raw[header_idx + 1:]:
        if not any(str(c or "").strip() for c in row):
            continue

        date_val = str(row[date_col] if date_col is not None and date_col < len(row) else "").strip()
        desc = str(row[desc_col] if desc_col is not None and desc_col < len(row) else "").strip()

        if not date_val or not desc:
            continue

        debit_raw   = row[debit_col] if debit_col is not None and debit_col < len(row) else ""
        credit_raw  = row[credit_col] if credit_col is not None and credit_col < len(row) else ""
        amount_raw  = row[amount_col] if amount_col is not None and amount_col < len(row) else ""
        balance_raw = row[balance_col] if balance_col is not None and balance_col < len(row) else ""

        dr = _amt(debit_raw)
        cr = _amt(credit_raw)
        amt = dr or cr or _amt(amount_raw)

        clean_desc = _strip_au_prefix(desc)

        if dr > 0 and cr == 0:
            txn_type = "debit"
        elif cr > 0 and dr == 0:
            txn_type = "credit"
        else:
            txn_type = _infer_txn_type(clean_desc)

        bal = _amt(balance_raw) if str(balance_raw or "").strip() else None

        records.append({
            "txn_date": _safe_parse_excel_date(date_val),
            "description": clean_desc,
            "narration": clean_desc,
            "amount": amt,
            "txn_type": txn_type,
            "raw_debit": dr if txn_type == "debit" else None,
            "raw_credit": cr if txn_type == "credit" else None,
            "balance": bal,
        })

    if opening_balance is None and records:
        first = next(
            (r for r in records
             if r.get("balance") is not None and r.get("amount") and r.get("txn_type") in ("debit", "credit")),
            None
        )
        if first:
            bal = float(first["balance"])
            amt = float(first["amount"])
            opening_balance = bal - amt if first["txn_type"] == "credit" else bal + amt

    statement_from_date = None
    statement_to_date = None
    explicit_closing_balance = _extract_explicit_closing_balance_from_rows(rows_raw[:60])

    try:
        header_rows = rows_raw[:25] if rows_raw else []
        statement_from_date, statement_to_date = _extract_statement_period_from_header_rows(header_rows)
        print(f"  [AU PERIOD] Header-derived period: {statement_from_date} to {statement_to_date}")
    except Exception as e:
        print(f"  [AU PERIOD] Header extraction failed: {e}")
        statement_from_date, statement_to_date = None, None


    closing_balance = _resolve_statement_closing_balance(
        records, opening_balance, explicit_closing_balance, "_parse_au_rows"
    )

    return {
        "records": records,
        "opening_balance": opening_balance,
        "closing_balance": closing_balance,
        "statement_from_date": statement_from_date,
        "statement_to_date": statement_to_date,
    }

def parse_excel_statement(file_bytes: bytes, filename: str = "", password: Optional[str] = None) -> Dict[str, Any]:
    def _log_parse_return(records, opening_balance, closing_balance, statement_from_date, statement_to_date):
        print("[PARSE RETURN]", {
            "statement_from_date": statement_from_date,
            "statement_to_date": statement_to_date,
            "opening_balance": opening_balance,
            "closing_balance": closing_balance,
            "record_count": len(records or [])
        }, flush=True)

    # ── PDF SUPPORT ──
    if (filename or "").lower().endswith(".pdf") or _is_pdf_bytes(file_bytes):
        pdf_result = parse_pdf_statement(file_bytes, filename, password=password)
        if isinstance(pdf_result, list):
            res = {
                "records": pdf_result,
                "opening_balance": None,
                "closing_balance": None,
                "statement_from_date": None,
                "statement_to_date": None,
            }
            _log_parse_return(res.get("records"), res.get("opening_balance"), res.get("closing_balance"), res.get("statement_from_date"), res.get("statement_to_date"))
            return res
        res = {
            "records": pdf_result.get("records", []),
            "opening_balance": pdf_result.get("opening_balance"),
            "closing_balance": pdf_result.get("closing_balance"),
            "statement_from_date": pdf_result.get("statement_from_date"),
            "statement_to_date": pdf_result.get("statement_to_date"),
        }
        _log_parse_return(res.get("records"), res.get("opening_balance"), res.get("closing_balance"), res.get("statement_from_date"), res.get("statement_to_date"))
        return res

    lower_name = (filename or "").lower()
    print(f"  [EXCEL] Parsing '{filename}' ({len(file_bytes)} bytes)")

    if _looks_like_encrypted_office_file(file_bytes, filename):
        if password:
            raise ValueError(
                "Password-protected Excel files are not yet decryptable in this runtime. "
                "Please unlock and save the workbook as a normal .xlsx, then upload that file."
            )
        raise ValueError(
            "Excel file appears to be password-protected. Entering a password is supported for PDFs, "
            "but encrypted Excel workbooks still need to be unlocked and saved as a normal .xlsx before upload."
        )

    parse_attempts: List[str] = []
    rows_raw: List[List[Any]] = []

    is_legacy_xls = (
        (lower_name.endswith(".xls") and not lower_name.endswith(".xlsx") and not lower_name.endswith(".xlsm"))
        or _is_xls_bytes(file_bytes)
    )

    # ── 1. Legacy .xls ────────────────────────────────────────────────────────
    if is_legacy_xls:
        print(f"  [EXCEL] Detected legacy .xls format for '{filename}'")
        xls_parsed = False

        try:
            import xlrd
            from xlrd import xldate_as_datetime, XL_CELL_DATE, XL_CELL_EMPTY

            book = xlrd.open_workbook(file_contents=file_bytes)
            print(f"  [XLS] xlrd opened workbook: {book.nsheets} sheets: {book.sheet_names()}")

            best_sheet = book.sheet_by_index(0)
            best_nrows = best_sheet.nrows
            for si in range(book.nsheets):
                s = book.sheet_by_index(si)
                print(f"  [XLS] xlrd sheet '{s.name}': {s.nrows} rows")
                if s.nrows > best_nrows:
                    best_nrows = s.nrows
                    best_sheet = s
            print(f"  [XLS] xlrd selected sheet: '{best_sheet.name}'")

            def _cell_val(cell):
                if cell.ctype == XL_CELL_DATE:
                    try:
                        return xldate_as_datetime(cell.value, book.datemode).strftime("%d/%m/%Y")
                    except Exception:
                        return cell.value
                if cell.ctype == XL_CELL_EMPTY:
                    return None
                return cell.value

            rows_raw = [
                [_cell_val(best_sheet.cell(r, c)) for c in range(best_sheet.ncols)]
                for r in range(best_sheet.nrows)
            ]
            xls_parsed = True
            print(f"  [XLS] xlrd extracted {len(rows_raw)} rows")
            parse_attempts.append(f"xlrd: OK ({len(rows_raw)} rows from '{best_sheet.name}')")
        except ImportError:
            msg = "xlrd not installed"
            print(f"  [XLS] {msg} — trying LibreOffice fallback")
            parse_attempts.append(f"xlrd: {msg}")
        except Exception as e:
            msg = f"xlrd failed: {e}"
            print(f"  [XLS] {msg} — trying LibreOffice fallback")
            parse_attempts.append(msg)

        if not xls_parsed:
            try:
                converted_bytes = _convert_xls_bytes_to_xlsx_bytes(file_bytes)
                file_bytes = converted_bytes
                parse_attempts.append("LibreOffice conversion: OK")
                print(f"  [XLS] LibreOffice conversion succeeded for '{filename}'")
            except ValueError as conv_err:
                parse_attempts.append(f"LibreOffice conversion: {conv_err}")
                raise ValueError("Excel parse failed. " + " | ".join(parse_attempts)) from conv_err

    # ── 2. Standard .xlsx / .xlsm via openpyxl ───────────────────────────────
    looks_like_xlsx = (
        lower_name.endswith((".xlsx", ".xlsm", ".xltx", ".xltm"))
        or _is_xlsx_bytes(file_bytes)
        or file_bytes[:2] == b"PK"
    )

    if not rows_raw and looks_like_xlsx:
        try:
            import openpyxl
            print(f"  [XLSX] Using openpyxl parser for '{filename}'")
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
            print(f"  [OPENPYXL] Opened workbook; sheets: {wb.sheetnames}")
            ws = _select_best_sheet_from_workbook(wb)
            rows_raw = [list(r) for r in ws.iter_rows(values_only=True)]
            print(f"  [OPENPYXL] Extracted {len(rows_raw)} rows from '{ws.title}'")
            print(f"  [XLSX] Final rows extracted: {len(rows_raw)}")
            parse_attempts.append(f"openpyxl: OK ({len(rows_raw)} rows from '{ws.title}')")
        except Exception as e:
            msg = f"openpyxl failed: {e}"
            print(f"  [OPENPYXL] {msg}")
            parse_attempts.append(msg)

        if not rows_raw:
            try:
                import openpyxl
                print(f"  [XLSX] Retrying openpyxl non-read-only mode for '{filename}'")
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
                ws = _select_best_sheet_from_workbook(wb)
                rows_raw = [list(r) for r in ws.iter_rows(values_only=True)]
                print(f"  [OPENPYXL] Non-read-only extracted {len(rows_raw)} rows from '{ws.title}'")
                parse_attempts.append(f"openpyxl(non-read-only): OK ({len(rows_raw)} rows from '{ws.title}')")
            except Exception as e:
                msg = f"openpyxl(non-read-only) failed: {e}"
                print(f"  [OPENPYXL] {msg}")
                parse_attempts.append(msg)

        if not rows_raw:
            try:
                print(f"  [XLSX] Trying pandas.read_excel fallback for '{filename}'")
                xl = pd.ExcelFile(io.BytesIO(file_bytes), engine="openpyxl")
                best_name = xl.sheet_names[0]
                best_df = None
                best_score = -1

                for sname in xl.sheet_names:
                    try:
                        df = xl.parse(sname, header=None, dtype=object)
                        score = 0
                        for _, row in df.head(200).iterrows():
                            if _is_transaction_like_row(row.tolist()):
                                score += 1
                        print(f"  [PANDAS] Sheet '{sname}': {score} transaction-like rows")
                        if score > best_score:
                            best_score = score
                            best_name = sname
                            best_df = df
                    except Exception as se:
                        print(f"  [PANDAS] Sheet '{sname}' failed: {se}")

                if best_df is not None:
                    rows_raw = best_df.where(pd.notnull(best_df), None).values.tolist()
                    print(f"  [PANDAS] Extracted {len(rows_raw)} rows from '{best_name}'")
                    parse_attempts.append(f"pandas: OK ({len(rows_raw)} rows from '{best_name}')")
            except Exception as e:
                msg = f"pandas failed: {e}"
                print(f"  [PANDAS] {msg}")
                parse_attempts.append(msg)

    # ── 3. XML fallback for non-standard XLSX ────────────────────────────────
    if not rows_raw and looks_like_xlsx:
        try:
            rows_raw = _parse_xlsx_xml(file_bytes)
            print(f"  [XML] Extracted {len(rows_raw)} rows")
            print(f"  [XLSX] Final rows extracted: {len(rows_raw)}")
            parse_attempts.append(f"XML fallback: OK ({len(rows_raw)} rows)")
        except Exception as e:
            msg = f"XML fallback failed: {e}"
            print(f"  [XML] {msg}")
            parse_attempts.append(msg)

    if not rows_raw:
        raise ValueError("Excel parse failed. " + " | ".join(parse_attempts))

    # ── Bank-specific early parsers ───────────────────────────────────────────
    if _looks_like_axis_statement(rows_raw, filename):
        axis_result = _parse_axis_rows(rows_raw, filename)
        if axis_result.get("records"):
            print(f"  [AXIS] Parsed {len(axis_result['records'])} transactions from '{filename}'")
            _log_parse_return(
                axis_result.get("records"),
                axis_result.get("opening_balance"),
                axis_result.get("closing_balance"),
                axis_result.get("statement_from_date"),
                axis_result.get("statement_to_date"),
            )
            return axis_result

    if _looks_like_au_statement(rows_raw, filename):
        au_result = _parse_au_rows(rows_raw, filename)
        if au_result.get("records"):
            print(f"  [AU] Parsed {len(au_result['records'])} transactions from '{filename}'")
            _log_parse_return(
                au_result.get("records"),
                au_result.get("opening_balance"),
                au_result.get("closing_balance"),
                au_result.get("statement_from_date"),
                au_result.get("statement_to_date"),
            )
            return au_result

    max_width = max((len(r) for r in rows_raw), default=0)
    rows = [list(r) + [None] * (max_width - len(r)) for r in rows_raw]

    # ── Helper: parse statement period from sheet header text ────────────────
    def _safe_parse_statement_date(text: Any) -> Optional[str]:
        s = str(text or "").strip()
        if not s:
            return None

        s = s.replace(".", "-").replace("/", "-")
        s = re.sub(r"\s+", "-", s)

        # Prefer day-first for Indian bank statements
        for fmt in (
            "%d-%m-%Y",
            "%d-%m-%y",
            "%d-%b-%Y",
            "%d-%b-%y",
            "%d-%B-%Y",
            "%d-%B-%y",
        ):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except Exception:
                pass

        # Explicit month-first fallback only if day-first failed
        for fmt in (
            "%m-%d-%Y",
            "%m-%d-%y",
        ):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except Exception:
                pass

        try:
            return _safe_parse_excel_date(s)
        except Exception:
            return None

    # ── Header detection ─────────────────────────────────────────────────────
    _DATE_VARIANTS = {
        "date", "txn date", "tran date", "trans date", "transaction date",
        "value date", "value dt", "posting date", "chq date", "cheque date",
        "book date", "entry date",
    }
    _DESC_VARIANTS = {
        "narration", "description", "particulars", "details", "remarks",
        "transaction details", "trans details", "transaction narration",
        "trans narration", "narration description",
    }
    _DEBIT_VARIANTS = {
        "debit", "withdrawal", "withdrawals", "withdrawl", "debit amount",
        "debit amt", "dr amount", "dr", "dr amt", "withdrawal amt",
        "withdrawal amt.", "paid out", "debit(inr)",
    }
    _CREDIT_VARIANTS = {
        "credit", "deposit", "deposits", "credit amount", "credit amt",
        "cr", "cr amount", "deposit amt", "deposit amt.", "paid in",
        "credit(inr)",
    }
    _BALANCE_VARIANTS = {
        "balance", "running balance", "closing balance", "bal", "outstanding",
        "avl bal", "available balance",
    }
    _AMOUNT_VARIANTS = {
        "amount", "txn amount", "net amount", "transaction amount", "amt",
    }

    def _norm(x: Any) -> str:
        return _normalize_header_label(x)

    header_idx: Optional[int] = None

    for i, row in enumerate(rows[:50]):
        vals = [_norm(c) for c in row]
        has_date = any(
            v in _DATE_VARIANTS or
            (v not in ("", "balance", "amount", "dr", "cr") and "date" in v)
            for v in vals
        )
        has_desc = any(
            v in _DESC_VARIANTS or
            any(w in v for w in ("narrat", "descri", "particu", "detail", "remark"))
            for v in vals
        )
        if has_date and has_desc:
            header_idx = i
            print(f"  [HEADER] Found formal header at row {i}: {[c for c in row if c is not None]}")
            break

    if header_idx is None:
        for i, row in enumerate(rows[:50]):
            vals = [_norm(c) for c in row]
            non_empty = [v for v in vals if v]

            has_dateish = any(v in _DATE_VARIANTS or "date" in v for v in non_empty)
            has_descish = any(
                v in _DESC_VARIANTS or any(w in v for w in ("narrat", "descri", "particu", "detail", "remark"))
                for v in non_empty
            )
            has_moneyish = any(
                v in _DEBIT_VARIANTS or v in _CREDIT_VARIANTS or v in _AMOUNT_VARIANTS or v in _BALANCE_VARIANTS
                for v in non_empty
            )

            if has_dateish and (has_descish or has_moneyish):
                header_idx = i
                print(f"  [HEADER] Using guarded fallback header at row {i}: {[c for c in row if c is not None]}")
                break

    if header_idx is None:
        for i, row in enumerate(rows[:30]):
            if _is_transaction_like_row(row):
                header_idx = max(0, i - 1)
                print(f"  [HEADER] Inferred header at row {header_idx} (first txn row at {i})")
                break

    if header_idx is None:
        raise ValueError(
            "Excel parse failed: Workbook was read but no transaction table could be detected. "
            + " | ".join(parse_attempts)
        )

    headers = [_norm(h) for h in rows[header_idx]]
    print(f"  [HEADER] Normalized headers: {headers}")

    def fc_exact(candidates: List[str]) -> Optional[int]:
        for kw in candidates:
            for i, h in enumerate(headers):
                if h == kw:
                    return i
        return None

    def fc_contains(candidates: List[str], exclude_cols: Optional[List[int]] = None) -> Optional[int]:
        exclude_cols = exclude_cols or []
        for kw in candidates:
            for i, h in enumerate(headers):
                if i in exclude_cols:
                    continue
                if kw in h:
                    return i
        return None

    date_col = fc_exact(list(_DATE_VARIANTS))
    if date_col is None:
        date_col = fc_contains(["date"])

    desc_col = fc_exact(list(_DESC_VARIANTS) + ["narration / description"])
    if desc_col is None:
        desc_col = fc_contains(["narrat", "descri", "particu", "detail", "remark"])

    type_col = fc_exact(["type", "txn type", "transaction type", "dr cr", "dr  cr", "cr dr"])

    skip = [c for c in [desc_col] if c is not None]
    debit_candidates = list(_DEBIT_VARIANTS)
    if type_col is None:
        debit_candidates.append("dr")
    debit_col = fc_exact(debit_candidates)
    if debit_col is None:
        debit_col = fc_contains(
            ["withdrawal", "debit amount", "debit amt", "dr amount", "debit inr", "paid out"],
            exclude_cols=skip,
        )

    credit_col = fc_exact(list(_CREDIT_VARIANTS))
    if credit_col is None:
        credit_col = fc_contains(
            ["deposit", "credit amount", "credit amt", "paid in", "credit inr"],
            exclude_cols=skip,
        )

    amount_col = fc_exact(list(_AMOUNT_VARIANTS))
    if amount_col is None:
        amount_col = fc_contains(
            ["amount", "amt"],
            exclude_cols=skip + [c for c in [debit_col, credit_col] if c is not None],
        )

    balance_col = fc_exact(list(_BALANCE_VARIANTS))
    if balance_col is None:
        balance_col = fc_contains(
            ["balance", "bal", "closing balance", "running balance"],
            exclude_cols=[c for c in [date_col, desc_col, debit_col, credit_col, amount_col, type_col] if c is not None],
        )

    print(
        f"  [COLS] date={date_col}, desc={desc_col}, type={type_col}, "
        f"debit={debit_col}, credit={credit_col}, amount={amount_col}, balance={balance_col}"
    )

    def safe_cell(row: List[Any], i: Optional[int]) -> Any:
        if i is None or i < 0 or i >= len(row):
            return None
        return row[i]

    def try_amt(val: Any) -> float:
        return _safe_parse_amount(val)

    def try_amt_signed(val: Any) -> float:
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip()
        su = s.upper()
        is_debit = su.endswith("DR") or su.endswith(" DR")
        is_credit = su.endswith("CR") or su.endswith(" CR")
        if su.startswith("(") and su.endswith(")"):
            is_debit = True
        magnitude = _safe_parse_amount(val)
        if is_debit:
            return -magnitude
        if is_credit:
            return magnitude
        return -magnitude if s.lstrip().startswith("-") else magnitude

    has_amounts = False
    for row in rows[header_idx + 1: header_idx + 51]:
        if (
            try_amt(safe_cell(row, debit_col)) > 0
            or try_amt(safe_cell(row, credit_col)) > 0
            or try_amt(safe_cell(row, amount_col)) > 0
        ):
            has_amounts = True
            break

    JUNK = re.compile(
        r"^(micr[\s/]*code|ifsc[\s/]*code|reward\s+points|total\s+credits"
        r"|total\s+debits|closing\s+balance|statement\s+summary"
        r"|[\*]+|▼|▶|◆|►|•)\s*$",
        re.I,
    )

    # ── Opening balance from explicit B/F rows ───────────────────────────────
    opening_balance = None
    bf_row_indexes = set()

    for idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 1):
        joined = " ".join(str(c or "") for c in row).strip().lower()
        if re.search(r"(?:^|\s)(b/?f|b\.f\.|opening\s+balance|brought\s+forward)(?:\s|$)", joined, re.I):
            bal_candidate = try_amt(safe_cell(row, balance_col)) if balance_col is not None else 0.0
            if bal_candidate > 0:
                opening_balance = bal_candidate
            else:
                nums = [try_amt(c) for c in row if try_amt(c) > 0]
                if nums:
                    opening_balance = max(nums)

            if opening_balance:
                print(f"  [OPENING] Captured opening balance: ₹{opening_balance}")
            bf_row_indexes.add(idx)

    DATE_RE = re.compile(r"\d")
    records = []

    for row_i, row in enumerate(rows[header_idx + 1:], start=header_idx + 1):
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        desc = str(safe_cell(row, desc_col) or "").strip()
        if not desc or JUNK.match(desc):
            continue
        if re.match(r"^[\s▼▶◆►•]+", desc):
            continue

        date_val = safe_cell(row, date_col)
        if date_val is not None:
            from datetime import date as _date_type
            if isinstance(date_val, (_date_type, datetime)):
                pass
            elif isinstance(date_val, float):
                pass
            elif isinstance(date_val, str) and not DATE_RE.search(str(date_val).strip()):
                continue

        dr = try_amt(safe_cell(row, debit_col))
        cr = try_amt(safe_cell(row, credit_col))

        if has_amounts:
            if row_i in bf_row_indexes:
                continue
            elif dr > 0 and cr == 0:
                amount, ttype = dr, "debit"
            elif cr > 0 and dr == 0:
                amount, ttype = cr, "credit"
            elif amount_col is not None:
                signed = try_amt_signed(safe_cell(row, amount_col))
                amount = abs(signed)
                if type_col is not None:
                    raw_type = str(safe_cell(row, type_col) or "").strip().lower()
                    ttype = "credit" if raw_type in ("credit", "cr") else "debit"
                elif signed < 0:
                    ttype = "debit"
                elif signed > 0:
                    ttype = "credit"
                else:
                    ttype = _infer_txn_type(desc)
            elif type_col is not None:
                raw_type = str(safe_cell(row, type_col) or "").strip().lower()
                ttype = "credit" if raw_type in ("credit", "cr") else "debit"
                amount = 0.0
            else:
                amount, ttype = 0.0, _infer_txn_type(desc)
        else:
            if type_col is not None:
                raw_type = str(safe_cell(row, type_col) or "").strip().lower()
                ttype = "credit" if raw_type in ("credit", "cr") else "debit"
            else:
                ttype = _infer_txn_type(desc)
            amount = 0.0

        ttype = derive_txn_type(
            raw_debit=safe_cell(row, debit_col),
            raw_credit=safe_cell(row, credit_col),
            raw_type=safe_cell(row, type_col),
            raw_amount=safe_cell(row, amount_col),
            narration=desc,
            parser_txn_type=ttype,
        )

        bal_raw = safe_cell(row, balance_col)
        bal_val = try_amt(bal_raw) if bal_raw is not None else None

        records.append({
            "txn_date": _safe_parse_excel_date(date_val),
            "description": desc,
            "narration": desc,
            "amount": amount,
            "txn_type": ttype,
            "raw_debit": safe_cell(row, debit_col),
            "raw_credit": safe_cell(row, credit_col),
            "balance": bal_val if bal_val else None,
        })

    print(f"  [EXCEL] Parsed {len(records)} transactions from '{filename}'")

    if not records:
        raise ValueError(
            "Workbook read succeeded but no valid transaction table was detected. "
            "Header row found at index " + str(header_idx) + " but zero data rows passed filters. "
            + " | ".join(parse_attempts)
        )

    # ── Opening balance fallback from first running balance ──────────────────
    if opening_balance is None and records:
        first = next(
            (
                r for r in records
                if r.get("balance") is not None and r.get("amount") and r.get("txn_type") in ("debit", "credit")
            ),
            None
        )
        if first:
            bal = float(first["balance"])
            amt = float(first["amount"])
            opening_balance = bal - amt if first["txn_type"] == "credit" else bal + amt
            print(f"  [OPENING] Derived opening balance from first transaction: ₹{opening_balance}")

     # ── Period and closing balance ───────────────────────────────────────────
    statement_from_date = None
    statement_to_date = None
    explicit_closing_balance = _extract_explicit_closing_balance_from_rows(rows_raw[:80])

    def _extract_statement_period_from_header_rows(header_rows: List[List[Any]]) -> Tuple[Optional[str], Optional[str]]:
        """
        Extract ONE real statement-period pair from header rows.
        Do not merge start from one line and end from another line.
        """
        lines = []
        for row in header_rows or []:
            txt = " ".join(str(c or "").strip() for c in row if str(c or "").strip())
            if txt:
                lines.append(re.sub(r"\s+", " ", txt).strip())

        candidates: List[Tuple[str, str, str]] = []

        strong_patterns = [
            r'(?:statement\s*period|period)\s*[:\-]?\s*(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})\s*(?:to|\-)\s*[:\-]?\s*(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})',
            r'(?:from)\s*[:\-]?\s*(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})\s*(?:to)\s*[:\-]?\s*(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})',
        ]

        for line in lines:
            s = re.sub(r'\s+', ' ', line).strip()
            for pat in strong_patterns:
                m = re.search(pat, s, flags=re.I)
                if m:
                    d1 = _safe_parse_statement_date(m.group(1))
                    d2 = _safe_parse_statement_date(m.group(2))
                    if d1 and d2 and d1 <= d2:
                        candidates.append((d1, d2, s))

        if not candidates:
            for line in lines:
                s = re.sub(r'\s+', ' ', line).strip()
                m = re.search(
                    r'(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})\s*(?:to|\-)\s*[:\-]?\s*(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})',
                    s,
                    flags=re.I
                )
                if m:
                    d1 = _safe_parse_statement_date(m.group(1))
                    d2 = _safe_parse_statement_date(m.group(2))
                    if d1 and d2 and d1 <= d2:
                        candidates.append((d1, d2, s))

        if not candidates:
            return None, None

        def _score(item):
            d1, d2, line = item
            start_day1 = 1 if d1.endswith("-01") else 0
            try:
                span_days = (datetime.strptime(d2, "%Y-%m-%d") - datetime.strptime(d1, "%Y-%m-%d")).days
            except Exception:
                span_days = -1
            return (start_day1, span_days, len(line))

        best = max(candidates, key=_score)
        print(f"  [PERIOD] Selected header period line: {best[2]}")
        return best[0], best[1]

    try:
        header_rows = rows_raw[:25] if rows_raw else []
        statement_from_date, statement_to_date = _extract_statement_period_from_header_rows(header_rows)
        print(f"  [PERIOD] Header-derived period: {statement_from_date} to {statement_to_date}")
    except Exception as e:
        print(f"  [PERIOD] Header extraction failed: {e}")
        statement_from_date, statement_to_date = None, None


    closing_balance = _resolve_statement_closing_balance(
        records, opening_balance, explicit_closing_balance, "parse_excel_statement"
    )

    res = {
        "records": records,
        "opening_balance": opening_balance,
        "closing_balance": closing_balance,
        "statement_from_date": statement_from_date,
        "statement_to_date": statement_to_date,
    }
    _log_parse_return(res.get("records"), res.get("opening_balance"), res.get("closing_balance"), res.get("statement_from_date"), res.get("statement_to_date"))
    return res

# ══════════════════════════════════════════════════════════════════════════════
# LEDGER & BALANCE SHEET GENERATORS
# ══════════════════════════════════════════════════════════════════════════════

def generate_ledger_report(db: DBStore, user_id: str) -> str:
    rows = db.get_ledger_summary(user_id)
    structure: Dict[str, Any] = {}
    for r in rows:
        b, s, g, a = r["book"], r["section"], r["grp"], r["account"]
        structure.setdefault(b, {}).setdefault(s, {}).setdefault(g, {}).setdefault(a, {"credit":0,"debit":0,"cnt":0})
        structure[b][s][g][a][r["txn_type"]] = structure[b][s][g][a].get(r["txn_type"],0) + r["total"]
        structure[b][s][g][a]["cnt"] += r["cnt"]
    all_txns = db.get_txns(user_id, limit=999999)

    # Get opening balance — use get_users() since plain DBStore has no get_user()
    _all_users = db.get_users()
    user = next((u for u in _all_users if u["id"] == user_id), None)
    opening_balance = float((user or {}).get("opening_balance") or 0)

    bank_debits = max(opening_balance, 0.0)
    bank_credits = abs(min(opening_balance, 0.0))
    for tx in all_txns:
        amt = float(tx.get("amount", 0) or 0)
        if tx.get("txn_type") == "credit":
            bank_debits += amt
        elif tx.get("txn_type") == "debit":
            bank_credits += amt

    bank_balance = bank_debits - bank_credits
    W = 90
    sym = "₹"
    def bar(c="─"): return c * W
    def hdr1(t): return f"\n{'═'*W}\n  {t}\n{'═'*W}"
    def amtf(v): return f"{sym}{v:>16,.2f}"
    def normal_balance(section: str, data: Dict[str, Any]) -> float:
        debit = float(data.get("debit", 0) or 0)
        credit = float(data.get("credit", 0) or 0)
        if section in ("Assets", "Expenditure"):
            return debit - credit
        if section in ("Liabilities", "Equity", "Income"):
            return credit - debit
        return debit + credit

    structure.setdefault("BALANCE_SHEET", {}) \
         .setdefault("Assets", {}) \
         .setdefault("Current Assets", {}) \
         .setdefault("Bank Balance", {"credit": 0.0, "debit": 0.0, "cnt": 0})

    structure["BALANCE_SHEET"]["Assets"]["Current Assets"]["Bank Balance"]["debit"] = bank_debits
    structure["BALANCE_SHEET"]["Assets"]["Current Assets"]["Bank Balance"]["credit"] = bank_credits
    structure["BALANCE_SHEET"]["Assets"]["Current Assets"]["Bank Balance"]["cnt"] = len(all_txns)

    users = db.get_users()
    uname = next((u["name"] for u in users if u["id"] == user_id), "Account Holder")

    total_income = total_expenditure = total_assets = total_liabilities = total_equity = 0.0
    lines = [
        "=" * W,
        f"{'HNI  —  COMPLETE LEDGER REPORT':^{W}}",
        f"{'Account Holder: ' + uname:^{W}}",
        f"{'Generated: ' + datetime.now().strftime('%d-%b-%Y  %H:%M'):^{W}}",
        "=" * W,
        f"\n  {'Accounting Basis: Cash Basis (bank statement = cash book)':^{W-4}}",
        f"  {'Debits = expenses already paid. Liabilities = only outstanding amounts.':^{W-4}}\n",
    ]
    
    # ── Book 1: INCOME & EXPENDITURE ─────────────────────────────────────────
    lines.append(hdr1("BOOK I  —  INCOME & EXPENDITURE ACCOUNT  (P&L)"))
    for section in ["Income", "Expenditure"]:
        if "INCOME_EXPENSE" not in structure:
            continue
        if section not in structure["INCOME_EXPENSE"]:
            continue

        lines.append(f"\n  ▌ {section.upper()}")
        sec_total = 0.0

        for grp, accounts in structure["INCOME_EXPENSE"][section].items():
            grp_total = 0.0
            grp_lines = []

            for acct, data in accounts.items():
                amt = normal_balance(section, data)
                if abs(amt) < 0.005:
                    continue

                grp_total += amt
                attr = next((v for k, v in ATTRIBUTION.items() if LEDGER_MAP.get(k, ("", "", "", ""))[3] == acct), "")

                grp_lines.append(f"      ├ {acct:<55} {amtf(amt)}  [{data['cnt']} txns]")
                if attr:
                    grp_lines.append(f"      │  ↳ {attr}")

            if not grp_lines:
                continue

            lines.append(f"\n    ┌── {grp}")
            lines.extend(grp_lines)
            lines.append(f"    └── {grp:<55} {amtf(grp_total)}")
            sec_total += grp_total

        if section == "Income":
            total_income = sec_total
            lines.append(f"\n  {bar('-')}")
            lines.append(f"  {'TOTAL INCOME':<71} {amtf(sec_total)}")
            lines.append(f"  {bar('-')}")
        else:
            total_expenditure = sec_total
            lines.append(f"\n  {bar('-')}")
            lines.append(f"  {'TOTAL EXPENDITURE':<71} {amtf(sec_total)}")
            lines.append(f"  {bar('-')}")

    net_surplus = total_income - total_expenditure
    lines.append(f"\n  {bar('═')}")
    lines.append(f"  {'NET SURPLUS (Income – Expenditure)':<71} {amtf(net_surplus)}")
    lines.append(f"  {bar('═')}")
    lines.append(f"  (This surplus increases owner equity / retained earnings)\n")

    # ── Book 2: BALANCE SHEET ─────────────────────────────────────────────────
    lines.append(hdr1("BOOK II  —  BALANCE SHEET  (Net Worth Statement)"))

    for bs_section in ["Assets", "Liabilities", "Equity"]:
        if "BALANCE_SHEET" not in structure:
            continue
        if bs_section not in structure["BALANCE_SHEET"]:
            continue

        lines.append(f"\n  ▌ {bs_section.upper()}")
        sec_total = 0.0

        for grp, accounts in structure["BALANCE_SHEET"][bs_section].items():
            grp_total = 0.0
            grp_lines = []

            for acct, data in accounts.items():
                amt = normal_balance(bs_section, data)
                if abs(amt) < 0.005:
                    continue

                grp_total += amt
                attr = next((v for k, v in ATTRIBUTION.items() if LEDGER_MAP.get(k, ("", "", "", ""))[3] == acct), "")

                grp_lines.append(f"      ├ {acct:<55} {amtf(amt)}  [{data['cnt']} txns]")
                if attr:
                    grp_lines.append(f"      │  ↳ {attr}")

            if not grp_lines:
                continue

            lines.append(f"\n    ┌── {grp}")
            lines.extend(grp_lines)
            lines.append(f"    └── {grp:<55} {amtf(grp_total)}")
            sec_total += grp_total

        if bs_section == "Assets":
            total_assets = sec_total
        if bs_section == "Liabilities":
            total_liabilities = sec_total
        if bs_section == "Equity":
            total_equity = sec_total

        lines.append(f"\n  {bar('-')}")
        lines.append(f"  {'TOTAL ' + bs_section.upper():<71} {amtf(sec_total)}")
        lines.append(f"  {bar('-')}")

    # FIX C6+C9: Include net P&L in equity for balance check
    retained_earnings = net_surplus
    total_equity_before_adjustment = total_equity + retained_earnings
    balance_gap = total_assets - (total_liabilities + total_equity_before_adjustment)
    balancing_adjustment = balance_gap if abs(balance_gap) >= 0.01 else 0.0
    total_equity_with_pl = total_equity_before_adjustment + balancing_adjustment

    # Balance check
    lines.append(hdr1("BOOK III  —  BALANCE CHECK"))
    lines.append(f"  {'Total Assets (A)':<55} {amtf(total_assets)}")
    lines.append(f"  {'Total Liabilities (B)':<55} {amtf(total_liabilities)}")
    lines.append(f"  {'Total Equity from BS entries (C)':<55} {amtf(total_equity)}")
    lines.append(f"  {'Retained Earnings from P&L (D)':<55} {amtf(retained_earnings)}")
    if balancing_adjustment:
        lines.append(f"  {'Final Balancing Adjustment (E)':<55} {amtf(balancing_adjustment)}")
    lines.append(f"  {bar('-')}")
    lines.append(f"  {'Total Equity incl. P&L and Adjustment':<55} {amtf(total_equity_with_pl)}")
    balance_gap = total_assets - (total_liabilities + total_equity_with_pl)
    lines.append(f"  {bar('═')}")
    lines.append(f"  {'BALANCE GAP  Assets – (Liabilities + Equity)':<55} {amtf(balance_gap)}")
    lines.append(f"  {bar('═')}")
    if abs(balance_gap) < 0.01:
        lines.append(f"  ✅ BALANCE SHEET IN BALANCE")
    else:
        lines.append(f"  ⚠️ Balance gap = {amtf(balance_gap)} — check suspense or unclassified items")

    # Suspense summary
    if "SUSPENSE" in structure:
        lines.append(hdr1("SUSPENSE ACCOUNTS  (Requires Manual Review)"))
        for section, grps in structure["SUSPENSE"].items():
            for grp, accounts in grps.items():
                for acct, data in accounts.items():
                    amt = data.get("credit",0) + data.get("debit",0)
                    lines.append(f"  ► {acct:<60} {amtf(amt)}  [{data['cnt']} txns]")

    lines.append(f"\n{'═'*W}")
    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════════════
# WEB SERVER
# ══════════════════════════════════════════════════════════════════════════════

from http.server import BaseHTTPRequestHandler, HTTPServer

_db = DBStore(DB_PATH)
_current_user_id: Optional[str] = None

HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>HNI Ledger v0655</title>
<!-- BUILD:20260331_065500 -->
<style>

:root{--blue:#1565c0;--green:#1b5e20;--red:#b71c1c;--amber:#f57f17;--teal:#004d40;
  --bg:#f0f2f5;--card:#fff;--border:#dde1e7;--mono:'Courier New',monospace}
*{box-sizing:border-box}
body{margin:0;font-family:'Segoe UI',Arial,sans-serif;background:var(--bg);color:#1a1a2e}
header{background:linear-gradient(135deg,#0d47a1,#1565c0,#0d47a1);color:#fff;padding:18px 28px;display:flex;align-items:center;gap:16px;box-shadow:0 2px 12px rgba(0,0,0,.2)}
header h1{margin:0;font-size:22px;font-weight:800;letter-spacing:.5px}
header .sub{font-size:12px;opacity:.75;margin-top:3px;letter-spacing:.3px}
.tabs{display:flex;background:#fff;border-bottom:2px solid var(--border);padding:0 28px;gap:2px;box-shadow:0 1px 4px rgba(0,0,0,.06)}
.tab{padding:14px 20px;cursor:pointer;border-bottom:3px solid transparent;font-weight:700;color:#777;font-size:13px;transition:.15s;letter-spacing:.3px}
.tab.active{color:var(--blue);border-bottom-color:var(--blue)}
.pane{display:none;padding:22px 28px}.pane.active{display:block}
.card{
background:rgba(255,255,255,.95);
border-radius:18px;
padding:22px;
margin-bottom:18px;
box-shadow:0 10px 30px rgba(0,0,0,.08);
border:1px solid #e8edf5;
backdrop-filter:blur(8px);
}
.card h3{margin:0 0 16px;font-size:14px;color:#333;border-bottom:1px solid #f0f0f0;padding-bottom:10px;font-weight:700;letter-spacing:.3px}
label{font-size:12.5px;font-weight:700;color:#555;display:block;margin-bottom:5px;margin-top:12px;text-transform:uppercase;letter-spacing:.5px}
input,select,textarea{width:100%;padding:10px 14px;border:1.5px solid var(--border);border-radius:8px;font-size:13px;font-family:inherit;outline:none;transition:border .2s;background:#fafafa}
input:focus,select:focus,textarea:focus{border-color:var(--blue);background:#fff;box-shadow:0 0 0 3px rgba(21,101,192,.1)}
textarea{resize:vertical;font-family:var(--mono);font-size:12px}
.btn{display:inline-flex;align-items:center;gap:7px;padding:10px 20px;margin:4px 4px 4px 0;border:none;border-radius:10px;font-size:13px;font-weight:800;cursor:pointer;transition:all .2s ease;letter-spacing:.3px;box-shadow:0 6px 14px rgba(0,0,0,.08)}
.btn:hover{transform:translateY(-2px);box-shadow:0 10px 20px rgba(0,0,0,.12)}
.bp{background:#1565c0;color:#fff}.bp:hover{background:#0d47a1}
.bg{background:#1b5e20;color:#fff}.bg:hover{background:#145a14}
.bo{background:#fff;color:#1565c0;border:1.5px solid #1565c0}.bo:hover{background:#e3f2fd}
.grid2{display:grid;grid-template-columns:1fr 1fr;gap:14px}
.grid3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:14px}
.alert{padding:12px 16px;border-radius:8px;font-size:13px;margin-bottom:14px;border-left:4px solid;font-weight:500}
.alert-info{background:#e3f2fd;color:#1565c0;border-color:#1565c0}
.alert-ok{background:#e8f5e9;color:#1b5e20;border-color:#1b5e20}
.alert-err{background:#ffebee;color:#b71c1c;border-color:#b71c1c}
table{width:100%;border-collapse:collapse;font-size:12px}
th{background:#f5f7fa;padding:10px 12px;text-align:left;border-bottom:2px solid var(--border);font-weight:800;color:#444;white-space:nowrap;text-transform:uppercase;font-size:11px;letter-spacing:.5px}
td{padding:9px 12px;border-bottom:1px solid #f1f3f5;vertical-align:top}
tr:hover td{background:#f8faff}
.badge{display:inline-block;padding:2px 9px;border-radius:10px;font-size:10.5px;font-weight:800;white-space:nowrap;letter-spacing:.3px}
.b-bs{background:#e8f5e9;color:#1b5e20}
.b-ie{background:#e3f2fd;color:#1565c0}
.b-cap{background:#fff3e0;color:#e65100}
.b-sus{background:#fce4ec;color:#880e4f}
.b-cr{background:#e8f5e9;color:#1b5e20}
.b-dr{background:#ffebee;color:#b71c1c}
.book-tag{display:inline-block;padding:3px 10px;border-radius:6px;font-size:10px;font-weight:800;letter-spacing:.5px}
.book-bs{background:#e8f5e9;color:#1b5e20}
.book-ie{background:#e3f2fd;color:#1565c0}
.book-cap{background:#fff8e1;color:#f57f17}
.book-sus{background:#fce4ec;color:#880e4f}
.drop-zone{border:2px dashed #b0bec5;border-radius:12px;padding:36px;text-align:center;cursor:pointer;background:#fafbff;transition:.2s}
.drop-zone:hover,.drop-zone.dragover{border-color:var(--blue);background:#e8f0fe}
.drop-zone input[type=file]{display:none}
.stat-grid{display:grid;grid-template-columns:repeat(5,1fr);gap:12px;margin-bottom:18px}
.stat-card{background:#fff;border-radius:12px;padding:16px;text-align:center;box-shadow:0 1px 8px rgba(0,0,0,.08);border:1px solid rgba(0,0,0,.05)}
.stat-card .val{font-size:19px;font-weight:800}
.stat-card .lbl{font-size:11px;color:#999;margin-top:4px;text-transform:uppercase;letter-spacing:.5px}
pre#ledger-out{background:#0d1117;color:#e6edf3;padding:20px;border-radius:12px;overflow:auto;max-height:650px;font-family:var(--mono);font-size:11.5px;line-height:1.7;white-space:pre-wrap;border:1px solid #30363d}
.section-head{background:linear-gradient(90deg,#0d47a1,#1565c0);color:#fff;padding:11px 16px;border-radius:8px 8px 0 0;font-weight:800;font-size:13px;display:flex;justify-content:space-between;align-items:center;cursor:pointer;user-select:none;margin-bottom:0}
.section-body{border:1px solid #e0e0e0;border-top:none;border-radius:0 0 8px 8px;margin-bottom:12px}
.grp-head{background:#f5f7fa;padding:9px 20px;font-weight:700;font-size:12.5px;color:#444;display:flex;justify-content:space-between;cursor:pointer;user-select:none;border-bottom:1px solid #eee}
.acct-head{padding:8px 28px;font-weight:600;font-size:12px;color:#555;display:flex;justify-content:space-between;cursor:pointer;user-select:none;border-bottom:1px solid #f5f5f5;background:#fafbff}
.acct-note{font-size:11px;color:#888;padding:4px 36px 8px;font-style:italic;border-bottom:1px solid #f5f5f5}
@media(max-width:700px){.grid2,.grid3,.stat-grid{grid-template-columns:1fr}}
/* Trace modal */
#trace-modal-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.45);z-index:9000;align-items:center;justify-content:center}
#trace-modal-overlay.open{display:flex}
#trace-modal{background:#fff;border-radius:14px;width:92%;max-width:900px;max-height:85vh;display:flex;flex-direction:column;box-shadow:0 8px 40px rgba(0,0,0,.25);overflow:hidden}
#trace-modal-head{background:linear-gradient(135deg,#0d47a1,#1565c0);color:#fff;padding:14px 20px;display:flex;justify-content:space-between;align-items:center}
#trace-modal-head h4{margin:0;font-size:14px;font-weight:800}
#trace-modal-close{background:none;border:none;color:#fff;font-size:20px;cursor:pointer;line-height:1;padding:0 4px}
#trace-modal-body{padding:16px;overflow-y:auto;flex:1}
#trace-modal-summary{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:12px}
.trace-chip{background:#e3f2fd;border:1.5px solid #90caf9;border-radius:8px;padding:6px 14px;font-size:12px;font-weight:700;color:#1565c0}
#trace-modal-table{width:100%;border-collapse:collapse;font-size:11.5px}
#trace-modal-table th{background:#f5f7fa;padding:8px 10px;text-align:left;border-bottom:2px solid #e0e0e0;font-weight:800;color:#444;text-transform:uppercase;font-size:10px;letter-spacing:.4px}
#trace-modal-table td{padding:7px 10px;border-bottom:1px solid #f1f3f5;vertical-align:top}
#trace-modal-table tr:hover td{background:#f8faff}
</style>
</head>
<!-- TRACE MODAL -->
<div id="trace-modal-overlay" onclick="if(event.target===this)closeTrace()">
  <div id="trace-modal">
    <div id="trace-modal-head">
      <h4 id="trace-modal-title">Transaction Trace</h4>
      <button id="trace-modal-close" onclick="closeTrace()">✕</button>
    </div>
    <div id="trace-modal-body">
      <div id="trace-modal-summary"></div>
      <div id="trace-modal-loading" style="text-align:center;padding:30px;color:#aaa">Loading transactions…</div>
      <div id="trace-modal-content" style="display:none;overflow-x:auto">
        <table id="trace-modal-table">
          <thead><tr>
            <th>Date</th><th>Counterparty</th><th style="text-align:right">Amount</th><th>Actions</th>
          </tr></thead>
          <tbody id="trace-modal-rows"></tbody>
        </table>
      </div>
      <div id="trace-modal-empty" style="display:none;text-align:center;padding:30px;color:#aaa">No transactions found for this bucket.</div>
    </div>
  </div>
</div>
<body>
<header>
  <div style="font-size:36px">📒</div>
  <div>
    <h1>HNI Ledger Classifier</h1>
    <div class="sub" id="hdr-sub">3-Book Accounting · Balance Sheet · Income &amp; Expenditure · Capital</div>
  </div>
</header>
<div class="tabs">
  <div class="tab active" onclick="tab('classify')">📋 Upload</div>
  <div class="tab" onclick="tab('review')">🔍 Review &amp; Approve</div>
  <div class="tab" onclick="tab('statements')">📑 Statements</div>
  <div class="tab" onclick="tab('ledger')">📊 Ledger</div>
  <div class="tab" onclick="tab('report')">⚖️ Full Report</div>
  <div class="tab" onclick="tab('trading-cashflow')">💹 Trading Cash Flow</div>
  <div class="tab" onclick="tab('setup')">⚙️ Setup</div>
</div>

<!-- CLASSIFY -->
<div class="pane active" id="pane-classify">
  <div id="status" class="alert alert-info">Ready — create a user in Setup, then upload a bank statement.</div>

  <div class="card">
    <h3>📂 Upload Bank Statement (.xlsx / .xls / .xlsm / .pdf)</h3>
    <div class="drop-zone" id="dropZone" onclick="document.getElementById('xlFile').click()"
         ondragover="event.preventDefault();this.classList.add('dragover')"
         ondragleave="this.classList.remove('dragover')" ondrop="handleDrop(event)">
      <!-- ── PDF SUPPORT (NEW) ── accept PDF uploads alongside Excel -->
      <input type="file" id="xlFile" accept=".xlsx,.xls,.xlsm,.pdf" onchange="setFile(event.target.files[0])">
      <div style="font-size:44px;margin-bottom:10px">📄</div>
      <p style="font-weight:700;margin:0 0 6px">Click to upload or drag &amp; drop</p>
      <small style="color:#888">HDFC / ICICI / Axis / SBI / Union Bank / PNB statement supported · Auto-classifies into 3-book ledger</small>
    </div>
    <div id="file-preview" style="display:none;margin-top:10px;padding:10px 16px;background:#f5f7fa;border-radius:8px;font-size:13px"></div>
    <div class="grid2" style="margin-top:14px">
      <div>
        <label style="font-size:12px;font-weight:700;color:#555;text-transform:uppercase;letter-spacing:.5px">Bank Account Name</label>
        <input id="acct_name" placeholder="e.g. HDFC Savings, ICICI Salary" style="margin-top:6px;margin-bottom:10px">
      </div>
      <div>
        <label style="font-size:12px;font-weight:700;color:#555;text-transform:uppercase;letter-spacing:.5px">File Password (if protected)</label>
        <input id="file_password" type="password" placeholder="Enter password only if the file is locked" style="margin-top:6px;margin-bottom:10px">
      </div>
    </div>
    <div style="margin-top:4px">
      <button class="btn bp" id="uploadBtn" onclick="uploadForReview()" disabled>🔍 Classify &amp; Review</button>
      <button class="btn bo" onclick="clearUpload()">✕ Clear</button>
    </div>
    <div id="profile-gate-msg" style="display:none;margin-top:10px;font-size:13px;color:#b71c1c;font-weight:700"></div>
  </div>

  <div class="card" id="uploaded-files-card">
    <h3>📁 Uploaded Files</h3>
    <div id="uploaded-files-body" style="overflow-x:auto">
      <p style="color:#bbb;font-size:13px">Log in to see previously uploaded statements.</p>
    </div>
  </div>

</div>

<!-- LEDGER -->
<div class="pane" id="pane-ledger">
  <div class="stat-grid" id="stat-cards"></div>
  <div class="card">
    <h3>Filter</h3>
    <div class="grid4">
      <div><label>Book</label>
        <select id="f_book"><option value="">All Books</option>
          <option value="BALANCE_SHEET">Balance Sheet</option>
          <option value="INCOME_EXPENSE">Income &amp; Expenditure</option>
          <option value="SUSPENSE">Suspense</option>
        </select></div>
      <div><label>Type</label>
        <select id="f_type"><option value="">All</option><option>credit</option><option>debit</option></select></div>
      <div><label>Search narration</label><input id="f_search" placeholder="SWIGGY, ZERODHA, SIP…"></div>
      <div><label>Counterparty</label><input id="f_counterparty" placeholder="Arinjay, Swiggy, Zerodha…"></div>
      <div><label>Account</label>
        <select id="f_account"><option value="">All Accounts</option></select>
      </div>
    </div>
    <div style="margin-top:12px">
      <button class="btn bp" onclick="loadLedger()">🔍 Filter</button>
      <button class="btn bo" onclick="loadLedger(true)">Show All</button>
    </div>
  </div>
  <div class="card">
    <h3>Filtered Ledger Entries <span id="ledger-cnt" style="color:#aaa;font-weight:400"></span></h3>
    <div id="ledger-table" style="overflow-x:auto"><p style="color:#bbb;font-size:13px">Apply a filter to see results here.</p></div>
  </div>
</div>

<!-- REVIEW & APPROVE -->
<div class="pane" id="pane-review">
  <div id="review-status" class="alert alert-info" style="display:none"></div>

  <!-- Approval action bar -->
  <div id="review-action-bar" class="card" style="display:none">
    <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:12px">
      <div>
        <span style="font-size:22px;font-weight:800;color:#1565c0" id="rv-total">0</span>
        <span style="font-size:13px;color:#555;margin-left:6px">transactions staged</span>
        &nbsp;·&nbsp;
        <span style="color:#f57f17;font-weight:700" id="rv-reclassified">0 reclassified</span>
        &nbsp;·&nbsp;
        <span style="color:#b71c1c;font-weight:700" id="rv-anomaly">0 anomalies</span>
      </div>
      <div style="display:flex;gap:10px;flex-wrap:wrap">
        <button class="btn bg" onclick="approveAll()">✅ Approve All &amp; Post to Ledger</button>
        <button class="btn bo" style="color:#b71c1c;border-color:#b71c1c" onclick="discardPending()">🗑 Discard Batch</button>
      </div>
    </div>
    <div style="margin-top:10px;font-size:12px;color:#888">
      Each row has a <strong>dropdown</strong> to pick a new category. Change as needed, then click <strong>Approve All</strong> to commit to ledger.
    </div>
    <!-- Quick filter bar -->
    <div style="margin-top:12px;display:flex;gap:8px;flex-wrap:wrap;align-items:center">
      <input id="rv-search" placeholder="🔍 Filter by narration or counterparty…"
        oninput="filterReview()"
        style="flex:1;min-width:200px;padding:7px 12px;border:1.5px solid #e0e0e0;border-radius:7px;font-size:12px">
      <select id="rv-filter-status" onchange="filterReview()"
        style="padding:7px 10px;border:1.5px solid #e0e0e0;border-radius:7px;font-size:12px">
        <option value="">All Statuses</option>
        <option value="PENDING">Pending</option>
        <option value="RECLASSIFIED">Reclassified</option>
      </select>
      <select id="rv-filter-book" onchange="filterReview()"
        style="padding:7px 10px;border:1.5px solid #e0e0e0;border-radius:7px;font-size:12px">
        <option value="">All Books</option>
        <option value="INCOME_EXPENSE">Income &amp; Exp</option>
        <option value="BALANCE_SHEET">Balance Sheet</option>
        <option value="SUSPENSE">Suspense</option>
      </select>
    </div>
    <div id="add-rule-panel" style="margin:12px 0;padding:10px;background:#fff3e0;border-radius:6px;display:flex;gap:8px;align-items:center;flex-wrap:wrap">
    <strong style="color:#e65100">➕ Add Classifier Rule:</strong>
    <input id="rule-pattern" placeholder="Narration pattern (e.g. SWIGGY)" style="flex:1;min-width:180px;padding:4px 8px;border:1px solid #ccc;border-radius:4px"/>
    <select id="rule-match-mode" style="padding:4px 8px;border:1px solid #ccc;border-radius:4px">
        <option value="contains">contains</option>
        <option value="startswith">startswith</option>
        <option value="regex">regex</option>
    </select>
    <select id="rule-txn-type" style="padding:4px 8px;border:1px solid #ccc;border-radius:4px">
        <option value="debit">debit</option>
        <option value="credit">credit</option>
        <option value="both">both</option>
    </select>
    <select id="rule-ledger-key" style="flex:1;min-width:160px;padding:4px 8px;border:1px solid #ccc;border-radius:4px"></select>
    <button onclick="addCustomRule()" style="background:#e65100;color:#fff;border:none;padding:5px 14px;border-radius:4px;cursor:pointer">Save Rule</button>
    <span id="rule-save-msg" style="color:green;font-size:12px"></span>
    </div>
  </div>

  <!-- Review rows container -->
  <div id="review-rows-container" style="display:none">
    <div id="review-rows"></div>
  </div>

  <div id="review-empty" style="text-align:center;padding:60px;color:#aaa">
    <div style="font-size:48px;margin-bottom:12px">📋</div>
    <div style="font-size:16px;font-weight:700">No pending transactions</div>
    <div style="font-size:13px;margin-top:6px">Upload a bank statement in the <strong>Upload</strong> tab first.</div>
  </div>
</div>

<!-- FINANCIAL STATEMENTS -->
<div class="pane" id="pane-statements">
  <!-- Filter bar -->
  <div class="card" style="padding:16px 20px;margin-bottom:12px">
    <div style="display:flex;gap:12px;flex-wrap:wrap;align-items:flex-end">
      <div>
        <label style="font-size:11px;font-weight:700;color:#555;text-transform:uppercase;letter-spacing:.5px;display:block;margin-bottom:4px">From</label>
        <input type="date" id="stmt_date_from" style="padding:7px 10px;width:150px">
      </div>
      <div>
        <label style="font-size:11px;font-weight:700;color:#555;text-transform:uppercase;letter-spacing:.5px;display:block;margin-bottom:4px">To</label>
        <input type="date" id="stmt_date_to" style="padding:7px 10px;width:150px">
      </div>
      <div>
        <label style="font-size:11px;font-weight:700;color:#555;text-transform:uppercase;letter-spacing:.5px;display:block;margin-bottom:4px">Account</label>
        <select id="stmt_account" onchange="onAccountChange()" style="padding:7px 10px;min-width:160px">
          <option value="">All Accounts</option>
        </select>
      </div>
      <div>
        <label style="font-size:11px;font-weight:700;color:#555;text-transform:uppercase;letter-spacing:.5px;display:block;margin-bottom:4px">Year</label>
        <select id="stmt_year" onchange="applyYearFilter()" style="padding:7px 10px;min-width:140px">
          <option value="">All Years</option>
        </select>
      </div>
      <button class="btn bp" onclick="refreshStmts()" style="padding:7px 18px">🔄 Generate</button>
      <button class="btn bo" onclick="clearStmtFilters()" style="padding:7px 14px;font-size:12px">✕ Clear</button>
      <button class="btn bo" onclick="loadYearlyStatements()" style="padding:7px 14px;font-size:12px">📅 Load Yearly</button>
    </div>
  </div>

  <div id="yearly-reports-wrap" class="card" style="display:block;margin-bottom:12px"></div>

  <!-- Sub-nav pills -->
  <div style="display:flex;gap:8px;margin-bottom:18px;flex-wrap:wrap;align-items:center">
    <button id="stmt-pill-income"  class="btn bp" onclick="showStmt('income')" style="border-radius:20px;padding:7px 18px">📈 Income Statement</button>
    <button id="stmt-pill-balance" class="btn bg" onclick="showStmt('balance')" style="border-radius:20px;padding:7px 18px">🏦 Balance Sheet</button>
    <button id="stmt-pill-pl"      class="btn bo" onclick="showStmt('pl')" style="border-radius:20px;padding:7px 18px;border-color:#7b1fa2;color:#7b1fa2">💼 Profit &amp; Loss</button>
    <button class="btn bo" onclick="showStmt('all')" style="border-radius:20px;padding:7px 14px;margin-left:auto;font-size:12px">Show All</button>
  </div>

  <!-- Income Statement card -->
  <div id="stmt-income" class="card" style="display:none">
    <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:14px;flex-wrap:wrap;gap:8px">
      <div>
        <h3 style="margin:0 0 2px">📈 Income Statement</h3>
        <div style="font-size:12px;color:#888" id="stmt-income-period">Individual — Income &amp; Expenditure</div>
      </div>
      <div style="display:flex;gap:8px;flex-wrap:wrap">
        <button class="btn bo" style="font-size:11px;padding:5px 12px" onclick="copyStmt('income')">📋 Copy</button>
        <button class="btn bo" style="font-size:11px;padding:5px 12px" onclick="downloadStmt('income')">📥 Download</button>
        <button class="btn bo" id="btn-copy-income-acct" style="font-size:11px;padding:5px 12px;display:none" onclick="copyStmt('income_acct')">📋 Copy (This Account)</button>
        <button class="btn bo" id="btn-download-income-acct" style="font-size:11px;padding:5px 12px;display:none" onclick="downloadStmt('income_acct')">📥 Download (This Account)</button>
        <button class="btn bo" style="font-size:11px;padding:5px 12px" onclick="downloadReportPdf('summary','income_statement')">🧾 Download Summary PDF</button>
        <button class="btn bo" style="font-size:11px;padding:5px 12px" onclick="downloadReportPdf('detailed','income_statement')">📄 Download Income PDF</button>
      </div>
    </div>
    <div id="stmt-income-chips" style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:14px"></div>
    <div id="stmt-income-split" style="display:none">
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px">
        <div>
          <div style="background:#e3f2fd;border-radius:8px;padding:8px 14px;margin-bottom:10px;font-size:12px;font-weight:700;color:#1565c0" id="stmt-income-acct-label">📄 This Account</div>
          <div id="stmt-income-acct-chips" style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:10px"></div>
          <div id="stmt-income-acct-body" style="font-size:11.5px"></div>
        </div>
        <div>
          <div style="background:#f3e5f5;border-radius:8px;padding:8px 14px;margin-bottom:10px;font-size:12px;font-weight:700;color:#6a1b9a">📊 All Accounts Combined</div>
          <div id="stmt-income-all-chips" style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:10px"></div>
          <div id="stmt-income-all-body" style="font-size:11.5px"></div>
        </div>
      </div>
    </div>
    <div id="stmt-income-single">
      <div id="stmt-income-body"></div>
    </div>
  </div>

  <!-- Balance Sheet card -->
  <div id="stmt-balance" class="card" style="display:none">
    <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:14px;flex-wrap:wrap;gap:8px">
      <div>
        <h3 style="margin:0 0 2px">🏦 Balance Sheet / Net Worth</h3>
        <div style="font-size:12px;color:#888" id="stmt-balance-period">Individual — Assets · Liabilities · Equity</div>
      </div>
      <div style="display:flex;gap:8px;flex-wrap:wrap">
        <button class="btn bo" style="font-size:11px;padding:5px 12px" onclick="copyStmt('balance')">📋 Copy</button>
        <button class="btn bo" style="font-size:11px;padding:5px 12px" onclick="downloadStmt('balance')">📥 Download</button>
        <button class="btn bo" id="btn-copy-balance-acct" style="font-size:11px;padding:5px 12px;display:none" onclick="copyStmt('balance_acct')">📋 Copy (This Account)</button>
        <button class="btn bo" id="btn-download-balance-acct" style="font-size:11px;padding:5px 12px;display:none" onclick="downloadStmt('balance_acct')">📥 Download (This Account)</button>
        <button class="btn bo" style="font-size:11px;padding:5px 12px" onclick="downloadReportPdf('summary','balance_sheet')">🧾 Download Summary PDF</button>
        <button class="btn bo" style="font-size:11px;padding:5px 12px" onclick="downloadReportPdf('detailed','balance_sheet')">📄 Download Balance Sheet PDF</button>
      </div>
    </div>
    <div id="stmt-balance-chips" style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:14px"></div>
    <div id="stmt-balance-split" style="display:none">
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px">
        <div>
          <div style="background:#e3f2fd;border-radius:8px;padding:8px 14px;margin-bottom:10px;font-size:12px;font-weight:700;color:#1565c0" id="stmt-balance-acct-label">📄 This Account</div>
          <div id="stmt-balance-acct-chips" style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:10px"></div>
          <div id="stmt-balance-acct-body" style="font-size:11.5px"></div>
        </div>
        <div>
          <div style="background:#f3e5f5;border-radius:8px;padding:8px 14px;margin-bottom:10px;font-size:12px;font-weight:700;color:#6a1b9a">📊 All Accounts Combined</div>
          <div id="stmt-balance-all-chips" style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:10px"></div>
          <div id="stmt-balance-all-body" style="font-size:11.5px"></div>
        </div>
      </div>
    </div>
    <div id="stmt-balance-single">
      <div id="stmt-balance-body"></div>
    </div>
  </div>

  <!-- Profit & Loss card -->
  <div id="stmt-pl" class="card" style="display:none">
    <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:14px;flex-wrap:wrap;gap:8px">
      <div>
        <h3 style="margin:0 0 2px">💼 Profit &amp; Loss Statement</h3>
        <div style="font-size:12px;color:#888" id="stmt-pl-period">Organisation — Revenue · Operating · Financial · Tax</div>
      </div>
      <div style="display:flex;gap:8px;flex-wrap:wrap">
        <button class="btn bo" style="font-size:11px;padding:5px 12px" onclick="copyStmt('pl')">📋 Copy</button>
        <button class="btn bo" style="font-size:11px;padding:5px 12px" onclick="downloadStmt('pl')">📥 Download</button>
        <button class="btn bo" id="btn-copy-pl-acct" style="font-size:11px;padding:5px 12px;display:none" onclick="copyStmt('pl_acct')">📋 Copy (This Account)</button>
        <button class="btn bo" id="btn-download-pl-acct" style="font-size:11px;padding:5px 12px;display:none" onclick="downloadStmt('pl_acct')">📥 Download (This Account)</button>
      </div>
    </div>
    <div id="stmt-pl-chips" style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:14px"></div>
    <div id="stmt-pl-split" style="display:none">
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px">
        <div>
          <div style="background:#e3f2fd;border-radius:8px;padding:8px 14px;margin-bottom:10px;font-size:12px;font-weight:700;color:#1565c0" id="stmt-pl-acct-label">📄 This Account</div>
          <div id="stmt-pl-acct-chips" style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:10px"></div>
          <div id="stmt-pl-acct-body" style="font-size:11.5px"></div>
        </div>
        <div>
          <div style="background:#f3e5f5;border-radius:8px;padding:8px 14px;margin-bottom:10px;font-size:12px;font-weight:700;color:#6a1b9a">📊 All Accounts Combined</div>
          <div id="stmt-pl-all-chips" style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:10px"></div>
          <div id="stmt-pl-all-body" style="font-size:11.5px"></div>
        </div>
      </div>
    </div>
    <div id="stmt-pl-single">
      <div id="stmt-pl-body"></div>
    </div>
  </div>

  <div id="stmt-empty" style="text-align:center;padding:60px;color:#aaa">
    <div style="font-size:48px;margin-bottom:12px">📑</div>
    <div style="font-size:16px;font-weight:700">No statements yet</div>
    <div style="font-size:13px;margin-top:6px">Approve a transaction batch first — statements generate automatically.</div>
  </div>
</div>

<!-- REPORT -->
<div class="pane" id="pane-report">
  <div class="card">
    <h3>Generate Full Ledger Report</h3>
    <p style="font-size:13px;color:#666;margin:0 0 14px">
      Produces Book I (Income &amp; Expenditure), Book II (Balance Sheet), Book III (Balance Check) with attributions.
    </p>
    <button class="btn bg" onclick="genReport()">📒 Generate Full Ledger Report</button>
    <button class="btn bo" onclick="document.getElementById('ledger-out').textContent=''" style="margin-left:8px">Clear</button>
  </div>
  <div class="card"><pre id="ledger-out">Click Generate to produce the full 3-book ledger report.</pre></div>
</div>

<!-- TRADING CASH FLOW -->
<div class="pane" id="pane-trading-cashflow">
  <div class="card">
    <h3>💹 Trading Account — Cash Flow Statement</h3>
    <p style="font-size:13px;color:#666;margin:0 0 14px">
      Indirect-method cash flow for 5paisa / broker ledger. Upload the trading ledger XLS first, then generate the statement.
    </p>

    <!-- Upload section -->
    <div style="border:1px solid #e0e0e0;border-radius:10px;padding:16px;margin-bottom:16px;background:#fafbff">
      <div style="font-weight:800;font-size:13px;color:#1565c0;margin-bottom:10px">📂 Step 1 — Upload Trading Ledger (XLS)</div>
      <div class="drop-zone" id="tcf-drop-zone" onclick="document.getElementById('tcf-file').click()"
           ondragover="event.preventDefault();this.classList.add('dragover')"
           ondragleave="this.classList.remove('dragover')" ondrop="tcfHandleDrop(event)"
           style="padding:20px">
        <!-- ── PDF SUPPORT (NEW) ── accept PDF uploads alongside Excel -->
        <input type="file" id="tcf-file" accept=".xls,.xlsx,.xlsm,.pdf" onchange="tcfSetFile(event.target.files[0])">
        <div style="font-size:32px;margin-bottom:6px">📄</div>
        <p style="font-weight:700;margin:0 0 4px;font-size:13px">Click to upload or drag &amp; drop</p>
        <small style="color:#888">5paisa / broker XLS ledger — auto-parsed via ZIP+XML</small>
      </div>
      <div id="tcf-file-preview" style="display:none;margin-top:8px;padding:8px 14px;background:#f5f7fa;border-radius:8px;font-size:13px"></div>
      <div style="margin-top:10px">
        <input id="tcf-account-id" placeholder="Account ID (e.g. 58439330)" style="width:220px;display:inline-block">
        &nbsp;
        <button class="btn bp" id="tcf-upload-btn" onclick="tcfUpload()" disabled>⬆️ Upload &amp; Parse</button>
        <span id="tcf-upload-msg" style="font-size:12px;margin-left:10px;color:#666"></span>
      </div>
    </div>

    <!-- Generate section -->
    <div style="border:1px solid #e0e0e0;border-radius:10px;padding:16px;background:#f9fff9">
      <div style="font-weight:800;font-size:13px;color:#1b5e20;margin-bottom:10px">📊 Step 2 — Generate Cash Flow Statement</div>
      <div style="display:flex;gap:12px;flex-wrap:wrap;align-items:flex-end">
        <div>
          <label style="font-size:11px;font-weight:700;color:#555;text-transform:uppercase;letter-spacing:.5px;display:block;margin-bottom:4px">From</label>
          <input type="date" id="tcf-date-from" style="padding:7px 10px;width:150px">
        </div>
        <div>
          <label style="font-size:11px;font-weight:700;color:#555;text-transform:uppercase;letter-spacing:.5px;display:block;margin-bottom:4px">To</label>
          <input type="date" id="tcf-date-to" style="padding:7px 10px;width:150px">
        </div>
        <div>
          <label style="font-size:11px;font-weight:700;color:#555;text-transform:uppercase;letter-spacing:.5px;display:block;margin-bottom:4px">Account ID</label>
          <input id="tcf-gen-account" placeholder="58439330" style="padding:7px 10px;width:140px">
        </div>
        <button class="btn bg" onclick="loadTradingCashflow()" style="padding:7px 18px">🔄 Generate Statement</button>
      </div>
    </div>
  </div>

  <!-- Summary chips -->
  <div id="tcf-chips" style="display:none;display:flex;gap:12px;flex-wrap:wrap;margin-bottom:16px;padding:0 2px"></div>

  <!-- Report output -->
  <div class="card" id="tcf-report-card" style="display:none">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px">
      <h3 style="margin:0">Cash Flow Report</h3>
      <button class="btn bo" style="font-size:11px;padding:5px 12px"
              onclick="_copyText(document.getElementById('tcf-report-out').textContent, this)">📋 Copy</button>
    </div>
    <pre id="tcf-report-out" style="background:#0d1117;color:#e6edf3;padding:20px;border-radius:12px;overflow:auto;max-height:650px;font-family:var(--mono);font-size:11.5px;line-height:1.7;white-space:pre-wrap;border:1px solid #30363d"></pre>
  </div>
  <div id="tcf-empty" style="text-align:center;padding:60px;color:#aaa">
    <div style="font-size:48px;margin-bottom:12px">💹</div>
    <div style="font-size:16px;font-weight:700">No trading cash flow yet</div>
    <div style="font-size:13px;margin-top:6px">Upload a 5paisa ledger XLS and click Generate.</div>
  </div>
</div>

<!-- SETUP -->
<div class="pane" id="pane-setup">
  <div id="active-user-banner" style="display:none;background:linear-gradient(135deg,#e8f5e9,#c8e6c9);border:1px solid #a5d6a7;border-radius:12px;padding:14px 20px;margin-bottom:16px;justify-content:space-between;align-items:center">
    <div>
      <div style="font-weight:800;font-size:14px;color:#1b5e20">✅ Active Session</div>
      <div id="active-user-info" style="font-size:13px;color:#2e7d32;margin-top:3px"></div>
    </div>
    <button class="btn bo" onclick="logoutUser()" style="font-size:12px;padding:7px 16px">Switch User</button>
  </div>

  <div class="card" id="user-auth-card">
    <h3>🧭 Classification Setup</h3>
    <p style="font-size:13px;color:#666;margin:0 0 14px">Step 1: Create Account. Step 2: Basic Classification Profile. This improves salary, family loan, broker, and own account classification.</p>
    <div style="display:flex;gap:0;border-bottom:2px solid #f0f0f0;margin-bottom:16px">
      <div id="tab-login" onclick="authTab('login')" style="padding:10px 22px;cursor:pointer;font-weight:800;font-size:13px;color:var(--blue);border-bottom:3px solid var(--blue);letter-spacing:.3px">🔑 Log In</div>
      <div id="tab-create" onclick="authTab('create')" style="padding:10px 22px;cursor:pointer;font-weight:700;font-size:13px;color:#aaa;border-bottom:3px solid transparent;letter-spacing:.3px">➕ New Account</div>
    </div>



    <div id="panel-login">
  <div class="grid2">
    <div><label>Name</label><input id="l_name" placeholder="Enter your name" autocomplete="name"></div>
    <div><label>Phone</label><input id="l_phone" placeholder="9876543210"></div>
  </div>

  <div class="grid2">
    <div><label>Email</label><input id="l_email" placeholder="mani@example.com"></div>
    <div><label>Password *</label><input id="l_password" type="password" placeholder="Enter password" autocomplete="current-password"></div>



 



  </div>

  <div style="margin-top:16px;display:flex;align-items:center;gap:14px">
    <button class="btn bp" onclick="loginUser()">🔑 Log In</button>
    <div id="login-msg" style="font-size:13px"></div>
  </div>

  <div style="margin-top:10px">
    <button type="button" class="btn bo" onclick="alert('Password reset feature coming soon')">
      Forgot Password?
    </button>
  </div>
</div>

<div id="panel-create" style="display:none">
  <div class="grid2">
    <div><label>Full Name *</label><input id="u_name" placeholder="Enter your full name"></div>
    <div><label>Email</label><input id="u_email" placeholder="mani@example.com"></div>
  </div>

  <div class="grid2">
    <div><label>Phone</label><input id="u_phone" placeholder="9876543210"></div>
    <div><label>Bank</label><input id="u_bank" value="HDFC Bank"></div>
  </div>

  <div class="grid2">
    <div>
      <label>Password *</label>
      <div style="position:relative">
        <input id="u_password" type="password" placeholder="Min 8 chars and 1 special character" autocomplete="new-password" style="padding-right:40px">
        <span onclick="toggleCreatePassword()" style="position:absolute;right:12px;top:12px;cursor:pointer">👁️</span>
      </div>
    </div>

    <div>
      <label>Confirm Password *</label>
      <div style="position:relative">
        <input id="u_password_confirm" type="password" placeholder="Re-enter password" autocomplete="new-password" style="padding-right:40px">
        <span onclick="toggleConfirmPassword()" style="position:absolute;right:12px;top:12px;cursor:pointer">👁️</span>
      </div>
    </div>
  </div>

  <div style="margin-top:16px">
    <label style="text-transform:uppercase;letter-spacing:.5px;font-size:12px;font-weight:700;color:#555;margin-bottom:10px;display:block">Account Type *</label>

    <div style="display:flex;gap:24px">
      <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-size:13px;font-weight:600;text-transform:none;letter-spacing:0;margin:0;color:#333">
        <input type="radio" name="u_type" value="INDIVIDUAL" style="width:auto;accent-color:#1565c0" checked onchange="document.getElementById('u_user_type').value=this.value">
        👤 Individual
      </label>

      <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-size:13px;font-weight:600;text-transform:none;letter-spacing:0;margin:0;color:#333">
        <input type="radio" name="u_type" value="ORGANISATION" style="width:auto;accent-color:#1565c0" onchange="document.getElementById('u_user_type').value=this.value">
        🏢 Organisation
      </label>
    </div>

    <input type="hidden" id="u_user_type" value="INDIVIDUAL">
  </div>

  <div style="margin-top:16px;display:flex;align-items:center;gap:14px">
    <button class="btn bg" onclick="createUser()">➕ Create Account</button>
    <div id="create-msg" style="font-size:13px"></div>
  </div>
</div>

<div id="profile-setup-card" style="margin-top:18px;border-top:1px solid #eef1f5;padding-top:18px">
  <h3>Step 2 — Basic Classification Profile</h3>
  <p style="font-size:13px;color:#666;margin:0 0 14px">This improves salary, family loan, broker, and own account classification. Save anytime and edit later.</p>
  <div id="profile_gate_banner" class="alert alert-info" style="display:none;margin-bottom:14px"></div>
  <div id="profile_logged_out_msg" class="alert alert-info" style="display:none;margin-bottom:14px">Create account first to save your classification profile</div>

  <div id="profile-basic-section" class="grid3">
    <div>
      <label>Entity Type</label>
      <select id="pf_entity_type" onchange="toggleProfileSections()">
        <option value="">Select</option>
        <option value="INDIVIDUAL">Individual</option>
        <option value="NRI">NRI</option>
        <option value="HUF">HUF</option>
      </select>
    </div>
    <div>
      <label>Legal Name</label>
      <input id="pf_legal_name" placeholder="Account holder / legal entity name">
    </div>
    <div id="pf_huf_name_wrap" style="display:none">
      <label>HUF Name</label>
      <input id="pf_huf_name" placeholder="Sarju Garg HUF" oninput="toggleProfileSections()">
    </div>
  </div>



    <div class="grid3">
      <div>
        <label>Date Of Birth</label>
        <input id="pf_dob" type="date">
      </div>
      <div style="display:flex;align-items:flex-end">
        <label style="display:flex;align-items:center;gap:8px;margin:0;text-transform:none;font-size:13px;color:#333">
          <input id="pf_family_toggle" type="checkbox" style="width:auto" onchange="toggleProfileSections()">
          Family transactions present
        </label>
      </div>
      <div style="display:flex;align-items:flex-end">
        <label style="display:flex;align-items:center;gap:8px;margin:0;text-transform:none;font-size:13px;color:#333">
          <input id="pf_credit_cards" type="checkbox" style="width:auto">
          Has credit cards
        </label>
      </div>
    </div>

    <div class="grid3">
      <div style="display:flex;align-items:flex-end">
        <label style="display:flex;align-items:center;gap:8px;margin:0;text-transform:none;font-size:13px;color:#333">
          <input id="pf_salaried_toggle" type="checkbox" style="width:auto" onchange="toggleProfileSections()">
          Salaried
        </label>
      </div>
      <div id="pf_employer_wrap" style="display:none">
        <label>Employer Name</label>
        <input id="pf_employer_name" placeholder="Employer / payroll counterparty">
      </div>
      <div style="display:flex;align-items:flex-end">
        <label style="display:flex;align-items:center;gap:8px;margin:0;text-transform:none;font-size:13px;color:#333">
          <input id="pf_consultancy_toggle" type="checkbox" style="width:auto" onchange="toggleProfileSections()">
          Consultancy income
        </label>
      </div>
    </div>

    <div class="grid3">
      <div style="display:flex;align-items:flex-end">
        <label style="display:flex;align-items:center;gap:8px;margin:0;text-transform:none;font-size:13px;color:#333">
          <input id="pf_trading_toggle" type="checkbox" style="width:auto" onchange="toggleProfileSections()">
          Trading / broker activity
        </label>
      </div>
      <div style="display:flex;align-items:flex-end">
        <label style="display:flex;align-items:center;gap:8px;margin:0;text-transform:none;font-size:13px;color:#333">
          <input id="pf_multi_bank_toggle" type="checkbox" style="width:auto" onchange="toggleProfileSections()">
          Multiple bank accounts
        </label>
      </div>
      <div style="display:flex;align-items:flex-end">
        <label style="display:flex;align-items:center;gap:8px;margin:0;text-transform:none;font-size:13px;color:#333">
          <input id="pf_rental_income" type="checkbox" style="width:auto">
          Rental income
        </label>
      </div>
    </div>

    <div id="pf_family_section" style="margin-top:16px">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">
        <div style="font-weight:800;font-size:13px;color:#444">Family / Known Family Counterparties</div>
        <button class="btn bo" style="font-size:11px;padding:5px 12px" onclick="addProfileCounterpartyRow('family')">＋ Add Family Row</button>
      </div>
      <div style="overflow:auto">
        <table>
          <thead><tr><th>Display Name</th><th>Relationship</th><th>Default Treatment</th><th>Default Ledger Key</th><th>Txn Direction Hint</th><th>Notes</th><th></th></tr></thead>
          <tbody id="pf_family_rows"></tbody>
        </table>
      </div>
    </div>

    <div id="pf_consultancy_section" style="margin-top:18px;display:none">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">
        <div style="font-weight:800;font-size:13px;color:#444">Consultancy Counterparties</div>
        <button class="btn bo" style="font-size:11px;padding:5px 12px" onclick="addProfileCounterpartyRow('consultancy')">＋ Add Consultancy Row</button>
      </div>
      <div style="overflow:auto">
        <table>
          <thead><tr><th>Display Name</th><th>Relationship</th><th>Default Treatment</th><th>Default Ledger Key</th><th>Txn Direction Hint</th><th>Notes</th><th></th></tr></thead>
          <tbody id="pf_consultancy_rows"></tbody>
        </table>
      </div>
    </div>

    <div id="pf_broker_section" style="margin-top:18px;display:none">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">
        <div style="font-weight:800;font-size:13px;color:#444">Broker Counterparties</div>
        <button class="btn bo" style="font-size:11px;padding:5px 12px" onclick="addProfileCounterpartyRow('broker')">＋ Add Broker Row</button>
      </div>
      <div style="overflow:auto">
        <table>
          <thead><tr><th>Display Name</th><th>Relationship</th><th>Default Treatment</th><th>Default Ledger Key</th><th>Txn Direction Hint</th><th>Notes</th><th></th></tr></thead>
          <tbody id="pf_broker_rows"></tbody>
        </table>
      </div>
    </div>

    <div id="pf_accounts_section" style="margin-top:18px;display:none">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">
        <div style="font-weight:800;font-size:13px;color:#444">Known Accounts</div>
        <button class="btn bo" style="font-size:11px;padding:5px 12px" onclick="addKnownAccountRow()">＋ Add Account Row</button>
      </div>
      <div style="overflow:auto">
        <table>
          <thead><tr><th>Account Label</th><th>Institution Name</th><th>Account Mask</th><th>Account Type</th><th>Ownership Type</th><th></th></tr></thead>
          <tbody id="pf_account_rows"></tbody>
        </table>
      </div>
    </div>

    <div style="margin-top:16px;display:flex;gap:10px;flex-wrap:wrap;align-items:center">
      <button id="pf_save_btn" class="btn bp" onclick="saveProfileQuestionnaire()">💾 Save Profile</button>
      <button id="pf_reload_btn" class="btn bo" onclick="loadProfileQuestionnaire()">↻ Reload Saved Profile</button>
      <button id="pf_apply_approved_btn" class="btn bo" onclick="applyProfileToApprovedSafe()">Apply to Approved Safely</button>
      <div id="pf_msg" style="font-size:13px;color:#666"></div>
    </div>
  </div>
  </div>

  <div class="card">
    <h3>📚 Dictionary Setup</h3>
    <p style="font-size:13px;color:#666;margin:0 0 12px">Browse every ledger key — which book it belongs to, what it means, and who it's attributed to.</p>
    <button class="btn bp" onclick="loadDictionary()">📚 Open Ledger Dictionary</button>
    <div id="dict-out" style="margin-top:14px;display:none">
      <input id="dict-search" placeholder="🔍 Search keys, descriptions…" oninput="filterDict()"
        style="width:100%;padding:8px 12px;border:1.5px solid #e0e0e0;border-radius:7px;font-size:12px;margin-bottom:10px">
      <div id="dict-body"></div>
    </div>
  </div>

  <div class="card">
    <h3>📖 Accounting Basis Explained</h3>
    <div style="font-size:13px;color:#444;line-height:1.7">
      <p><strong>This system uses 3-book accounting:</strong></p>
      <p>📗 <strong>Book I — Income &amp; Expenditure:</strong> All revenues received and all expenses paid (cash basis). Bank debits = expenses already settled, not "liabilities".</p>
      <p>📘 <strong>Book II — Balance Sheet:</strong> Only true balance-sheet items — Assets (investments, loans given, land, FDs) and Liabilities (outstanding loan principal, credit card balance).</p>
      <p>📙 <strong>Book III — Balance Check:</strong> Net Worth = Assets − Liabilities + Equity + Net Surplus from P&amp;L.</p>
    </div>
  </div>

  <div class="card">
    <h3>🔍 System Check</h3>
    <button class="btn bo" onclick="checkSys()">Run Check</button>
    <pre id="sys-out" style="margin-top:12px;background:#f5f7fa;padding:14px;border-radius:8px;font-size:12px;font-family:var(--mono);white-space:pre-wrap"></pre>
  </div>

  <div class="card" id="opening-balance-card">
  <h3>💰 Opening Balance</h3>
  <p style="font-size:13px;color:#666;margin:0 0 14px">
    Set the bank account balance on the date of your <strong>first uploaded statement</strong>.
    This is the B/F (Brought Forward) figure from your statement's first row.
  </p>
    <div class="grid2">
      <div><label>Opening Balance (₹)</label>
        <input id="ob_amount" type="number" step="0.01" placeholder="e.g. 13665.69"></div>
      <div style="display:flex;align-items:flex-end;padding-bottom:2px">
        <button id="ob_save_btn" class="btn bp" onclick="setOpeningBalance()" style="width:100%">💾 Save Opening Balance</button>
      </div>
    </div>
    <div id="ob-msg" style="font-size:13px;margin-top:10px"></div>
  </div>

  <div class="card" id="manage-data-card">
    <h3>🗂 Manage Data</h3>
    <p style="font-size:13px;color:#666;margin:0 0 14px">Delete transaction data by account or clear everything. Your user account is never deleted.</p>
    <div id="manage-accounts-list"><p style="font-size:13px;color:#aaa">Log in to manage data.</p></div>
    <div style="margin-top:14px;border-top:1px solid #f0f0f0;padding-top:14px">
      <button id="delete_all_btn" class="btn bo" style="color:#b71c1c;border-color:#b71c1c" onclick="deleteAllData()">🗑 Delete All Data</button>
      <button id="delete_user_btn" class="btn bo" style="color:#8b0000;border-color:#8b0000" onclick="deleteUserCompletely()">⛔ Delete User</button>
    </div>
    <div id="manage-msg" style="font-size:13px;margin-top:10px"></div>
  </div>
</div>

<script>
// ── Tab routing ──────────────────────────────────────────────────────────────
const TABS = ['classify','review','statements','ledger','report','trading-cashflow','setup'];
function tab(n){
  TABS.forEach((t,i)=>{
    const tabEls = document.querySelectorAll('.tab');
    if(tabEls[i]) tabEls[i].classList.toggle('active',t===n);
    const pane = document.getElementById('pane-'+t);
    if(pane) pane.classList.toggle('active',t===n);
  });
  if(n==='ledger') loadStats();
  if(n==='classify') refreshUploadedFiles();
  if(n==='setup') refreshProfileGate({preserveBanner:true});
  if(n==='review'){
    if(Array.isArray(_reviewRowsPrefetched)){ loadReview(); }
    else if(!_allReviewRows.length){ loadReview(); }
    else { _ensureLedgerKeys().then(()=>filterReview()); }
  }
  if(n==='statements') showStmtTab();
}
function setStatus(msg,type='info'){const el=document.getElementById('status');if(el){el.textContent=msg;el.className='alert alert-'+type;}}

// ── Ledger display helpers ───────────────────────────────────────────────────
const BOOK_CLASS={'BALANCE_SHEET':'book-bs','INCOME_EXPENSE':'book-ie','CAPITAL':'book-cap','SUSPENSE':'book-sus'};
const BOOK_LABEL={'BALANCE_SHEET':'Balance Sheet','INCOME_EXPENSE':'Income & Exp','CAPITAL':'Capital','SUSPENSE':'Suspense'};
const SEC_ICON={'Income':'📈','Expenditure':'💸','Assets':'🏦','Liabilities':'📜','Equity':'👤','Suspense':'⚠️'};
let _uid=0;
let _activeUserId='';
let _profileGateState={ok:false,profile_required:false,is_complete:false,missing_reason:''};

function _profileGateText(reason=''){
  return reason
    ? `Complete basic classification profile before importing statements. ${reason}`
    : 'Complete basic classification profile before importing statements';
}

function _setProfileBanner(msg='', kind='info'){
  const el=document.getElementById('profile_gate_banner');
  if(!el) return;
  if(!msg){
    el.style.display='none';
    el.textContent='';
    el.className='alert alert-info';
    return;
  }
  el.style.display='block';
  el.textContent=msg;
  el.className=`alert alert-${kind}`;
}

function _focusProfileSetup(message=''){
  tab('setup');
  if(message) _setProfileBanner(message, 'info');
  const card=document.getElementById('profile-setup-card');
  if(card) card.scrollIntoView({behavior:'smooth', block:'start'});
  const first=document.getElementById('pf_entity_type') || document.getElementById('pf_legal_name');
  if(first && !first.disabled) setTimeout(()=>first.focus(), 250);
}

function _setProfileInputsEnabled(enabled){
  document.querySelectorAll('#profile-setup-card input,#profile-setup-card select,#profile-setup-card textarea,#profile-setup-card button')
    .forEach(el=>{ el.disabled = !enabled; });
  const loggedOut=document.getElementById('profile_logged_out_msg');
  if(loggedOut) loggedOut.style.display = enabled ? 'none' : 'block';
  if(!enabled) _profileMsg('Create account first to save your classification profile', '#b71c1c');
}

function _applySetupInteractionState(){
  const loggedIn=!!_activeUserId;
  _setProfileInputsEnabled(loggedIn);
  const obBtn=document.getElementById('ob_save_btn');
  const obInput=document.getElementById('ob_amount');
  if(obBtn) obBtn.disabled=!loggedIn;
  if(obInput) obInput.disabled=!loggedIn;
  ['delete_all_btn','delete_user_btn'].forEach(id=>{
    const el=document.getElementById(id);
    if(el) el.disabled=!loggedIn;
  });
  const manageList=document.getElementById('manage-accounts-list');
  if(manageList && !loggedIn){
    manageList.innerHTML='<p style="font-size:13px;color:#aaa">Log in to manage data.</p>';
  }
}




let _tcfFile = null;

function _applyProfileGate(){
  const blocked=!!_profileGateState.profile_required;
  const noUser=!_activeUserId;
  const uploadBtn=document.getElementById('uploadBtn');
  const tradingUploadBtn=document.getElementById('tcf-upload-btn');
  const tradingMsg=document.getElementById('tcf-upload-msg');
  const gateMsg=document.getElementById('profile-gate-msg');
  const dropZone=document.getElementById('dropZone');
  if(uploadBtn) uploadBtn.disabled = noUser || blocked || !_file;
  if(tradingUploadBtn) tradingUploadBtn.disabled = noUser || blocked || !_tcfFile;
  if(gateMsg){
    gateMsg.style.display = (noUser || blocked) ? 'block' : 'none';
    gateMsg.textContent = noUser ? 'Log in or create an account before importing statements.' : (blocked ? _profileGateText(_profileGateState.missing_reason || '') : '');
  }
  if(tradingMsg && (noUser || blocked)){
    tradingMsg.innerHTML = `<span style="color:var(--red)">${noUser ? 'Log in before uploading a trading ledger.' : _profileGateText(_profileGateState.missing_reason || '')}</span>`;
  }else if(tradingMsg && !blocked && !_tcfFile){
    tradingMsg.textContent = '';
  }
  if(dropZone){
    dropZone.style.opacity = (noUser || blocked) ? '0.72' : '1';
    dropZone.style.borderStyle = blocked ? 'dashed' : 'dashed';
  }
  const advancedCollapsed=!!_activeUserId && blocked;
  ['pf_family_section','pf_consultancy_section','pf_broker_section','pf_accounts_section'].forEach(id=>{
    const el=document.getElementById(id);
    if(el && advancedCollapsed) el.style.display='none';
  });
}

async function refreshProfileGate(opts={}){
  const preserveBanner=!!opts.preserveBanner;
  if(!_activeUserId){
    _profileGateState={ok:false,profile_required:false,is_complete:false,missing_reason:''};
    if(!preserveBanner) _setProfileBanner('Create account first to save your classification profile', 'info');
    _applySetupInteractionState();
    _applyProfileGate();
    return _profileGateState;
  }
  try{
    const r=await fetch('/profile-status');
    const d=r.ok ? await r.json() : {};
    _profileGateState={
      ok: !!d.ok,
      profile_required: !!d.profile_required,
      is_complete: !!d.is_complete,
      missing_reason: d.missing_reason || ''
    };
  }catch(_e){
    _profileGateState={ok:false,profile_required:false,is_complete:false,missing_reason:''};
  }
  if(_profileGateState.profile_required){
    _setProfileBanner(opts.message || _profileGateText(_profileGateState.missing_reason), 'info');
    if(opts.focusProfile) _focusProfileSetup(opts.message || _profileGateText(_profileGateState.missing_reason));
  }else if(!preserveBanner){
    _setProfileBanner('Basic classification profile complete. Statement import is enabled.', 'ok');
  }
  _applySetupInteractionState();
  _applyProfileGate();
  return _profileGateState;
}

function renderResults(rows){
  // renderResults is now a no-op — results go directly to the Review tab
  // Kept for backward compatibility
  if(rows && rows.length){ _allReviewRows=rows; renderReviewRows(rows); }
}
function toggleEl(id){const el=document.getElementById(id);if(el)el.style.display=el.style.display==='none'?'block':'none';}

// ── Upload → stage for review ─────────────────────────────────────────────────
let _file=null;
function handleDrop(e){e.preventDefault();document.getElementById('dropZone').classList.remove('dragover');const f=e.dataTransfer.files[0];if(f)setFile(f);}
function setFile(f){_file=f;const p=document.getElementById('file-preview');p.style.display='block';p.innerHTML=`📎 <strong>${f.name}</strong> · ${(f.size/1024).toFixed(1)} KB &nbsp; <a href="/download-upload/${encodeURIComponent(f.name)}" download style="font-size:12px;color:#1565c0">📥 Download</a>`;_applyProfileGate();}
function clearUpload(){_file=null;document.getElementById('xlFile').value='';document.getElementById('file-preview').style.display='none';const fp=document.getElementById('file_password');if(fp)fp.value='';_applyProfileGate();setStatus('Ready — upload a bank statement.','info');}

function _fmtUploadedDate(s){
  if(!s) return '';
  const d=new Date(s);
  if(Number.isNaN(d.getTime())) return String(s).slice(0,19).replace('T',' ');
  return d.toLocaleString(undefined,{year:'numeric',month:'short',day:'2-digit',hour:'2-digit',minute:'2-digit'});
}
function _fmtPeriod(row){
  const a=row.statement_from_date||'', b=row.statement_to_date||'';
  return a && b ? `${a} to ${b}` : (a || b || '—');
}
async function refreshUploadedFiles(){
  const box=document.getElementById('uploaded-files-body');
  if(!box) return;
  if(!_activeUserId){
    box.innerHTML='<p style="color:#bbb;font-size:13px">Log in to see previously uploaded statements.</p>';
    return;
  }
  box.innerHTML='<p style="color:#888;font-size:13px">Loading uploaded files…</p>';
  try{
    const r=await fetch('/uploaded-files');
    const d=r.ok ? await r.json() : {files:[]};
    const files=Array.isArray(d.files) ? d.files : [];
    if(!files.length){
      box.innerHTML='<p style="color:#bbb;font-size:13px">No uploaded files yet.</p>';
      return;
    }
    const rows=files.map(f=>{
      const name=_escapeHtmlAttr(f.file_name||'Uploaded statement');
      const acct=_escapeHtmlAttr(f.account_id||'main');
      const status=_escapeHtmlAttr(f.import_status||'');
      const period=_escapeHtmlAttr(_fmtPeriod(f));
      const uploaded=_escapeHtmlAttr(_fmtUploadedDate(f.created_at));
      const count=Number(f.approved_rows||0) ? `${f.approved_rows} approved` : `${f.staged_rows||0} staged`;
      const btn=f.download_available
        ? `<button class="btn bo" style="font-size:11px;padding:5px 10px" onclick="downloadUploadedFile('${_escapeHtmlAttr(f.id)}')">Download</button>`
        : `<span style="color:#aaa;font-size:12px">Unavailable</span>`;
      return `<tr>
        <td style="padding:8px;border-bottom:1px solid #eee;font-weight:700">${name}<div style="font-size:11px;color:#999;font-weight:400">${_escapeHtmlAttr(count)}</div></td>
        <td style="padding:8px;border-bottom:1px solid #eee">${acct}</td>
        <td style="padding:8px;border-bottom:1px solid #eee">${uploaded}</td>
        <td style="padding:8px;border-bottom:1px solid #eee">${period}</td>
        <td style="padding:8px;border-bottom:1px solid #eee">${status}</td>
        <td style="padding:8px;border-bottom:1px solid #eee;text-align:right">${btn}</td>
      </tr>`;
    }).join('');
    box.innerHTML=`<table style="width:100%;border-collapse:collapse;font-size:12px">
      <thead><tr style="text-align:left;color:#666;background:#f8fafc">
        <th style="padding:8px">File</th><th style="padding:8px">Account</th><th style="padding:8px">Uploaded</th>
        <th style="padding:8px">Statement Period</th><th style="padding:8px">Status</th><th style="padding:8px;text-align:right">Action</th>
      </tr></thead><tbody>${rows}</tbody></table>`;
  }catch(e){
    box.innerHTML=`<p style="color:#b71c1c;font-size:13px">Could not load uploaded files: ${_escapeHtmlAttr(e.message||e)}</p>`;
  }
}
function downloadUploadedFile(id){
  if(!id) return;
  window.location.href='/download-uploaded-file?id='+encodeURIComponent(id);
}

async function uploadForReview(){
  if(!_activeUserId){
    setStatus('Log in or create an account before importing statements.','err');
    tab('setup');
    return;
  }
  if(_profileGateState.profile_required){
    setStatus(_profileGateText(_profileGateState.missing_reason),'err');
    _focusProfileSetup(_profileGateText(_profileGateState.missing_reason));
    return;
  }
  if(!_file){setStatus('Select a file first','err');return;}
  const uploadBtn=document.getElementById('uploadBtn');
  setStatus('Classifying transactions — please wait…','info');
  if(uploadBtn){uploadBtn.disabled=true;uploadBtn.textContent='⏳ Working…';}
  await _ensureLedgerKeys();
  const acctName=(document.getElementById('acct_name').value||'').trim()||'main';
  const filePassword=(document.getElementById('file_password').value||'').trim();
  const fd=new FormData();fd.append('file',_file);fd.append('account_id',acctName);if(filePassword)fd.append('password',filePassword);
  try{
    const controller = new AbortController();
    const timeoutId = setTimeout(()=>controller.abort(), 120000);
    const r=await fetch('/upload-pending',{method:'POST',body:fd,signal:controller.signal});
    clearTimeout(timeoutId);
    const raw = await r.text();
    let d = {};
    try { d = raw ? JSON.parse(raw) : {}; }
    catch(parseErr) {
      console.error('Server response parse error:', parseErr, 'Raw:', raw);
      setStatus('Server error: invalid JSON response. Check server logs.','err');
      return;
    }
    if(!r.ok && !d.error){ d.error = 'Upload failed with HTTP ' + r.status; }
    if(d.profile_required){
      await refreshProfileGate({focusProfile:true, message:_profileGateText(d.missing_reason || '')});
      setStatus('❌ ' + _profileGateText(d.missing_reason || ''), 'err');
      return;
    }
    if(d.error){
      setStatus('❌ ' + d.error + (d.warnings && d.warnings.length ? '  |  ' + d.warnings.join(' | ') : ''), 'err');
      return;
    }
    if(d.ok === false){
      const warn = (d.warnings && d.warnings.length) ? d.warnings.join(' | ') : 'No transactions found.';
      setStatus('⚠️ ' + warn, 'err');
      return;
    }
    if(d.warnings && d.warnings.length){ setStatus('⚠️ ' + d.warnings.join(' | '), 'info'); }
    if(!d.staged || d.staged === 0){
      setStatus('⚠️ No transactions found in the file. Check file format or that it has data rows.', 'info');
      return;
    }
    _allReviewRows = Array.isArray(d.transactions) ? d.transactions : [];
    _reviewVisibleCount = _reviewPageSize;
    _reviewRowsPrefetched = _allReviewRows.slice();
    setStatus(`✅ ${d.staged} transactions classified (account: ${d.account_id||acctName}) — review & approve in the Review tab`,'ok');
    loadAccountDropdowns();
    refreshUploadedFiles();
    tab('review');
    clearUpload();
  }catch(e){
    const msg = (e && e.name==='AbortError') ? 'Upload timed out. Check server logs and try again.' : ('Upload failed: '+e);
    setStatus(msg,'err');
  }
  finally{if(uploadBtn){uploadBtn.disabled=false;uploadBtn.textContent='🔍 Classify & Review';}}
}

// ── Review tab ────────────────────────────────────────────────────────────────
let _allReviewRows = [];
let _reviewRowsPrefetched = null;
let _ledgerKeys = [];   // loaded once from /ledger-keys
let _ledgerKeyOptionsBaseHtml = '';
let _reviewPageSize = 100;
let _reviewVisibleCount = 100;

function _invalidateLedgerKeyOptionsCache(){ _ledgerKeyOptionsBaseHtml = ''; }

async function _ensureLedgerKeys(){
  if(_ledgerKeys.length) return;
  try{
    const r = await fetch('/ledger-keys');
    _ledgerKeys = await r.json();
    _invalidateLedgerKeyOptionsCache();
  }catch(e){ _ledgerKeys = []; }
}

async function loadReview(){
  await _ensureLedgerKeys();
  let rows = [];
  if(Array.isArray(_reviewRowsPrefetched)){
    rows = _reviewRowsPrefetched.slice();
    _reviewRowsPrefetched = null;
    if(rows.length){ _allReviewRows = rows; _reviewVisibleCount = _reviewPageSize; renderReviewRows(rows); return; }
  }
  try{
    const r = await fetch('/pending?limit=500');
    if(!r.ok){
      const errText = await r.text();
      console.error('GET /pending failed:', r.status, errText);
      showReviewEmpty();
      const rs=document.getElementById('review-status');
      if(rs){rs.style.display='block';rs.className='alert alert-err';rs.textContent='Failed to load pending transactions (HTTP '+r.status+').';}
      return;
    }
    const data = await r.json();
    if(!Array.isArray(data)){
      console.error('GET /pending returned non-array:', data);
      if(data && data.error){ showReviewError(data.error); return; }
      showReviewEmpty();
      return;
    }
    rows = data;
  }catch(e){
    console.error('loadReview error:', e);
    showReviewEmpty();
    return;
  }
  if(rows && rows.length){ _allReviewRows = rows; _reviewVisibleCount = _reviewPageSize; renderReviewRows(rows); } else showReviewEmpty();
}

function showReviewError(msg){
  const ab=document.getElementById('review-action-bar'); if(ab)ab.style.display='none';
  const rc=document.getElementById('review-rows-container'); if(rc)rc.style.display='none';
  const re=document.getElementById('review-empty'); if(re)re.style.display='none';
  const rs=document.getElementById('review-status');
  if(rs){rs.style.display='block';rs.className='alert alert-err';rs.textContent='Error: '+msg;}
}

function showReviewEmpty(msg=''){
  const actionBar=document.getElementById('review-action-bar');
  const rowsContainer=document.getElementById('review-rows-container');
  const empty=document.getElementById('review-empty');
  const rs=document.getElementById('review-status');
  if(actionBar)actionBar.style.display='none';
  if(rowsContainer)rowsContainer.style.display='none';
  if(empty){
    empty.style.display='block';
    empty.innerHTML = `<div style="text-align:center;padding:30px;color:#888;font-size:13px">${msg || 'No transactions to review yet.'}</div>`;
  }
  if(rs)rs.style.display='none';
}

function filterReview(){
  const q=(document.getElementById('rv-search').value||'').toLowerCase();
  const st=document.getElementById('rv-filter-status').value;
  const bk=document.getElementById('rv-filter-book').value;
  const filtered=_allReviewRows.filter(r=>{
    if(st && r.status!==st) return false;
    if(bk && r.book!==bk) return false;
    if(q){
      const narr=(r.narration||'').toLowerCase();
      const cp=(r.counterparty||'').toLowerCase();
      if(!narr.includes(q)&&!cp.includes(q)) return false;
    }
    return true;
  });
  _reviewVisibleCount = _reviewPageSize;
  renderReviewRows(filtered, /*skipCounterUpdate=*/true);
}

function _buildKeyOptions(currentKey){
  // Group options by book > section
  const groups={};
  _ledgerKeys.forEach(k=>{
    const g=k.book+'___'+k.section;
    if(!groups[g]) groups[g]={label:`${k.book.replace('_',' ')} › ${k.section}`, opts:[]};
    groups[g].opts.push(k);
  });
  let html='<option value="">— pick category —</option>';
  for(const [, grp] of Object.entries(groups)){
    html+=`<optgroup label="${grp.label}">`;
    grp.opts.forEach(k=>{
      const sel = k.key===currentKey ? 'selected' : '';
      html+=`<option value="${k.key}" ${sel}>${k.key} · ${k.account}</option>`;
    });
    html+='</optgroup>';
  }
  return html;
}

function downloadStmt(which){
  const map = {
    income:      { id: 'stmt-income-body',      file: 'income_statement.txt' },
    income_acct: { id: 'stmt-income-acct-body', file: 'income_statement_this_account.txt' },
    balance:     { id: 'stmt-balance-body',     file: 'balance_sheet.txt' },
    balance_acct:{ id: 'stmt-balance-acct-body',file: 'balance_sheet_this_account.txt' },
    pl:          { id: 'stmt-pl-body',          file: 'profit_loss.txt' },
    pl_acct:     { id: 'stmt-pl-acct-body',     file: 'profit_loss_this_account.txt' },
  };
  const cfg = map[which];
  if(!cfg){ alert('Unknown statement type'); return; }

  const el = document.getElementById(cfg.id);
  if(!el){ alert('Statement not found'); return; }

  const text = (el.innerText || el.textContent || '').trim();
  if(!text){
    alert('No statement data to download.');
    return;
  }

  const blob = new Blob([text], {type:'text/plain;charset=utf-8'});
  const url = URL.createObjectURL(blob);
  const a = document.createElement('a');
  a.href = url;
  a.download = cfg.file;
  document.body.appendChild(a);
  a.click();
  a.remove();
  setTimeout(() => URL.revokeObjectURL(url), 1000);
}

async function downloadYearPdf(report, year){
  const account_id = (document.getElementById('stmt_account') || {value:''}).value || '';
  const date_from = `${year}-01-01`;
  const date_to   = `${year}-12-31`;

  const qs = new URLSearchParams({
    kind: 'detailed',
    report: report || '',
    date_from,
    date_to,
    account_id
  });

  try{
    const res = await fetch('/download-report-pdf?' + qs.toString(), {
      method: 'GET',
      credentials: 'same-origin'
    });

    if(!res.ok){
      const msg = await res.text();
      alert('PDF download failed: ' + msg);
      return;
    }

    const blob = await res.blob();
    const cd = res.headers.get('Content-Disposition') || '';
    let filename = `${report || 'report'}_${year}.pdf`;
    const m = cd.match(/filename=\"?([^"]+)\"?/i);
    if(m && m[1]) filename = m[1];

    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url;
    a.download = filename;
    document.body.appendChild(a);
    a.click();
    a.remove();
    setTimeout(() => URL.revokeObjectURL(url), 1500);
  }catch(e){
    alert('PDF download failed: ' + e);
  }
}

async function onAccountChange(){
  await populateYearDropdown();
  await refreshStmts();
}

async function viewYearStatement(report, fyStart){
  const fy = parseInt(fyStart, 10);
  const df = document.getElementById('stmt_date_from');
  const dt = document.getElementById('stmt_date_to');
  const yr = document.getElementById('stmt_year');

  if(df) df.value = `${fy}-04-01`;
  if(dt) dt.value = `${fy + 1}-03-31`;
  if(yr) yr.value = String(fy);

  await refreshStmts();

  if(report === 'income_statement') showStmt('income');
  else if(report === 'balance_sheet') showStmt('balance');
  else if(report === 'profit_loss') showStmt('pl');
}

async function loadYearlyStatements(){
  const account_id=(document.getElementById('stmt_account')||{value:''}).value||'';
  const qs = new URLSearchParams({account_id});
  const r = await fetch('/financial-reports-yearly?' + qs.toString(), {credentials:'same-origin'});
  const d = await r.json();
  if(d.error){ alert(d.error); return; }

  const years = Object.keys(d).sort().reverse();
  let html = '';
  for(const yr of years){
    const fy = parseInt(yr, 10);
    const fyLabel = `FY ${fy}-${String(fy + 1).slice(-2)}`;
    html += `
      <div style="border:1px solid #ddd;border-radius:10px;padding:10px;margin:8px 0;background:#fff">
        <div style="font-weight:700;margin-bottom:8px">${fyLabel}</div>
        <div style="display:flex;gap:8px;flex-wrap:wrap">
          <button onclick="downloadYearPdf('income_statement','${yr}')">Income PDF</button>
          <button onclick="downloadYearPdf('balance_sheet','${yr}')">Balance Sheet PDF</button>
          <button onclick="viewYearStatement('income_statement','${yr}')">View Income</button>
          <button onclick="viewYearStatement('balance_sheet','${yr}')">View Balance Sheet</button>
        </div>
      </div>
    `;
  }
  document.getElementById('yearly-reports-wrap').innerHTML = html || '<div>No yearly reports found</div>';
}

async function populateYearDropdown(){
  const account_id=(document.getElementById('stmt_account')||{value:''}).value||'';
  const sel=document.getElementById('stmt_year');
  if(!sel) return;

  try{
    const qs = new URLSearchParams({account_id});
    const r = await fetch('/financial-reports-yearly?' + qs.toString(), {credentials:'same-origin'});
    const d = await r.json();
    if(!r.ok || d.error){
      sel.innerHTML = '<option value="">All Financial Years</option>';
      return;
    }

    const years = Object.keys(d).sort().reverse();
    const current = sel.value || '';
    let html = '<option value="">All Financial Years</option>';

    years.forEach(yr => {
      const fy = parseInt(yr, 10);
      const fyLabel = `FY ${fy}-${String(fy + 1).slice(-2)}`;
      html += `<option value="${fy}" ${current===String(fy)?'selected':''}>${fyLabel}</option>`;
    });

    sel.innerHTML = html;
  }catch(e){
    sel.innerHTML = '<option value="">All Financial Years</option>';
  }
}

function applyYearFilter(){
  const fyStart = parseInt((document.getElementById('stmt_year')||{value:''}).value || '', 10);
  const df = document.getElementById('stmt_date_from');
  const dt = document.getElementById('stmt_date_to');

  if(!df || !dt) return;

  if(!fyStart){
    df.value = '';
    dt.value = '';
  }else{
    const fyEnd = fyStart + 1;
    df.value = `${fyStart}-04-01`;
    dt.value = `${fyEnd}-03-31`;
  }
  refreshStmts();
}

function downloadYearPdf(report, fyStart){
  const account_id = (document.getElementById('stmt_account') || {value:''}).value || '';
  const fy = parseInt(fyStart, 10);
  const date_from = `${fy}-04-01`;
  const date_to   = `${fy + 1}-03-31`;

  downloadReportPdf('detailed', report, date_from, date_to, account_id);
}

function renderReviewRows(rows, skipCounterUpdate=false){
  if(!Array.isArray(rows)) rows=[];
  const originalRows = rows.slice();
  if(!skipCounterUpdate){
    _allReviewRows=originalRows;
    document.getElementById('rv-total').textContent=originalRows.length;
    document.getElementById('rv-reclassified').textContent=originalRows.filter(r=>r.status==='RECLASSIFIED').length+' reclassified';
    document.getElementById('rv-anomaly').textContent=originalRows.filter(r=>r.is_anomaly).length+' anomalies';
  }
  document.getElementById('review-empty').style.display='none';
  document.getElementById('review-action-bar').style.display='block';
  document.getElementById('review-rows-container').style.display='block';

  // FIX B5: Guard — if ledger keys failed to load, dropdowns will be disabled
  const keysAvailable = Array.isArray(_ledgerKeys) && _ledgerKeys.length > 0;
  if(!keysAvailable){
    console.warn('renderReviewRows: ledger keys not loaded — dropdowns will be disabled');
  }

  const container=document.getElementById('review-rows');
  if(!originalRows.length){container.innerHTML='<div style="text-align:center;padding:30px;color:#aaa;font-size:13px">No rows match your filter.</div>';return;}
  const rowsToRender = originalRows.slice(0, Math.max(_reviewVisibleCount, _reviewPageSize));

  const BOOK_BADGE={
    BALANCE_SHEET:'background:#e3f2fd;color:#0d47a1;border:1px solid #bbdefb',
    INCOME_EXPENSE:'background:#e8f5e9;color:#1b5e20;border:1px solid #c8e6c9',
    SUSPENSE:'background:#fff3e0;color:#e65100;border:1px solid #ffe0b2',
  };

  container.innerHTML = rowsToRender.map((r,i)=>{
    const activeKey = r.reclassified_key||r.predicted_ledger_key||'';
    const bb = BOOK_BADGE[r.book]||'background:#f5f5f5;color:#555;border:1px solid #ddd';
    const isReclass = r.status==='RECLASSIFIED';
    const isAnomaly = r.is_anomaly;
    const amt = parseFloat(r.amount||0).toLocaleString('en-IN',{minimumFractionDigits:2});
    const amtColor = r.txn_type==='credit'?'#1b5e20':'#b71c1c';
    const narrDisplay = (r.narration || r.description || '—').trim() || '—';
    const cpRaw = (r.counterparty||'').trim();
    const cpDisplay = cpRaw || '—';
    const statusDot = isReclass
      ? '<span style="display:inline-block;width:8px;height:8px;border-radius:50%;background:#f57f17;margin-right:5px"></span>'
      : '<span style="display:inline-block;width:8px;height:8px;border-radius:50%;background:#1565c0;margin-right:5px"></span>';
    const anomalyBadge = isAnomaly ? '<span style="font-size:10px;background:#fce4ec;color:#c62828;border:1px solid #ef9a9a;border-radius:4px;padding:1px 6px;margin-left:6px">⚠ anomaly</span>' : '';

    return `<div class="card" id="rvr-${r.id}" style="margin-bottom:10px;padding:14px 18px;${isReclass?'border-left:4px solid #f57f17':'border-left:4px solid #e0e0e0'}">
      <!-- Row header: index, date, amount, type, status -->
      <div style="display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:8px;margin-bottom:10px">
        <div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap">
          <span style="font-size:11px;color:#aaa;font-weight:600;min-width:22px">#${i+1}</span>
          <span style="font-size:12px;color:#666;font-weight:600;font-family:monospace">${r.txn_date||'—'}</span>
          <span class="badge ${r.txn_type==='credit'?'b-cr':'b-dr'}">${r.txn_type||''}</span>
          <span style="font-size:14px;font-weight:800;color:${amtColor}">₹${amt}</span>
          ${anomalyBadge}
        </div>
        <div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap">
          <span style="font-size:10px;padding:2px 8px;border-radius:5px;font-weight:700;${bb}">${r.book||''}</span>
          <span style="font-size:11px;font-weight:700">${statusDot}${r.status}</span>
        </div>
      </div>

      <!-- Compact review row -->
      <div style="margin-bottom:8px">
        <div style="font-size:12px;color:#555;display:flex;align-items:flex-start;gap:6px;flex-wrap:nowrap">
          <span style="font-size:10px;background:#eef3ff;border-radius:4px;padding:1px 7px;color:#4a6fa5;font-weight:600;text-transform:uppercase;letter-spacing:.3px;flex:0 0 auto">Narration</span>
          <span style="font-weight:500;color:#333;word-break:break-word;line-height:1.45">${narrDisplay}</span>
        </div>
      </div>
      <div style="margin-bottom:10px">
        <div style="font-size:12px;color:#555;display:flex;align-items:center;gap:6px;flex-wrap:wrap">
          <span style="font-size:10px;background:#f0f0f0;border-radius:4px;padding:1px 7px;color:#777;font-weight:600;text-transform:uppercase;letter-spacing:.3px">Counterparty</span>
          <span style="font-weight:600;color:#333">${cpDisplay}</span>
        </div>
      </div>

      <!-- Reclassify dropdown -->
      <div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap">
        <span style="display:none"></span>
        ${keysAvailable
          ? `<select id="rk-${r.id}"
              style="flex:1;min-width:220px;max-width:520px;padding:7px 10px;border:1.5px solid ${isReclass?'#f57f17':'#e0e0e0'};border-radius:7px;font-size:12px;font-family:monospace;background:#fff"
              onchange="doReclassifyDropdown('${r.id}', this.value)">
              ${_buildKeyOptions(activeKey)}
            </select>`
          : `<select disabled style="flex:1;min-width:220px;padding:7px 10px;border:1.5px solid #ddd;border-radius:7px;font-size:12px;background:#f5f5f5;color:#aaa">
              <option>Categories loading — retry Review tab</option>
             </select>`
        }
        <button class="btn bo" onclick="saveCustomRuleFromRow('${r.id}')">💾 Save as rule</button>
        <span style="font-size:11px;color:#aaa">Current: <code style="background:#f0f4ff;padding:1px 6px;border-radius:3px;color:#1565c0">${activeKey}</code></span>
      </div>
    </div>`;
  }).join('');
  if(originalRows.length > rowsToRender.length){
    container.innerHTML += `<div style="text-align:center;padding:14px"><button class="btn bo" onclick="_reviewVisibleCount += _reviewPageSize; filterReview();">Show more (${originalRows.length - rowsToRender.length} remaining)</button></div>`;
  }
}

async function doReclassifyDropdown(id, newKey){
  if(!newKey) return;

  const sel = document.getElementById('rk-' + id);
  if(sel) sel.disabled = true;

  try{
    const r = await fetch('/reclassify', {
      method: 'POST',
      headers: {'Content-Type':'application/json'},
      body: JSON.stringify({ txn_id: id, new_key: newKey })
    });

    const d = await r.json();

    if(d.error){
      alert('❌ ' + d.error);
      if(sel) sel.disabled = false;
      return;
    }

    const idx = _allReviewRows.findIndex(x => x.id === id);
    if(idx >= 0){
      _allReviewRows[idx] = {
        ..._allReviewRows[idx],
        ...d,
        reclassified_key: newKey,
        status: 'RECLASSIFIED'
      };
    }

    document.getElementById('rv-reclassified').textContent =
      _allReviewRows.filter(r => r.status === 'RECLASSIFIED').length + ' reclassified';

    filterReview();

  }catch(e){
    alert('Error: ' + e);
    if(sel) sel.disabled = false;
  }
}

async function saveCustomRuleFromRow(id){
  const row = _allReviewRows.find(x => x.id === id);
  if(!row){ alert('Row not found'); return; }

  const sel = document.getElementById('rk-' + id);
  const activeKey =
    (sel && sel.value ? sel.value : '') ||
    row.reclassified_key ||
    row.predicted_ledger_key ||
    '';

  if(!activeKey){ alert('Pick a category first'); return; }

  const suggested = (row.counterparty || row.narration || '').trim();
  const pattern = prompt(
    'Save auto-classification rule.\n\nEnter text to match in future narrations:',
    suggested
  );
  if(pattern === null) return;

  const match_mode = 'contains';
  const txn_type = row.txn_type || '';

  try{
    const r = await fetch('/add-custom-rule', {
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body: JSON.stringify({
        pattern,
        ledger_key: activeKey,
        txn_type,
        match_mode,
        priority: 100
      })
    });
    const d = await r.json();
    if(d.error){ alert('❌ ' + d.error); return; }
    alert(`✅ Rule saved\n\n"${pattern}" → ${activeKey}`);
  }catch(e){
    alert('Failed to save rule: ' + e);
  }
}

async function addCustomRule() {
  const pattern = document.getElementById('rule-pattern').value.trim();
  const ledger_key = document.getElementById('rule-ledger-key').value;
  const match_mode = document.getElementById('rule-match-mode').value;
  const txn_type = document.getElementById('rule-txn-type').value;
  if (!pattern || !ledger_key) { alert('Pattern and ledger key required'); return; }
  const r = await fetch('/add-custom-rule', {
    method: 'POST', headers: {'Content-Type':'application/json'},
    body: JSON.stringify({pattern, ledger_key, match_mode, txn_type, priority: 50})
  });
  const d = await r.json();
  document.getElementById('rule-save-msg').textContent = d.ok ? '✅ Rule saved!' : ('❌ ' + d.error);
  setTimeout(() => document.getElementById('rule-save-msg').textContent = '', 3000);
}
// Populate ledger key dropdown on load
fetch('/ledger-keys').then(r=>r.json()).then(keys=>{
  const sel = document.getElementById('rule-ledger-key');
  if (!sel) return;
  keys.forEach(k=>{ const o=document.createElement('option'); o.value=k.key; o.textContent=k.key; sel.appendChild(o); });
});

async function approveAll(){
  if(!confirm('Approve all transactions and post to ledger?\n\nReclassified rows will use the new category. All others keep the predicted category.')) return;
  const btn=document.querySelector('#review-action-bar .btn.bg');
  if(btn){btn.disabled=true;btn.textContent='⏳ Posting…';}
  try{
    const r=await fetch('/approve',{method:'POST',headers:{'Content-Type':'application/json'},body:'{}'});
    const d=await r.json();
    if(d.error){alert('Error: '+d.error);return;}
    _allReviewRows=[];
    showReviewEmpty();
    const rs=document.getElementById('review-status');
    rs.style.display='block';rs.className='alert alert-ok';
    rs.textContent=`✅ ${d.approved} transactions posted. Opening Statements…`;
    if(d.reports)renderStatements(d.reports);
    refreshUploadedFiles();
    setTimeout(()=>tab('statements'),700);
  }catch(e){alert('Approve failed: '+e);}
  finally{if(btn){btn.disabled=false;btn.textContent='✅ Approve All & Post to Ledger';}}
}

async function discardPending(){
  if(!confirm('Discard this batch? All staged transactions will be deleted.')) return;
  await fetch('/discard-pending',{method:'POST',headers:{'Content-Type':'application/json'},body:'{}'});
  showReviewEmpty();
}

// ── Financial Statements ──────────────────────────────────────────────────────
let _lastReports={};

function showStmt(which){
  ['income','balance','pl'].forEach(k=>{
    const el=document.getElementById('stmt-'+k);
    if(el)el.style.display=(which==='all'||which===k)?'block':'none';
  });
  document.getElementById('stmt-empty').style.display='none';
}

function _extractChips(text){
  /* Scan the raw report text for key totals and return chip data */
  const chips=[];
  const patterns=[
    {re:/TOTAL INCOME\s+₹([\d,]+\.\d+)/,label:'Total Income',color:'#1b5e20',bg:'#e8f5e9'},
    {re:/TOTAL EXPENSES\s+₹([\d,]+\.\d+)/,label:'Total Expenses',color:'#b71c1c',bg:'#ffebee'},
    {re:/NET INCOME[^\n]+₹([\d,]+\.\d+)/,label:'Net Income',color:'#0d47a1',bg:'#e3f2fd'},
    {re:/TOTAL REVENUE\s+₹([\d,]+\.\d+)/,label:'Revenue',color:'#1b5e20',bg:'#e8f5e9'},
    {re:/NET PROFIT[^\n]+₹([\d,]+\.\d+)/,label:'Net Profit',color:'#0d47a1',bg:'#e3f2fd'},
    {re:/ESTIMATED NET WORTH[^\n]+₹([\d,]+\.\d+)/,label:'Net Worth',color:'#6a1b9a',bg:'#f3e5f5'},
    {re:/TOTAL ASSETS\s+₹([\d,]+\.\d+)/,label:'Total Assets',color:'#0d47a1',bg:'#e3f2fd'},
    {re:/TOTAL LIABILITIES\s+₹([\d,]+\.\d+)/,label:'Total Liabilities',color:'#b71c1c',bg:'#ffebee'},
  ];
  patterns.forEach(({re,label,color,bg})=>{
    const m=text.match(re);
    if(m)chips.push({label,value:'₹'+m[1],color,bg});
  });
  return chips;
}

function _renderChips(chips,containerId){
  const el=document.getElementById(containerId);
  if(!el)return;
  el.innerHTML=chips.map(c=>`
    <div style="background:${c.bg};border:1.5px solid ${c.color}33;border-radius:10px;padding:8px 16px;min-width:120px">
      <div style="font-size:11px;color:#666;font-weight:600;letter-spacing:.3px;text-transform:uppercase">${c.label}</div>
      <div style="font-size:15px;font-weight:800;color:${c.color};margin-top:2px">${c.value}</div>
    </div>`).join('');
}
function renderStatements(reports, acctReports, selectedAcctId){
  // reports      = combined (all accounts) data
  // acctReports  = per-account data (null if no account selected)
  // selectedAcctId = account_id string or null
  reports = reports || {};
  _lastReports = reports;
  if(acctReports) _lastReports._acct = acctReports;
  if(selectedAcctId) _lastReports._acctId = selectedAcctId;

  const incomeWrap   = document.getElementById('stmt-income');
  const incomeBody   = document.getElementById('stmt-income-body');
  const balanceWrap  = document.getElementById('stmt-balance');
  const balanceBody  = document.getElementById('stmt-balance-body');
  const plWrap       = document.getElementById('stmt-pl');
  const plBody       = document.getElementById('stmt-pl-body');
  const emptyWrap    = document.getElementById('stmt-empty');

  const isSplit = !!(acctReports && selectedAcctId);
  _toggleStmtAccountButtons(isSplit);
  // ── Period label: extract from statement text (Fix 1) ──────────────────────
  const from = (document.getElementById('stmt_date_from') || {}).value || '';
  const to   = (document.getElementById('stmt_date_to')   || {}).value || '';
  const acct = selectedAcctId || (document.getElementById('stmt_account') || {}).value || '';
  const fallbackLabel = [
    from && to ? `${from} to ${to}` : (from ? `From ${from}` : (to ? `Up to ${to}` : 'All dates')),
    acct ? `Account: ${acct}` : 'All accounts'
  ].join(' • ');
  const incomePeriod  = document.getElementById('stmt-income-period');
  const balancePeriod = document.getElementById('stmt-balance-period');
  const plPeriod      = document.getElementById('stmt-pl-period');
  const _pLabel = (txt, suffix) => {
    const p = _extractStmtPeriod(txt);
    return (p || `Period: ${fallbackLabel}`) + (suffix ? `  |  ${suffix}` : '');
  };
  const combinedSuffix = isSplit ? `Showing: All Accounts vs ${selectedAcctId}` : '';
  const incomeTxt  = isSplit ? ((acctReports && acctReports.income_statement) || '') : reports.income_statement;
  const balanceTxt = isSplit ? ((acctReports && acctReports.balance_sheet) || '') : reports.balance_sheet;
  const plTxt      = isSplit ? ((acctReports && acctReports.profit_loss) || '') : reports.profit_loss;
  console.log('[STMT PERIOD TEXT SOURCE]', {isSplit, selectedAcctId, incomeTxt, balanceTxt, plTxt});

  if(incomePeriod)  incomePeriod.textContent  = _pLabel(incomeTxt,  combinedSuffix);
  if(balancePeriod) balancePeriod.textContent = _pLabel(balanceTxt, combinedSuffix);
  if(plPeriod)      plPeriod.textContent      = _pLabel(plTxt,      combinedSuffix);

  if(incomeWrap)  incomeWrap.style.display  = 'none';
  if(balanceWrap) balanceWrap.style.display = 'none';
  if(plWrap)      plWrap.style.display      = 'none';

  // Helper: toggle split vs single view
  function _toggleSplitView(key, show){
    const splitEl  = document.getElementById(`stmt-${key}-split`);
    const singleEl = document.getElementById(`stmt-${key}-single`);
    if(splitEl)  splitEl.style.display  = show ? 'block' : 'none';
    if(singleEl) singleEl.style.display = show ? 'none'  : 'block';
  }

  let any = false;
  try {
    // ── Income Statement ──────────────────────────────────────────────────────
    if ((reports.income_statement || (acctReports && acctReports.income_statement)) && incomeWrap) {
      incomeWrap.style.display = 'block';
      _toggleSplitView('income', isSplit);
      if(isSplit){
        // Per-account panel
        const acctLbl = document.getElementById('stmt-income-acct-label');
        if(acctLbl) acctLbl.textContent = `📄 ${selectedAcctId}`;
        document.getElementById('stmt-income-acct-body').innerHTML = stmtHtml(acctReports.income_statement || '(no data)');
        _renderChips(_extractChips(acctReports.income_statement || ''), 'stmt-income-acct-chips');
        // Combined panel
        document.getElementById('stmt-income-all-body').innerHTML = stmtHtml(reports.income_statement || '(no data)');
        _renderChips(_extractChips(reports.income_statement || ''), 'stmt-income-all-chips');
        // Top-level chips = combined
        _renderChips(_extractChips(reports.income_statement || ''), 'stmt-income-chips');
      } else {
        if(incomeBody) incomeBody.innerHTML = stmtHtml(reports.income_statement);
        _renderChips(_extractChips(reports.income_statement), 'stmt-income-chips');
      }
      any = true;
    }

    // ── Balance Sheet ─────────────────────────────────────────────────────────
    if ((reports.balance_sheet || (acctReports && acctReports.balance_sheet)) && balanceWrap) {
      balanceWrap.style.display = 'block';
      _toggleSplitView('balance', isSplit);
      if(isSplit){
        const acctLbl = document.getElementById('stmt-balance-acct-label');
        if(acctLbl) acctLbl.textContent = `📄 ${selectedAcctId}`;
        document.getElementById('stmt-balance-acct-body').innerHTML = stmtHtml(acctReports.balance_sheet || '(no data)');
        _renderChips(_extractChips(acctReports.balance_sheet || ''), 'stmt-balance-acct-chips');
        document.getElementById('stmt-balance-all-body').innerHTML = stmtHtml(reports.balance_sheet || '(no data)');
        _renderChips(_extractChips(reports.balance_sheet || ''), 'stmt-balance-all-chips');
        _renderChips(_extractChips(reports.balance_sheet || ''), 'stmt-balance-chips');
      } else {
        if(balanceBody) balanceBody.innerHTML = stmtHtml(reports.balance_sheet);
        _renderChips(_extractChips(reports.balance_sheet), 'stmt-balance-chips');
      }
      any = true;
    }

    // ── Profit & Loss ─────────────────────────────────────────────────────────
    if ((reports.profit_loss || (acctReports && acctReports.profit_loss)) && plWrap) {
      plWrap.style.display = 'block';
      _toggleSplitView('pl', isSplit);
      if(isSplit){
        const acctLbl = document.getElementById('stmt-pl-acct-label');
        if(acctLbl) acctLbl.textContent = `📄 ${selectedAcctId}`;
        document.getElementById('stmt-pl-acct-body').innerHTML = stmtHtml(acctReports.profit_loss || '(no data)');
        _renderChips(_extractChips(acctReports.profit_loss || ''), 'stmt-pl-acct-chips');
        document.getElementById('stmt-pl-all-body').innerHTML = stmtHtml(reports.profit_loss || '(no data)');
        _renderChips(_extractChips(reports.profit_loss || ''), 'stmt-pl-all-chips');
        _renderChips(_extractChips(reports.profit_loss || ''), 'stmt-pl-chips');
      } else {
        if(plBody) plBody.innerHTML = stmtHtml(reports.profit_loss);
        _renderChips(_extractChips(reports.profit_loss), 'stmt-pl-chips');
      }
      any = true;
    }

  } catch (e) {
    console.error('renderStatements failed', e, reports);
    if (emptyWrap) {
      emptyWrap.style.display = 'block';
      emptyWrap.innerHTML = `<div style="color:#b71c1c;font-size:13px">Statement rendering failed: ${String(e)}</div>`;
    }
    return;
  }

  if (emptyWrap) {
    emptyWrap.style.display = any ? 'none' : 'block';
    if (!any) {
      const err = reports.error ? `<div style="color:#b71c1c;font-size:13px;margin-top:6px">${reports.error}</div>` : '';
      emptyWrap.innerHTML = `<div>No statement data available for the selected filters.</div>${err}`;
    }
  }
}

function _parseTraceMeta(line){
  const m=line.match(/\[TRACE:([^\]]+)\]/);
  if(!m) return null;
  const parts=m[1].split(':');
  if(parts[0]==='bank_balance') return {type:'bank_balance'};
  if(parts.length>=4) return {type:'bucket',book:parts[0],section:parts[1],grp:parts[2],account:parts.slice(3).join(':')};
  if(parts.length>=3) return {type:'bucket',book:parts[0],section:parts[1],grp:parts[2]};
  return null;
}

// ── Statement renderer with counterparty grouping ─────────────────────────────
let _stmtDetailCounter = 0; // unique IDs for expand/collapse toggles

function _parseDetailLine(line) {
  // Matches:  │  2025-12-14  Counterparty Name  |  narration...  CR  ₹ 1,234.56
  // Note: _fmt() pads with spaces after ₹ so we allow \s* between ₹ and digits
  const m = line.match(/^\s*│\s+(\d{4}-\d{2}-\d{2})\s+(.+?)\s+\|\s+(.+?)\s+(CR|DR)\s+(₹[\s\d,]+\.\d+)\s*$/);
  if (m) return { date: m[1], counterparty: m[2].trim(), narration: m[3].trim(), dir: m[4], amount: m[5].trim() };
  // Fallback: synthetic lines like bank balance breakdown
  const m2 = line.match(/^\s*│\s+(.+)$/);
  if (m2) return { raw: m2[1].trim() };
  return null;
}

function _amountToNum(s) {
  return parseFloat((s||'').replace(/[₹,\s]/g,'')) || 0;
}

function _escapeHtmlAttr(s){
  return String(s||'').replace(/&/g,'&amp;').replace(/"/g,'&quot;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
}
function openTraceFromBtn(btn){
  try{
    const title = btn.getAttribute('data-trace-title') || 'Trace';
    const raw = btn.getAttribute('data-trace-params') || '{}';
    const params = JSON.parse(raw);
    return openTrace(title, params);
  }catch(e){
    console.error('trace button failed', e);
    alert('Trace could not be opened for this line.');
  }
}

function _renderDetailGroup(lines, blockId) {
  const parsed = lines.map(_parseDetailLine).filter(Boolean);
  if (!parsed.length) return '';

  const rawLines = parsed.filter(p => p.raw);
  const txnLines = parsed.filter(p => p.date);

  // Group by counterparty — case-insensitive key, preserve first-seen label
  const groupMap = {};
  txnLines.forEach(t => {
    const raw   = (t.counterparty || '').trim();
    const label = raw || 'Unknown / Unmapped';
    const key   = label.toLowerCase();
    if (!groupMap[key]) groupMap[key] = { label, net: 0, hasCR: false, hasDR: false };
    const amt = _amountToNum(t.amount);
    // CR = positive (inflow), DR = negative (outflow)
    groupMap[key].net += (t.dir === 'CR' ? amt : -amt);
    if (t.dir === 'CR') groupMap[key].hasCR = true;
    else                groupMap[key].hasDR = true;
  });

  // Sort by absolute net descending
  const sorted = Object.values(groupMap)
    .sort((a, b) => Math.abs(b.net) - Math.abs(a.net));

  let html = `<div id="${blockId}" style="display:none;margin:2px 0 6px 32px;border-left:2px solid #e3f2fd;padding-left:10px;padding-top:4px;padding-bottom:4px">`;

  // Synthetic/raw lines (bank balance note etc.)
  rawLines.forEach(r => {
    html += `<div style="font-size:11px;color:#888;padding:2px 0;font-family:monospace">${r.raw.replace(/&/g,'&amp;').replace(/</g,'&lt;')}</div>`;
  });

  // One row per counterparty — name left, net amount right, no drill-down
  sorted.forEach(g => {
    const isPositive = g.net >= 0;
    const amtColor   = isPositive ? '#1b5e20' : '#b71c1c';
    const amtFmt     = '₹' + Math.abs(g.net).toLocaleString('en-IN', {minimumFractionDigits:2});
    const sign       = isPositive ? '+' : '−';
    const dirLabel   = (g.hasCR && g.hasDR) ? 'NET' : g.hasCR ? 'CR' : 'DR';
    html += `<div style="display:flex;justify-content:space-between;align-items:center;padding:4px 6px;border-bottom:1px solid #f5f5f5;border-radius:3px" onmouseover="this.style.background='#f9f9f9'" onmouseout="this.style.background=''">
      <span style="font-size:12px;color:#333;font-weight:600">${g.label.replace(/&/g,'&amp;')}</span>
      <span style="display:flex;align-items:center;gap:8px;flex-shrink:0">
        <span style="font-size:10px;color:#aaa">${dirLabel}</span>
        <span style="font-size:12px;font-weight:800;color:${amtColor};min-width:90px;text-align:right">${sign} ${amtFmt}</span>
      </span>
    </div>`;
  });

  html += '</div>';
  return html;
}

function stmtHtml(text){
  const lines=(text||'').split('\n');
  let html='<div style="font-family:\'Courier New\',monospace;font-size:12px;line-height:1.9;overflow-x:auto">';

  // We process lines in two passes:
  // When we hit a ├ account line, we collect all following │ detail lines
  // and render them as a grouped expandable block.
  let i = 0;
  while (i < lines.length) {
    const line = lines[i];
    const traceMeta = _parseTraceMeta(line);
    const e = line.replace(/\s*\[TRACE:[^\]]+\]/g,'').replace(/&/g,'&amp;').replace(/</g,'&lt;');

    let traceBtn = '';
    if (traceMeta) {
      let onclick = '', label = '🔍 Trace';
      if (traceMeta.type === 'bank_balance') {
        const params = {account:'Bank Balance'};
        traceBtn = `<button type="button" data-trace-title="${_escapeHtmlAttr('Bank Balance (All Transactions)')}" data-trace-params="${_escapeHtmlAttr(JSON.stringify(params))}" onclick="openTraceFromBtn(this)" style="font-size:10px;padding:1px 7px;margin-left:8px;background:#e3f2fd;border:1px solid #90caf9;border-radius:4px;color:#1565c0;cursor:pointer;font-family:sans-serif;vertical-align:middle">🔍 Bank Bal</button>`;
      } else {
        const t = `${traceMeta.section} › ${traceMeta.grp}${traceMeta.account ? ' › ' + traceMeta.account : ''}`;
        const params = {book:traceMeta.book, section:traceMeta.section, grp:traceMeta.grp};
        if (traceMeta.account) params.account = traceMeta.account;
        traceBtn = `<button type="button" data-trace-title="${_escapeHtmlAttr(t)}" data-trace-params="${_escapeHtmlAttr(JSON.stringify(params))}" onclick="openTraceFromBtn(this)" style="font-size:10px;padding:1px 7px;margin-left:8px;background:#e3f2fd;border:1px solid #90caf9;border-radius:4px;color:#1565c0;cursor:pointer;font-family:sans-serif;vertical-align:middle">🔍 Trace</button>`;
      }
    }

    // ── Account line (├) — collect following detail lines (│) ────────────────
    if (/├/.test(line)) {
      // Gather all consecutive │ lines after this ├ line
      const detailLines = [];
      let j = i + 1;
      while (j < lines.length && /^\s*│/.test(lines[j])) {
        detailLines.push(lines[j]);
        j++;
      }

      const blockId = `stmt-detail-${++_stmtDetailCounter}`;
      const hasDetail = detailLines.length > 0;

      let expandBtn = '';
      if (hasDetail) {
        expandBtn = `<button onclick="toggleEl('${blockId}')" style="font-size:10px;padding:1px 7px;margin-left:8px;background:#f5f5f5;border:1px solid #ddd;border-radius:4px;color:#555;cursor:pointer;font-family:sans-serif;vertical-align:middle">▼ detail</button>`;
      }

      html += `<div style="color:#333">${e}${traceBtn}${expandBtn}</div>`;

      if (hasDetail) {
        html += _renderDetailGroup(detailLines, blockId);
      }

      i = j; // skip the │ lines we already consumed
      continue;
    }

    // ── All other line types ──────────────────────────────────────────────────
    if (/^\s*│/.test(line)) {
      // Orphan │ line (not preceded by a ├) — skip, already consumed above
      i++; continue;
    }

    if      (/^[═=]{5,}/.test(e.trim()))                                                      html+=`<div style="color:#1565c0;font-weight:800">${e}</div>`;
    else if (/^[─\-]{5,}/.test(e.trim()))                                                     html+=`<div style="color:#ddd">${e}</div>`;
    else if (/(TOTAL|NET INCOME|NET PROFIT|NET WORTH|EBITDA|EBIT)/.test(e)&&/₹/.test(e))     html+=`<div style="font-weight:800;color:#1b5e20;background:#f0fff4;padding:1px 4px;border-radius:3px">${e}</div>`;
    else if (/(▌|INCOME|EXPENSES|REVENUE|ASSETS|LIABILITIES|EQUITY|TAXATION)/.test(e)&&e.length<80) html+=`<div style="font-weight:800;color:#1565c0;margin-top:6px">${e}</div>`;
    else                                                                                       html+=`<div style="color:#555">${e}</div>`;

    i++;
  }
  return html+'</div>';
}

function _extractStmtPeriod(text){
  const lines=String(text||'').split('\n').map(s=>s.trim()).filter(Boolean);
  const p=lines.find(l=>/^Period:/i.test(l));
  if(p) return p;
  const a=lines.find(l=>/^All transactions/i.test(l));
  return a||'';
}

function _copyText(text, btn){
  if(!text){alert('No data to copy.');return;}
  function _fallback(){
    const ta=document.createElement('textarea');
    ta.value=text;ta.style.position='fixed';ta.style.opacity='0';
    document.body.appendChild(ta);ta.focus();ta.select();
    try{document.execCommand('copy');
      if(btn){const o=btn.textContent;btn.textContent='✅ Copied!';setTimeout(()=>{btn.textContent=o;},1500);}
    }catch(e){alert('Copy failed — please select and copy manually.');}
    document.body.removeChild(ta);
  }
  if(navigator.clipboard&&window.isSecureContext){
    navigator.clipboard.writeText(text).then(()=>{
      if(btn){const o=btn.textContent;btn.textContent='✅ Copied!';setTimeout(()=>{btn.textContent=o;},1500);}
    }).catch(_fallback);
  } else { _fallback(); }
}
function copyStmt(which){
  const map={income:'income_statement',balance:'balance_sheet',pl:'profit_loss',
             income_acct:'income_statement',balance_acct:'balance_sheet',pl_acct:'profit_loss'};
  const isAcct = which.endsWith('_acct');
  const src = isAcct ? (_lastReports._acct||{}) : _lastReports;
  const key = map[which]||which;
  const text = src[key]||'';
  _copyText(text, event.target);
}

function downloadStmt(which){
  const map = {
    income: 'income_statement',
    balance: 'balance_sheet',
    pl: 'profit_loss',
    income_acct: 'income_statement',
    balance_acct: 'balance_sheet',
    pl_acct: 'profit_loss'
  };

  const isAcct = which.endsWith('_acct');
  const src = isAcct ? (_lastReports._acct || {}) : _lastReports;
  const key = map[which] || which;
  const text = src[key] || '';

  if(!text){
    alert('No statement data to download.');
    return;
  }

  const filenameMap = {
    income: 'income_statement.txt',
    balance: 'balance_sheet.txt',
    pl: 'profit_loss.txt',
    income_acct: 'income_statement_this_account.txt',
    balance_acct: 'balance_sheet_this_account.txt',
    pl_acct: 'profit_loss_this_account.txt'
  };

  const blob = new Blob([text], { type: 'text/plain;charset=utf-8' });
  const url = URL.createObjectURL(blob);
  const a = document.createElement('a');
  a.href = url;
  a.download = filenameMap[which] || 'statement.txt';
  document.body.appendChild(a);
  a.click();
  a.remove();
  setTimeout(() => URL.revokeObjectURL(url), 1000);
}

function _toggleStmtAccountButtons(isSplit){
  [
    'btn-copy-income-acct',
    'btn-download-income-acct',
    'btn-copy-balance-acct',
    'btn-download-balance-acct',
    'btn-copy-pl-acct',
    'btn-download-pl-acct'
  ].forEach(id => {
    const el = document.getElementById(id);
    if(el) el.style.display = isSplit ? '' : 'none';
  });
}

async function downloadReportPdf(kind, reportName='', date_from_override='', date_to_override='', account_override=''){
  const date_from = date_from_override || (document.getElementById('stmt_date_from') || {value:''}).value || '';
  const date_to   = date_to_override   || (document.getElementById('stmt_date_to')   || {value:''}).value || '';
  const account_id = account_override  || (document.getElementById('stmt_account')    || {value:''}).value || '';

  const qs = new URLSearchParams({
    kind: kind || 'detailed',
    date_from,
    date_to,
    account_id
  });

  if(reportName){
    qs.set('report', reportName);
  }

  try{
    const res = await fetch('/download-report-pdf?' + qs.toString(), {
      method: 'GET',
      credentials: 'same-origin'
    });

    if(!res.ok){
      const msg = await res.text();
      console.error('PDF download failed:', res.status, msg);
      alert('PDF download failed: ' + msg);
      return;
    }

    const blob = await res.blob();
    if(!blob || blob.size === 0){
      alert('PDF download failed: empty file returned.');
      return;
    }

    const cd = res.headers.get('Content-Disposition') || '';
    let filename = 'report.pdf';
    const m = cd.match(/filename=\"?([^"]+)\"?/i);
    if(m && m[1]) filename = m[1];

    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url;
    a.download = filename;
    document.body.appendChild(a);
    a.click();
    a.remove();
    setTimeout(() => URL.revokeObjectURL(url), 1500);
  }catch(e){
    console.error('PDF download exception:', e);
    alert('PDF download failed: ' + e);
  }
}

async function refreshStmts(){
  const date_from=(document.getElementById('stmt_date_from')||{value:''}).value||null;
  const date_to  =(document.getElementById('stmt_date_to'  )||{value:''}).value||null;
  const account_id=(document.getElementById('stmt_account'  )||{value:''}).value||null;

  if (date_from && date_to && date_from > date_to) {
    alert('From date cannot be greater than To date');
    return;
  }
  try{
    // Always fetch combined (all accounts) report
    const rAll=await fetch('/financial-reports',{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({date_from,date_to,account_id:null})
    });
    const dAll=await rAll.json();
    if(!rAll.ok||dAll.error){ renderStatements({error:dAll.error||('HTTP '+rAll.status)},null,null); return; }
    // If a specific account is selected, also fetch per-account report
    if(account_id){
      const rAcct=await fetch('/financial-reports',{
        method:'POST',
        headers:{'Content-Type':'application/json'},
        body:JSON.stringify({date_from,date_to,account_id})
      });
      const dAcct=await rAcct.json();
      if(rAcct.ok&&!dAcct.error){ renderStatements(dAll||{},dAcct||{},account_id); return; }
    }
    renderStatements(dAll||{},null,null);
  }catch(e){
    renderStatements({error:'Refresh failed: '+e},null,null);
  }
}

function clearStmtFilters(){
  const df=document.getElementById('stmt_date_from'); if(df) df.value='';
  const dt=document.getElementById('stmt_date_to');   if(dt) dt.value='';
  const ac=document.getElementById('stmt_account');   if(ac) ac.value='';
  const yr=document.getElementById('stmt_year');      if(yr) yr.value='';
  refreshStmts();
}

async function showStmtTab(){
  await loadAccountDropdowns();
  await populateYearDropdown();
  await refreshStmts();
}

// ── Manual classify (paste) removed because the quick classify UI is disabled.
function classifyPasted(){
  setStatus('Quick classify has been removed. Use file upload and review instead.','info');
}

function renderLedgerTable(rows){
  const el=document.getElementById('ledger-table');
  if(!el)return;
  if(!rows||!rows.length){
    el.innerHTML='<p style="color:#bbb;font-size:13px">No ledger entries match this filter.</p>';
    return;
  }
  let html='<table><thead><tr><th>Date</th><th>Narration</th><th>Counterparty</th><th>Book</th><th>Section</th><th>Group</th><th>Account</th><th style="text-align:right">Amount</th><th>Type</th><th>Conf</th><th>Source</th></tr></thead><tbody>';
  rows.forEach(r=>{
    const amt='₹'+parseFloat(r.amount||0).toLocaleString('en-IN',{minimumFractionDigits:2});
    const bookLbl=BOOK_LABEL[r.book]||r.book||'—';
    const cp=r.counterparty||'—';
    html+=`<tr>
      <td style="white-space:nowrap">${r.txn_date||''}</td>
      <td style="max-width:320px;word-break:break-word">${r.narration||r.description||''}</td>
      <td style="max-width:160px;word-break:break-word">${cp}</td>
      <td><span class="book-tag ${BOOK_CLASS[r.book]||'book-sus'}">${bookLbl}</span></td>
      <td>${r.section||'—'}</td>
      <td>${r.grp||r.group||'—'}</td>
      <td>${r.account||'—'}</td>
      <td style="text-align:right;font-weight:800">${amt}</td>
      <td><span class="badge ${r.txn_type==='credit'?'b-cr':'b-dr'}">${r.txn_type||''}</span></td>
      <td style="font-size:11px;color:#777">${Math.round((r.confidence||0)*100)}%</td>
      <td style="font-size:11px;color:#777">${r.source||''}</td>
    </tr>`;
  });
  html+='</tbody></table>';
  el.innerHTML=html;
}

// ── Ledger ────────────────────────────────────────────────────────────────────
async function loadStats(){
  const host = document.getElementById('stat-cards');
  if(!host) return;
  try{
    const r = await fetch('/stats');
    const d = await r.json();
    const items=[{k:'income',l:'Total Income',c:'#1b5e20'},{k:'expenditure',l:'Total Expenditure',c:'#b71c1c'},{k:'assets',l:'Total Assets',c:'#0d47a1'},{k:'net_surplus',l:'Net Surplus',c:'#e65100'},{k:'txns',l:'Transactions',c:'#444'}];
    host.innerHTML=items.map(({k,l,c})=>{const v=d[k]||0;const disp=k==='txns'?v:`₹${parseFloat(v).toLocaleString('en-IN',{minimumFractionDigits:0})}`;return `<div class="stat-card"><div class="val" style="color:${c}">${disp}</div><div class="lbl">${l}</div></div>`;}).join('');
  }catch(e){
    host.innerHTML = `<div style="color:#b71c1c;font-size:13px">Stats failed to load</div>`;
  }
}

async function loadLedger(all=false){
  const bookEl = document.getElementById('f_book');
  const typeEl = document.getElementById('f_type');
  const searchEl = document.getElementById('f_search');
  const cntEl = document.getElementById('ledger-cnt');
  if(!bookEl || !typeEl || !searchEl) return;
  const book=bookEl.value;
  const type=typeEl.value;
  const search=searchEl.value.trim();
  const counterparty=(document.getElementById('f_counterparty')||{value:''}).value.trim();
  const account_id=(document.getElementById('f_account')||{value:''}).value.trim();
  try{
    const r=await fetch('/ledger',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({book:book||null,txn_type:type||null,search:search||null,counterparty:counterparty||null,account_id:account_id||null,limit:all?2000:500})});
    const rows=await r.json();
    if(cntEl) cntEl.textContent=`(${Array.isArray(rows)?rows.length:0})`;
    renderLedgerTable((Array.isArray(rows)?rows:[]).map(r=>({...r,narration:r.narration||r.description||''})));
  }catch(e){
    if(cntEl) cntEl.textContent='(0)';
    renderLedgerTable([]);
  }
}


// ── Account dropdowns + manage data ──────────────────────────────────────────
async function loadAccountDropdowns(){
  try{
    const r=await fetch('/accounts');
    if(!r.ok) return;
    const accounts=await r.json();
    // Populate stmt_account dropdown
    const sd=document.getElementById('stmt_account');
    if(sd){
      const cur=sd.value;
      sd.innerHTML='<option value="">All Accounts</option>'+
        accounts.map(a=>`<option value="${a.account_id}">${a.account_id} (${a.txn_count} txns)</option>`).join('');
      if(cur) sd.value=cur;
    }
    // Populate ledger f_account dropdown
    const ld=document.getElementById('f_account');
    if(ld){
      const cur2=ld.value;
      ld.innerHTML='<option value="">All Accounts</option>'+
        accounts.map(a=>`<option value="${a.account_id}">${a.account_id}</option>`).join('');
      if(cur2) ld.value=cur2;
    }
    // Populate manage data list
    const ml=document.getElementById('manage-accounts-list');
    if(ml){
      if(!accounts.length){
        ml.innerHTML='<p style="font-size:13px;color:#aaa">No accounts yet.</p>';
      } else {
        ml.innerHTML=accounts.map(a=>`
          <div style="display:flex;justify-content:space-between;align-items:center;padding:8px 0;border-bottom:1px solid #f5f5f5">
            <div>
              <span style="font-weight:700;font-size:13px">${a.account_id}</span>
              <span style="font-size:11px;color:#888;margin-left:10px">${a.txn_count} txns · ${a.date_from||''} to ${a.date_to||''}</span>
            </div>
            <button class="btn bo" style="font-size:11px;padding:4px 12px;color:#b71c1c;border-color:#b71c1c"
              onclick="deleteAccountData('${a.account_id}')">🗑 Delete</button>
          </div>`).join('');
      }
    }
  }catch(e){}
}

async function deleteAccountData(account_id){
  if(!_activeUserId){ setStatus('Log in to manage data.','err'); return; }
  if(!confirm(`Delete ALL data for account '${account_id}'?\nThis cannot be undone.`)) return;
  try{
    const r=await fetch('/delete-account-data',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({account_id})});
    const d=await r.json();
    const msg=document.getElementById('manage-msg');
    if(msg) msg.innerHTML=`<span style="color:${d.error?'#b71c1c':'#1b5e20'}">${d.msg||d.error}</span>`;
    await loadAccountDropdowns();
    _lastReports={};
  }catch(e){alert('Delete failed: '+e);}
}

async function deleteAllData(){
  if(!_activeUserId){ setStatus('Log in to manage data.','err'); return; }
  if(!confirm('Delete ALL transaction data for this user?\nYour login account is preserved. This cannot be undone.')) return;
  try{
    const r=await fetch('/delete-all-data',{method:'POST',headers:{'Content-Type':'application/json'},body:'{}'});
    const d=await r.json();
    const msg=document.getElementById('manage-msg');
    if(msg) msg.innerHTML=`<span style="color:${d.error?'#b71c1c':'#1b5e20'}">${d.msg||d.error}</span>`;
    await loadAccountDropdowns();
    _lastReports={};
    document.getElementById('stmt-income') && (document.getElementById('stmt-income').style.display='none');
    document.getElementById('stmt-balance') && (document.getElementById('stmt-balance').style.display='none');
    document.getElementById('stmt-pl') && (document.getElementById('stmt-pl').style.display='none');
    document.getElementById('stmt-empty').style.display='block';
  }catch(e){alert('Delete failed: '+e);}
}

async function deleteUserCompletely(){
  if(!_activeUserId){ setStatus('Log in to manage data.','err'); return; }
  const infoEl = document.getElementById('active-user-info');
  const txt = (infoEl && infoEl.textContent) || '';
  const userName = txt.split(' · ')[0].trim();
  if(!userName){ alert('No active user found.'); return; }

  if(!confirm(`Delete USER '${userName}' completely?\nThis deletes login, accounts, statements, imports, and review data.\nThis cannot be undone.`)) return;

  const typed = prompt(`Type the exact user name to confirm deletion:\n${userName}`);
  if(typed !== userName){
    alert('Typed confirmation did not match. Delete cancelled.');
    return;
  }

  try{
    const r = await fetch('/delete-user',{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({confirm_name:userName})
    });
    const d = await r.json();
    if(d.error){ alert(d.error); return; }
    alert(d.msg || 'User deleted.');
    logoutUser();
    location.reload();
  }catch(e){
    alert('Delete failed: ' + e);
  }
}

// ── Full Report ───────────────────────────────────────────────────────────────
async function genReport(){document.getElementById('ledger-out').textContent='Generating…';const r=await fetch('/report');document.getElementById('ledger-out').textContent=await r.text();}
async function checkSys(){const r=await fetch('/check');document.getElementById('sys-out').textContent=await r.text();}

// ── Ledger Dictionary ─────────────────────────────────────────────────────────
let _dictData = [];
async function loadDictionary(){
  const out = document.getElementById('dict-out');
  const body = document.getElementById('dict-body');
  if(!out || !body) return;

  if(_dictData.length){
    out.style.display='block';
    renderDict(_dictData);
    return;
  }

  body.innerHTML='<p style="color:#aaa;font-size:13px">Loading…</p>';
  out.style.display='block';
  try{
    const r=await fetch('/ledger-dictionary');
    const d=await r.json();
    if(d.error){
      body.innerHTML=`<p style="color:var(--red)">${d.error}</p>`;
      return;
    }
    _dictData=d.dictionary||[];
    renderDict(_dictData);
  }catch(e){
    body.innerHTML=`<p style="color:var(--red)">Failed: ${e}</p>`;
  }
}
function filterDict(){
  const el=document.getElementById('dict-search');
  renderDict(_filterDictGroups((el&&el.value)||''));
}

function _renderDictInto(bodyId, groups){
  const body=document.getElementById(bodyId);
  if(!body) return;
  body.style.display='block';
  body.style.visibility='visible';
  body.style.minHeight='120px';
  const BOOK_COLOR={'BALANCE_SHEET':'#1b5e20','INCOME_EXPENSE':'#1565c0','SUSPENSE':'#880e4f','CAPITAL':'#e65100'};
  if(!groups || !groups.length){
    body.innerHTML = '<p style="color:#aaa;font-size:13px">No matching entries.</p>';
    return;
  }

  let html='';
  groups.forEach(g=>{
    const bc=BOOK_COLOR[g.book]||'#555';
    html+=`<div style="margin-bottom:14px;border-radius:10px;overflow:hidden;border:1px solid #e8e8e8">
      <div style="background:${bc};color:#fff;padding:9px 14px;font-size:12px;font-weight:800;letter-spacing:.4px">
        ${String(g.book||'').replace('_',' ')} › ${g.section} › ${g.group}
      </div>
      <table style="width:100%;border-collapse:collapse;font-size:11.5px">
        <thead><tr style="background:#f5f7fa">
          <th style="padding:7px 10px;text-align:left;border-bottom:1px solid #e8e8e8;width:28%">Key</th>
          <th style="padding:7px 10px;text-align:left;border-bottom:1px solid #e8e8e8;width:24%">Account Name</th>
          <th style="padding:7px 10px;text-align:left;border-bottom:1px solid #e8e8e8">Description</th>
        </tr></thead><tbody>`;
    (g.entries||[]).forEach(e=>{
      html+=`<tr style="border-bottom:1px solid #f3f3f3">
        <td style="padding:7px 10px;font-family:monospace;color:${bc};font-size:11px;font-weight:700">${e.key}</td>
        <td style="padding:7px 10px;font-weight:600;color:#333">${e.account||'—'}</td>
        <td style="padding:7px 10px;color:#666">${e.description||'—'}</td>
      </tr>`;
    });
    html+=`</tbody></table></div>`;
  });
  body.innerHTML = html;
}

function renderDict(groups){ _renderDictInto('dict-body', groups); }
function _filterDictGroups(q){
  q=(q||'').toLowerCase();
  if(!q) return _dictData;
  return _dictData.map(g=>({...g,entries:(g.entries||[]).filter(e=>
    (e.key||'').toLowerCase().includes(q)||
    (e.account||'').toLowerCase().includes(q)||
    (e.description||'').toLowerCase().includes(q)||
    (e.attribution||'').toLowerCase().includes(q)
  )})).filter(g=>g.entries.length>0);
}

// ── Profile Questionnaire ────────────────────────────────────────────────────
function _escHtml(v){
  return String(v ?? '')
    .replace(/&/g,'&amp;')
    .replace(/</g,'&lt;')
    .replace(/>/g,'&gt;')
    .replace(/"/g,'&quot;')
    .replace(/'/g,'&#39;');
}
function _profileMsg(msg='', color='#666'){
  const el=document.getElementById('pf_msg');
  if(el){ el.textContent=msg; el.style.color=color; }
}
function _profileNoteParts(raw){
  const txt=String(raw||'');
  const m=txt.match(/^\[default_treatment=([^\]]*)\]\s*(.*)$/);
  return { default_treatment: m ? m[1] : '', notes: m ? (m[2]||'') : txt };
}
function _profileJoinNotes(defaultTreatment, notes){
  const dt=String(defaultTreatment||'').trim();
  const tail=String(notes||'').trim();
  return dt ? `[default_treatment=${dt}]${tail ? ' ' + tail : ''}` : tail;
}
function _counterpartyRowHtml(kind, row={}){
  const meta=_profileNoteParts(row.notes||'');
  return `<tr data-kind="${_escHtml(kind)}">
    <td><input data-f="display_name" value="${_escHtml(row.display_name||'')}" placeholder="Sarju Garg HUF"></td>
    <td><input data-f="relationship" value="${_escHtml(row.relationship||'')}" placeholder="${kind==='family'?'Mother / Brother':'Client / Broker'}"></td>
    <td><input data-f="default_treatment" value="${_escHtml(meta.default_treatment||'')}" placeholder="family loan / salary / trading"></td>
    <td><input data-f="default_ledger_key" value="${_escHtml(row.default_ledger_key||'')}" placeholder="liability_loan_outstanding"></td>
    <td><input data-f="txn_direction_hint" value="${_escHtml(row.txn_direction_hint||'')}" placeholder="credit / debit / both"></td>
    <td><input data-f="notes" value="${_escHtml(meta.notes||'')}" placeholder="Optional notes"></td>
    <td><button class="btn bo" style="font-size:11px;padding:4px 10px" onclick="this.closest('tr').remove();toggleProfileSections()">✕</button></td>
  </tr>`;
}
function _knownAccountRowHtml(row={}){
  return `<tr>
    <td><input data-f="account_label" value="${_escHtml(row.account_label||'')}" placeholder="HDFC Salary"></td>
    <td><input data-f="institution_name" value="${_escHtml(row.institution_name||'')}" placeholder="HDFC Bank"></td>
    <td><input data-f="account_mask" value="${_escHtml(row.account_mask||'')}" placeholder="XX1234"></td>
    <td><input data-f="account_type" value="${_escHtml(row.account_type||'')}" placeholder="Savings"></td>
    <td><input data-f="ownership_type" value="${_escHtml(row.ownership_type||'')}" placeholder="Self / Joint"></td>
    <td><button class="btn bo" style="font-size:11px;padding:4px 10px" onclick="this.closest('tr').remove();toggleProfileSections()">✕</button></td>
  </tr>`;
}
function addProfileCounterpartyRow(kind, row={}){
  const body=document.getElementById(`pf_${kind}_rows`);
  if(body) body.insertAdjacentHTML('beforeend', _counterpartyRowHtml(kind, row));
  toggleProfileSections();
}
function addKnownAccountRow(row={}){
  const body=document.getElementById('pf_account_rows');
  if(body) body.insertAdjacentHTML('beforeend', _knownAccountRowHtml(row));
  toggleProfileSections();
}
function _collectProfileRows(bodyId, mapper){
  const body=document.getElementById(bodyId);
  if(!body) return [];
  return Array.from(body.querySelectorAll('tr')).map(mapper).filter(Boolean);
}
function _resetProfileQuestionnaire(){
  ['pf_entity_type','pf_legal_name','pf_huf_name','pf_dob','pf_employer_name'].forEach(id=>{
    const el=document.getElementById(id); if(el) el.value='';
  });
  ['pf_family_toggle','pf_salaried_toggle','pf_consultancy_toggle','pf_trading_toggle','pf_multi_bank_toggle','pf_credit_cards','pf_rental_income'].forEach(id=>{
    const el=document.getElementById(id); if(el) el.checked=false;
  });
  ['pf_family_rows','pf_consultancy_rows','pf_broker_rows','pf_account_rows'].forEach(id=>{
    const el=document.getElementById(id); if(el) el.innerHTML='';
  });
  _profileMsg('');
  toggleProfileSections();
}
function toggleProfileSections(){
  const entity=(document.getElementById('pf_entity_type')||{value:''}).value;
  const hufName=(document.getElementById('pf_huf_name')||{value:''}).value.trim();
  const familyOn=(document.getElementById('pf_family_toggle')||{checked:false}).checked || !!document.querySelector('#pf_family_rows tr');
  const salariedOn=(document.getElementById('pf_salaried_toggle')||{checked:false}).checked;
  const consultancyOn=(document.getElementById('pf_consultancy_toggle')||{checked:false}).checked || !!document.querySelector('#pf_consultancy_rows tr');
  const tradingOn=(document.getElementById('pf_trading_toggle')||{checked:false}).checked || !!document.querySelector('#pf_broker_rows tr');
  const multiBankOn=(document.getElementById('pf_multi_bank_toggle')||{checked:false}).checked || !!document.querySelector('#pf_account_rows tr');
  const hufWrap=document.getElementById('pf_huf_name_wrap');
  const employerWrap=document.getElementById('pf_employer_wrap');
  const familySec=document.getElementById('pf_family_section');
  const consultancySec=document.getElementById('pf_consultancy_section');
  const brokerSec=document.getElementById('pf_broker_section');
  const accountsSec=document.getElementById('pf_accounts_section');
  if(hufWrap) hufWrap.style.display=(entity==='HUF' || !!hufName)?'block':'none';
  if(employerWrap) employerWrap.style.display=salariedOn?'block':'none';
  if(familySec) familySec.style.display=familyOn?'block':'none';
  if(consultancySec) consultancySec.style.display=consultancyOn?'block':'none';
  if(brokerSec) brokerSec.style.display=tradingOn?'block':'none';
  if(accountsSec) accountsSec.style.display=multiBankOn?'block':'none';
  if(!_activeUserId || _profileGateState.profile_required){
    [familySec, consultancySec, brokerSec, accountsSec].forEach(el=>{ if(el) el.style.display='none'; });
  }
}

function _validateProfilePayload(payload){
  const entityType=String(payload.entity_type||'').trim().toUpperCase();
  if(!entityType) return 'Entity type is required';
  if(!['INDIVIDUAL','NRI','HUF'].includes(entityType)) return 'Entity type must be one of INDIVIDUAL, NRI, HUF';
  if(!String(payload.legal_name||'').trim()) return 'Legal name is required';
  if(entityType==='HUF' && !String(payload.huf_name||'').trim()) return 'HUF name is required when Entity Type is HUF';
  if(payload.is_salaried && !String(payload.employer_name||'').trim()) return 'Employer name is required when Salaried is selected';
  return '';
}

async function loadProfileQuestionnaire(){
  if(!_activeUserId){
    _resetProfileQuestionnaire();
    _setProfileInputsEnabled(false);
    _setProfileBanner('Create account first to save your classification profile', 'info');
    return;
  }
  _profileMsg('Loading profile…');
  try{
    const [rp, rc, ra] = await Promise.all([
      fetch('/profile'),
      fetch('/known-counterparties?all=1'),
      fetch('/known-accounts?all=1')
    ]);
    const profile = rp.ok ? await rp.json() : {};
    const cpData = rc.ok ? await rc.json() : {rows:[]};
    const acctData = ra.ok ? await ra.json() : {rows:[]};

    _resetProfileQuestionnaire();

    const entityType = String(profile.entity_type||'').toUpperCase();
    const entityEl = document.getElementById('pf_entity_type');
    if(entityEl) entityEl.value = entityType || (Number(profile.is_nri||0) ? 'NRI' : '');
    const legalNameEl = document.getElementById('pf_legal_name'); if(legalNameEl) legalNameEl.value = profile.legal_name || '';
    const hufNameEl = document.getElementById('pf_huf_name'); if(hufNameEl) hufNameEl.value = profile.huf_name || '';
    const dobEl = document.getElementById('pf_dob'); if(dobEl) dobEl.value = profile.dob || '';
    const employerEl = document.getElementById('pf_employer_name'); if(employerEl) employerEl.value = profile.employer_name || '';

    ['pf_family_toggle','pf_salaried_toggle','pf_consultancy_toggle','pf_trading_toggle','pf_multi_bank_toggle','pf_credit_cards','pf_rental_income'].forEach((id)=>{
      const map = {
        pf_family_toggle: 'has_family_transactions',
        pf_salaried_toggle: 'is_salaried',
        pf_consultancy_toggle: 'has_consultancy',
        pf_trading_toggle: 'has_trading',
        pf_multi_bank_toggle: 'has_multiple_bank_accounts',
        pf_credit_cards: 'has_credit_cards',
        pf_rental_income: 'has_rental_income'
      };
      const el=document.getElementById(id);
      if(el) el.checked = Number(profile[map[id]] || 0) === 1;
    });

    const rows = Array.isArray(cpData.rows) ? cpData.rows : [];
    rows.filter(r => String(r.party_type||'').toLowerCase()==='family').forEach(r => addProfileCounterpartyRow('family', r));
    rows.filter(r => String(r.party_type||'').toLowerCase()==='consultancy').forEach(r => addProfileCounterpartyRow('consultancy', r));
    rows.filter(r => String(r.party_type||'').toLowerCase()==='broker').forEach(r => addProfileCounterpartyRow('broker', r));
    (Array.isArray(acctData.rows) ? acctData.rows : []).forEach(r => addKnownAccountRow(r));

    _profileMsg('Profile loaded.', '#1b5e20');
    await refreshProfileGate({preserveBanner:true});
    toggleProfileSections();
  }catch(e){
    _profileMsg('Profile load failed: ' + (e.message || e), '#b71c1c');
  }
}
async function saveProfileQuestionnaire(){
  if(!_activeUserId){
    _profileMsg('Create account first to save your classification profile', '#b71c1c');
    return;
  }
  const entityType=(document.getElementById('pf_entity_type')||{value:''}).value.trim();
  const payload = {
    entity_type: entityType,
    legal_name: (document.getElementById('pf_legal_name')||{value:''}).value.trim(),
    huf_name: (document.getElementById('pf_huf_name')||{value:''}).value.trim(),
    dob: (document.getElementById('pf_dob')||{value:''}).value,
    is_nri: entityType === 'NRI',
    has_family_transactions: !!(document.getElementById('pf_family_toggle')||{checked:false}).checked,
    is_salaried: !!(document.getElementById('pf_salaried_toggle')||{checked:false}).checked,
    employer_name: (document.getElementById('pf_employer_name')||{value:''}).value.trim(),
    has_consultancy: !!(document.getElementById('pf_consultancy_toggle')||{checked:false}).checked,
    has_trading: !!(document.getElementById('pf_trading_toggle')||{checked:false}).checked,
    has_multiple_bank_accounts: !!(document.getElementById('pf_multi_bank_toggle')||{checked:false}).checked,
    has_credit_cards: !!(document.getElementById('pf_credit_cards')||{checked:false}).checked,
    has_rental_income: !!(document.getElementById('pf_rental_income')||{checked:false}).checked,
  };
  const validationError=_validateProfilePayload(payload);
  if(validationError){
    _profileMsg(validationError, '#b71c1c');
    return;
  }

  const familyRows = _collectProfileRows('pf_family_rows', tr => {
    const get = f => (tr.querySelector(`[data-f="${f}"]`)||{value:''}).value.trim();
    const display_name = get('display_name');
    if(!display_name) return null;
    return {
      display_name,
      relationship: get('relationship'),
      party_type: 'family',
      default_ledger_key: get('default_ledger_key'),
      txn_direction_hint: get('txn_direction_hint'),
      notes: _profileJoinNotes(get('default_treatment'), get('notes')),
      is_active: 1
    };
  });
  const consultancyRows = _collectProfileRows('pf_consultancy_rows', tr => {
    const get = f => (tr.querySelector(`[data-f="${f}"]`)||{value:''}).value.trim();
    const display_name = get('display_name');
    if(!display_name) return null;
    return {
      display_name,
      relationship: get('relationship'),
      party_type: 'consultancy',
      default_ledger_key: get('default_ledger_key'),
      txn_direction_hint: get('txn_direction_hint'),
      notes: _profileJoinNotes(get('default_treatment'), get('notes')),
      is_active: 1
    };
  });
  const brokerRows = _collectProfileRows('pf_broker_rows', tr => {
    const get = f => (tr.querySelector(`[data-f="${f}"]`)||{value:''}).value.trim();
    const display_name = get('display_name');
    if(!display_name) return null;
    return {
      display_name,
      relationship: get('relationship') || 'broker',
      party_type: 'broker',
      default_ledger_key: get('default_ledger_key'),
      txn_direction_hint: get('txn_direction_hint'),
      notes: _profileJoinNotes(get('default_treatment'), get('notes')),
      is_active: 1
    };
  });
  const accountRows = _collectProfileRows('pf_account_rows', tr => {
    const get = f => (tr.querySelector(`[data-f="${f}"]`)||{value:''}).value.trim();
    const row = {
      account_label: get('account_label'),
      institution_name: get('institution_name'),
      account_mask: get('account_mask'),
      account_type: get('account_type'),
      ownership_type: get('ownership_type'),
      is_active: 1
    };
    return (row.account_label || row.institution_name || row.account_mask) ? row : null;
  });

  _profileMsg('Saving profile…');
  try{
    const [rp, rc, ra] = await Promise.all([
      fetch('/save-profile', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify(payload)}),
      fetch('/save-known-counterparties', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({rows:[...familyRows,...consultancyRows,...brokerRows]})}),
      fetch('/save-known-accounts', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({rows:accountRows})}),
    ]);
    const profileRes = await rp.json();
    const cpRes = await rc.json();
    const acctRes = await ra.json();
    if(!rp.ok || profileRes.error || !rc.ok || cpRes.error || !ra.ok || acctRes.error){
      throw new Error(profileRes.error || cpRes.error || acctRes.error || 'Profile save failed');
    }
    let applyMsg='';
    try{
      const ar=await fetch('/apply-profile-to-pending', {method:'POST', headers:{'Content-Type':'application/json'}, body:'{}'});
      const ad=ar.ok ? await ar.json() : {};
      if(ad && Number(ad.updated||0)>0) applyMsg=` Applied profile to ${ad.updated} pending transaction${Number(ad.updated)===1?'':'s'}.`;
    }catch(_e){}
    _profileMsg('Profile saved.' + applyMsg, '#1b5e20');
    await refreshProfileGate();
    toggleProfileSections();
    if(typeof loadReview === 'function') loadReview();
  }catch(e){
    _profileMsg('Profile save failed: ' + (e.message || e), '#b71c1c');
  }
}
async function applyProfileToApprovedSafe(){
  if(!_activeUserId){
    _profileMsg('Create account first to save your classification profile', '#b71c1c');
    return;
  }
  if(!confirm('Apply profile hints to already approved transactions where it is safe to do so?')) return;
  _profileMsg('Applying profile to approved transactions…');
  try{
    const r=await fetch('/apply-profile-to-approved-safe', {method:'POST', headers:{'Content-Type':'application/json'}, body:'{}'});
    const d=await r.json();
    if(!r.ok || d.error) throw new Error(d.error || 'Apply failed');
    _profileMsg(d.msg || `Reviewed ${d.reviewed||0}; updated ${d.updated||0}.`, '#1b5e20');
    _lastReports={};
  }catch(e){
    _profileMsg('Apply failed: ' + (e.message || e), '#b71c1c');
  }
}
_resetProfileQuestionnaire();
_applySetupInteractionState();

// ── Auth ──────────────────────────────────────────────────────────────────────
function authTab(t){
  const loginPanel = document.getElementById('panel-login');
  const createPanel = document.getElementById('panel-create');

  if(loginPanel){
    loginPanel.style.display = t === 'login' ? 'block' : 'none';
  }

  if(createPanel){
    createPanel.style.display = t === 'create' ? 'block' : 'none';
    createPanel.hidden = false;
  }

  ['login','create'].forEach(x=>{
    const tab = document.getElementById('tab-' + x);
    if(tab){
      tab.style.color = t === x ? 'var(--blue)' : '#aaa';
      tab.style.borderBottom = t === x ? '2px solid var(--blue)' : 'none';
    }
  });
}
function _setActiveUser(d, opts={}){
  _activeUserId=d.id || '';
  const banner = document.getElementById('active-user-banner');
  const info   = document.getElementById('active-user-info');
  const hdr    = document.getElementById('hdr-sub');

  if(banner) banner.style.display='flex';

  const ut = d.user_type ? ` · ${d.user_type}` : '';
  if(info){
    info.textContent = `${d.name}${d.phone ? ' · ' + d.phone : ''}${d.email ? ' · ' + d.email : ''}${ut} · ${d.txn_count || 0} transactions`;
  }

  // Keep setup visible after login
  const loginPanel  = document.getElementById('panel-login');
  const createPanel = document.getElementById('panel-create');
  const tabWrap     = document.getElementById('tab-login')?.parentElement;

  if(loginPanel)  loginPanel.style.display = 'none';
  if(createPanel) createPanel.style.display = 'none';
  if(tabWrap)     tabWrap.style.display = 'none';

  const authCard = document.getElementById('user-auth-card');
  if(authCard) authCard.style.display = 'block';

  if(hdr){
    hdr.textContent = 'Active: ' + d.name + (d.user_type ? ' (' + d.user_type + ')' : '');
  }

  setStatus(`✅ ${d.msg}`,'ok');
  _applySetupInteractionState();
  refreshUploadedFiles();
  if(!opts.skipProfileLoad){
    loadProfileQuestionnaire()
      .then(()=>refreshProfileGate({focusProfile: !!d.profile_required, message: d.msg, preserveBanner: false}))
      .catch(()=>refreshProfileGate({focusProfile: !!d.profile_required, message: d.msg}));
  }
}
function _validEmail(v){
  return !v || /^[^\s@]+@[^\s@]+\.[^\s@]+$/.test(v);
}
function _cleanPhone(v){
  return (v || '').replace(/\D+/g,'');
}
function _validPhone(v){
  return !v || (/^\d{8,15}$/.test(v));
}

async function loginUser(){
  const name=(document.getElementById('l_name')?.value || '').trim();
  const phone=_cleanPhone((document.getElementById('l_phone')?.value || '').trim());
  const email=((document.getElementById('l_email')?.value || '').trim().toLowerCase());
  const password=(document.getElementById('l_password')?.value || '');

  const msgEl = document.getElementById('login-msg');
  if(msgEl) msgEl.textContent='';

  if(!name && !phone && !email){
    if(msgEl) msgEl.innerHTML='<span style="color:var(--red)">Enter at least one field</span>';
    return;
  }
  if(phone && !_validPhone(phone)){
    if(msgEl) msgEl.innerHTML='<span style="color:var(--red)">Phone must be 8 to 15 digits</span>';
    return;
  }
  if(email && !_validEmail(email)){
    if(msgEl) msgEl.innerHTML='<span style="color:var(--red)">Enter a valid email</span>';
    return;
  }
  if(!password){
    if(msgEl) msgEl.innerHTML='<span style="color:var(--red)">Enter password</span>';
    return;
  }

  if(msgEl) msgEl.textContent='Signing in…';

  try{
    const r = await fetch('/login',{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({name,phone,email,password})
    });

    let d = {};
    try{ d = await r.json(); } catch(_e){ throw new Error(`Server returned ${r.status}`); }

    if(!r.ok || d.error){
      if(msgEl) msgEl.innerHTML=`<span style="color:var(--red)">${d.error || 'Login failed'}</span>`;
      return;
    }

    _setActiveUser(d);
    if(msgEl) msgEl.innerHTML='<span style="color:#1b5e20">Logged in successfully</span>';
    try{ await loadAccountDropdowns(); }catch(_e){}
  }catch(e){
    if(msgEl) msgEl.innerHTML=`<span style="color:var(--red)">Login failed: ${e.message || e}</span>`;
  }
}

windows.forgotPassword = function(){
  const msgEl = document.getElementById('login-msg');

  if(msgEl){
    msgEl.innerHTML =
      '<span style="color:#b26a00">Password reset feature is not available yet. Please contact admin/support.</span>';
  }else{
    alert('Password reset feature is not available yet. Please contact admin/support.');
  }
}

function toggleCreatePassword() {
  const p = document.getElementById("u_password");
  p.type = p.type === "password" ? "text" : "password";
}

function toggleConfirmPassword() {
  const p = document.getElementById("u_password_confirm");
  p.type = p.type === "password" ? "text" : "password";
}



async function createUser(){
  const name=document.getElementById('u_name').value.trim();
  const email=document.getElementById('u_email').value.trim().toLowerCase();
  const phone=_cleanPhone(document.getElementById('u_phone').value.trim());
  const password=(document.getElementById('u_password').value || '');
  const confirmPassword=(document.getElementById('u_password_confirm').value || '');
  const user_type=document.getElementById('u_user_type').value||'INDIVIDUAL';

  if(name.length < 2){
    document.getElementById('create-msg').innerHTML='<span style="color:var(--red)">Name must be at least 2 characters</span>';
    return;
  }
  if(phone && !_validPhone(phone)){
    document.getElementById('create-msg').innerHTML='<span style="color:var(--red)">Phone must be 8 to 15 digits</span>';
    return;
  }
  if(email && !_validEmail(email)){
    document.getElementById('create-msg').innerHTML='<span style="color:var(--red)">Enter a valid email</span>';
    return;
  }
  if(!phone && !email){
    document.getElementById('create-msg').innerHTML='<span style="color:var(--red)">Provide at least phone or email</span>';
    return;
  }
  if(password.length < 8){
    document.getElementById('create-msg').innerHTML='<span style="color:var(--red)">Password must be at least 8 characters</span>';
    return;
  }
  if(!/[!@#$%^&*()_\-+=\[{\]};:'",.<>/?\\|`~]/.test(password)){
    document.getElementById('create-msg').innerHTML='<span style="color:var(--red)">Password must include one special character</span>';
    return;
  }
  if(password !== confirmPassword){
    document.getElementById('create-msg').innerHTML='<span style="color:var(--red)">Passwords do not match</span>';
    return;
  }

  document.getElementById('create-msg').textContent='Creating…';

  try{
    const r=await fetch('/create-user',{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({name,email,phone,password,user_type})
    });

    let d={};
    try{ d=await r.json(); }catch(_e){ throw new Error(`Server returned ${r.status}`); }

    if(!r.ok || d.error){
      document.getElementById('create-msg').innerHTML=`<span style="color:var(--red)">${d.error||'Create failed'}</span>`;
      return;
    }

    _setActiveUser(d, {skipProfileLoad:true});
    document.getElementById('create-msg').innerHTML='<span style="color:#1b5e20">Account created. Complete your basic classification profile to continue.</span>';
    try{ await loadProfileQuestionnaire(); }catch(_e){}
    const legalNameEl=document.getElementById('pf_legal_name');
    if(legalNameEl && !legalNameEl.value.trim()) legalNameEl.value=name;
    _applySetupInteractionState();
    _focusProfileSetup('Account created. Complete your basic classification profile to continue.');
    try{ await refreshProfileGate({focusProfile:true, message:'Account created. Complete your basic classification profile to continue.'}); }catch(_e){}
    try{ await loadAccountDropdowns(); }catch(_e){}
  }catch(e){
    document.getElementById('create-msg').innerHTML=`<span style="color:var(--red)">Create failed: ${e.message || e}</span>`;
  }
}

async function forgotPassword(){
  const email=((document.getElementById('fp_email')?.value || '').trim().toLowerCase());
  const phone=_cleanPhone((document.getElementById('fp_phone')?.value || '').trim());
  const msgEl=document.getElementById('forgot-msg');

  if(!email && !phone){
    if(msgEl) msgEl.innerHTML='<span style="color:var(--red)">Enter email or phone</span>';
    return;
  }

  try{
    const r=await fetch('/forgot-password',{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({email,phone})
    });
    const d=await r.json();
    if(!r.ok || d.error){
      if(msgEl) msgEl.innerHTML=`<span style="color:var(--red)">${d.error || 'Request failed'}</span>`;
      return;
    }
    if(msgEl) msgEl.innerHTML=`<span style="color:#1b5e20">Reset token: ${d.reset_token}</span>`;
  }catch(e){
    if(msgEl) msgEl.innerHTML=`<span style="color:var(--red)">Request failed: ${e.message || e}</span>`;
  }
}

async function resetPassword(){
  const token=(document.getElementById('rp_token')?.value || '').trim();
  const new_password=(document.getElementById('rp_password')?.value || '');
  const msgEl=document.getElementById('reset-msg');

  if(!token){
    if(msgEl) msgEl.innerHTML='<span style="color:var(--red)">Enter reset token</span>';
    return;
  }
  if(new_password.length < 8){
    if(msgEl) msgEl.innerHTML='<span style="color:var(--red)">Password must be at least 8 characters</span>';
    return;
  }
  if(!/[!@#$%^&*()_\-+=\[{\]};:'",.<>/?\\|`~]/.test(new_password)){
    if(msgEl) msgEl.innerHTML='<span style="color:var(--red)">Password must include one special character</span>';
    return;
  }

  try{
    const r=await fetch('/reset-password',{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({token,new_password})
    });
    const d=await r.json();
    if(!r.ok || d.error){
      if(msgEl) msgEl.innerHTML=`<span style="color:var(--red)">${d.error || 'Reset failed'}</span>`;
      return;
    }
    if(msgEl) msgEl.innerHTML='<span style="color:#1b5e20">Password reset successfully</span>';
  }catch(e){
    if(msgEl) msgEl.innerHTML=`<span style="color:var(--red)">Reset failed: ${e.message || e}</span>`;
  }
}

function logoutUser(){
  _activeUserId='';
  _profileGateState={ok:false,profile_required:false,is_complete:false,missing_reason:''};
  const banner = document.getElementById('active-user-banner');
  const authCard = document.getElementById('user-auth-card');
  const hdr = document.getElementById('hdr-sub');

  if(banner) banner.style.display='none';
  if(authCard) authCard.style.display='block';

  const loginPanel  = document.getElementById('panel-login');
  const createPanel = document.getElementById('panel-create');
  const tabWrap     = document.getElementById('tab-login')?.parentElement;

  if(tabWrap) tabWrap.style.display = 'flex';
  if(loginPanel) loginPanel.style.display = 'block';
  if(createPanel) createPanel.style.display = 'none';

  if(hdr) hdr.textContent='3-Book Accounting · Balance Sheet · Income & Expenditure · Capital';

  if(document.getElementById('l_name')) document.getElementById('l_name').value='';
  if(document.getElementById('l_phone')) document.getElementById('l_phone').value='';
  if(document.getElementById('l_email')) document.getElementById('l_email').value='';
  if(document.getElementById('l_password')) document.getElementById('l_password').value='';
  if(document.getElementById('u_name')) document.getElementById('u_name').value='';
  if(document.getElementById('u_email')) document.getElementById('u_email').value='';
  if(document.getElementById('u_phone')) document.getElementById('u_phone').value='';

  if(document.getElementById('login-msg')) document.getElementById('login-msg').textContent='';
  if(document.getElementById('create-msg')) document.getElementById('create-msg').textContent='';
  if(document.getElementById('manage-msg')) document.getElementById('manage-msg').textContent='';
  _resetProfileQuestionnaire();
  _setProfileBanner('Create account first to save your classification profile', 'info');
  _applySetupInteractionState();
  _applyProfileGate();
  refreshUploadedFiles();

  fetch('/logout',{method:'POST'}).catch(()=>{});
}

// ── Trace Modal ───────────────────────────────────────────────────────────────

async function openTrace(title, params) {
  window._lastTraceParams = params || {};
  await _ensureLedgerKeys();
  document.getElementById('trace-modal-title').textContent = title;
  document.getElementById('trace-modal-loading').style.display = 'block';
  document.getElementById('trace-modal-content').style.display = 'none';
  document.getElementById('trace-modal-empty').style.display = 'none';
  document.getElementById('trace-modal-summary').innerHTML = '';
  document.getElementById('trace-modal-overlay').classList.add('open');

  try {
    const merged = Object.assign({}, params || {});
    const df = (document.getElementById('stmt_date_from')||{}).value || '';
    const dt = (document.getElementById('stmt_date_to')||{}).value || '';
    const acct = (document.getElementById('stmt_account')||{}).value || '';
    if(df && !merged.date_from) merged.date_from = df;
    if(dt && !merged.date_to) merged.date_to = dt;
    if(acct && !merged.account_id) merged.account_id = acct;
    window._lastTraceParams = merged;
    document.getElementById('trace-modal-title').textContent = merged.account || title || 'Transaction Trace';
    const qs = new URLSearchParams(merged).toString();
    const r = await fetch('/trace?' + qs);
    const d = await r.json();
    if (d.error) { alert(d.error); closeTrace(); return; }

    const rows = d.transactions || [];
    window._lastTraceRows = Array.isArray(rows) ? rows : [];
    document.getElementById('trace-modal-loading').style.display = 'none';

    // Summary chips
    const isComputedTrace = !!(
      d.synthetic === true ||
      d.editable === false ||
      (rows || []).some(r => r && (r.synthetic === true || r.editable === false))
    );
    const totalCredit = rows.filter(r=>r.txn_type==='credit').reduce((s,r)=>s+r.amount,0);
    const totalDebit  = rows.filter(r=>r.txn_type==='debit').reduce((s,r)=>s+r.amount,0);
    const net = d.synthetic ? rows.reduce((s,r)=>s+(r.signed||0),0) : (totalCredit - totalDebit);
    const fmt = v => '₹'+Math.abs(v).toLocaleString('en-IN',{minimumFractionDigits:2});
    let chips = `<div class="trace-chip">${rows.length} ${isComputedTrace ? 'computed rows' : 'transactions'}</div>`;
    if (!isComputedTrace) {
      chips += `<div class="trace-chip" style="background:#e8f5e9;border-color:#a5d6a7;color:#1b5e20">Credits: ${fmt(totalCredit)}</div>`;
      chips += `<div class="trace-chip" style="background:#ffebee;border-color:#ef9a9a;color:#b71c1c">Debits: ${fmt(totalDebit)}</div>`;
    }
    chips += `<div class="trace-chip" style="background:#f3e5f5;border-color:#ce93d8;color:#6a1b9a">Net: ${fmt(net)} ${net>=0?'(+)':'(-)'}</div>`;
    if (isComputedTrace) chips += '<div class="trace-chip" style="background:#fff3e0;border-color:#ffb74d;color:#e65100">Computed statement summary — not a transaction list</div>';
    document.getElementById('trace-modal-summary').innerHTML = chips;

    if (!rows.length) {
      document.getElementById('trace-modal-empty').style.display = 'block';
      return;
    }

    const showSigned = d.synthetic;
    const headerRow = document.querySelector('#trace-modal-table thead tr');
    if (headerRow) {
      headerRow.innerHTML = isComputedTrace
        ? '<th>Line</th><th>Computed Summary</th><th style="text-align:right">Amount</th>'
        : '<th>Date</th><th>Counterparty</th><th style="text-align:right">Amount</th><th>Actions</th>';
    }
    document.getElementById('trace-modal-rows').innerHTML = rows.map(r => {
      const amt = parseFloat(r.amount||0);
      const amtColor = r.txn_type==='credit' ? '#1b5e20' : '#b71c1c';
      const narrDisplay = (r.narration || r.description || '—').trim() || '—';
      const rowComputed = isComputedTrace || r.synthetic === true || r.editable === false;
      if (rowComputed) {
        const signedVal = parseFloat(r.signed || r.amount || 0);
        const signedColor = signedVal >= 0 ? '#1b5e20' : '#b71c1c';
        return `<tr>
          <td style=\"white-space:nowrap;font-family:monospace;font-size:11px\">${r.id||'computed'}</td>
          <td style=\"min-width:320px;max-width:520px;word-break:break-word;font-size:11px;color:#555\">
            <div style="font-weight:700;color:#333;margin-bottom:4px">${narrDisplay}</div>
            <div style="font-size:11px;color:#777">${r.note || 'Computed statement summary — not a transaction list'}</div>
          </td>
          <td style=\"text-align:right;font-weight:700;color:${signedColor};white-space:nowrap\">${fmt(signedVal)}</td>
        </tr>`;
      }
      return `<tr>
        <td style=\"white-space:nowrap;font-family:monospace;font-size:11px\">${r.txn_date||'—'}</td>
        <td style=\"min-width:320px;max-width:420px;word-break:break-word;font-size:11px;color:#555\">
          <div style="margin-bottom:8px">
            <div style="font-size:12px;color:#555;display:flex;align-items:flex-start;gap:6px;flex-wrap:nowrap">
              <span style="font-size:10px;background:#eef3ff;border-radius:4px;padding:1px 7px;color:#4a6fa5;font-weight:600;text-transform:uppercase;letter-spacing:.3px;flex:0 0 auto">Narration</span>
              <span style="font-weight:500;color:#333;word-break:break-word;line-height:1.45">${narrDisplay}</span>
            </div>
          </div>
          <div style="font-size:12px;color:#555;display:flex;align-items:flex-start;gap:6px;flex-wrap:nowrap">
            <span style="font-size:10px;background:#f0f0f0;border-radius:4px;padding:1px 7px;color:#777;font-weight:600;text-transform:uppercase;letter-spacing:.3px;flex:0 0 auto">Counterparty</span>
            <span style="font-weight:500;color:#333;word-break:break-word;line-height:1.45">${(r.counterparty||'—')}</span>
          </div>
        </td>
        <td style=\"text-align:right;font-weight:700;color:${amtColor};white-space:nowrap\">${fmt(amt)}</td>
        <td style=\"min-width:340px;white-space:nowrap\">
          <div style=\"display:flex;align-items:center;gap:8px;flex-wrap:wrap\">
            <select id=\"trace-rk-${r.id||''}\" onchange=\"reclassifyApprovedFromTrace('${r.id||''}', this.value)\"
              style=\"width:210px;padding:5px 8px;border:1px solid #ddd;border-radius:6px;font-size:11px\">
              ${_buildKeyOptions(r.ledger_key || '')}
            </select>
            <button type=\"button\"
              onclick=\"saveCustomRuleFromTraceRow('${r.id||''}')\"
              style=\"padding:5px 9px;border:1px solid #d6d6d6;border-radius:6px;background:#fff;cursor:pointer;font-size:11px\">
              💾 Save as rule
            </button>
          </div>
        </td>
      </tr>`;
    }).join('');

    document.getElementById('trace-modal-content').style.display = 'block';
  } catch(e) {
    alert('Trace failed: ' + e);
    closeTrace();
  }
}

async function saveCustomRuleFromTraceRow(id){
  const rows = Array.isArray(window._lastTraceRows) ? window._lastTraceRows : [];
  const row = rows.find(x => String(x.id||'') === String(id||''));
  if(!row){ alert('Row not found'); return; }

  const sel = document.getElementById('trace-rk-' + id);
  const activeKey =
    (sel && sel.value ? sel.value : '') ||
    row.reclassified_key ||
    row.ledger_key ||
    row.predicted_ledger_key ||
    '';

  if(!activeKey){ alert('Pick a category first'); return; }

  const suggested = (row.counterparty || row.narration || '').trim();
  const pattern = prompt('Enter rule pattern:', suggested);
  if(pattern === null) return;

  try{
    const r = await fetch('/add-custom-rule', {
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body: JSON.stringify({
        pattern,
        ledger_key: activeKey,
        txn_type: row.txn_type || '',
        match_mode: 'contains',
        priority: 100
      })
    });
    const d = await r.json();
    if(d.error){ alert(d.error); return; }
    alert('Rule saved');
  }catch(e){
    alert('Failed: ' + e);
  }
}

function _needsStockName(key){
  const k = String(key || '').toLowerCase().trim();
  return k.includes('asset_investment_equity')
      || k.includes('assets_investment_equity')
      || k.includes('investment_equity');
}

function _sanitizeStockName(v){
  return String(v || '')
    .trim()
    .toUpperCase()
    .replace(/[^A-Z0-9.\-& ]/g, '')
    .replace(/\s+/g, ' ');
}

async function reclassifyApprovedFromTrace(txnId, newKey) {
  if (!txnId || !newKey) return;

  if (_needsStockName(newKey)) {
    // Remove any stale modal
    const stale = document.getElementById('_stock-modal-overlay');
    if (stale) stale.remove();

    const overlay = document.createElement('div');
    overlay.id = '_stock-modal-overlay';
    overlay.style.cssText = 'position:fixed;inset:0;background:rgba(0,0,0,.52);z-index:99999;display:flex;align-items:center;justify-content:center';
    overlay.innerHTML = `
      <div style="background:#fff;border-radius:14px;padding:28px 32px;width:440px;max-width:96vw;box-shadow:0 8px 40px rgba(0,0,0,.3);font-family:inherit">
        <h4 style="margin:0 0 18px;font-size:15px;font-weight:800;color:#0d47a1">📊 Equity Share Details</h4>
        <label style="font-size:11.5px;font-weight:700;color:#555;text-transform:uppercase;letter-spacing:.4px;display:block;margin-bottom:4px">Share Symbol <span style="color:#b71c1c">*</span></label>
        <input id="_sm-stock" placeholder="e.g. INFY, TCS, RELIANCE" autocomplete="off"
          style="width:100%;padding:9px 12px;border:1.5px solid #bbb;border-radius:7px;font-size:13px;margin-bottom:14px;box-sizing:border-box">
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:14px">
          <div>
            <label style="font-size:11.5px;font-weight:700;color:#555;text-transform:uppercase;letter-spacing:.4px;display:block;margin-bottom:4px">Transaction Type</label>
            <select id="_sm-type" style="width:100%;padding:9px 10px;border:1.5px solid #bbb;border-radius:7px;font-size:13px">
              <option value="BUY">BUY</option>
              <option value="SELL">SELL</option>
            </select>
          </div>
          <div>
            <label style="font-size:11.5px;font-weight:700;color:#555;text-transform:uppercase;letter-spacing:.4px;display:block;margin-bottom:4px">Price per Share (₹)</label>
            <input id="_sm-price" placeholder="e.g. 1500.00" type="number" step="0.01" min="0"
              style="width:100%;padding:9px 12px;border:1.5px solid #bbb;border-radius:7px;font-size:13px;box-sizing:border-box">
          </div>
        </div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:22px">
          <div>
            <label style="font-size:11.5px;font-weight:700;color:#555;text-transform:uppercase;letter-spacing:.4px;display:block;margin-bottom:4px">Quantity</label>
            <input id="_sm-qty" placeholder="e.g. 10" type="number" step="1" min="0"
              style="width:100%;padding:9px 12px;border:1.5px solid #bbb;border-radius:7px;font-size:13px;box-sizing:border-box">
          </div>
          <div>
            <label style="font-size:11.5px;font-weight:700;color:#555;text-transform:uppercase;letter-spacing:.4px;display:block;margin-bottom:4px">TDS (₹, or 0)</label>
            <input id="_sm-tds" placeholder="e.g. 25.00" type="number" step="0.01" min="0"
              style="width:100%;padding:9px 12px;border:1.5px solid #bbb;border-radius:7px;font-size:13px;box-sizing:border-box">
          </div>
        </div>
        <div id="_sm-err" style="font-size:12px;color:#b71c1c;min-height:18px;margin-bottom:10px"></div>
        <div style="display:flex;gap:10px;justify-content:flex-end">
          <button onclick="document.getElementById('_stock-modal-overlay').remove()"
            style="padding:9px 20px;border:1.5px solid #ccc;border-radius:7px;font-size:13px;cursor:pointer;background:#fff;font-weight:600">Cancel</button>
          <button id="_sm-save-btn"
            onclick="_saveStockReclassify('${txnId}','${newKey}')"
            style="padding:9px 20px;background:#1565c0;color:#fff;border:none;border-radius:7px;font-size:13px;font-weight:700;cursor:pointer">💾 Save</button>
        </div>
      </div>`;
    document.body.appendChild(overlay);
    // Focus share symbol input
    setTimeout(() => { const el = document.getElementById('_sm-stock'); if(el) el.focus(); }, 60);
    return;
  }

  // Non-equity key: post directly with no stock fields
  await _postReclassifyApproved(txnId, newKey, '', '', '', '', '');
}

async function _saveStockReclassify(txnId, newKey) {
  const stockName  = _sanitizeStockName((document.getElementById('_sm-stock') || {}).value || '');
  const tradeType  = ((document.getElementById('_sm-type')  || {}).value || 'BUY').trim().toUpperCase();
  const tradePrice = ((document.getElementById('_sm-price') || {}).value || '').trim();
  const tradeQty   = ((document.getElementById('_sm-qty')   || {}).value || '').trim();
  const tradeTds   = ((document.getElementById('_sm-tds')   || {}).value || '').trim();
  const errEl      = document.getElementById('_sm-err');

  if (!stockName) {
    if (errEl) errEl.textContent = 'Share symbol is required (e.g. INFY).';
    return;
  }
  if (errEl) errEl.textContent = '';

  const btn = document.getElementById('_sm-save-btn');
  if (btn) { btn.disabled = true; btn.textContent = '⏳ Saving…'; }

  await _postReclassifyApproved(txnId, newKey, stockName, tradeType, tradePrice, tradeQty, tradeTds);

  const ov = document.getElementById('_stock-modal-overlay');
  if (ov) ov.remove();
}

async function _postReclassifyApproved(txnId, newKey, stockName, tradeType, tradePrice, tradeQty, tradeTds) {
  try {
    const r = await fetch('/reclassify-approved', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({
        txn_id:      txnId,
        new_key:     newKey,
        stock_name:  stockName  || '',
        trade_type:  tradeType  || '',
        trade_price: tradePrice || '',
        trade_qty:   tradeQty   || '',
        trade_tds:   tradeTds   || '',
      })
    });
    const d = await r.json();
    if (d.error) { alert('❌ ' + d.error); return; }

    // Refresh statements silently then re-open trace with same params
    try { await refreshStmts(); } catch(_) {}
    const title = document.getElementById('trace-modal-title').textContent;
    const currentParams = window._lastTraceParams || {};
    openTrace(title, currentParams);
  } catch (e) {
    alert('Reclassify failed: ' + e);
  }
}

function closeTrace() {
  document.getElementById('trace-modal-overlay').classList.remove('open');
}

// Parse TRACE markers embedded in statement text lines
// Format: [TRACE:book:section:grp:acct]  or  [TRACE:bank_balance]
function _parseTraceMeta(line) {
  const m = line.match(/\[TRACE:([^\]]+)\]/);
  if (!m) return null;
  const parts = m[1].split(':');
  if (parts[0] === 'bank_balance') return { type: 'bank_balance' };
  if (parts.length >= 4) return { type: 'bucket', book: parts[0], section: parts[1], grp: parts[2], account: parts.slice(3).join(':') };
  if (parts.length >= 3) return { type: 'bucket', book: parts[0], section: parts[1], grp: parts[2] };
  return null;
}

async function setOpeningBalance(){
  if(!_activeUserId){
    const msg = document.getElementById('ob-msg');
    if(msg) msg.innerHTML='<span style="color:var(--red)">Log in before saving opening balance.</span>';
    return;
  }
  const rawVal = document.getElementById('ob_amount').value;
  const amt = parseFloat(rawVal);
  const msg = document.getElementById('ob-msg');
  if (!rawVal || isNaN(amt)) {
    msg.innerHTML='<span style="color:var(--red)">❌ Enter a valid number (can be 0 if no prior balance).</span>';
    return;
  }
  msg.textContent = 'Saving…';
  try {
    const r = await fetch('/set-opening-balance', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({amount: amt})
    });
    const d = await r.json();
    if (d.error) { msg.innerHTML=`<span style="color:var(--red)">❌ ${d.error}</span>`; return; }
    msg.innerHTML=`<span style="color:var(--green)">✅ Opening balance saved: ₹${amt.toLocaleString('en-IN',{minimumFractionDigits:2})}. Regenerate Statements to reflect it.</span>`;
  } catch(e) {
    msg.innerHTML=`<span style="color:var(--red)">❌ ${e}</span>`;
  }
}



function tcfSetFile(f){
  _tcfFile = f;
  const prev = document.getElementById('tcf-file-preview');
  if(f){
    prev.style.display='block';
    prev.innerHTML = `📄 <strong>${f.name}</strong> &nbsp;(${(f.size/1024).toFixed(1)} KB)`;
  } else {
    prev.style.display='none';
  }
  _applyProfileGate();
}

function tcfHandleDrop(e){
  e.preventDefault();
  document.getElementById('tcf-drop-zone').classList.remove('dragover');
  const f = e.dataTransfer.files[0];
  if(f) tcfSetFile(f);
}

async function tcfUpload(){
  if(!_activeUserId){
    const msg = document.getElementById('tcf-upload-msg');
    if(msg) msg.innerHTML = '<span style="color:var(--red)">Log in before uploading a trading ledger.</span>';
    tab('setup');
    return;
  }
  if(_profileGateState.profile_required){
    const msg = document.getElementById('tcf-upload-msg');
    if(msg) msg.innerHTML = `<span style="color:var(--red)">${_profileGateText(_profileGateState.missing_reason || '')}</span>`;
    _focusProfileSetup(_profileGateText(_profileGateState.missing_reason));
    return;
  }
  if(!_tcfFile){ alert('Select a file first.'); return; }
  const acctId = document.getElementById('tcf-account-id').value.trim() || '5paisa';
  const msg = document.getElementById('tcf-upload-msg');
  msg.textContent = '⏳ Uploading & parsing…';
  const fd = new FormData();
  fd.append('file', _tcfFile);
  fd.append('account_id', acctId);
  try {
    const r = await fetch('/upload-trading-ledger', {method:'POST', body: fd});
    const d = await r.json();
    if(d.profile_required){
      msg.innerHTML=`<span style="color:var(--red)">❌ ${_profileGateText(d.missing_reason || '')}</span>`;
      await refreshProfileGate({focusProfile:true, message:_profileGateText(d.missing_reason || '')});
      return;
    }
    if(d.error){ msg.innerHTML=`<span style="color:var(--red)">❌ ${d.error}</span>`; return; }
    msg.innerHTML = `<span style="color:var(--green)">✅ ${d.rows_inserted} rows stored (${d.rows_skipped||0} skipped). Account: ${d.account_id}</span>`;
    document.getElementById('tcf-gen-account').value = acctId;
  } catch(e){ msg.innerHTML=`<span style="color:var(--red)">❌ ${e}</span>`; }
}

function _tcfChip(label, value, color){
  const v = typeof value === 'number'
    ? '₹' + Math.abs(value).toLocaleString('en-IN',{minimumFractionDigits:2})
    : value;
  return `<div class="trace-chip" style="background:${color};border-color:#aaa;min-width:120px">
    <div style="font-size:10px;color:#666;font-weight:600;text-transform:uppercase;letter-spacing:.4px">${label}</div>
    <div style="font-size:15px;font-weight:800;margin-top:2px">${v}</div>
  </div>`;
}

async function loadTradingCashflow(){
  const date_from = (document.getElementById('tcf-date-from')||{}).value || null;
  const date_to   = (document.getElementById('tcf-date-to'  )||{}).value || null;
  const account_id= (document.getElementById('tcf-gen-account')||{}).value.trim() || null;

  document.getElementById('tcf-empty').style.display='none';
  document.getElementById('tcf-report-card').style.display='none';
  document.getElementById('tcf-chips').style.display='none';

  try {
    const r = await fetch('/trading-cashflow', {
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body: JSON.stringify({date_from, date_to, account_id})
    });
    const d = await r.json();
    if(d.error){
      document.getElementById('tcf-empty').style.display='block';
      document.getElementById('tcf-empty').querySelector('div:last-child').textContent = '❌ ' + d.error;
      return;
    }

    // Render summary chips
    const s = d.summary || {};
    const fmt = v => typeof v==='number' ? '₹'+Math.abs(v).toLocaleString('en-IN',{minimumFractionDigits:2}) : (v||'—');
    const chips = document.getElementById('tcf-chips');
    chips.innerHTML = [
      _tcfChip('Opening Balance', s.opening, '#e3f2fd'),
      _tcfChip('Net Operating',   s.net_operating,
               s.net_operating >= 0 ? '#e8f5e9' : '#ffebee'),
      _tcfChip('Net Financing',   s.net_financing,
               s.net_financing >= 0 ? '#e8f5e9' : '#ffebee'),
      _tcfChip('Closing Balance', s.closing_calculated, '#fff3e0'),
      _tcfChip('Reconciliation',  s.reconciliation_ok ? '✅ Matches' : '⚠️ Difference: ₹'+Math.abs(s.reconciliation_diff||0).toLocaleString('en-IN',{minimumFractionDigits:2}),
               s.reconciliation_ok ? '#e8f5e9' : '#fff3e0'),
    ].join('');
    chips.style.display='flex';

    // Render report text
    document.getElementById('tcf-report-out').textContent = d.report || '';
    document.getElementById('tcf-report-card').style.display='block';
  } catch(e){
    document.getElementById('tcf-empty').style.display='block';
    document.getElementById('tcf-empty').querySelector('div:last-child').textContent = '❌ ' + e;
  }
}

</script>
</body>
</html>"""


def _parse_multipart(body: bytes, ct: str) -> Dict[str, Any]:
    m = re.search(r"boundary=([^\s;]+)", ct)
    if not m:
        return {}
    boundary = ("--" + m.group(1)).encode()
    parts = body.split(boundary)
    result: Dict[str, Any] = {}
    for part in parts:
        if b"\r\n\r\n" not in part:
            continue
        header_raw, _, content = part.partition(b"\r\n\r\n")
        content = content.rstrip(b"\r\n--")
        hs = header_raw.decode("utf-8", errors="replace")
        nm = re.search(r'name="([^"]+)"', hs)
        if not nm:
            continue
        name = nm.group(1)
        fm = re.search(r'filename="([^"]*)"', hs)
        if fm:
            result[name] = {"filename": fm.group(1), "content": content}
        else:
            result[name] = content.decode("utf-8", errors="replace").strip()
    return result

_EMAIL_RE = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")
_PHONE_RE = re.compile(r"^\d{8,15}$")
_PASSWORD_SPECIAL_RE = re.compile(r"[!@#$%^&*()_\-+=\[{\]};:'\",.<>/?\\|`~]")

def _norm_email(v: str) -> str:
    return (v or "").strip().lower()

def _norm_phone(v: str) -> str:
    return re.sub(r"\D+", "", (v or "").strip())

def _hash_password(password: str) -> str:
    password = str(password or "")
    salt = os.urandom(16)
    iterations = 200000
    dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations)
    return f"pbkdf2_sha256${iterations}${salt.hex()}${dk.hex()}"

def _verify_password(password: str, stored: str) -> bool:
    try:
        algo, iter_s, salt_hex, hash_hex = str(stored or "").split("$", 3)
        if algo != "pbkdf2_sha256":
            return False
        iterations = int(iter_s)
        salt = bytes.fromhex(salt_hex)
        expected = bytes.fromhex(hash_hex)
        actual = hashlib.pbkdf2_hmac("sha256", str(password or "").encode("utf-8"), salt, iterations)
        return hmac.compare_digest(actual, expected)
    except Exception:
        return False

def _validate_password(password: str) -> Optional[str]:
    password = str(password or "")
    if len(password) < 8:
        return "Password must be at least 8 characters."
    if password.strip() != password:
        return "Password should not start or end with spaces."
    if not _PASSWORD_SPECIAL_RE.search(password):
        return "Password must include at least one special character."
    return None

def _validate_auth_fields(name: str, email: str, phone: str, create: bool = False) -> Optional[str]:
    name = (name or "").strip()
    email = _norm_email(email)
    phone = _norm_phone(phone)

    if create and len(name) < 2:
        return "Name must be at least 2 characters."
    if email and not _EMAIL_RE.match(email):
        return "Invalid email format."
    if phone and not _PHONE_RE.match(phone):
        return "Phone must contain 10 digits."
    if create and not email and not phone:
        return "Provide at least phone or email."
    if not create and not name and not email and not phone:
        return "Enter at least one field."
    return None

class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args): pass

    def _text(self, code, body):
        b = body.encode("utf-8"); self.send_response(code)
        self.send_header("Content-Type", "text/plain; charset=utf-8")
        self.send_header("Content-Length", len(b)); self.end_headers(); self.wfile.write(b)

    def _json(self, code, obj):
        b = json.dumps(obj, ensure_ascii=False).encode("utf-8"); self.send_response(code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", len(b))
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.send_header("Pragma", "no-cache")
        self.end_headers(); self.wfile.write(b)

    def _html(self, body):
        import time
        b = body.encode("utf-8"); self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", len(b))
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        self.send_header("ETag", str(time.time()))
        self.end_headers(); self.wfile.write(b)

    def _body_bytes(self): return self.rfile.read(int(self.headers.get("Content-Length", 0)))
    def _body(self):
        raw = self._body_bytes().decode("utf-8", errors="replace")
        try: return json.loads(raw) if raw else {}
        except: return {}

    def do_GET(self):
        global _current_user_id
        if handle_get_extension(self, self.path, _current_user_id, _wflow, _edb):
            return
        if self.path == "/":
            self._html(HTML)
        elif self.path == "/check":
            ml = get_ml(); cl = get_cl()
            users = _db.get_users()
            lines = [
                "SYSTEM STATUS",
                f"  DB path         : {DB_PATH}",
                f"  Active user     : {_current_user_id or 'None'}",
                f"  ML model        : {'✅ Ready' if ml.ready else '❌ Not ready'} ({len(ml.labels)} categories)",
                f"  KMeans          : {'✅ Ready' if cl.ready else '❌ Not ready'} ({len(cl.cluster_labels)} clusters)",
                f"  Users in DB     : {len(users)}",
                "",
                "ACCOUNTING MODEL:",
                "  Bank debit = expense ALREADY PAID → Expenditure (not Liability)",
                "  Liabilities = only true outstanding amounts (loan principal, CC balance)",
                "  Balance Sheet = Assets / Liabilities / Equity (net worth snapshot)",
                "  Income & Expenditure = P&L for the period",
            ]
            self._text(200, "\n".join(lines))

        elif self.path.startswith("/download-upload/"):
            # Serve any file the user uploaded (for download)
            import urllib.parse
            fname = urllib.parse.unquote(self.path[len("/download-upload/"):])
            # Sanitize: strip path separators
            fname = os.path.basename(fname)
            fpath = os.path.join("/mnt/user-data/uploads", fname)
            if not os.path.exists(fpath):
                self._text(404, "File not found"); return
            import mimetypes
            mime, _ = mimetypes.guess_type(fpath)
            mime = mime or "application/octet-stream"
            with open(fpath, "rb") as f:
                data = f.read()
            self.send_response(200)
            self.send_header("Content-Type", mime)
            self.send_header("Content-Disposition", f'attachment; filename="{fname}"')
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return 
        
        elif self.path == "/report":
            if not _current_user_id: self._text(400, "No active user."); return
            self._text(200, generate_ledger_report(_db, _current_user_id))
        elif self.path == "/stats":
            if not _current_user_id: self._json(200, {}); return
            rows = _db.get_ledger_summary(_current_user_id)
            def _normal_row_balance(r):
                total = float(r.get("total") or 0)
                section = r.get("section")
                txn_type = r.get("txn_type")
                if section in ("Assets", "Expenditure"):
                    return total if txn_type == "debit" else -total
                if section in ("Liabilities", "Equity", "Income"):
                    return total if txn_type == "credit" else -total
                return total

            income = sum(
                _normal_row_balance(r)
                for r in rows
                if r["book"] == "INCOME_EXPENSE" and r["section"] == "Income"
            )
            expenditure = sum(
                _normal_row_balance(r)
                for r in rows
                if r["book"] == "INCOME_EXPENSE" and r["section"] == "Expenditure"
            )
            assets = sum(
                _normal_row_balance(r)
                for r in rows
                if r["book"] == "BALANCE_SHEET" and r["section"] == "Assets"
            )
            txns = len(_db.get_txns(_current_user_id, limit=9999))
            net_surplus = income - expenditure
            self._json(200, {"income": income, "expenditure": expenditure,
                             "assets": assets, "net_surplus": net_surplus, "txns": txns})
        elif self.path == "/ledger-keys":
            # Return all valid ledger keys with book/section/account for the reclassify dropdown
            keys = [
                {"key": k, "book": v[0], "section": v[1], "grp": v[2], "account": v[3]}
                for k, v in LEDGER_MAP.items()
            ]
            self._json(200, keys)
        # ── PDF SUPPORT (NEW) ── Library availability endpoint
        elif self.path == "/pdf-support":
            self._json(200, _pdf_library_status())
        else:
            self._text(404, "Not Found")

    def do_POST(self):
        global _current_user_id

        # ── Auth routes: handled here so _current_user_id is set in this scope ──
        if self.path == "/create-user":
            p = self._body()
            name = (p.get("name") or "").strip()
            email = _norm_email(p.get("email", ""))
            phone = _norm_phone(p.get("phone", ""))
            password = str(p.get("password") or "")
            user_type = (p.get("user_type") or "INDIVIDUAL").upper()

            err = _validate_auth_fields(name, email, phone, create=True)
            if err:
                self._json(400, {"error": err})
                return

            pw_err = _validate_password(password)
            if pw_err:
                self._json(400, {"error": pw_err})
                return

            if user_type not in ("INDIVIDUAL", "ORGANISATION"):
                user_type = "INDIVIDUAL"

            if phone:
                u = _edb.find_user(phone=phone)
                if u:
                    self._json(409, {"error": "Phone already exists for another user."})
                    return
            if email:
                u = _edb.find_user(email=email)
                if u:
                    self._json(409, {"error": "Email already exists for another user."})
                    return

            uid = str(uuid.uuid4())
            pw_hash = _hash_password(password)
            _edb.create_user(uid, name, email, phone, user_type, pw_hash)
            _edb.ensure_user_profile_exists(uid)

            if p.get("opening_balance"):
                _edb.set_opening_balance(uid, float(p.get("opening_balance", 0)))

            _current_user_id = uid
            self._json(200, {
                "status":"created",
                "id":uid,
                "name":name,
                "email":email,
                "phone":phone,
                "user_type":user_type,
                "txn_count":0,
                "profile_required": True,
                "next_step":"complete_classification_profile",
                "msg":"Account created. Complete your basic classification profile to continue."
            })
            return

        elif self.path == "/login":
            p = self._body()
            name = (p.get("name") or "").strip()
            email = _norm_email(p.get("email", ""))
            phone = _norm_phone(p.get("phone", ""))
            password = str(p.get("password") or "")

            err = _validate_auth_fields(name, email, phone, create=False)
            if err:
                self._json(400, {"error": err})
                return

            if not password:
                self._json(400, {"error": "Password is required."})
                return

            with _edb._conn() as c:
                sql = "SELECT * FROM users WHERE 1=1"
                params = []
                if phone:
                    sql += " AND phone=?"
                    params.append(phone)
                if email:
                    sql += " AND LOWER(email)=LOWER(?)"
                    params.append(email)
                if name:
                    sql += " AND LOWER(name)=LOWER(?)"
                    params.append(name)
                sql += " ORDER BY created_at DESC"
                matches = [dict(r) for r in c.execute(sql, tuple(params)).fetchall()]

            if not matches:
                self._json(404, {"error":"User not found. Check name/phone/email or create account."})
                return
            if len(matches) > 1:
                self._json(409, {"error":"Ambiguous login. Enter phone or email along with name for an exact match."})
                return

            user = matches[0]
            stored_hash = user.get("password_hash") or ""
            if not stored_hash or not _verify_password(password, stored_hash):
                self._json(401, {"error":"Incorrect password."})
                return

            _current_user_id = user["id"]
            cnt = len(_edb.get_txns(user["id"], limit=9999))
            user_type = user.get("user_type") or "INDIVIDUAL"
            self._json(200, {
                "status":"logged_in",
                "id":user["id"],
                "name":user["name"],
                "email":user.get("email",""),
                "phone":user.get("phone",""),
                "user_type":user_type,
                "txn_count":cnt,
                "msg":f"Welcome back, {user['name']}! {cnt} transactions on record."
            })
            return
 
        elif self.path == "/forgot-password":
            p = self._body()
            email = _norm_email(p.get("email", ""))
            phone = _norm_phone(p.get("phone", ""))

            if not email and not phone:
                self._json(400, {"error": "Provide email or phone."})
                return

            user = _edb.find_user_for_reset(email=email, phone=phone)
            if not user:
                self._json(404, {"error": "No user found for the provided email or phone."})
                return

            token = secrets.token_urlsafe(24)
            expires_at = (datetime.utcnow() + timedelta(minutes=15)).isoformat()
            _edb.set_reset_token(user["id"], token, expires_at)

            # For now return token in API response since no email/SMS sender is wired yet
            self._json(200, {
                "status": "ok",
                "msg": "Reset token generated.",
                "reset_token": token, 
                "expires_at": expires_at
            })
            return

        elif self.path == "/reset-password":
            p = self._body()
            token = str(p.get("token") or "").strip()
            new_password = str(p.get("new_password") or "")

            if not token:
                self._json(400, {"error": "Reset token is required."})
                return

            pw_err = _validate_password(new_password)
            if pw_err:
                self._json(400, {"error": pw_err})
                return

            ok = _edb.reset_password_with_token(token, _hash_password(new_password))
            if not ok:
                self._json(400, {"error": "Invalid or expired reset token."})
                return

            self._json(200, {"status":"ok", "msg":"Password reset successfully."})
            return
        
        elif self.path == "/logout":
            _current_user_id = None; self._json(200, {"status":"ok"}); return

        # ── Extension routes (never touch _current_user_id) ───────────────────
        if handle_post_extension(self, self.path, _wflow, _edb, _current_user_id):
            return

        elif self.path == "/upload":
            if not _current_user_id:
                self._json(400, {"error": "No active user. Create one in Setup."})
                return
            is_complete, missing_reason = is_min_profile_complete(_current_user_id)
            if not is_complete:
                self._json(400, {
                    "ok": False,
                    "error": "Complete basic classification profile before importing statements",
                    "profile_required": True,
                    "missing_reason": missing_reason,
                })
                return

            ct = self.headers.get("Content-Type", "")
            body = self._body_bytes()
            parts = _parse_multipart(body, ct)
            file_part = parts.get("file")

            if not file_part:
                self._json(400, {"error": "No file received"})
                return

            if isinstance(file_part, dict):
                file_bytes = file_part.get("content", b"")
                filename = file_part.get("filename", "")
            else:
                file_bytes = file_part
                filename = ""

            password = str(parts.get("password") or "").strip() or None

            try:
                parsed = parse_excel_statement(file_bytes, filename, password=password)
                records = parsed.get("records", [])
                opening_balance = parsed.get("opening_balance")
                closing_balance = parsed.get("closing_balance")
                statement_from_date = parsed.get("statement_from_date")
                statement_to_date = parsed.get("statement_to_date")
                print(f"[PERIOD PARSED] statement_from_date={statement_from_date} statement_to_date={statement_to_date}", flush=True)
            except Exception as e:
                self._json(400, {"error": f"Parse error: {e}"})
                return

            # Do not auto-write users.opening_balance from uploaded files.
            # File-level opening balances are already stored in raw_import_batches
            # and should be picked at report time based on selected account/period.
            # Keep users.opening_balance only as a manual fallback set from Setup.

            _wflow.discard_pending(_current_user_id)
            account_id = str(parts.get("account_id") or "main").strip() or "main"

            print(f"[UPLOAD] parse returned {len(records)} records for user={_current_user_id} account={account_id}")
            if records:
                print(f"[UPLOAD] First record sample: {records[0]}")

            try:
                print(f"[PERIOD STAGED] statement_from_date={statement_from_date} statement_to_date={statement_to_date}", flush=True)
                res = _wflow.process_import_batch(
                    user_id=_current_user_id,
                    account_id=account_id,
                    statement_type="bank_statement",
                    file_name=filename,
                    file_bytes=file_bytes,
                    raw_records=records,
                    statement_from_date=statement_from_date,
                    statement_to_date=statement_to_date,
                    opening_balance=opening_balance,
                    closing_balance=closing_balance,
                )
                print(f"[UPLOAD] process_import_batch result: {res}")
                pending = _wflow.get_pending_transactions(_current_user_id) or []
                print(f"[UPLOAD] get_pending_transactions returned {len(pending)} rows")
                res["transactions"] = pending
                res["staged"] = res.get("staged", 0)
                if opening_balance is not None:
                    res["opening_balance"] = float(opening_balance)
                self._json(200, res)
            except Exception as e:
                import traceback
                tb = traceback.format_exc()
                print(f"[UPLOAD ERROR] {e}\n{tb}")
                self._json(
                    500,
                    {
                        "error": f"Classification/staging error: {e}",
                        "detail": tb,
                        "staged": 0,
                        "transactions": [],
                    },
                )
            return

        elif self.path == "/classify":
            if not _current_user_id:
                self._json(400, {"error":"No active user."})
                return
            is_complete, missing_reason = is_min_profile_complete(_current_user_id)
            if not is_complete:
                self._json(400, {
                    "ok": False,
                    "error": "Complete basic classification profile before importing statements",
                    "profile_required": True,
                    "missing_reason": missing_reason,
                })
                return

            p = self._body()
            lines_raw = [l.strip() for l in (p.get("text","") or "").splitlines() if l.strip()]
            amount = float(p.get("amount",0) or 0)
            ttype = p.get("txn_type") or None

            _wflow.discard_pending(_current_user_id)
            results = []

            for ln in lines_raw:
                pending_row = _wflow.classify_and_stage(
                    _current_user_id,
                    narration=ln,
                    amount=amount,
                    txn_type=ttype,
                    txn_date=datetime.now().strftime("%Y-%m-%d"),
                )
                results.append(pending_row or {})

            self._json(200, results)

        elif self.path == "/ledger":
            if not _current_user_id: self._json(200, []); return
            p = self._body()
            rows = _db.get_txns(_current_user_id, limit=p.get("limit",500))
            filtered = []
            for r in rows:
                if p.get("book") and r["book"] != p["book"]: continue
                if p.get("txn_type") and r["txn_type"] != p["txn_type"]: continue
                if p.get("search") and p["search"].upper() not in (r["narration"] or "").upper(): continue
                if p.get("counterparty") and p["counterparty"].upper() not in (r.get("counterparty", "") or "").upper(): continue
                filtered.append(dict(r))
            self._json(200, filtered)

        else:
            self._text(404, "Not Found")


# ── HNI Extension: user_type, review workflow, typed statements ───────────────
from hni_ledger_extension import (
    extend_db, ExtendedDBStore, ReviewWorkflow,
    handle_get_extension, handle_post_extension,
    is_min_profile_complete,
)
extend_db(_db)
_edb   = ExtendedDBStore(_db)
_wflow = ReviewWorkflow(_edb)
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  HNI LEDGER CLASSIFIER  —  3-Book Accounting Edition")
    print("=" * 60)
    print(f"  DB    : {DB_PATH}")
    print()
    print("  ACCOUNTING MODEL:")
    print("  ✅ Bank debits = expenses ALREADY PAID → Expenditure")
    print("  ✅ Liabilities = only outstanding amounts (loans, CC)")
    print("  ✅ Balance Sheet = Assets / Liabilities / Equity")
    print("  ✅ Book I: Income & Expenditure (P&L)")
    print("  ✅ Book II: Balance Sheet (net worth)")
    print("  ✅ Book III: Balance check")
    print()

    print("  Loading ML models …")
    get_ml()
    get_cl()

    users = _db.get_users()
    global _current_user_id
    _current_user_id = None
    print("  No active user. Please log in from Setup.")

    port = 8082
    server = ThreadingHTTPServer(("0.0.0.0", port), Handler)
    print(f"\n  ✅ Server → http://localhost:{port}")
    print("  Press Ctrl+C to stop\n")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        server.shutdown()
        print("\n  Server stopped.")


if __name__ == "__main__":
    main()
